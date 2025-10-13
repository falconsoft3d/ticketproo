from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User, Group
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum, F
from django.db import models
from django.http import HttpResponse, Http404, JsonResponse
from django.utils import timezone
from django.conf import settings
from django.views.decorators.http import require_http_methods
from datetime import datetime, date, timedelta
import os
import json

# Imports para OpenAI
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    print("OpenAI library not available")

from .models import (
    Ticket, TicketAttachment, Category, TicketComment, UserProfile, 
    UserNote, TimeEntry, Project, Company, SystemConfiguration, Document, UrlManager, WorkOrder, Task,
    DailyTaskSession, DailyTaskItem, ChatRoom, ChatMessage, ContactFormSubmission,
    Opportunity, OpportunityStatus, OpportunityNote, OpportunityStatusHistory,
    Meeting, MeetingAttendee, MeetingQuestion, Contact,
    BlogCategory, BlogPost, BlogComment, AIChatSession, AIChatMessage, Concept,
    Exam, ExamQuestion, ExamAttempt, ExamAnswer, ContactoWeb, PublicDocumentUpload,
    Employee, JobApplicationToken, EmployeePayroll, Agreement, AgreementSignature,
    LandingPage, LandingPageSubmission, WorkOrderTask, WorkOrderTaskTimeEntry, SharedFile, SharedFileDownload,
    Recording, RecordingPlayback, MultipleDocumentation, MultipleDocumentationItem,
    TaskSchedule, ScheduleTask, ScheduleComment, SatisfactionSurvey
)
from .forms import (
    TicketForm, AgentTicketForm, UserManagementForm, UserEditForm, 
    TicketAttachmentForm, CategoryForm, UserTicketForm, UserTicketEditForm, 
    TicketCommentForm, UserNoteForm, TimeEntryStartForm, TimeEntryEndForm, 
    TimeEntryEditForm, ProjectForm, CompanyForm, SystemConfigurationForm, DocumentForm,
    UrlManagerForm, UrlManagerFilterForm, WorkOrderForm, WorkOrderFilterForm, TaskForm,
    ChatMessageForm, ChatRoomForm, AIChatSessionForm, AIChatMessageForm, ConceptForm,
    EmployeeHiringOpinionForm, EmployeePayrollForm, AgreementForm, AgreementSignatureForm, AgreementPublicForm,
    LandingPageForm, LandingPageSubmissionForm, PublicCompanyTicketForm, WorkOrderTaskForm, 
    WorkOrderTaskTimeEntryForm, WorkOrderTaskBulkForm, SharedFileForm, PublicSharedFileForm,
    TaskScheduleForm, ScheduleTaskForm, ScheduleCommentForm
)
from .utils import is_agent, is_regular_user, is_teacher, can_manage_courses, get_user_role, assign_user_to_group

def home_view(request):
    """Vista para la página de inicio de TicketProo - accesible para todos"""
    # Estadísticas generales del sistema
    total_tickets = Ticket.objects.count()
    total_users = Ticket.objects.values('created_by').distinct().count()
    
    # Estadísticas por estado si el usuario está autenticado
    user_tickets_count = 0
    user_role = None
    tickets_by_status = {}
    
    if request.user.is_authenticated:
        user_role = get_user_role(request.user)
        
        if is_agent(request.user):
            # Para agentes: estadísticas de todos los tickets
            tickets_by_status = {
                'open': Ticket.objects.filter(status='open').count(),
                'in_progress': Ticket.objects.filter(status='in_progress').count(),
                'resolved': Ticket.objects.filter(status='resolved').count(),
                'closed': Ticket.objects.filter(status='closed').count(),
            }
            user_tickets_count = Ticket.objects.count()
        else:
            # Para usuarios regulares: solo sus tickets
            user_tickets = Ticket.objects.filter(created_by=request.user)
            tickets_by_status = {
                'open': user_tickets.filter(status='open').count(),
                'in_progress': user_tickets.filter(status='in_progress').count(),
                'resolved': user_tickets.filter(status='resolved').count(),
                'closed': user_tickets.filter(status='closed').count(),
            }
            user_tickets_count = user_tickets.count()
    
    # Obtener conceptos activos para mostrar en el home
    concepts = Concept.objects.filter(is_active=True)[:10]  # Máximo 10 conceptos
    
    # Obtener los últimos 4 artículos del blog
    from .models import BlogPost
    latest_blog_posts = BlogPost.objects.filter(
        status='published'
    ).order_by('-created_at')[:4]
    
    context = {
        'total_tickets': total_tickets,
        'total_users': total_users,
        'user_tickets_count': user_tickets_count,
        'user_role': user_role,
        'tickets_by_status': tickets_by_status,
        'is_authenticated': request.user.is_authenticated,
        'concepts': concepts,
        'latest_blog_posts': latest_blog_posts,
    }
    return render(request, 'tickets/home.html', context)

def register_view(request):
    """Vista para registro de nuevos usuarios"""
    # Verificar si el registro está habilitado en la configuración
    config = SystemConfiguration.get_config()
    
    if not config.allow_user_registration:
        messages.error(request, 'El registro de nuevos usuarios está deshabilitado. Contacta a un administrador.')
        return redirect('login')
    
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            username = form.cleaned_data.get('username')
            
            # Asignar automáticamente al grupo de Usuarios
            assign_user_to_group(user, 'Usuarios')
            
            messages.success(request, f'Cuenta creada para {username}. Ya puedes iniciar sesión.')
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

@login_required
def dashboard_view(request):
    """Vista principal del dashboard de TicketProo"""
    user_role = get_user_role(request.user)
    
    if is_agent(request.user):
        # Estadísticas para agentes (todos los tickets)
        all_tickets = Ticket.objects.all()
        total_tickets = all_tickets.count()
        open_tickets = all_tickets.filter(status='open').count()
        in_progress_tickets = all_tickets.filter(status='in_progress').count()
        resolved_tickets = all_tickets.filter(status='resolved').count()
        recent_tickets = all_tickets[:5]
        
        # Estadísticas adicionales para agentes
        unassigned_tickets = all_tickets.filter(assigned_to__isnull=True).count()
        my_assigned_tickets = all_tickets.filter(assigned_to=request.user).count()
        
        # Obtener horas diarias
        daily_hours = 0
        try:
            daily_hours = request.user.userprofile.get_daily_hours()
        except AttributeError:
            # Crear UserProfile si no existe
            from tickets.models import UserProfile
            profile, created = UserProfile.objects.get_or_create(user=request.user)
            daily_hours = profile.get_daily_hours()
        
        # Órdenes de trabajo pendientes de aprobación (solo primeras 5)
        pending_work_orders = WorkOrder.objects.filter(status='draft').order_by('-created_at')[:5]
        
        context = {
            'total_tickets': total_tickets,
            'open_tickets': open_tickets,
            'in_progress_tickets': in_progress_tickets,
            'resolved_tickets': resolved_tickets,
            'recent_tickets': recent_tickets,
            'unassigned_tickets': unassigned_tickets,
            'my_assigned_tickets': my_assigned_tickets,
            'pending_work_orders': pending_work_orders,
            'user_role': user_role,
            'is_agent': True,
            'daily_hours': daily_hours,
        }
        
        # Agregar conceptos activos
        concepts = Concept.objects.filter(is_active=True)[:10]
        context['concepts'] = concepts
    else:
        # Estadísticas para usuarios regulares (sus tickets + tickets de empresa + proyectos)
        user_projects = request.user.assigned_projects.all()
        user_company = None
        
        # Obtener empresa del usuario si tiene perfil
        try:
            user_company = request.user.profile.company
        except:
            pass
        
        # Construir query base
        query_conditions = Q(created_by=request.user)  # Sus propios tickets
        
        # Agregar tickets de proyectos asignados
        if user_projects.exists():
            query_conditions |= Q(project__in=user_projects)
        
        # Agregar tickets de la empresa del usuario
        if user_company:
            query_conditions |= Q(company=user_company)
        
        user_tickets = Ticket.objects.filter(query_conditions).distinct()
        total_tickets = user_tickets.count()
        open_tickets = user_tickets.filter(status='open').count()
        in_progress_tickets = user_tickets.filter(status='in_progress').count()
        resolved_tickets = user_tickets.filter(status='resolved').count()
        recent_tickets = user_tickets.order_by('-created_at')[:5]
        
        # Estadísticas adicionales para mostrar el contexto
        own_tickets = Ticket.objects.filter(created_by=request.user).count()
        company_tickets = 0
        if user_company:
            company_tickets = Ticket.objects.filter(company=user_company).exclude(created_by=request.user).count()
        
        # Obtener horas diarias
        daily_hours = 0
        try:
            daily_hours = request.user.userprofile.get_daily_hours()
        except AttributeError:
            # Crear UserProfile si no existe
            from tickets.models import UserProfile
            profile, created = UserProfile.objects.get_or_create(user=request.user)
            daily_hours = profile.get_daily_hours()
        
        context = {
            'total_tickets': total_tickets,
            'open_tickets': open_tickets,
            'in_progress_tickets': in_progress_tickets,
            'resolved_tickets': resolved_tickets,
            'recent_tickets': recent_tickets,
            'own_tickets': own_tickets,
            'company_tickets': company_tickets,
            'user_company': user_company,
            'user_role': user_role,
            'is_agent': False,
            'daily_hours': daily_hours,
        }
    
    return render(request, 'tickets/dashboard.html', context)

@login_required
def ticket_list_view(request):
    """Vista para listar tickets según el rol del usuario"""
    if is_agent(request.user):
        # Los agentes ven todos los tickets
        tickets = Ticket.objects.all()
    else:
        # Los usuarios regulares ven sus propios tickets Y tickets de proyectos asignados Y tickets de su empresa
        user_projects = request.user.assigned_projects.all()
        user_company = None
        
        # Obtener empresa del usuario si tiene perfil
        try:
            user_company = request.user.profile.company
        except:
            pass
        
        # Construir query base
        query_conditions = Q(created_by=request.user)  # Sus propios tickets
        
        # Agregar tickets de proyectos asignados
        if user_projects.exists():
            query_conditions |= Q(project__in=user_projects)
        
        # Agregar tickets de la empresa del usuario
        if user_company:
            query_conditions |= Q(company=user_company)
        
        tickets = Ticket.objects.filter(query_conditions).distinct()
    
    # Filtros
    status_filter = request.GET.get('status')
    priority_filter = request.GET.get('priority')
    search = request.GET.get('search')
    assigned_filter = request.GET.get('assigned_to')  # Nuevo filtro para agentes
    
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    
    if priority_filter:
        tickets = tickets.filter(priority=priority_filter)
    
    if search:
        tickets = tickets.filter(
            Q(title__icontains=search) | 
            Q(description__icontains=search)
        )
    
    # Filtro adicional para agentes
    if is_agent(request.user) and assigned_filter:
        if assigned_filter == 'unassigned':
            tickets = tickets.filter(assigned_to__isnull=True)
        elif assigned_filter == 'me':
            tickets = tickets.filter(assigned_to=request.user)
    
    # Paginación
    paginator = Paginator(tickets, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'search': search,
        'assigned_filter': assigned_filter,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'is_agent': is_agent(request.user),
        'user_role': get_user_role(request.user),
    }
    return render(request, 'tickets/ticket_list.html', context)

@login_required
def ticket_export_excel(request):
    """Vista para exportar tickets a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from datetime import datetime
    
    # Obtener los mismos tickets que se muestran en la lista
    if is_agent(request.user):
        tickets = Ticket.objects.all()
    else:
        user_projects = request.user.assigned_projects.all()
        user_company = None
        
        try:
            user_company = request.user.profile.company
        except:
            pass
        
        query_conditions = Q(created_by=request.user)
        
        if user_projects.exists():
            query_conditions |= Q(project__in=user_projects)
        
        if user_company:
            query_conditions |= Q(company=user_company)
        
        tickets = Ticket.objects.filter(query_conditions).distinct()
    
    # Aplicar los mismos filtros que en la lista
    status_filter = request.GET.get('status')
    priority_filter = request.GET.get('priority')
    search = request.GET.get('search')
    assigned_filter = request.GET.get('assigned_to')
    
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    
    if priority_filter:
        tickets = tickets.filter(priority=priority_filter)
    
    if search:
        tickets = tickets.filter(
            Q(title__icontains=search) | 
            Q(description__icontains=search)
        )
    
    if is_agent(request.user) and assigned_filter:
        if assigned_filter == 'unassigned':
            tickets = tickets.filter(assigned_to__isnull=True)
        elif assigned_filter == 'me':
            tickets = tickets.filter(assigned_to=request.user)
    
    # Crear el libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    
    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Encabezados
    headers = [
        'Número', 'Título', 'Descripción', 'Categoría', 'Tipo', 'Empresa', 
        'Creador', 'Asignado a', 'Prioridad', 'Estado', 'Aprobación', 
        'Antigüedad (horas)', 'Fecha Creación', 'Fecha Actualización'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Datos de los tickets
    for row, ticket in enumerate(tickets.order_by('-created_at'), 2):
        ws.cell(row=row, column=1, value=ticket.ticket_number)
        ws.cell(row=row, column=2, value=ticket.title)
        ws.cell(row=row, column=3, value=ticket.description)
        ws.cell(row=row, column=4, value=ticket.category.name if ticket.category else "Sin categoría")
        ws.cell(row=row, column=5, value=ticket.get_ticket_type_display())
        ws.cell(row=row, column=6, value=ticket.company.name if ticket.company else "Sin empresa")
        ws.cell(row=row, column=7, value=ticket.created_by.get_full_name() if ticket.created_by.get_full_name() else ticket.created_by.username)
        ws.cell(row=row, column=8, value=ticket.assigned_to.get_full_name() if ticket.assigned_to and ticket.assigned_to.get_full_name() else (ticket.assigned_to.username if ticket.assigned_to else "Sin asignar"))
        ws.cell(row=row, column=9, value=ticket.get_priority_display())
        ws.cell(row=row, column=10, value=ticket.get_status_display())
        ws.cell(row=row, column=11, value="Aprobado" if ticket.is_approved else "Pendiente")
        ws.cell(row=row, column=12, value=round(ticket.get_age_in_hours(), 1))
        ws.cell(row=row, column=13, value=ticket.created_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=14, value=ticket.updated_at.strftime('%d/%m/%Y %H:%M'))
    
    # Ajustar ancho de columnas
    column_widths = [12, 30, 50, 15, 12, 20, 20, 20, 12, 12, 12, 15, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo con fecha
    filename = f'tickets_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar el libro en la respuesta
    wb.save(response)
    
    return response

@login_required
def ticket_create_view(request):
    """Vista para crear un nuevo ticket"""
    # Usar formulario apropiado según el rol
    FormClass = AgentTicketForm if is_agent(request.user) else UserTicketForm
    
    if request.method == 'POST':
        form = FormClass(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            # Los usuarios regulares siempre crean tickets con estado "open"
            if not is_agent(request.user):
                ticket.status = 'open'
            ticket.save()
            messages.success(request, 'Ticket creado exitosamente.')
            return redirect('ticket_detail', pk=ticket.pk)
    else:
        form = FormClass()
    
    context = {
        'form': form,
        'is_agent': is_agent(request.user),
        'user_role': get_user_role(request.user),
    }
    
    return render(request, 'tickets/ticket_create.html', context)

@login_required
def ticket_detail_view(request, pk):
    """Vista para ver los detalles de un ticket"""
    if is_agent(request.user):
        # Los agentes pueden ver cualquier ticket
        ticket = get_object_or_404(Ticket, pk=pk)
    else:
        # Los usuarios pueden ver sus propios tickets O tickets de su empresa O tickets de proyectos asignados
        user_projects = request.user.assigned_projects.all()
        user_company = None
        
        # Obtener empresa del usuario si tiene perfil
        try:
            user_company = request.user.profile.company
        except:
            pass
        
        # Construir query de acceso
        query_conditions = Q(created_by=request.user)  # Sus propios tickets
        
        # Agregar tickets de proyectos asignados
        if user_projects.exists():
            query_conditions |= Q(project__in=user_projects)
        
        # Agregar tickets de la empresa del usuario
        if user_company:
            query_conditions |= Q(company=user_company)
        
        ticket = get_object_or_404(Ticket, pk=pk)
        
        # Verificar que el usuario tenga acceso al ticket
        if not Ticket.objects.filter(pk=pk).filter(query_conditions).exists():
            messages.error(request, 'No tienes permisos para ver este ticket.')
            return redirect('ticket_list')
    
    # Manejar autoasignación para agentes
    if request.method == 'POST' and 'assign_to_me' in request.POST and is_agent(request.user):
        ticket.assigned_to = request.user
        ticket.save()
        messages.success(request, f'Te has asignado el ticket #{ticket.pk}.')
        return redirect('ticket_detail', pk=pk)
    
    # Manejar activar compartir público (solo agentes)
    if request.method == 'POST' and 'enable_public_share' in request.POST and is_agent(request.user):
        ticket.is_public_shareable = True
        ticket.save()
        messages.success(request, 'Compartir público activado. Ahora puedes compartir este ticket mediante un enlace.')
        return redirect('ticket_detail', pk=pk)
    
    # Manejar desactivar compartir público (solo agentes)
    if request.method == 'POST' and 'disable_public_share' in request.POST and is_agent(request.user):
        ticket.is_public_shareable = False
        ticket.save()
        messages.warning(request, 'Compartir público desactivado. El enlace ya no será accesible.')
        return redirect('ticket_detail', pk=pk)
    
    # Manejar regenerar token (solo agentes)
    if request.method == 'POST' and 'regenerate_token' in request.POST and is_agent(request.user):
        ticket.regenerate_public_token()
        messages.info(request, 'Enlace público regenerado. El enlace anterior ya no es válido.')
        return redirect('ticket_detail', pk=pk)
    
    # Formulario para agregar comentarios
    comment_form = TicketCommentForm()
    if request.method == 'POST' and 'add_comment' in request.POST:
        comment_form = TicketCommentForm(request.POST)
        if comment_form.is_valid():
            comment = comment_form.save(commit=False)
            comment.ticket = ticket
            comment.user = request.user
            comment.save()
            messages.success(request, 'Comentario agregado exitosamente.')
            return redirect('ticket_detail', pk=pk)
    
    # Formulario para subir adjuntos
    attachment_form = TicketAttachmentForm()
    if request.method == 'POST' and 'upload_attachment' in request.POST:
        attachment_form = TicketAttachmentForm(request.POST, request.FILES)
        if attachment_form.is_valid():
            attachment = attachment_form.save(commit=False)
            attachment.ticket = ticket
            attachment.uploaded_by = request.user
            attachment.original_filename = attachment.file.name
            attachment.file_size = attachment.file.size
            attachment.save()
            messages.success(request, 'Adjunto subido exitosamente.')
            return redirect('ticket_detail', pk=pk)
    
    # Obtener comentarios del ticket
    comments = ticket.comments.all().order_by('created_at')
    
    context = {
        'ticket': ticket,
        'comments': comments,
        'comment_form': comment_form,
        'attachment_form': attachment_form,
        'is_agent': is_agent(request.user),
        'user_role': get_user_role(request.user),
    }
    return render(request, 'tickets/ticket_detail.html', context)

@login_required
def ticket_edit_view(request, pk):
    """Vista para editar un ticket"""
    if is_agent(request.user):
        # Los agentes pueden editar cualquier ticket
        ticket = get_object_or_404(Ticket, pk=pk)
        FormClass = AgentTicketForm
    else:
        # Los usuarios pueden editar sus propios tickets O tickets de su empresa O tickets de proyectos asignados
        user_projects = request.user.assigned_projects.all()
        user_company = None
        
        # Obtener empresa del usuario si tiene perfil
        try:
            user_company = request.user.profile.company
        except:
            pass
        
        # Construir query de acceso
        query_conditions = Q(created_by=request.user)  # Sus propios tickets
        
        # Agregar tickets de proyectos asignados
        if user_projects.exists():
            query_conditions |= Q(project__in=user_projects)
        
        # Agregar tickets de la empresa del usuario
        if user_company:
            query_conditions |= Q(company=user_company)
        
        ticket = get_object_or_404(Ticket, pk=pk)
        
        # Verificar que el usuario tenga acceso al ticket
        if not Ticket.objects.filter(pk=pk).filter(query_conditions).exists():
            messages.error(request, 'No tienes permisos para editar este ticket.')
            return redirect('ticket_list')
        
        FormClass = UserTicketEditForm
    
    if request.method == 'POST':
        form = FormClass(request.POST, instance=ticket)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ticket actualizado exitosamente.')
            return redirect('ticket_detail', pk=ticket.pk)
    else:
        form = FormClass(instance=ticket)
    
    context = {
        'form': form, 
        'ticket': ticket,
        'is_agent': is_agent(request.user),
        'user_role': get_user_role(request.user),
    }
    return render(request, 'tickets/ticket_edit.html', context)

@login_required
def ticket_delete_view(request, pk):
    """Vista para eliminar un ticket"""
    if is_agent(request.user):
        # Los agentes pueden eliminar cualquier ticket
        ticket = get_object_or_404(Ticket, pk=pk)
    else:
        # Los usuarios solo pueden eliminar sus propios tickets
        ticket = get_object_or_404(Ticket, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        ticket.delete()
        messages.success(request, 'Ticket eliminado exitosamente.')
        return redirect('ticket_list')
    
    context = {
        'ticket': ticket,
        'is_agent': is_agent(request.user),
        'user_role': get_user_role(request.user),
    }
    return render(request, 'tickets/ticket_delete.html', context)


@login_required
def ticket_approve_view(request, pk):
    """Vista para aprobar/desaprobar un ticket"""
    # Verificar acceso al ticket
    if is_agent(request.user):
        # Los agentes pueden aprobar cualquier ticket
        ticket = get_object_or_404(Ticket, pk=pk)
    else:
        # Los usuarios pueden aprobar tickets donde están involucrados
        user_company = getattr(request.user.profile, 'company', None) if hasattr(request.user, 'profile') else None
        user_projects = Project.objects.filter(assigned_users=request.user)
        
        query_conditions = Q(created_by=request.user)
        
        if user_projects.exists():
            query_conditions |= Q(project__in=user_projects)
        
        if user_company:
            query_conditions |= Q(company=user_company)
        
        ticket = get_object_or_404(Ticket, pk=pk)
        
        # Verificar que el usuario tenga acceso al ticket
        if not Ticket.objects.filter(pk=pk).filter(query_conditions).exists():
            messages.error(request, 'No tienes permisos para aprobar este ticket.')
            return redirect('ticket_list')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            if not ticket.is_approved:
                ticket.approve_ticket(request.user)
                messages.success(request, f'Ticket #{ticket.ticket_number} aprobado exitosamente.')
            else:
                messages.info(request, 'Este ticket ya estaba aprobado.')
        
        elif action == 'unapprove':
            if ticket.is_approved:
                ticket.unapprove_ticket()
                messages.success(request, f'Aprobación del ticket #{ticket.ticket_number} removida.')
            else:
                messages.info(request, 'Este ticket no estaba aprobado.')
        
        return redirect('ticket_detail', pk=ticket.pk)
    
    # Si no es POST, redirigir al detalle del ticket
    return redirect('ticket_detail', pk=ticket.pk)

@login_required
def logout_view(request):
    """Vista personalizada para logout"""
    logout(request)
    messages.info(request, 'Has cerrado sesión exitosamente.')
    return redirect('home')

# ========== VISTAS DE GESTIÓN DE USUARIOS ==========

def agent_required(user):
    """Decorador para verificar si el usuario es agente"""
    return is_agent(user)

def course_manager_required(user):
    """Decorador para verificar si el usuario puede gestionar cursos (Agente o Profesor)"""
    return can_manage_courses(user)

@login_required
@user_passes_test(agent_required, login_url='dashboard')
def user_management_view(request):
    """Vista para listar todos los usuarios (solo para agentes)"""
    users = User.objects.select_related('profile').prefetch_related('groups').all().order_by('username')
    
    # Filtros
    role_filter = request.GET.get('role')
    status_filter = request.GET.get('status')
    search = request.GET.get('search')
    
    if role_filter:
        if role_filter == 'agents':
            users = users.filter(groups__name='Agentes')
        elif role_filter == 'users':
            users = users.filter(groups__name='Usuarios')
    
    if status_filter:
        if status_filter == 'active':
            users = users.filter(is_active=True)
        elif status_filter == 'inactive':
            users = users.filter(is_active=False)
    
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Agregar información de actividad actual para cada usuario
    for user in users:
        # Buscar si tiene una jornada activa (sin fecha de salida)
        current_entry = TimeEntry.objects.filter(
            user=user,
            fecha_salida__isnull=True
        ).select_related('project', 'ticket', 'work_order').first()
        
        user.current_entry = current_entry
        user.is_working = current_entry is not None
        
        # Calcular estadísticas de tiempo del usuario de manera simple
        from django.utils import timezone
        from datetime import timedelta
        
        # Tiempo total trabajado
        completed_entries = TimeEntry.objects.filter(
            user=user,
            fecha_salida__isnull=False
        )
        
        total_seconds = 0
        for entry in completed_entries:
            duration = entry.fecha_salida - entry.fecha_entrada
            total_seconds += duration.total_seconds()
        
        user.total_hours = round(total_seconds / 3600, 1)
        
        # Tiempo trabajado este mes
        now = timezone.now()
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        month_entries = TimeEntry.objects.filter(
            user=user,
            fecha_entrada__gte=start_of_month,
            fecha_salida__isnull=False
        )
        
        month_seconds = 0
        for entry in month_entries:
            duration = entry.fecha_salida - entry.fecha_entrada
            month_seconds += duration.total_seconds()
        
        user.month_hours = round(month_seconds / 3600, 1)
        
        # Tiempo trabajado esta semana
        start_of_week = now - timedelta(days=now.weekday())
        start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
        
        week_entries = TimeEntry.objects.filter(
            user=user,
            fecha_entrada__gte=start_of_week,
            fecha_salida__isnull=False
        )
        
        week_seconds = 0
        for entry in week_entries:
            duration = entry.fecha_salida - entry.fecha_entrada
            week_seconds += duration.total_seconds()
        
        user.week_hours = round(week_seconds / 3600, 1)
        
        # Tiempo trabajado hoy
        today = now.date()
        today_entries = TimeEntry.objects.filter(
            user=user,
            fecha_entrada__date=today,
            fecha_salida__isnull=False
        )
        
        today_seconds = 0
        for entry in today_entries:
            duration = entry.fecha_salida - entry.fecha_entrada
            today_seconds += duration.total_seconds()
        
        user.today_hours = round(today_seconds / 3600, 1)
    
    # Paginación
    paginator = Paginator(users, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'search': search,
    }
    return render(request, 'tickets/user_management.html', context)

@login_required
@user_passes_test(agent_required, login_url='dashboard')
def user_create_view(request):
    """Vista para crear nuevos usuarios (solo para agentes)"""
    if request.method == 'POST':
        form = UserManagementForm(request.POST)
        if form.is_valid():
            user = form.save()
            role = form.cleaned_data['role']
            messages.success(request, f'Usuario "{user.username}" creado exitosamente como {role[:-1]}.')
            return redirect('user_management')
    else:
        form = UserManagementForm()
    
    return render(request, 'tickets/user_create.html', {'form': form})

@login_required
@user_passes_test(agent_required, login_url='dashboard')
def user_edit_view(request, user_id):
    """Vista para editar usuarios existentes (solo para agentes)"""
    user_to_edit = get_object_or_404(User, pk=user_id)
    
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user_to_edit)
        if form.is_valid():
            user = form.save()
            
            # Actualizar rol del usuario
            new_role = form.cleaned_data['role']
            user.groups.clear()  # Remover todos los grupos
            try:
                group = Group.objects.get(name=new_role)
                user.groups.add(group)
            except Group.DoesNotExist:
                pass
            
            messages.success(request, f'Usuario "{user.username}" actualizado exitosamente.')
            return redirect('user_management')
    else:
        form = UserEditForm(instance=user_to_edit)
    
    context = {
        'form': form,
        'user_to_edit': user_to_edit,
    }
    return render(request, 'tickets/user_edit.html', context)

@login_required
@user_passes_test(agent_required, login_url='dashboard')
def user_toggle_status_view(request, user_id):
    """Vista para activar/desactivar usuarios (solo para agentes)"""
    user_to_toggle = get_object_or_404(User, pk=user_id)
    
    if user_to_toggle == request.user:
        messages.error(request, 'No puedes desactivar tu propia cuenta.')
    else:
        user_to_toggle.is_active = not user_to_toggle.is_active
        user_to_toggle.save()
        
        status = "activado" if user_to_toggle.is_active else "desactivado"
        messages.success(request, f'Usuario "{user_to_toggle.username}" {status} exitosamente.')
    
    return redirect('user_management')


@login_required
def download_attachment_view(request, attachment_id):
    """Vista para descargar adjuntos"""
    attachment = get_object_or_404(TicketAttachment, pk=attachment_id)
    
    # Verificar permisos
    if is_agent(request.user) or attachment.ticket.created_by == request.user:
        if attachment.file:
            try:
                with open(attachment.file.path, 'rb') as f:
                    response = HttpResponse(f.read(), content_type='application/octet-stream')
                    response['Content-Disposition'] = f'attachment; filename="{attachment.original_filename}"'
                    return response
            except FileNotFoundError:
                messages.error(request, 'El archivo no se encontró en el servidor.')
        else:
            messages.error(request, 'No hay archivo asociado.')
    else:
        messages.error(request, 'No tienes permisos para descargar este archivo.')
    
    return redirect('ticket_detail', pk=attachment.ticket.pk)


@login_required
def delete_attachment_view(request, attachment_id):
    """Vista para eliminar adjuntos"""
    attachment = get_object_or_404(TicketAttachment, pk=attachment_id)
    ticket_pk = attachment.ticket.pk
    
    # Verificar permisos (solo el que subió el archivo o agentes pueden eliminarlo)
    if is_agent(request.user) or attachment.uploaded_by == request.user:
        attachment.delete()
        messages.success(request, 'Adjunto eliminado exitosamente.')
    else:
        messages.error(request, 'No tienes permisos para eliminar este archivo.')
    
    return redirect('ticket_detail', pk=ticket_pk)


@login_required
@user_passes_test(agent_required, login_url='dashboard')
def unassign_ticket_view(request, pk):
    """Vista para que los agentes desasignen tickets"""
    ticket = get_object_or_404(Ticket, pk=pk)
    ticket.assigned_to = None
    ticket.save()
    messages.success(request, f'Ticket #{ticket.pk} desasignado exitosamente.')
    return redirect('ticket_detail', pk=pk)


# ================================
# VISTAS PARA GESTIÓN DE CATEGORÍAS
# ================================

@login_required
@user_passes_test(is_agent)
def category_list_view(request):
    """Vista para listar todas las categorías (solo agentes)"""
    categories = Category.objects.all().order_by('name')
    
    # Agregar contador de tickets por categoría
    for category in categories:
        category.ticket_count = category.tickets.count()
    
    context = {
        'categories': categories,
        'page_title': 'Gestión de Categorías'
    }
    return render(request, 'tickets/category_list.html', context)


@login_required
@user_passes_test(is_agent)
def category_create_view(request):
    """Vista para crear una nueva categoría (solo agentes)"""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Categoría "{category.name}" creada exitosamente.')
            return redirect('category_list')
    else:
        form = CategoryForm()
    
    context = {
        'form': form,
        'page_title': 'Crear Categoría',
        'form_title': 'Nueva Categoría'
    }
    return render(request, 'tickets/category_form.html', context)


@login_required
@user_passes_test(is_agent)
def category_edit_view(request, pk):
    """Vista para editar una categoría existente (solo agentes)"""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Categoría "{category.name}" actualizada exitosamente.')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    
    context = {
        'form': form,
        'category': category,
        'page_title': 'Editar Categoría',
        'form_title': f'Editar: {category.name}'
    }
    return render(request, 'tickets/category_form.html', context)


@login_required
@user_passes_test(is_agent)
def category_delete_view(request, pk):
    """Vista para eliminar una categoría (solo agentes)"""
    category = get_object_or_404(Category, pk=pk)
    ticket_count = category.tickets.count()
    
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Categoría "{category_name}" eliminada exitosamente.')
        return redirect('category_list')
    
    context = {
        'category': category,
        'ticket_count': ticket_count,
        'page_title': 'Eliminar Categoría'
    }
    return render(request, 'tickets/category_delete.html', context)


@login_required
@user_passes_test(is_agent)
def category_detail_view(request, pk):
    """Vista para ver detalles de una categoría y sus tickets (solo agentes)"""
    category = get_object_or_404(Category, pk=pk)
    tickets = category.tickets.all().order_by('-created_at')
    
    # Paginación de tickets
    paginator = Paginator(tickets, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'category': category,
        'tickets': page_obj,
        'ticket_count': tickets.count(),
        'page_title': f'Categoría: {category.name}'
    }
    return render(request, 'tickets/category_detail.html', context)


def public_ticket_view(request, token):
    """Vista pública para mostrar un ticket mediante su token único"""
    try:
        ticket = get_object_or_404(Ticket, public_share_token=token, is_public_shareable=True)
    except:
        # Si el ticket no existe o no está habilitado para compartir
        context = {
            'error_message': 'El enlace del ticket no es válido o ha expirado.',
            'error_type': 'not_found'
        }
        return render(request, 'tickets/public_ticket.html', context, status=404)
    
    # Obtener comentarios del ticket
    comments = ticket.comments.all().order_by('created_at')
    
    # Obtener adjuntos del ticket
    attachments = ticket.attachments.all().order_by('-uploaded_at')
    
    # Verificar si ya tiene aprobación
    has_approval = hasattr(ticket, 'client_approval')
    
    context = {
        'ticket': ticket,
        'comments': comments,
        'attachments': attachments,
        'has_approval': has_approval,
        'approval': ticket.client_approval if has_approval else None,
        'token': token,  # Agregar el token al contexto
        'page_title': f'Ticket Público: {ticket.ticket_number}',
    }
    return render(request, 'tickets/public_ticket.html', context)


def public_ticket_approve(request, token):
    """Vista para que los clientes aprueben tickets públicos"""
    try:
        ticket = get_object_or_404(Ticket, public_share_token=token, is_public_shareable=True)
    except:
        messages.error(request, 'Ticket no encontrado o no es público.')
        return redirect('ticket_list')
    
    # Verificar si ya tiene aprobación
    if hasattr(ticket, 'client_approval'):
        messages.warning(request, 'Este ticket ya ha sido aprobado anteriormente.')
        return redirect('public_ticket', token=token)
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            client_name = request.POST.get('client_name', '').strip()
            client_email = request.POST.get('client_email', '').strip()
            notes = request.POST.get('notes', '').strip()
            
            # Validaciones básicas
            if not client_name or not client_email:
                messages.error(request, 'Nombre y correo electrónico son obligatorios.')
                return redirect('public_ticket', token=token)
            
            # Validar formato de email
            from django.core.validators import validate_email
            from django.core.exceptions import ValidationError
            try:
                validate_email(client_email)
            except ValidationError:
                messages.error(request, 'Por favor ingresa un correo electrónico válido.')
                return redirect('public_ticket', token=token)
            
            # Obtener IP del cliente
            def get_client_ip(request):
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                if x_forwarded_for:
                    ip = x_forwarded_for.split(',')[0]
                else:
                    ip = request.META.get('REMOTE_ADDR')
                return ip
            
            # Crear la aprobación
            from .models import TicketApproval
            approval = TicketApproval.objects.create(
                ticket=ticket,
                client_name=client_name,
                client_email=client_email,
                notes=notes,
                ip_address=get_client_ip(request)
            )
            
            # Actualizar el estado de aprobación del ticket
            ticket.is_approved = True
            ticket.approved_by = None  # Aprobado por el cliente, no por un usuario del sistema
            ticket.approved_at = approval.approved_at
            ticket.save(update_fields=['is_approved', 'approved_by', 'approved_at'])
            
            # Agregar un comentario informando de la aprobación
            TicketComment.objects.create(
                ticket=ticket,
                user=ticket.created_by,  # Sistema
                content=f"✅ TICKET APROBADO POR EL CLIENTE\n\n" +
                       f"Cliente: {client_name}\n" +
                       f"Email: {client_email}\n" +
                       f"Fecha: {approval.approved_at.strftime('%d/%m/%Y %H:%M')}\n" +
                       (f"Comentarios: {notes}\n" if notes else "") +
                       f"IP: {approval.ip_address}\n\n" +
                       f"El cliente ha confirmado su satisfacción con la solución proporcionada."
            )
            
            messages.success(
                request, 
                f'¡Gracias {client_name}! Tu aprobación ha sido registrada exitosamente. '
                f'El equipo ha sido notificado de tu satisfacción con la solución.'
            )
            
            return redirect('public_ticket', token=token)
            
        except Exception as e:
            messages.error(request, f'Error al procesar la aprobación: {str(e)}')
            return redirect('public_ticket', token=token)
    
    # Si es GET, redirigir al ticket público
    return redirect('public_ticket', token=token)
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def public_ticket_survey(request, token):
    """Vista para que los clientes llenen encuesta de satisfacción"""
    try:
        ticket = get_object_or_404(Ticket, public_share_token=token, is_public_shareable=True)
    except:
        messages.error(request, 'Ticket no encontrado o no es público.')
        return redirect('ticket_list')
    
    # Verificar si ya tiene encuesta de satisfacción
    if hasattr(ticket, 'satisfaction_survey'):
        messages.info(request, 'Ya has enviado una encuesta de satisfacción para este ticket.')
        return redirect('public_ticket', token=token)
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            client_name = request.POST.get('client_name', '').strip()
            client_email = request.POST.get('client_email', '').strip()
            
            # Calificaciones (obligatorias)
            overall_satisfaction = request.POST.get('overall_satisfaction')
            resolution_quality = request.POST.get('resolution_quality')
            response_time = request.POST.get('response_time')
            communication = request.POST.get('communication')
            
            # Comentarios (opcionales)
            what_went_well = request.POST.get('what_went_well', '').strip()
            what_could_improve = request.POST.get('what_could_improve', '').strip()
            additional_comments = request.POST.get('additional_comments', '').strip()
            
            # Recomendación
            would_recommend = request.POST.get('would_recommend') == 'true'
            recommendation_reason = request.POST.get('recommendation_reason', '').strip()
            
            # Validaciones básicas
            if not client_name or not client_email:
                messages.error(request, 'Nombre y correo electrónico son obligatorios.')
                return redirect('public_ticket_survey', token=token)
            
            if not all([overall_satisfaction, resolution_quality, response_time, communication]):
                messages.error(request, 'Todas las calificaciones son obligatorias.')
                return redirect('public_ticket_survey', token=token)
            
            # Validar formato de email
            from django.core.validators import validate_email
            from django.core.exceptions import ValidationError
            try:
                validate_email(client_email)
            except ValidationError:
                messages.error(request, 'Por favor ingresa un correo electrónico válido.')
                return redirect('public_ticket_survey', token=token)
            
            # Validar que las calificaciones sean números válidos
            try:
                overall_satisfaction = int(overall_satisfaction)
                resolution_quality = int(resolution_quality)
                response_time = int(response_time)
                communication = int(communication)
                
                # Validar rango 1-5
                if not all(1 <= rating <= 5 for rating in [overall_satisfaction, resolution_quality, response_time, communication]):
                    raise ValueError("Las calificaciones deben estar entre 1 y 5")
                    
            except (ValueError, TypeError):
                messages.error(request, 'Las calificaciones deben ser números válidos entre 1 y 5.')
                return redirect('public_ticket_survey', token=token)
            
            # Obtener IP del cliente
            def get_client_ip(request):
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                if x_forwarded_for:
                    ip = x_forwarded_for.split(',')[0]
                else:
                    ip = request.META.get('REMOTE_ADDR')
                return ip
            
            # Crear la encuesta de satisfacción
            survey = SatisfactionSurvey.objects.create(
                ticket=ticket,
                client_name=client_name,
                client_email=client_email,
                overall_satisfaction=overall_satisfaction,
                resolution_quality=resolution_quality,
                response_time=response_time,
                communication=communication,
                what_went_well=what_went_well,
                what_could_improve=what_could_improve,
                additional_comments=additional_comments,
                would_recommend=would_recommend,
                recommendation_reason=recommendation_reason,
                ip_address=get_client_ip(request)
            )
            
            # Agregar un comentario informando de la encuesta
            TicketComment.objects.create(
                ticket=ticket,
                user=ticket.created_by,  # Sistema
                content=f"📊 ENCUESTA DE SATISFACCIÓN COMPLETADA\n\n" +
                       f"Cliente: {client_name}\n" +
                       f"Email: {client_email}\n" +
                       f"Fecha: {survey.submitted_at.strftime('%d/%m/%Y %H:%M')}\n\n" +
                       f"📈 CALIFICACIONES:\n" +
                       f"• Satisfacción General: {overall_satisfaction}/5\n" +
                       f"• Calidad de la Solución: {resolution_quality}/5\n" +
                       f"• Tiempo de Respuesta: {response_time}/5\n" +
                       f"• Comunicación: {communication}/5\n" +
                       f"• Promedio: {survey.average_rating}/5 - {survey.rating_summary}\n\n" +
                       f"🎯 RECOMENDACIÓN: {'Sí' if would_recommend else 'No'}\n" +
                       (f"Razón: {recommendation_reason}\n\n" if recommendation_reason else "\n") +
                       (f"✅ Lo que funcionó bien:\n{what_went_well}\n\n" if what_went_well else "") +
                       (f"🔧 Qué se podría mejorar:\n{what_could_improve}\n\n" if what_could_improve else "") +
                       (f"💬 Comentarios adicionales:\n{additional_comments}\n\n" if additional_comments else "") +
                       f"IP: {survey.ip_address}"
            )
            
            messages.success(
                request, 
                f'¡Gracias {client_name}! Tu encuesta de satisfacción ha sido enviada exitosamente. '
                f'Tu retroalimentación es muy valiosa para nosotros.'
            )
            
            return redirect('public_ticket', token=token)
            
        except Exception as e:
            messages.error(request, f'Error al procesar la encuesta: {str(e)}')
            return redirect('public_ticket_survey', token=token)
    
    # Si es GET, mostrar el formulario de encuesta
    context = {
        'ticket': ticket,
        'token': token,
        'rating_choices': SatisfactionSurvey.RATING_CHOICES,
        'resolution_quality_choices': SatisfactionSurvey.RESOLUTION_QUALITY_CHOICES,
        'response_time_choices': SatisfactionSurvey.RESPONSE_TIME_CHOICES,
    }
    
    return render(request, 'tickets/public_ticket_survey.html', context)


def public_ticket_upload(request, token):
    """Vista para que los clientes suban archivos a tickets públicos"""
    try:
        ticket = get_object_or_404(Ticket, public_share_token=token, is_public_shareable=True)
    except:
        messages.error(request, 'Ticket no encontrado o no es público.')
        return redirect('ticket_list')
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            client_name = request.POST.get('client_name', '').strip()
            client_email = request.POST.get('client_email', '').strip()
            description = request.POST.get('description', '').strip()
            
            # Validaciones básicas
            if not client_name or not client_email:
                messages.error(request, 'Nombre y correo electrónico son obligatorios.')
                return redirect('public_ticket_upload', token=token)
            
            # Validar formato de email
            from django.core.validators import validate_email
            from django.core.exceptions import ValidationError
            try:
                validate_email(client_email)
            except ValidationError:
                messages.error(request, 'Por favor ingresa un correo electrónico válido.')
                return redirect('public_ticket_upload', token=token)
            
            # Verificar que se haya subido al menos un archivo
            files = request.FILES.getlist('attachments')
            if not files:
                messages.error(request, 'Debes subir al menos un archivo.')
                return redirect('public_ticket_upload', token=token)
            
            # Validar cada archivo
            max_file_size = 10 * 1024 * 1024  # 10MB
            allowed_extensions = ['.pdf', '.doc', '.docx', '.txt', '.jpg', '.jpeg', '.png', '.gif', '.zip', '.rar', '.xlsx', '.xls', '.ppt', '.pptx']
            
            for file in files:
                # Validar tamaño
                if file.size > max_file_size:
                    messages.error(request, f'El archivo "{file.name}" excede el tamaño máximo permitido (10MB).')
                    return redirect('public_ticket_upload', token=token)
                
                # Validar extensión
                file_extension = os.path.splitext(file.name.lower())[1]
                if file_extension not in allowed_extensions:
                    messages.error(request, f'El archivo "{file.name}" tiene una extensión no permitida. Extensiones permitidas: {", ".join(allowed_extensions)}')
                    return redirect('public_ticket_upload', token=token)
            
            # Obtener IP del cliente
            def get_client_ip(request):
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                if x_forwarded_for:
                    ip = x_forwarded_for.split(',')[0]
                else:
                    ip = request.META.get('REMOTE_ADDR')
                return ip
            
            # Crear usuario temporal para los adjuntos (usar el creador del ticket)
            upload_user = ticket.created_by
            
            # Subir cada archivo
            uploaded_files = []
            for file in files:
                attachment = TicketAttachment.objects.create(
                    ticket=ticket,
                    file=file,
                    original_filename=file.name,
                    uploaded_by=upload_user,
                    file_size=file.size
                )
                uploaded_files.append(attachment)
            
            # Crear comentario informativo sobre la subida de archivos
            files_list = '\n'.join([f"• {att.original_filename} ({att.get_file_size_display()})" for att in uploaded_files])
            
            TicketComment.objects.create(
                ticket=ticket,
                user=upload_user,
                content=f"📎 ARCHIVOS SUBIDOS POR EL CLIENTE\n\n" +
                       f"Cliente: {client_name}\n" +
                       f"Email: {client_email}\n" +
                       f"Fecha: {timezone.now().strftime('%d/%m/%Y %H:%M')}\n" +
                       f"IP: {get_client_ip(request)}\n\n" +
                       f"Archivos subidos ({len(uploaded_files)}):\n{files_list}\n\n" +
                       (f"Descripción: {description}\n\n" if description else "") +
                       f"Los archivos han sido adjuntados al ticket y están disponibles para revisión."
            )
            
            messages.success(
                request, 
                f'¡Gracias {client_name}! Se {"ha" if len(uploaded_files) == 1 else "han"} subido {len(uploaded_files)} archivo{"" if len(uploaded_files) == 1 else "s"} exitosamente al ticket.'
            )
            
            return redirect('public_ticket', token=token)
            
        except Exception as e:
            messages.error(request, f'Error al subir los archivos: {str(e)}')
            return redirect('public_ticket_upload', token=token)
    
    # Si es GET, mostrar el formulario de subida
    context = {
        'ticket': ticket,
        'token': token,
        'max_file_size_mb': 10,
        'allowed_extensions': ['.pdf', '.doc', '.docx', '.txt', '.jpg', '.jpeg', '.png', '.gif', '.zip', '.rar', '.xlsx', '.xls', '.ppt', '.pptx'],
    }
    
    return render(request, 'tickets/public_ticket_upload.html', context)


@login_required
def user_profile_view(request):
    """Vista para mostrar y actualizar el perfil del usuario"""
    from . import utils
    
    # Obtener o crear el perfil del usuario
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    # Asegurar que el perfil tenga un token público
    if not profile.public_token:
        import uuid
        profile.public_token = uuid.uuid4()
        profile.save()
    
    # Importar formularios aquí para evitar importación circular
    from .forms import UserProfileForm, CustomPasswordChangeForm
    
    profile_form = UserProfileForm(instance=profile, user=request.user)
    password_form = CustomPasswordChangeForm(request.user)
    
    if request.method == 'POST':
        if 'update_profile' in request.POST:
            # Actualizar perfil
            profile_form = UserProfileForm(
                request.POST, 
                instance=profile, 
                user=request.user
            )
            if profile_form.is_valid():
                profile_form.save(user=request.user)
                messages.success(request, 'Tu perfil ha sido actualizado exitosamente.')
                return redirect('user_profile')
        
        elif 'update_contact_form_toggle' in request.POST:
            # Actualizar solo el toggle del formulario de contacto
            enable_form = request.POST.get('enable_public_contact_form') == 'true'
            profile.enable_public_contact_form = enable_form
            profile.save()
            return JsonResponse({'success': True})
        
        elif 'change_password' in request.POST:
            # Cambiar contraseña
            password_form = CustomPasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, user)  # Mantener la sesión activa
                messages.success(request, 'Tu contraseña ha sido cambiada exitosamente.')
                return redirect('user_profile')
        
        elif 'regenerate_token' in request.POST:
            # Regenerar token público
            new_token = profile.regenerate_token()
            messages.success(request, 'Tu token de acceso público ha sido regenerado.')
            return redirect('user_profile')
    
    # Construir URL pública para tareas
    if request.is_secure():
        protocol = 'https'
    else:
        protocol = 'http'
    
    public_task_url = f"{protocol}://{request.get_host()}/public/tasks/{profile.public_token}/"
    
    # Construir URL pública para formulario de contacto
    public_contact_form_url = None
    contact_form_stats = None
    if profile.enable_public_contact_form:
        public_contact_form_url = f"{protocol}://{request.get_host()}/contact/{profile.public_token}/"
        
        # Calcular estadísticas de formularios de contacto
        from .models import ContactFormSubmission
        submissions = ContactFormSubmission.objects.filter(submitted_by_user=request.user)
        contact_form_stats = {
            'pending': submissions.filter(status='pending').count(),
            'approved': submissions.filter(status='approved').count(),
            'rejected': submissions.filter(status='rejected').count(),
            'company_created': submissions.filter(status='company_created').count(),
            'total': submissions.count(),
        }
    
    context = {
        'profile_form': profile_form,
        'password_form': password_form,
        'user_profile': profile,
        'public_task_url': public_task_url,
        'public_contact_form_url': public_contact_form_url,
        'contact_form_stats': contact_form_stats,
        'is_agent': utils.is_agent(request.user),
        'page_title': 'Mi Perfil'
    }
    return render(request, 'tickets/user_profile.html', context)


# =====================================
# VIEWS PARA GESTIÓN DE NOTAS INTERNAS
# =====================================

@login_required
@user_passes_test(is_agent, login_url='dashboard')
def notes_list_view(request):
    """Lista todas las notas internas (solo para agentes)"""
    
    # Obtener parámetros de filtro
    search_query = request.GET.get('search', '')
    user_filter = request.GET.get('user', '')
    privacy_filter = request.GET.get('privacy', '')
    
    # Base queryset
    notes = UserNote.objects.select_related('user', 'created_by').prefetch_related('tickets')
    
    # Aplicar filtros
    if search_query:
        notes = notes.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query)
        )
    
    if user_filter:
        notes = notes.filter(user_id=user_filter)
    
    if privacy_filter == 'private':
        notes = notes.filter(is_private=True)
    elif privacy_filter == 'public':
        notes = notes.filter(is_private=False)
    
    # Paginación
    paginator = Paginator(notes, 15)  # 15 notas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener lista de usuarios para el filtro
    users_with_notes = User.objects.filter(notes__isnull=False).distinct().order_by('username')
    
    context = {
        'page_obj': page_obj,
        'users_with_notes': users_with_notes,
        'search_query': search_query,
        'user_filter': user_filter,
        'privacy_filter': privacy_filter,
        'page_title': 'Notas Internas'
    }
    return render(request, 'tickets/notes_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def note_detail_view(request, note_id):
    """Vista detallada de una nota (solo para agentes)"""
    
    note = get_object_or_404(UserNote, id=note_id)
    
    # Verificar permisos
    if not note.can_view(request.user):
        messages.error(request, 'No tienes permisos para ver esta nota.')
        return redirect('notes_list')
    
    context = {
        'note': note,
        'page_title': f'Nota: {note.title}'
    }
    return render(request, 'tickets/note_detail.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def note_create_view(request):
    """Crear nueva nota interna (solo para agentes)"""
    
    if request.method == 'POST':
        form = UserNoteForm(request.POST, current_user=request.user)
        if form.is_valid():
            note = form.save(created_by=request.user)
            messages.success(request, f'Nota "{note.title}" creada exitosamente.')
            return redirect('note_detail', note_id=note.id)
    else:
        form = UserNoteForm(current_user=request.user)
        
        # Si se pasa un usuario por parámetro, preseleccionarlo
        user_id = request.GET.get('user')
        if user_id:
            try:
                user = User.objects.get(id=user_id)
                form.fields['user'].initial = user
                # Cargar tickets del usuario seleccionado
                form.fields['tickets'].queryset = Ticket.objects.filter(
                    created_by=user
                ).order_by('-created_at')
            except User.DoesNotExist:
                pass
    
    context = {
        'form': form,
        'page_title': 'Crear Nota Interna'
    }
    return render(request, 'tickets/note_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def note_edit_view(request, note_id):
    """Editar nota interna existente (solo para agentes)"""
    
    note = get_object_or_404(UserNote, id=note_id)
    
    # Verificar permisos
    if not note.can_edit(request.user):
        messages.error(request, 'No tienes permisos para editar esta nota.')
        return redirect('note_detail', note_id=note.id)
    
    if request.method == 'POST':
        form = UserNoteForm(request.POST, instance=note, current_user=request.user)
        if form.is_valid():
            note = form.save()
            messages.success(request, f'Nota "{note.title}" actualizada exitosamente.')
            return redirect('note_detail', note_id=note.id)
    else:
        form = UserNoteForm(instance=note, current_user=request.user)
    
    context = {
        'form': form,
        'note': note,
        'page_title': f'Editar Nota: {note.title}'
    }
    return render(request, 'tickets/note_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def note_delete_view(request, note_id):
    """Eliminar nota interna (solo para agentes)"""
    
    note = get_object_or_404(UserNote, id=note_id)
    
    # Verificar permisos
    if not note.can_delete(request.user):
        messages.error(request, 'No tienes permisos para eliminar esta nota.')
        return redirect('note_detail', note_id=note.id)
    
    if request.method == 'POST':
        note_title = note.title
        note.delete()
        messages.success(request, f'Nota "{note_title}" eliminada exitosamente.')
        return redirect('notes_list')
    
    context = {
        'note': note,
        'page_title': f'Eliminar Nota: {note.title}'
    }
    return render(request, 'tickets/note_delete.html', context)


@login_required
def user_notes_view(request, user_id=None):
    """Vista de notas asociadas a un usuario específico"""
    
    # Si no se proporciona user_id, usar el usuario actual
    if user_id is None:
        target_user = request.user
    else:
        target_user = get_object_or_404(User, id=user_id)
    
    # Verificar permisos
    can_view_notes = False
    if is_agent(request.user):
        # Los agentes pueden ver notas de cualquier usuario
        can_view_notes = True
        notes = UserNote.objects.filter(user=target_user)
    elif request.user == target_user:
        # Los usuarios pueden ver sus propias notas (solo las no privadas)
        can_view_notes = True
        notes = UserNote.objects.filter(user=target_user, is_private=False)
    else:
        # Otros usuarios no pueden ver las notas
        messages.error(request, 'No tienes permisos para ver estas notas.')
        return redirect('dashboard')
    
    if not can_view_notes:
        messages.error(request, 'No tienes permisos para ver estas notas.')
        return redirect('dashboard')
    
    # Ordenar por fecha de creación descendente
    notes = notes.select_related('created_by').prefetch_related('tickets').order_by('-created_at')
    
    # Paginación
    paginator = Paginator(notes, 10)  # 10 notas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'target_user': target_user,
        'page_obj': page_obj,
        'is_own_notes': request.user == target_user,
        'can_create_notes': is_agent(request.user),
        'page_title': f'Notas de {target_user.get_full_name() or target_user.username}'
    }
    return render(request, 'tickets/user_notes.html', context)


# =====================================
# API PARA OBTENER TICKETS DE USUARIO
# =====================================

@login_required
@user_passes_test(is_agent, login_url='dashboard')
def api_user_tickets(request, user_id):
    """API para obtener tickets de un usuario específico (solo para agentes)"""
    
    try:
        user = User.objects.get(id=user_id)
        tickets = Ticket.objects.filter(created_by=user).order_by('-created_at')
        
        tickets_data = [
            {
                'id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'title': ticket.title,
                'status': ticket.get_status_display(),
                'created_at': ticket.created_at.strftime('%Y-%m-%d %H:%M')
            }
            for ticket in tickets
        ]
        
        return JsonResponse({
            'user': {
                'id': user.id,
                'username': user.username,
                'full_name': user.get_full_name()
            },
            'tickets': tickets_data
        })
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# =======================================
# VISTAS PARA CONTROL DE HORARIO
# =======================================

@login_required
@user_passes_test(is_agent, login_url='dashboard')
def time_clock_view(request):
    """Vista principal del control de horario para agentes"""
    
    # Obtener registro activo del usuario
    active_entry = TimeEntry.get_active_entry(request.user)
    
    # Obtener registros recientes (última semana)
    from datetime import timedelta
    # Calcular el inicio de la semana actual (lunes)
    now = timezone.now()
    start_week = now - timedelta(days=now.weekday())
    start_week = start_week.replace(hour=0, minute=0, second=0, microsecond=0)
    
    recent_entries = TimeEntry.objects.filter(
        user=request.user,
        fecha_entrada__gte=start_week
    ).order_by('-fecha_entrada')[:10]
    
    # Calcular estadísticas de la semana actual
    week_entries = TimeEntry.objects.filter(
        user=request.user,
        fecha_entrada__gte=start_week,
        fecha_salida__isnull=False  # Solo entradas completadas
    )
    total_minutos = sum(entry.duracion_trabajada for entry in week_entries)
    total_horas = round(total_minutos / 60, 2)
    
    context = {
        'active_entry': active_entry,
        'recent_entries': recent_entries,
        'total_horas_semana': total_horas,
        'can_start': not active_entry,
        'can_end': bool(active_entry),
        'page_title': 'Control de Horario'
    }
    
    return render(request, 'tickets/time_clock.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def time_start_work(request):
    """Iniciar jornada laboral"""
    
    # Verificar que no hay una jornada activa
    active_entry = TimeEntry.get_active_entry(request.user)
    if active_entry:
        messages.warning(request, 'Ya tienes una jornada activa desde las ' + 
                        active_entry.fecha_entrada.strftime('%H:%M'))
        return redirect('time_clock')
    
    if request.method == 'POST':
        form = TimeEntryStartForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                project = form.cleaned_data.get('project')
                ticket = form.cleaned_data.get('ticket')
                work_order = form.cleaned_data.get('work_order')
                task = form.cleaned_data.get('task')
                notas = form.cleaned_data.get('notas_entrada', '')
                
                entry = TimeEntry.create_entry(
                    user=request.user,
                    project=project,
                    ticket=ticket,
                    work_order=work_order,
                    task=task,
                    notas_entrada=notas
                )
                
                # Mensaje personalizado según los elementos seleccionados
                message_parts = [f'¡Jornada iniciada a las {entry.fecha_entrada.strftime("%H:%M")}']
                if project:
                    message_parts.append(f'en el proyecto "{project.name}"')
                if ticket:
                    message_parts.append(f'trabajando en el ticket "{ticket.ticket_number}"')
                if work_order:
                    message_parts.append(f'en la orden "{work_order.order_number}"')
                if task:
                    message_parts.append(f'en la tarea "{task.title}"')
                
                message = ' '.join(message_parts) + '!'
                messages.success(request, message)
                    
                return redirect('time_clock')
            except ValueError as e:
                messages.error(request, str(e))
                return redirect('time_clock')
    else:
        form = TimeEntryStartForm(user=request.user)
    
    context = {
        'form': form,
        'page_title': 'Iniciar Jornada'
    }
    return render(request, 'tickets/time_start.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def time_end_work(request):
    """Finalizar jornada laboral"""
    
    # Verificar que hay una jornada activa
    active_entry = TimeEntry.get_active_entry(request.user)
    if not active_entry:
        messages.error(request, 'No tienes una jornada activa para finalizar.')
        return redirect('time_clock')
    
    if request.method == 'POST':
        form = TimeEntryEndForm(request.POST)
        if form.is_valid():
            try:
                notas_salida = form.cleaned_data.get('notas_salida', '')
                active_entry.finalizar_jornada(notas_salida)
                
                duracion = active_entry.duracion_formateada
                messages.success(request, 
                    f'¡Jornada finalizada! Duración total: {duracion} horas')
                return redirect('time_clock')
            except ValueError as e:
                messages.error(request, str(e))
                return redirect('time_clock')
    else:
        form = TimeEntryEndForm()
    
    context = {
        'form': form,
        'active_entry': active_entry,
        'page_title': 'Finalizar Jornada'
    }
    return render(request, 'tickets/time_end.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def time_entries_list(request):
    """Lista de registros de horario del empleado"""
    
    # Filtros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    # Base queryset - solo registros del usuario actual
    entries = TimeEntry.objects.filter(user=request.user)
    
    # Aplicar filtros de fecha
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            entries = entries.filter(fecha_entrada__date__gte=fecha_desde_dt)
        except ValueError:
            messages.warning(request, 'Formato de fecha inválido')
    
    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            entries = entries.filter(fecha_entrada__date__lte=fecha_hasta_dt)
        except ValueError:
            messages.warning(request, 'Formato de fecha inválido')
    
    # Ordenar por fecha descendente
    entries = entries.order_by('-fecha_entrada')
    
    # Paginación
    paginator = Paginator(entries, 20)  # 20 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calcular totales
    if page_obj.object_list:
        total_minutos = sum(entry.duracion_trabajada for entry in entries if entry.fecha_salida)
        total_horas = round(total_minutos / 60, 2)
        total_dias = entries.filter(fecha_salida__isnull=False).count()
    else:
        total_horas = 0
        total_dias = 0
    
    context = {
        'page_obj': page_obj,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'total_horas': total_horas,
        'total_dias': total_dias,
        'page_title': 'Mis Registros de Horario'
    }
    
    return render(request, 'tickets/time_entries_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def time_entry_detail(request, entry_id):
    """Vista detallada de un registro de horario"""
    
    entry = get_object_or_404(TimeEntry, id=entry_id)
    
    # Verificar permisos - solo el propietario puede ver sus registros
    if not entry.can_view(request.user):
        messages.error(request, 'No tienes permisos para ver este registro.')
        return redirect('time_entries_list')
    
    context = {
        'entry': entry,
        'page_title': f'Registro del {entry.fecha_entrada.strftime("%d/%m/%Y")}'
    }
    
    return render(request, 'tickets/time_entry_detail.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def time_entry_edit(request, entry_id):
    """Editar notas de un registro de horario"""
    
    entry = get_object_or_404(TimeEntry, id=entry_id)
    
    # Verificar permisos
    if not entry.can_edit(request.user):
        messages.error(request, 'No tienes permisos para editar este registro.')
        return redirect('time_entry_detail', entry_id=entry.id)
    
    if request.method == 'POST':
        form = TimeEntryEditForm(request.POST, instance=entry, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro actualizado exitosamente.')
            return redirect('time_entry_detail', entry_id=entry.id)
    else:
        form = TimeEntryEditForm(instance=entry, user=request.user)
    
    context = {
        'form': form,
        'entry': entry,
        'page_title': f'Editar Registro del {entry.fecha_entrada.strftime("%d/%m/%Y")}'
    }
    
    return render(request, 'tickets/time_entry_edit.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def ajax_current_time_status(request):
    """API AJAX para obtener el estado actual del tiempo trabajado"""
    
    active_entry = TimeEntry.get_active_entry(request.user)
    
    if active_entry:
        return JsonResponse({
            'is_working': True,
            'start_time': active_entry.fecha_entrada.strftime('%H:%M'),
            'duration': active_entry.duracion_formateada,
            'minutes_worked': active_entry.duracion_trabajada
        })
    else:
        return JsonResponse({
            'is_working': False,
            'start_time': None,
            'duration': '00:00',
            'minutes_worked': 0
        })


# ===========================================
# VISTAS DE GESTIÓN DE PROYECTOS
# ===========================================

@login_required
@user_passes_test(is_agent, login_url='ticket_list')
def project_list(request):
    """Vista para listar todos los proyectos (solo agentes)"""
    projects = Project.objects.all()
    
    # Búsqueda
    search_query = request.GET.get('search', '')
    if search_query:
        projects = projects.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filtro por estado
    status_filter = request.GET.get('status', '')
    if status_filter:
        projects = projects.filter(status=status_filter)
    
    # Paginación
    paginator = Paginator(projects, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener configuración del sistema
    from tickets.models import SystemConfiguration
    system_config = SystemConfiguration.get_config()
    
    context = {
        'page_title': 'Gestión de Proyectos',
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'project_status_choices': Project.STATUS_CHOICES,
        'system_config': system_config,
    }
    
    return render(request, 'tickets/project_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='ticket_list')
def project_create(request):
    """Vista para crear un nuevo proyecto (solo agentes)"""
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(created_by=request.user)
            messages.success(request, f'Proyecto "{project.name}" creado exitosamente.')
            return redirect('project_list')
    else:
        form = ProjectForm()
    
    context = {
        'page_title': 'Crear Nuevo Proyecto',
        'form': form,
        'action': 'create'
    }
    
    return render(request, 'tickets/project_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='ticket_list')
def project_edit(request, project_id):
    """Vista para editar un proyecto existente (solo agentes)"""
    project = get_object_or_404(Project, id=project_id)
    
    # Verificar permisos
    if not project.can_edit(request.user):
        messages.error(request, 'No tienes permisos para editar este proyecto.')
        return redirect('project_list')
    
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, f'Proyecto "{project.name}" actualizado exitosamente.')
            return redirect('project_list')
    else:
        form = ProjectForm(instance=project)
    
    context = {
        'page_title': f'Editar Proyecto: {project.name}',
        'form': form,
        'project': project,
        'action': 'edit'
    }
    
    return render(request, 'tickets/project_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='ticket_list')
def project_detail(request, project_id):
    """Vista para ver detalles de un proyecto"""
    project = get_object_or_404(Project, id=project_id)
    
    # Obtener registros de tiempo de este proyecto
    time_entries = project.time_entries.filter(
        fecha_salida__isnull=False
    ).order_by('-fecha_entrada')[:10]
    
    # Obtener trabajadores activos en este proyecto
    active_workers = project.time_entries.filter(
        fecha_salida__isnull=True
    ).select_related('user')
    
    context = {
        'page_title': f'Proyecto: {project.name}',
        'project': project,
        'time_entries': time_entries,
        'active_workers': active_workers,
        'total_hours': project.get_total_hours(),
        'active_workers_count': project.get_active_workers_count(),
        'can_edit_project': project.can_edit(request.user) if request.user.is_authenticated else False,
    }
    
    return render(request, 'tickets/project_detail.html', context)


@login_required
@user_passes_test(is_agent, login_url='ticket_list')
def project_delete(request, project_id):
    """Vista para eliminar un proyecto (solo agentes)"""
    project = get_object_or_404(Project, id=project_id)
    
    # Verificar permisos
    if not project.can_delete(request.user):
        messages.error(request, 'No tienes permisos para eliminar este proyecto.')
        return redirect('project_list')
    
    # Verificar si hay registros de tiempo asociados
    if project.time_entries.exists():
        messages.error(request, 
            'No se puede eliminar el proyecto porque tiene registros de tiempo asociados. '
            'Considera marcarlo como inactivo en su lugar.')
        return redirect('project_detail', project_id=project.id)
    
    if request.method == 'POST':
        project_name = project.name
        project.delete()
        messages.success(request, f'Proyecto "{project_name}" eliminado exitosamente.')
        return redirect('project_list')
    
    context = {
        'page_title': f'Eliminar Proyecto: {project.name}',
        'project': project,
    }
    
    return render(request, 'tickets/project_delete.html', context)


@login_required
@user_passes_test(is_agent, login_url='ticket_list')
def project_generate_public_token(request, project_id):
    """Vista para generar o regenerar el token público de un proyecto"""
    project = get_object_or_404(Project, id=project_id)
    
    # Verificar permisos
    if not project.can_edit(request.user):
        messages.error(request, 'No tienes permisos para gestionar este proyecto.')
        return redirect('project_detail', project_id=project.id)
    
    if request.method == 'POST':
        # Generar nuevo token (esto también actualizará uno existente)
        import secrets
        project.public_share_token = secrets.token_urlsafe(32)
        project.save()
        
        messages.success(request, 
            'Enlace público generado exitosamente. '
            'Ahora puedes compartir este proyecto con información básica sin necesidad de iniciar sesión.')
    
    return redirect('project_detail', project_id=project.id)


@login_required
@user_passes_test(is_agent, login_url='ticket_list')
def project_duplicate(request, project_id):
    """Vista para duplicar un proyecto con todas sus tareas"""
    original_project = get_object_or_404(Project, id=project_id)
    
    # Verificar permisos
    if not original_project.can_edit(request.user):
        messages.error(request, 'No tienes permisos para duplicar este proyecto.')
        return redirect('project_detail', project_id=project_id)
    
    if request.method == 'POST':
        try:
            # Crear el proyecto duplicado
            duplicated_project = Project.objects.create(
                name=f"Copia de {original_project.name}",
                description=original_project.description,
                company=original_project.company,
                status=original_project.status,
                start_date=original_project.start_date,
                end_date=original_project.end_date,
                color=original_project.color,
                is_active=True,  # El proyecto duplicado siempre empieza activo
                created_by=request.user,
                # No copiamos el token público por seguridad
                public_share_token=None,
                # No copiamos precios ya que son específicos del proyecto original
                precio_hora=original_project.precio_hora,
                costo_hora=original_project.costo_hora,
            )
            
            # Duplicar las tareas del cronograma si existen
            original_schedules = TaskSchedule.objects.filter(project=original_project)
            
            for original_schedule in original_schedules:
                # Crear cronograma duplicado
                duplicated_schedule = TaskSchedule.objects.create(
                    title=f"Copia de {original_schedule.title}",
                    description=original_schedule.description,
                    project=duplicated_project,
                    company=original_schedule.company,
                    start_date=original_schedule.start_date,
                    end_date=original_schedule.end_date,
                    is_public=False,  # Por seguridad, no hacer público automáticamente
                    created_by=request.user,
                    # No copiamos el token público
                    public_token=None,
                )
                
                # Duplicar las tareas del cronograma
                original_tasks = ScheduleTask.objects.filter(schedule=original_schedule)
                task_map = {}  # Para mapear dependencias
                
                # Primer paso: crear todas las tareas sin dependencias
                for original_task in original_tasks:
                    duplicated_task = ScheduleTask.objects.create(
                        schedule=duplicated_schedule,
                        title=original_task.title,
                        description=original_task.description,
                        start_date=original_task.start_date,
                        end_date=original_task.end_date,
                        priority=original_task.priority,
                        assigned_to=original_task.assigned_to,
                        is_completed=False,  # Las tareas duplicadas empiezan como no completadas
                        progress_percentage=0,  # Progreso en 0
                    )
                    task_map[original_task.id] = duplicated_task
                
                # Segundo paso: establecer dependencias
                for original_task in original_tasks:
                    duplicated_task = task_map[original_task.id]
                    for dependency in original_task.dependencies.all():
                        if dependency.id in task_map:
                            duplicated_task.dependencies.add(task_map[dependency.id])
            
            messages.success(
                request, 
                f'Proyecto "{duplicated_project.name}" duplicado exitosamente con todas sus tareas. '
                f'Se han copiado {TaskSchedule.objects.filter(project=duplicated_project).count()} cronogramas.'
            )
            
            # Redirigir al nuevo proyecto duplicado
            return redirect('project_detail', project_id=duplicated_project.id)
            
        except Exception as e:
            messages.error(request, f'Error al duplicar el proyecto: {str(e)}')
            return redirect('project_detail', project_id=project_id)
    
    # Si es GET, mostrar confirmación
    context = {
        'page_title': f'Duplicar Proyecto: {original_project.name}',
        'project': original_project,
        'schedules_count': TaskSchedule.objects.filter(project=original_project).count(),
        'total_tasks_count': ScheduleTask.objects.filter(schedule__project=original_project).count(),
    }
    
    return render(request, 'tickets/project_duplicate_confirm.html', context)


@login_required
@user_passes_test(is_agent, login_url='dashboard')
def attendance_overview(request):
    """Vista para que los agentes vean la asistencia de todos los empleados"""
    from django.db.models import Q
    from datetime import datetime, timedelta
    
    # Filtros de fecha
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    proyecto_id = request.GET.get('proyecto')
    usuario_id = request.GET.get('usuario')
    
    # Si no hay fechas, mostrar los últimos 7 días
    if not fecha_desde:
        fecha_desde = (timezone.now() - timedelta(days=7)).date()
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
    
    if not fecha_hasta:
        fecha_hasta = timezone.now().date()
    else:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Base queryset
    queryset = TimeEntry.objects.select_related('user', 'project').filter(
        fecha_entrada__date__gte=fecha_desde,
        fecha_entrada__date__lte=fecha_hasta
    ).order_by('-fecha_entrada')
    
    # Filtro por proyecto
    if proyecto_id:
        queryset = queryset.filter(project_id=proyecto_id)
    
    # Filtro por usuario
    if usuario_id:
        queryset = queryset.filter(user_id=usuario_id)
    
    # Obtener estadísticas
    total_entries = queryset.count()
    active_entries = queryset.filter(fecha_salida__isnull=True).count()
    
    # Calcular total de horas y promedio
    total_hours = 0
    completed_entries = 0
    completed_hours = 0
    
    for entry in queryset:
        if entry.fecha_salida:
            # Jornada completada
            duration = (entry.fecha_salida - entry.fecha_entrada).total_seconds()
            completed_hours += duration
            completed_entries += 1
        else:
            # Jornada activa - solo para el total
            duration = (timezone.now() - entry.fecha_entrada).total_seconds()
        
        total_hours += duration
    
    total_hours = round(total_hours / 3600, 2)
    
    # Calcular promedio por jornada (solo jornadas completadas)
    if completed_entries > 0:
        avg_hours_per_session = round(completed_hours / 3600 / completed_entries, 2)
    else:
        avg_hours_per_session = 0
    
    # Paginación
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Datos para filtros
    proyectos = Project.objects.filter(is_active=True).order_by('name')
    usuarios = User.objects.filter(
        groups__name='Agentes'
    ).order_by('first_name', 'username')
    
    context = {
        'page_title': 'Control de Asistencia General',
        'page_obj': page_obj,
        'total_entries': total_entries,
        'active_entries': active_entries,
        'total_hours': total_hours,
        'avg_hours_per_session': avg_hours_per_session,
        'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
        'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d'),
        'proyectos': proyectos,
        'usuarios': usuarios,
        'proyecto_seleccionado': proyecto_id,
        'usuario_seleccionado': usuario_id,
    }
    
    return render(request, 'tickets/attendance_overview.html', context)



# =============================================================================
# VISTAS PARA GESTIÓN DE EMPRESAS (SOLO AGENTES)
# =============================================================================

@login_required
@user_passes_test(is_agent, login_url='/')
def company_list_view(request):
    """Vista para listar todas las empresas"""
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', 'all')
    
    # Filtro base
    queryset = Company.objects.all().order_by('name')
    
    # Aplicar filtros
    if search_query:
        queryset = queryset.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if status_filter == 'active':
        queryset = queryset.filter(is_active=True)
    elif status_filter == 'inactive':
        queryset = queryset.filter(is_active=False)
    
    # Paginación
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    total_companies = Company.objects.count()
    active_companies = Company.objects.filter(is_active=True).count()
    
    context = {
        'page_title': 'Gestión de Empresas',
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'total_companies': total_companies,
        'active_companies': active_companies,
    }
    
    return render(request, 'tickets/company_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def company_create_view(request):
    """Vista para crear una nueva empresa"""
    if request.method == 'POST':
        form = CompanyForm(request.POST, request.FILES)
        if form.is_valid():
            company = form.save()
            messages.success(request, f'La empresa "{company.name}" ha sido creada exitosamente.')
            return redirect('company_list')
    else:
        form = CompanyForm()
    
    context = {
        'page_title': 'Crear Nueva Empresa',
        'form': form,
        'is_create': True,
    }
    
    return render(request, 'tickets/company_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def company_detail_view(request, company_id):
    """Vista para ver detalles de una empresa"""
    company = get_object_or_404(Company, id=company_id)
    
    # Generar public_token si no existe
    if not company.public_token:
        import uuid
        company.public_token = uuid.uuid4()
        company.save()
    
    # Estadísticas de la empresa
    total_tickets = company.tickets.count()
    open_tickets = company.tickets.filter(status='open').count()
    total_users = company.users.count()
    active_users = company.users.filter(user__is_active=True).count()
    
    # Tickets recientes
    recent_tickets = company.tickets.select_related(
        'created_by', 'assigned_to', 'category'
    ).order_by('-created_at')[:10]
    
    # Usuarios de la empresa
    company_users = company.users.select_related('user').order_by('user__first_name', 'user__username')
    
    context = {
        'page_title': f'Empresa: {company.name}',
        'company': company,
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'total_users': total_users,
        'active_users': active_users,
        'recent_tickets': recent_tickets,
        'company_users': company_users,
    }
    
    return render(request, 'tickets/company_detail.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def company_edit_view(request, company_id):
    """Vista para editar una empresa"""
    company = get_object_or_404(Company, id=company_id)
    
    if request.method == 'POST':
        form = CompanyForm(request.POST, request.FILES, instance=company)
        if form.is_valid():
            form.save()
            messages.success(request, f'La empresa "{company.name}" ha sido actualizada exitosamente.')
            return redirect('company_detail', company_id=company.id)
    else:
        form = CompanyForm(instance=company)
    
    context = {
        'page_title': f'Editar Empresa: {company.name}',
        'form': form,
        'company': company,
        'is_create': False,
    }
    
    return render(request, 'tickets/company_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def company_delete_view(request, company_id):
    """Vista para eliminar una empresa"""
    company = get_object_or_404(Company, id=company_id)
    
    # Verificar si la empresa tiene tickets o usuarios asociados
    has_tickets = company.tickets.exists()
    has_users = company.users.exists()
    
    if request.method == 'POST':
        if has_tickets or has_users:
            messages.error(
                request, 
                f'No se puede eliminar la empresa "{company.name}" porque tiene tickets o usuarios asociados. '
                'Desactívala en su lugar.'
            )
        else:
            company_name = company.name
            company.delete()
            messages.success(request, f'La empresa "{company_name}" ha sido eliminada exitosamente.')
            return redirect('company_list')
    
    context = {
        'page_title': f'Eliminar Empresa: {company.name}',
        'company': company,
        'has_tickets': has_tickets,
        'has_users': has_users,
    }
    
    return render(request, 'tickets/company_delete.html', context)


# =============================================================================
# VISTAS PARA CONFIGURACIÓN DEL SISTEMA (SOLO AGENTES)
# =============================================================================

@login_required
@user_passes_test(is_agent, login_url='/')
def system_configuration_view(request):
    """Vista para configurar el sistema"""
    config = SystemConfiguration.get_config()
    
    if request.method == 'POST':
        # Verificar si es una prueba de Telegram
        if 'test_telegram' in request.POST:
            from .telegram_utils import test_telegram_connection
            bot_token = request.POST.get('telegram_bot_token', '').strip()
            chat_id = request.POST.get('telegram_chat_id', '').strip()
            
            result = test_telegram_connection(bot_token, chat_id)
            
            if result['success']:
                messages.success(request, result['message'])
            else:
                messages.error(request, result['message'])
            
            return redirect('system_configuration')
        
        # Verificar si es una prueba de email
        if 'test_email' in request.POST:
            try:
                from .utils import send_contact_notification
                from .models import ContactoWeb
                
                # Crear un contacto de prueba temporal
                contacto_prueba = ContactoWeb(
                    nombre='Prueba del Sistema',
                    email='prueba@sistema.local',
                    telefono='+34 600 000 000',
                    empresa='Sistema TicketProo',
                    asunto='Prueba de notificación por email',
                    mensaje='Este es un mensaje de prueba generado desde la configuración del sistema.',
                    ip_address='127.0.0.1',
                    user_agent='System Configuration Test'
                )
                
                # Intentar enviar la notificación sin guardar el contacto
                result = send_contact_notification(contacto_prueba)
                
                if result:
                    messages.success(request, '✅ Email de prueba enviado exitosamente. Revisa los emails configurados.')
                else:
                    messages.error(request, '❌ Error enviando el email de prueba. Revisa la configuración.')
                    
            except Exception as e:
                messages.error(request, f'❌ Error en la prueba de email: {str(e)}')
            
            return redirect('system_configuration')
        
        # Procesar formulario normal
        form = SystemConfigurationForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, 'La configuración del sistema ha sido actualizada exitosamente.')
            return redirect('system_configuration')
    else:
        form = SystemConfigurationForm(instance=config)
    
    context = {
        'page_title': 'Configuración del Sistema',
        'form': form,
        'config': config,
    }
    
    return render(request, 'tickets/system_configuration.html', context)


# =============================================================================
# VISTAS PARA GESTIÓN DE DOCUMENTOS
# =============================================================================

@login_required
@user_passes_test(is_agent, login_url='/')
def document_list_view(request):
    """Vista para listar todos los documentos"""
    search_query = request.GET.get('search', '').strip()
    company_filter = request.GET.get('company', '')
    tag_filter = request.GET.get('tag', '').strip()
    
    # Filtro base
    queryset = Document.objects.select_related('created_by', 'company').order_by('-created_at')
    
    # Aplicar filtros
    if search_query:
        queryset = queryset.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(tags__icontains=search_query)
        )
    
    if company_filter:
        if company_filter == 'none':
            queryset = queryset.filter(company__isnull=True)
        else:
            queryset = queryset.filter(company_id=company_filter)
    
    if tag_filter:
        queryset = queryset.filter(tags__icontains=tag_filter)
    
    # Paginación
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Datos para filtros
    companies = Company.objects.filter(is_active=True).order_by('name')
    
    # Obtener todas las etiquetas únicas
    all_tags = set()
    for doc in Document.objects.exclude(tags=''):
        for tag in doc.get_tags_list():
            all_tags.add(tag)
    all_tags = sorted(list(all_tags))
    
    context = {
        'page_title': 'Gestión de Documentos',
        'page_obj': page_obj,
        'search_query': search_query,
        'company_filter': company_filter,
        'tag_filter': tag_filter,
        'companies': companies,
        'all_tags': all_tags,
        'total_documents': Document.objects.count(),
    }
    
    return render(request, 'tickets/document_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def document_create_view(request):
    """Vista para crear un nuevo documento"""
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.created_by = request.user
            document.save()
            messages.success(request, f'El documento "{document.title}" ha sido creado exitosamente.')
            return redirect('document_list')
    else:
        form = DocumentForm()
    
    context = {
        'page_title': 'Subir Nuevo Documento',
        'form': form,
        'is_create': True,
    }
    
    return render(request, 'tickets/document_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def document_detail_view(request, document_id):
    """Vista para ver detalles de un documento"""
    document = get_object_or_404(Document, id=document_id)
    
    context = {
        'page_title': f'Documento: {document.title}',
        'document': document,
    }
    
    return render(request, 'tickets/document_detail.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def document_edit_view(request, document_id):
    """Vista para editar un documento"""
    document = get_object_or_404(Document, id=document_id)
    
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES, instance=document)
        if form.is_valid():
            form.save()
            messages.success(request, f'El documento "{document.title}" ha sido actualizado exitosamente.')
            return redirect('document_detail', document_id=document.id)
    else:
        form = DocumentForm(instance=document)
    
    context = {
        'page_title': f'Editar Documento: {document.title}',
        'form': form,
        'document': document,
        'is_create': False,
    }
    
    return render(request, 'tickets/document_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def document_delete_view(request, document_id):
    """Vista para eliminar un documento"""
    document = get_object_or_404(Document, id=document_id)
    
    if request.method == 'POST':
        document_title = document.title
        # Eliminar archivo físico
        if document.file:
            document.file.delete(save=False)
        document.delete()
        messages.success(request, f'El documento "{document_title}" ha sido eliminado exitosamente.')
        return redirect('document_list')
    
    context = {
        'page_title': f'Eliminar Documento: {document.title}',
        'document': document,
    }
    
    return render(request, 'tickets/document_delete.html', context)


def document_public_view(request, token):
    """Vista pública para compartir documentos sin autenticación"""
    document = get_object_or_404(Document, public_share_token=token, is_public=True)
    
    context = {
        'document': document,
        'page_title': f'Documento Compartido: {document.title}',
    }
    
    return render(request, 'tickets/document_public.html', context)


def document_download_view(request, token):
    """Vista para descargar documentos públicos"""
    document = get_object_or_404(Document, public_share_token=token, is_public=True)
    
    # Incrementar contador de descargas
    document.increment_download_count()
    
    # Servir el archivo
    from django.http import HttpResponse, Http404
    import os
    
    if not document.file or not os.path.exists(document.file.path):
        raise Http404("Archivo no encontrado")
    
    response = HttpResponse(
        document.file.read(),
        content_type='application/octet-stream'
    )
    response['Content-Disposition'] = f'attachment; filename="{document.file.name}"'
    
    return response


@login_required
@user_passes_test(is_agent, login_url='/')
def document_download_private_view(request, document_id):
    """Vista para descargar documentos con autenticación"""
    document = get_object_or_404(Document, id=document_id)
    
    # Incrementar contador de descargas
    document.increment_download_count()
    
    # Servir el archivo
    from django.http import HttpResponse, Http404
    import os
    
    if not document.file or not os.path.exists(document.file.path):
        raise Http404("Archivo no encontrado")
    
    response = HttpResponse(
        document.file.read(),
        content_type='application/octet-stream'
    )
    response['Content-Disposition'] = f'attachment; filename="{document.file.name}"'
    
    return response


# VIEWS PARA GESTIÓN DE URLs CON CREDENCIALES

@login_required
@user_passes_test(is_agent, login_url='/')
def url_manager_list_view(request):
    """Lista todas las URLs gestionadas (solo para agentes)"""
    from .models import UrlManager
    from .forms import UrlManagerFilterForm
    from django.db.models import Q
    
    # Obtener filtros
    form = UrlManagerFilterForm(request.GET)
    urls = UrlManager.objects.all()
    
    # Aplicar filtros
    if form.is_valid():
        search = form.cleaned_data.get('search')
        category = form.cleaned_data.get('category')
        is_active = form.cleaned_data.get('is_active')
        
        if search:
                       urls = urls.filter(
                Q(title__icontains=search) |
                Q(url__icontains=search) |
                Q(description__icontains=search) |
                Q(username__icontains=search)
            )
        
        if category:
            urls = urls.filter(category__icontains=category)
        
        if is_active:
            urls = urls.filter(is_active=is_active == 'True')
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(urls, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener categorías únicas para el filtro
    categories = UrlManager.objects.values_list('category', flat=True).distinct().exclude(category__isnull=True).exclude(category='')
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'categories': categories,
        'search_query': form.cleaned_data.get('search', '') if form.is_valid() else '',
        'category_filter': form.cleaned_data.get('category', '') if form.is_valid() else '',
        'status_filter': form.cleaned_data.get('is_active', '') if form.is_valid() else '',
        'page_title': 'Gestión de URLs'
    }
    
    return render(request, 'tickets/url_manager_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def url_manager_create_view(request):
    """Crear nueva URL gestionada (solo para agentes)"""
    from .models import UrlManager
    from .forms import UrlManagerForm
    
    if request.method == 'POST':
        form = UrlManagerForm(request.POST)
        if form.is_valid():
            url_manager = form.save(commit=False)
            url_manager.created_by = request.user
            url_manager.save()
            
            messages.success(request, f'URL "{url_manager.title}" creada exitosamente.')
            return redirect('url_manager_list')
    else:
        form = UrlManagerForm()
    
    context = {
        'form': form,
        'page_title': 'Crear Nueva URL',
        'action': 'Crear'
    }
    
    return render(request, 'tickets/url_manager_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def url_manager_edit_view(request, url_id):
    """Editar URL gestionada (solo para agentes)"""
    from .models import UrlManager
    from .forms import UrlManagerForm
    
    url_manager = get_object_or_404(UrlManager, id=url_id)
    
    if request.method == 'POST':
        form = UrlManagerForm(request.POST, instance=url_manager)
        if form.is_valid():
            form.save()
            messages.success(request, f'URL "{url_manager.title}" actualizada exitosamente.')
            return redirect('url_manager_list')
    else:
        form = UrlManagerForm(instance=url_manager)
    
    context = {
        'form': form,
        'url_manager': url_manager,
        'page_title': f'Editar URL: {url_manager.title}',
        'action': 'Actualizar'
    }
    
    return render(request, 'tickets/url_manager_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def url_manager_detail_view(request, url_id):
    """Ver detalles de una URL gestionada (solo para agentes)"""
    from .models import UrlManager
    
    url_manager = get_object_or_404(UrlManager, id=url_id)
    
    # Marcar como accedida
    url_manager.mark_accessed()
    
    context = {
        'url_manager': url_manager,
        'page_title': f'Detalles: {url_manager.title}'
    }
    
    return render(request, 'tickets/url_manager_detail.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def url_manager_delete_view(request, url_id):
    """Eliminar URL gestionada (solo para agentes)"""
    from .models import UrlManager
    
    url_manager = get_object_or_404(UrlManager, id=url_id)
    
    if request.method == 'POST':
        title = url_manager.title
        url_manager.delete()
        messages.success(request, f'URL "{title}" eliminada exitosamente.')
        return redirect('url_manager_list')
    
    context = {
        'url_manager': url_manager,
        'page_title': f'Eliminar URL: {url_manager.title}'
    }
    
    return render(request, 'tickets/url_manager_delete.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def url_manager_password_view(request, url_id):
    """AJAX endpoint para obtener la contraseña desencriptada"""
    from .models import UrlManager
    from django.http import JsonResponse
    import logging
    
    logger = logging.getLogger(__name__)
    logger.info(f"Solicitando contraseña para URL ID: {url_id}")
    
    try:
        url_manager = get_object_or_404(UrlManager, id=url_id)
        logger.info(f"URL encontrada: {url_manager.title}")
        
        password = url_manager.get_password()
        logger.info(f"Contraseña obtenida exitosamente")
        
        return JsonResponse({
            'success': True,
            'password': password
        })
    except Exception as e:
        logger.error(f"Error al obtener contraseña: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


# ===========================================
# VISTAS PARA ÓRDENAS DE TRABAJO
# ===========================================

@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_list_view(request):
    """Vista para listar órdenes de trabajo"""
    # Obtener todas las órdenes de trabajo
    work_orders = WorkOrder.objects.all()
    
    # Aplicar filtros si existen
    form = WorkOrderFilterForm(request.GET)
    if form.is_valid():
        search = form.cleaned_data.get('search')
        company = form.cleaned_data.get('company')
        status = form.cleaned_data.get('status')
        priority = form.cleaned_data.get('priority')
        assigned_to = form.cleaned_data.get('assigned_to')
        
        if search:
            work_orders = work_orders.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search)
            )
        
        if company:
            work_orders = work_orders.filter(company=company)
        
        if status:
            work_orders = work_orders.filter(status=status)
        
        if priority:
            work_orders = work_orders.filter(priority=priority)
        
        if assigned_to:
            work_orders = work_orders.filter(assigned_to=assigned_to)
    
    # Agregar información de permisos a cada orden
    for work_order in work_orders:
        work_order.user_can_edit = work_order.can_edit(request.user)
        work_order.user_can_delete = work_order.can_delete(request.user)
    
    # Paginación
    paginator = Paginator(work_orders, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'filter_form': form,
        'page_title': 'Órdenes de Trabajo'
    }
    
    return render(request, 'tickets/work_order_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_create_view(request):
    """Vista para crear nuevas órdenes de trabajo"""
    if request.method == 'POST':
        form = WorkOrderForm(request.POST, request.FILES)
        if form.is_valid():
            work_order = form.save(commit=False)
            work_order.created_by = request.user
            work_order.save()
            messages.success(request, f'Orden de trabajo "{work_order.title}" creada exitosamente.')
            return redirect('work_order_detail', pk=work_order.pk)
    else:
        form = WorkOrderForm()
    
    context = {
        'form': form,
        'page_title': 'Crear Orden de Trabajo'
    }
    
    return render(request, 'tickets/work_order_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_detail_view(request, pk):
    """Vista para ver detalles de una orden de trabajo"""
    work_order = get_object_or_404(WorkOrder, pk=pk)
    
    context = {
        'work_order': work_order,
        'attachments': work_order.attachments.all(),
        'page_title': f'Orden: {work_order.order_number}'
    }
    
    return render(request, 'tickets/work_order_detail.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_edit_view(request, pk):
    """Vista para editar órdenes de trabajo"""
    work_order = get_object_or_404(WorkOrder, pk=pk)
    
    # Verificar permisos
    if not work_order.can_edit(request.user):
        messages.error(request, 'No tienes permisos para editar esta orden de trabajo.')
        return redirect('work_order_detail', pk=work_order.pk)
    
    if request.method == 'POST':
        form = WorkOrderForm(request.POST, request.FILES, instance=work_order)
        if form.is_valid():
            form.save(commit=False)
            messages.success(request, f'Orden de trabajo "{work_order.title}" actualizada exitosamente.')
            return redirect('work_order_detail', pk=work_order.pk)
    else:
        form = WorkOrderForm(instance=work_order)
    
    context = {
        'form': form,
        'work_order': work_order,
        'page_title': f'Editar: {work_order.order_number}'
    }
    
    return render(request, 'tickets/work_order_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_delete_view(request, pk):
    """Vista para eliminar órdenes de trabajo"""
    work_order = get_object_or_404(WorkOrder, pk=pk)
    
    # Verificar permisos
    if not work_order.can_delete(request.user):
        messages.error(request, 'No puedes eliminar esta orden de trabajo. Solo se pueden eliminar órdenes en estado borrador.')
        return redirect('work_order_detail', pk=work_order.pk)
    
    if request.method == 'POST':
        title = work_order.title
        work_order.delete()
        messages.success(request, f'Orden de trabajo "{title}" eliminada exitosamente.')
        return redirect('work_order_list')
    
    context = {
        'work_order': work_order,
        'page_title': f'Eliminar: {work_order.order_number}'
    }
    
    return render(request, 'tickets/work_order_delete.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_change_status_view(request, pk):
    """Vista AJAX para cambiar el estado de una orden de trabajo"""
    if request.method == 'POST':
        work_order = get_object_or_404(WorkOrder, pk=pk)
        new_status = request.POST.get('status')
        
        if new_status in ['draft', 'accepted', 'finished']:
            old_status = work_order.get_status_display()
            work_order.status = new_status
            
            # Si se marca como terminada, establecer fecha de finalización
            if new_status == 'finished':
                work_order.completed_at = timezone.now()
            else:
                work_order.completed_at = None
            
            work_order.save()
            
            new_status_display = work_order.get_status_display()
            messages.success(request, f'Estado de la orden cambiado de "{old_status}" a "{new_status_display}".')
            
            return JsonResponse({
                'success': True,
                'message': f'Estado actualizado a "{new_status_display}"',
                'new_status': new_status,
                'new_status_display': new_status_display
            })
        
        return JsonResponse({
            'success': False,
            'error': 'Estado inválido'
        })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def work_order_public_view(request, token):
    """Vista pública para órdenes de trabajo compartidas"""
    work_order = get_object_or_404(WorkOrder, public_share_token=token, is_public=True)
    
    # Manejar aprobación por POST
    if request.method == 'POST' and 'approve_work_order' in request.POST:
        if work_order.status == 'draft':
            work_order.status = 'accepted'
            work_order.save()
            messages.success(request, '¡Orden de trabajo aprobada exitosamente!')
        else:
            messages.info(request, 'Esta orden ya ha sido procesada.')
        
        return redirect('tickets:public_work_order', token=token)
    
    context = {
        'work_order': work_order,
        'attachments': work_order.attachments.all(),
        'page_title': f'Orden: {work_order.order_number}',
        'is_public_view': True
    }
    
    return render(request, 'tickets/work_order_public.html', context)


# ===========================================
# VISTAS PARA REPORTES
# ===========================================

@login_required
@user_passes_test(is_agent, login_url='/')
def daily_report_view(request):
    """Vista para generar reporte de parte diario"""
    
    # Obtener parámetros de fecha
    fecha_desde = request.GET.get('fecha_desde', date.today().strftime('%Y-%m-%d'))
    fecha_hasta = request.GET.get('fecha_hasta', date.today().strftime('%Y-%m-%d'))
    usuario_id = request.GET.get('usuario')
    
    # Convertir fechas
    try:
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    except ValueError:
        fecha_desde_obj = date.today()
        fecha_hasta_obj = date.today()
        fecha_desde = fecha_desde_obj.strftime('%Y-%m-%d')
        fecha_hasta = fecha_hasta_obj.strftime('%Y-%m-%d')
    
    # Obtener registros de tiempo
    time_entries = TimeEntry.objects.filter(
        fecha_entrada__date__gte=fecha_desde_obj,
        fecha_entrada__date__lte=fecha_hasta_obj
    )
    
    # Filtrar por usuario si se especifica
    usuario_nombre = None
    if usuario_id:
        time_entries = time_entries.filter(user_id=usuario_id)
        try:
            usuario = User.objects.get(id=usuario_id)
            usuario_nombre = usuario.get_full_name() or usuario.username
        except User.DoesNotExist:
            pass
    
    time_entries = time_entries.select_related('user', 'project', 'ticket', 'work_order', 'task').order_by('-fecha_entrada')
    
    # Agregar paginación
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    paginator = Paginator(time_entries, 25)  # 25 registros por página
    page = request.GET.get('page')
    
    try:
        time_entries_page = paginator.page(page)
    except PageNotAnInteger:
        # Si page no es un entero, mostrar la primera página
        time_entries_page = paginator.page(1)
    except EmptyPage:
        # Si page está fuera de rango, mostrar la última página
        time_entries_page = paginator.page(paginator.num_pages)
    
    # Calcular estadísticas (usando el queryset original, no paginado)
    total_entries = time_entries.count()
    entries_in_progress = time_entries.filter(fecha_salida__isnull=True).count()
    
    # Calcular total de horas completadas usando duracion_trabajada
    total_minutos = 0
    for entry in time_entries:
        total_minutos += entry.duracion_trabajada
    
    # Convertir a horas con formato decimal
    total_hours = round(total_minutos / 60, 1)
    
    # Convertir total de minutos a formato HH:MM
    total_horas_formateado = f"{total_minutos // 60:02d}:{total_minutos % 60:02d}"
    
    # Obtener usuarios únicos
    unique_users = time_entries.values('user').distinct().count()
    
    # Indicadores CRM para la cabecera del reporte
    from .models import Contact, Opportunity, Company, Ticket
    
    # Obtener conteos de elementos creados en el rango de fechas
    contactos_creados = Contact.objects.filter(
        created_at__date__gte=fecha_desde_obj,
        created_at__date__lte=fecha_hasta_obj
    ).count()
    
    oportunidades_creadas = Opportunity.objects.filter(
        created_at__date__gte=fecha_desde_obj,
        created_at__date__lte=fecha_hasta_obj
    ).count()
    
    tickets_creados = Ticket.objects.filter(
        created_at__date__gte=fecha_desde_obj,
        created_at__date__lte=fecha_hasta_obj
    ).count()
    
    empresas_creadas = Company.objects.filter(
        created_at__date__gte=fecha_desde_obj,
        created_at__date__lte=fecha_hasta_obj
    ).count()
    
    # Obtener usuarios para el filtro
    usuarios = User.objects.filter(is_active=True).order_by('first_name', 'last_name', 'username')
    
    context = {
        'time_entries': time_entries,
        'usuarios': usuarios,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'usuario_seleccionado': usuario_id,
        'usuario_nombre': usuario_nombre,
        'page_title': 'Reporte de Parte Diario',
        'total_entries': total_entries,
        'entries_in_progress': entries_in_progress,
        'total_hours': total_hours,
        'total_horas_formateado': total_horas_formateado,
        'unique_users': unique_users,
        # Indicadores CRM
        'contactos_creados': contactos_creados,
        'oportunidades_creadas': oportunidades_creadas,
        'tickets_creados': tickets_creados,
        'empresas_creadas': empresas_creadas,
    }
    
    return render(request, 'tickets/daily_report.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def daily_report_pdf(request):
    """Vista para generar PDF del reporte diario"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    import io
    
    # Obtener parámetros
    fecha_desde = request.GET.get('fecha_desde', date.today().strftime('%Y-%m-%d'))
    fecha_hasta = request.GET.get('fecha_hasta', date.today().strftime('%Y-%m-%d'))
    usuario_id = request.GET.get('usuario')
    
    # Convertir fechas
    try:
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    except ValueError:
        fecha_desde_obj = date.today()
        fecha_hasta_obj = date.today()
    
    # Obtener datos
    time_entries = TimeEntry.objects.filter(
        fecha_entrada__date__gte=fecha_desde_obj,
        fecha_entrada__date__lte=fecha_hasta_obj
    )
    
    if usuario_id:
        time_entries = time_entries.filter(user_id=usuario_id)
        usuario = User.objects.get(id=usuario_id)
        usuario_nombre = f" - {usuario.get_full_name() or usuario.username}"
    else:
        usuario_nombre = " - Todos los usuarios"
    
    time_entries = time_entries.select_related('user', 'project', 'ticket', 'work_order', 'task').order_by('user__username', 'fecha_entrada')
    
    # Crear PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    # Contenido del PDF
    story = []
    
    # Título
    title = f"Reporte de Parte Diario{usuario_nombre}"
    story.append(Paragraph(title, title_style))
    
    # Rango de fechas
    if fecha_desde_obj == fecha_hasta_obj:
        periodo = f"Fecha: {fecha_desde_obj.strftime('%d/%m/%Y')}"
    else:
        periodo = f"Período: {fecha_desde_obj.strftime('%d/%m/%Y')} al {fecha_hasta_obj.strftime('%d/%m/%Y')}"
    
    story.append(Paragraph(periodo, styles['Normal']))
    story.append(Spacer(1, 12))
    
    # Indicadores CRM para la cabecera del reporte
    from .models import Contact, Opportunity, Company, Ticket
    
    # Obtener conteos de elementos creados en el rango de fechas
    contactos_creados = Contact.objects.filter(
        created_at__date__gte=fecha_desde_obj,
        created_at__date__lte=fecha_hasta_obj
    ).count()
    
    oportunidades_creadas = Opportunity.objects.filter(
        created_at__date__gte=fecha_desde_obj,
        created_at__date__lte=fecha_hasta_obj
    ).count()
    
    tickets_creados = Ticket.objects.filter(
        created_at__date__gte=fecha_desde_obj,
        created_at__date__lte=fecha_hasta_obj
    ).count()
    
    empresas_creadas = Company.objects.filter(
        created_at__date__gte=fecha_desde_obj,
        created_at__date__lte=fecha_hasta_obj
    ).count()
    
    # Agregar indicadores CRM como una tabla
    indicadores_data = [
        ['INDICADORES CRM', ''],
        ['Total de Contactos Creados:', str(contactos_creados)],
        ['Total de Oportunidades Creadas:', str(oportunidades_creadas)],
        ['Total de Tickets Creados:', str(tickets_creados)],
        ['Total de Empresas Creadas:', str(empresas_creadas)],
    ]
    
    indicadores_table = Table(indicadores_data, colWidths=[3*inch, 1*inch])
    indicadores_table.setStyle(TableStyle([
        # Estilo para la cabecera
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Estilo para el contenido
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(indicadores_table)
    story.append(Spacer(1, 20))
    
    # Tabla de datos
    data = [['Usuario', 'Fecha', 'Asignación', 'Entrada', 'Salida', 'Duración']]
    
    total_minutos = 0  # Variable para acumular el total de minutos
    
    for entry in time_entries:
        # Construir asignación
        asignacion_parts = []
        if entry.project:
            asignacion_parts.append(f"Proyecto: {entry.project.name}")
        if entry.ticket:
            asignacion_parts.append(f"Ticket: #{entry.ticket.id}")
        if entry.work_order:
            asignacion_parts.append(f"Orden: #{entry.work_order.id}")
        if entry.task:
            asignacion_parts.append(f"Tarea: {entry.task.title}")
        
        asignacion = "\n".join(asignacion_parts) if asignacion_parts else "Sin asignación específica"
        
        # Formatear salida
        salida = entry.fecha_salida.strftime('%H:%M') if entry.fecha_salida else "En progreso"
        
        # Acumular minutos trabajados
        total_minutos += entry.duracion_trabajada
        
        data.append([
            entry.user.get_full_name() or entry.user.username,
            entry.fecha_entrada.strftime('%d/%m/%Y'),
            asignacion,
            entry.fecha_entrada.strftime('%H:%M'),
            salida,
            entry.duracion_formateada
        ])
    
    # Agregar fila de totales
    if time_entries.exists():
        # Convertir total de minutos a formato HH:MM
        total_horas = total_minutos // 60
        total_mins = total_minutos % 60
        total_formateado = f"{total_horas:02d}:{total_mins:02d}"
        
        data.append(['', '', '', '', 'TOTAL HORAS:', total_formateado])
    
    # Crear tabla con anchos ajustados para evitar cortes
    table = Table(data, colWidths=[1.3*inch, 0.9*inch, 1.8*inch, 0.7*inch, 1*inch, 1*inch])
    
    # Estilos base de la tabla
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    
    # Si hay datos y por tanto fila de totales, agregar estilo especial para la última fila
    if len(data) > 1:  # Header + al menos una fila de datos
        last_row = len(data) - 1
        table_style.extend([
            ('BACKGROUND', (0, last_row), (-1, last_row), colors.lightgrey),
            ('FONTNAME', (0, last_row), (-1, last_row), 'Helvetica-Bold'),
            ('FONTSIZE', (0, last_row), (-1, last_row), 10),
            ('TEXTCOLOR', (0, last_row), (-1, last_row), colors.black),
        ])
    
    table.setStyle(TableStyle(table_style))
    
    story.append(table)
    
    # Construir PDF
    doc.build(story)
    
    # Respuesta HTTP
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"parte_diario_{fecha_desde}_{fecha_hasta}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ==================== VISTAS DE TAREAS ====================

@login_required
def task_list_view(request):
    """Vista para listar tareas"""
    tasks = Task.objects.all()
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    assigned_to_me = request.GET.get('assigned_to_me', '')
    
    if search:
        tasks = tasks.filter(
            Q(title__icontains=search) | 
            Q(description__icontains=search)
        )
    
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    
    if priority_filter:
        tasks = tasks.filter(priority=priority_filter)
    
    if assigned_to_me:
        tasks = tasks.filter(assigned_users=request.user)
    
    # Paginación
    paginator = Paginator(tasks, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'assigned_to_me': assigned_to_me,
        'status_choices': Task.STATUS_CHOICES,
        'priority_choices': Task.PRIORITY_CHOICES,
    }
    return render(request, 'tickets/task_list.html', context)


@login_required
def task_create_view(request):
    """Vista para crear una nueva tarea"""
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.created_by = request.user
            task.save()
            form.save_m2m()  # Para guardar los ManyToMany
            
            messages.success(request, f'Tarea "{task.title}" creada exitosamente.')
            return redirect('task_detail', pk=task.pk)
    else:
        form = TaskForm()
    
    context = {
        'form': form,
        'title': 'Crear Nueva Tarea'
    }
    return render(request, 'tickets/task_form.html', context)


@login_required
def task_detail_view(request, pk):
    """Vista para ver detalles de una tarea"""
    task = get_object_or_404(Task, pk=pk)
    
    context = {
        'task': task,
    }
    return render(request, 'tickets/task_detail.html', context)


@login_required
def task_edit_view(request, pk):
    """Vista para editar una tarea"""
    task = get_object_or_404(Task, pk=pk)
    
    # Solo el creador o agentes pueden editar
    if not is_agent(request.user) and task.created_by != request.user:
        messages.error(request, 'No tienes permisos para editar esta tarea.')
        return redirect('task_detail', pk=task.pk)
    
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            task = form.save(commit=False)
            task.save()
            form.save_m2m()  # Para guardar los ManyToMany
            messages.success(request, f'Tarea "{task.title}" actualizada exitosamente.')
            return redirect('task_detail', pk=task.pk)
    else:
        form = TaskForm(instance=task)
    
    context = {
        'form': form,
        'task': task,
        'title': f'Editar Tarea: {task.title}'
    }
    return render(request, 'tickets/task_form.html', context)


@login_required
def task_delete_view(request, pk):
    """Vista para eliminar una tarea"""
    task = get_object_or_404(Task, pk=pk)
    
    # Solo el creador o agentes pueden eliminar
    if not is_agent(request.user) and task.created_by != request.user:
        messages.error(request, 'No tienes permisos para eliminar esta tarea.')
        return redirect('task_detail', pk=task.pk)
    
    if request.method == 'POST':
        task_title = task.title
        task.delete()
        messages.success(request, f'Tarea "{task_title}" eliminada exitosamente.')
        return redirect('task_list')
    
    context = {
        'task': task,
    }
    return render(request, 'tickets/task_delete.html', context)


@login_required
def task_toggle_status_view(request, pk):
    """Vista para cambiar el estado de una tarea"""
    task = get_object_or_404(Task, pk=pk)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Task.STATUS_CHOICES):
            old_status = task.get_status_display()
            task.status = new_status
            
            if new_status == 'completed':
                task.completed_at = timezone.now()
            elif task.status != 'completed' and task.completed_at:
                task.completed_at = None
            
            task.save()
            
            messages.success(request, f'Estado de la tarea "{task.title}" cambiado de {old_status} a {task.get_status_display()}.')
    
    return redirect('task_detail', pk=task.pk)


# ===========================================
# VISTAS PARA GESTIÓN DIARIA DE TAREAS
# ===========================================

def auto_load_daily_tasks(session, user, date):
    """
    Carga automáticamente tareas relevantes para el día en una nueva sesión.
    Prioriza:
    1. Tareas con fecha límite de hoy
    2. Tareas con alta prioridad sin fecha límite
    3. Tareas en progreso del usuario
    """
    # Tareas con fecha límite de hoy (máxima prioridad)
    tasks_due_today = Task.objects.filter(
        assigned_users=user,
        due_date=date,
        status__in=['pending', 'in_progress']
    ).exclude(
        status__in=['completed', 'cancelled']
    )
    
    # Tareas de alta prioridad sin fecha límite
    high_priority_tasks = Task.objects.filter(
        assigned_users=user,
        priority='high',
        due_date__isnull=True,
        status__in=['pending', 'in_progress']
    ).exclude(
        status__in=['completed', 'cancelled']
    )[:3]  # Máximo 3 tareas de alta prioridad
    
    # Tareas actualmente en progreso del usuario
    in_progress_tasks = Task.objects.filter(
        assigned_users=user,
        status='in_progress'
    ).exclude(
        status__in=['completed', 'cancelled']
    )
    
    # Combinar y eliminar duplicados
    auto_tasks = list(tasks_due_today) + list(high_priority_tasks) + list(in_progress_tasks)
    unique_tasks = []
    seen_ids = set()
    
    for task in auto_tasks:
        if task.id not in seen_ids:
            unique_tasks.append(task)
            seen_ids.add(task.id)
    
    # Crear items de la sesión diaria para estas tareas
    for order, task in enumerate(unique_tasks[:8], 1):  # Máximo 8 tareas auto-cargadas
        DailyTaskItem.objects.create(
            session=session,
            task=task,
            order=order
        )


@login_required
def daily_task_management(request):
    """Vista principal - Lista de gestiones de tareas de todos los usuarios"""
    # Obtener todas las gestiones de tareas ordenadas por fecha
    # Si el usuario es superuser o staff, ve todas las gestiones
    # Si no, solo ve las suyas
    if request.user.is_superuser or request.user.is_staff:
        task_sessions = DailyTaskSession.objects.all()
    else:
        task_sessions = DailyTaskSession.objects.filter(user=request.user)
    
    task_sessions = task_sessions.prefetch_related(
        'daily_task_items__task', 'user'
    ).order_by('-created_at')
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(task_sessions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'task_sessions': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
    }
    
    return render(request, 'tickets/daily_task_management.html', context)


@login_required
def create_task_session(request):
    """Crear nueva gestión de tareas"""
    if request.method == 'POST':
        # Obtener la fecha del formulario
        session_date = request.POST.get('date')
        notes = request.POST.get('notes', '')
        
        if session_date:
            # Crear nueva sesión
            session = DailyTaskSession.objects.create(
                user=request.user,
                date=session_date,
                notes=notes
            )
            
            # Cargar TODAS las tareas donde el usuario esté involucrado (sin filtros de fecha)
            # Incluir: asignadas al usuario O creadas por el usuario
            assigned_tasks = Task.objects.filter(
                Q(assigned_users=request.user) | Q(created_by=request.user)
            ).exclude(
                status__in=['cancelled']  # Solo excluir canceladas, no completadas
            ).order_by('priority', '-created_at').distinct()
            
            print(f"DEBUG: Usuario {request.user.username}")
            print(f"DEBUG: Tareas totales encontradas: {assigned_tasks.count()}")
            
            # Debug detallado
            all_tasks = Task.objects.all()
            print(f"DEBUG: Total de tareas en el sistema: {all_tasks.count()}")
            
            user_assigned = Task.objects.filter(assigned_users=request.user)
            user_created = Task.objects.filter(created_by=request.user)
            print(f"DEBUG: Tareas asignadas al usuario: {user_assigned.count()}")
            print(f"DEBUG: Tareas creadas por el usuario: {user_created.count()}")
            
            for task in assigned_tasks[:5]:  # Mostrar primeras 5 para debug
                assigned_users = [u.username for u in task.assigned_users.all()]
                print(f"  - {task.title} | Estado: {task.status} | Asignada a: {assigned_users} | Creada por: {task.created_by.username if task.created_by else 'N/A'}")
            
            # Crear items para todas las tareas asignadas
            for order, task in enumerate(assigned_tasks, 1):
                DailyTaskItem.objects.create(
                    session=session,
                    task=task,
                    order=order
                )
            
            messages.success(
                request, 
                f'Nueva gestión de tareas creada con {assigned_tasks.count()} tareas cargadas.'
            )
            return redirect('view_task_session', session_id=session.id)
        else:
            messages.error(request, 'Debes seleccionar una fecha.')
    
    # Pasar la fecha de hoy al template
    today = timezone.now().date()
    return render(request, 'tickets/create_task_session.html', {'today': today})


@login_required
def view_task_session(request, session_id):
    """Ver y gestionar una sesión específica de tareas (checklist)"""
    try:
        session = DailyTaskSession.objects.get(
            id=session_id,
            user=request.user
        )
    except DailyTaskSession.DoesNotExist:
        messages.error(request, 'Gestión de tareas no encontrada.')
        return redirect('daily_task_management')
    
    # Obtener items de la sesión
    session_items = session.daily_task_items.select_related('task').order_by('order')
    
    context = {
        'session': session,
        'session_items': session_items,
    }
    
    return render(request, 'tickets/view_task_session.html', context)


@login_required
def delete_task_session(request, session_id):
    """Eliminar una gestión de tareas"""
    try:
        session = DailyTaskSession.objects.get(id=session_id)
        
        # Verificar permisos: solo el propietario o superusuario puede eliminar
        if session.user != request.user and not request.user.is_superuser:
            messages.error(request, 'No tienes permisos para eliminar esta gestión.')
            return redirect('daily_task_management')
        
        if request.method == 'POST':
            session_date = session.date.strftime('%d/%m/%Y')
            session_user = session.user.get_full_name() or session.user.username
            session.delete()
            messages.success(request, f'Gestión de tareas del {session_date} de {session_user} eliminada correctamente.')
            return redirect('daily_task_management')
        
        context = {
            'session': session,
        }
        
        return render(request, 'tickets/delete_task_session.html', context)
        
    except DailyTaskSession.DoesNotExist:
        messages.error(request, 'Gestión de tareas no encontrada.')
        return redirect('daily_task_management')


@login_required
def add_task_to_daily_session(request):
    """Agregar una tarea a la sesión diaria actual"""
    if request.method == 'POST':
        task_id = request.POST.get('task_id')
        today = timezone.now().date()
        
        try:
            task = Task.objects.get(id=task_id, assigned_users=request.user)
            session, created = DailyTaskSession.objects.get_or_create(
                user=request.user,
                date=today
            )
            
            # Verificar que la tarea no esté ya en la sesión
            if not DailyTaskItem.objects.filter(session=session, task=task).exists():
                # Obtener el siguiente orden
                max_order = session.daily_task_items.aggregate(
                    max_order=models.Max('order')
                )['max_order'] or 0
                
                DailyTaskItem.objects.create(
                    session=session,
                    task=task,
                    order=max_order + 1
                )
                
                messages.success(request, f'Tarea "{task.title}" agregada a tu plan del día.')
            else:
                messages.warning(request, 'Esta tarea ya está en tu plan del día.')
                
        except Task.DoesNotExist:
            messages.error(request, 'Tarea no encontrada o no tienes permisos.')
    
    return redirect('daily_task_management')


@login_required
def toggle_daily_task_completion(request, item_id):
    """Marcar/desmarcar una tarea como completada en la sesión"""
    try:
        item = DailyTaskItem.objects.get(
            id=item_id,
            session__user=request.user
        )
        
        if item.completed:
            item.mark_pending()
            messages.info(request, f'Tarea "{item.task.title}" marcada como pendiente.')
        else:
            item.mark_completed()
            messages.success(request, f'¡Excelente! Tarea "{item.task.title}" completada.')
        
        # Redirigir a la vista de la sesión específica
        return redirect('view_task_session', session_id=item.session.id)
        
    except DailyTaskItem.DoesNotExist:
        messages.error(request, 'Item no encontrado.')
        return redirect('daily_task_management')


@login_required
def remove_task_from_daily_session(request, item_id):
    """Remover una tarea de la sesión diaria"""
    try:
        item = DailyTaskItem.objects.get(
            id=item_id,
            session__user=request.user
        )
        task_title = item.task.title
        item.delete()
        messages.info(request, f'Tarea "{task_title}" removida de tu plan del día.')
        
    except DailyTaskItem.DoesNotExist:
        messages.error(request, 'Item no encontrado.')
    
    return redirect('daily_task_management')


@login_required
def update_daily_session_notes(request):
    """Actualizar las notas de la sesión diaria"""
    if request.method == 'POST':
        today = timezone.now().date()
        notes = request.POST.get('notes', '')
        
        session, created = DailyTaskSession.objects.get_or_create(
            user=request.user,
            date=today
        )
        
        session.notes = notes
        session.save()
        
        messages.success(request, 'Notas del día actualizadas.')
    
    return redirect('daily_task_management')


@login_required
def daily_task_history(request):
    """Vista del historial de sesiones diarias"""
    sessions = DailyTaskSession.objects.filter(
        user=request.user
    ).prefetch_related('daily_task_items__task').order_by('-date')
    
    # Aplicar filtros
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    min_completion = request.GET.get('min_completion')
    
    if date_from:
        sessions = sessions.filter(date__gte=date_from)
    if date_to:
        sessions = sessions.filter(date__lte=date_to)
    
    # Calcular estadísticas de todas las sesiones (sin filtros)
    all_sessions = DailyTaskSession.objects.filter(user=request.user)
    total_sessions = all_sessions.count()
    total_tasks = sum(session.get_total_tasks() for session in all_sessions)
    total_completed_tasks = sum(session.get_completed_tasks() for session in all_sessions)
    
    if total_sessions > 0:
        average_completion = sum(session.get_completion_percentage() for session in all_sessions) / total_sessions
    else:
        average_completion = 0
    
    # Filtrar por porcentaje mínimo de finalización
    if min_completion:
        filtered_sessions = []
        for session in sessions:
            if session.get_completion_percentage() >= int(min_completion):
                filtered_sessions.append(session)
        sessions = filtered_sessions
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(sessions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'sessions': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'total_sessions': total_sessions,
        'total_tasks': total_tasks,
        'total_completed_tasks': total_completed_tasks,
        'average_completion': round(average_completion, 1),
    }
    
    return render(request, 'tickets/daily_task_history.html', context)


@login_required
def refresh_daily_tasks(request):
    """Recargar tareas relevantes del día en la sesión actual"""
    today = timezone.now().date()
    
    try:
        session = DailyTaskSession.objects.get(
            user=request.user,
            date=today
        )
        
        # Obtener tareas ya en la sesión para no duplicar
        existing_task_ids = session.daily_task_items.values_list('task_id', flat=True)
        
        # Tareas con fecha límite de hoy que no están en la sesión
        tasks_due_today = Task.objects.filter(
            assigned_users=request.user,
            due_date=today,
            status__in=['pending', 'in_progress']
        ).exclude(
            id__in=existing_task_ids
        ).exclude(
            status__in=['completed', 'cancelled']
        )
        
        # Tareas de alta prioridad sin fecha que no están en la sesión
        high_priority_tasks = Task.objects.filter(
            assigned_users=request.user,
            priority='high',
            due_date__isnull=True,
            status__in=['pending', 'in_progress']
        ).exclude(
            id__in=existing_task_ids
        ).exclude(
            status__in=['completed', 'cancelled']
        )[:2]  # Máximo 2 adicionales
        
        # Obtener el siguiente orden disponible
        max_order = session.daily_task_items.aggregate(
            max_order=models.Max('order')
        )['max_order'] or 0
        
        # Agregar nuevas tareas
        new_tasks = list(tasks_due_today) + list(high_priority_tasks)
        tasks_added = 0
        
        for task in new_tasks:
            max_order += 1
            DailyTaskItem.objects.create(
                session=session,
                task=task,
                order=max_order
            )
            tasks_added += 1
        
        if tasks_added > 0:
            messages.success(
                request, 
                f'Se agregaron {tasks_added} tarea(s) relevante(s) a tu plan del día.'
            )
        else:
            messages.info(
                request, 
                'No hay nuevas tareas relevantes para agregar hoy.'
            )
            
    except DailyTaskSession.DoesNotExist:
        messages.error(request, 'No tienes una sesión activa para hoy.')
    
    return redirect('daily_task_management')


# ===========================================
# VISTAS PARA CONTROL DIARIO DE TAREAS (MANTENEMOS LAS ANTERIORES)
# ===========================================

@login_required
def task_control_dashboard(request):
    """Dashboard diario de control de tareas para agentes"""
    # Obtener solo las tareas asignadas al usuario actual
    user_tasks = Task.objects.filter(
        assigned_users=request.user
    ).exclude(
        status__in=['cancelled']
    ).prefetch_related('assigned_users', 'created_by').order_by(
        'status', 'priority', 'due_date'
    )
    
    # Separar tareas por estado
    pending_tasks = user_tasks.filter(status='pending')
    in_progress_tasks = user_tasks.filter(status='in_progress')
    completed_tasks = user_tasks.filter(status='completed')
    
    # Estadísticas del día
    today = timezone.now().date()
    today_completed = user_tasks.filter(
        status='completed',
        completed_at__date=today
    ).count()
    
    overdue_tasks = [task for task in user_tasks if task.is_overdue()]
    
    context = {
        'pending_tasks': pending_tasks,
        'in_progress_tasks': in_progress_tasks,
        'completed_tasks': completed_tasks,
        'overdue_tasks': overdue_tasks,
        'today_completed': today_completed,
        'total_tasks': user_tasks.count(),
        'total_pending': pending_tasks.count(),
        'total_in_progress': in_progress_tasks.count(),
        'total_completed': completed_tasks.count(),
    }
    
    return render(request, 'tickets/task_control_dashboard.html', context)


@login_required
def task_control_complete(request, pk):
    """Marcar una tarea como completada desde el dashboard de control"""
    task = get_object_or_404(Task, pk=pk)
    
    # Verificar que el usuario esté asignado a la tarea
    if request.user not in task.assigned_users.all():
        messages.error(request, 'No tienes permisos para completar esta tarea.')
        return redirect('task_control_dashboard')
    
    if request.method == 'POST':
        old_status = task.get_status_display()
        task.mark_as_completed()
        
        messages.success(
            request,
            f'¡Excelente! Has completado la tarea "{task.title}".'
        )
    
    return redirect('task_control_dashboard')


@login_required
def task_control_start(request, pk):
    """Marcar una tarea como en progreso desde el dashboard de control"""
    task = get_object_or_404(Task, pk=pk)
    
    # Verificar que el usuario esté asignado a la tarea
    if request.user not in task.assigned_users.all():
        messages.error(request, 'No tienes permisos para modificar esta tarea.')
        return redirect('task_control_dashboard')
    
    if request.method == 'POST':
        if task.status == 'pending':
            task.status = 'in_progress'
            task.save()
            
            messages.success(
                request,
                f'Has comenzado a trabajar en la tarea "{task.title}".'
            )
        else:
            messages.warning(request, 'Esta tarea ya está en progreso o completada.')
    
    return redirect('task_control_dashboard')


@login_required
def task_control_pause(request, pk):
    """Pausar una tarea (marcar como pendiente) desde el dashboard de control"""
    task = get_object_or_404(Task, pk=pk)
    
    # Verificar que el usuario esté asignado a la tarea
    if request.user not in task.assigned_users.all():
        messages.error(request, 'No tienes permisos para modificar esta tarea.')
        return redirect('task_control_dashboard')
    
    if request.method == 'POST':
        if task.status == 'in_progress':
            task.status = 'pending'
            task.save()
            
            messages.info(
                request,
                f'Has pausado la tarea "{task.title}". Puedes retomarla cuando quieras.'
            )
        else:
            messages.warning(request, 'Solo puedes pausar tareas que estén en progreso.')
    
    return redirect('task_control_dashboard')


def public_task_view(request, token):
    """Vista pública para gestionar tareas sin autenticación"""
    try:
        user_profile = UserProfile.objects.get(public_token=token)
        user = user_profile.user
    except UserProfile.DoesNotExist:
        return render(request, 'tickets/public_tasks_error.html', {
            'error': 'Token no válido'
        })
    
    if request.method == 'POST':
        # Obtener las tareas seleccionadas (completadas)
        selected_task_ids = request.POST.getlist('selected_tasks')
        selected_task_ids = [int(task_id) for task_id in selected_task_ids]
        
        # Obtener TODAS las tareas asignadas al usuario (para crear la gestión completa)
        all_user_tasks = Task.objects.filter(
            Q(assigned_users=user) | Q(created_by=user)
        ).exclude(status__in=['cancelled']).distinct()
        
        if all_user_tasks.exists():
            # Crear nueva sesión con fecha de hoy
            from django.utils import timezone
            session = DailyTaskSession.objects.create(
                user=user,
                date=timezone.now().date(),
                notes='Gestión creada desde acceso público'
            )
            
            # Crear un item para CADA tarea asignada
            completed_count = 0
            order_counter = 1  # Contador para el orden
            for task in all_user_tasks:
                is_completed = task.id in selected_task_ids
                
                DailyTaskItem.objects.create(
                    session=session,
                    task=task,
                    order=order_counter,
                    completed=is_completed,
                    completed_at=timezone.now() if is_completed else None
                )
                
                order_counter += 1
                
                # Si la tarea fue marcada como completada, actualizar su estado
                if is_completed and task.status != 'completed':
                    task.status = 'completed'
                    task.save()
                    completed_count += 1
            
            return render(request, 'tickets/public_task_success.html', {
                'session': session,
                'user': user,
                'token': token,
                'task_count': all_user_tasks.count(),
                'completed_count': completed_count
            })
        else:
            error_message = 'No tienes tareas asignadas para crear una gestión'
    else:
        error_message = None
    
    # Obtener todas las tareas asignadas al usuario (incluyendo completadas)
    tasks = Task.objects.filter(
        Q(assigned_users=user) | Q(created_by=user)
    ).exclude(status='cancelled').distinct().order_by('-created_at')
    
    context = {
        'user': user,
        'tasks': tasks,
        'token': token,
        'error_message': error_message,
    }
    
    return render(request, 'tickets/public_tasks_simple.html', context)


def public_create_task_session(request, token):
    """Redirigir a la vista principal simplificada"""
    return redirect('public_task_view', token=token)


def public_project_view(request, token):
    """Vista pública para mostrar información del proyecto sin autenticación"""
    try:
        project = get_object_or_404(Project, public_share_token=token)
    except Project.DoesNotExist:
        return render(request, 'tickets/public_project_error.html', {
            'error': 'El enlace proporcionado no es válido o ha expirado.'
        })
    
    # Calcular estadísticas del proyecto usando los métodos del modelo
    total_hours = project.get_total_hours()
    tickets_count = project.get_tickets_count()
    resolved_tickets_count = project.get_resolved_tickets_count()
    completed_work_orders_count = project.get_completed_work_orders_count()
    
    # Obtener trabajadores del proyecto con sus datos de contacto
    workers = project.get_project_workers()
    
    # Crear lista de trabajadores con información de contacto
    workers_info = []
    for worker in workers:
        worker_info = {
            'name': f"{worker.first_name} {worker.last_name}".strip() or worker.username,
            'email': worker.email,
            'phone': getattr(worker.profile, 'phone', '') if hasattr(worker, 'profile') else '',
        }
        workers_info.append(worker_info)
    
    # Obtener tickets directamente asignados al proyecto para mostrar
    project_tickets = project.tickets.all().order_by('-created_at')[:10]
    
    # Obtener órdenes de trabajo directamente asignadas al proyecto para mostrar
    project_work_orders = project.work_orders.all().order_by('-created_at')[:10]
    
    context = {
        'project': project,
        'total_hours': total_hours,
        'tickets_count': tickets_count,
        'resolved_tickets_count': resolved_tickets_count,
        'completed_work_orders_count': completed_work_orders_count,
        'workers_info': workers_info,
        'project_tickets': project_tickets,
        'project_work_orders': project_work_orders,
    }
    
    return render(request, 'tickets/public_project.html', context)


# ===== VISTAS DEL CHAT =====

@login_required
def chat_list_view(request):
    """Vista para listar las salas de chat del usuario"""
    user_rooms = request.user.chat_rooms.all().order_by('-last_activity')
    
    # Agregar conteo de mensajes no leídos a cada sala
    for room in user_rooms:
        room.unread_count = room.get_unread_count_for_user(request.user)
    
    context = {
        'rooms': user_rooms,
        'page_title': 'Chat - Mensajes',
    }
    
    return render(request, 'tickets/chat_list.html', context)


@login_required
def chat_room_view(request, room_id):
    """Vista para una sala de chat específica"""
    room = get_object_or_404(ChatRoom, id=room_id, participants=request.user)
    
    # Marcar mensajes como leídos
    room.mark_as_read_for_user(request.user)
    
    # Obtener mensajes de la sala
    messages = room.messages.all().order_by('created_at')
    
    # Procesar envío de mensaje
    if request.method == 'POST':
        form = ChatMessageForm(request.POST, request.FILES)
        if form.is_valid():
            message = form.save(commit=False)
            message.room = room
            message.sender = request.user
            message.save()
            
            # Actualizar última actividad de la sala
            room.last_activity = timezone.now()
            room.save()
            
            # Redirigir para evitar reenvío
            return redirect('chat_room', room_id=room.id)
    else:
        form = ChatMessageForm()
    
    context = {
        'room': room,
        'messages': messages,
        'form': form,
        'page_title': f'Chat - {room}',
    }
    
    return render(request, 'tickets/chat_room.html', context)


@login_required
def chat_start_conversation(request, user_id):
    """Inicia una conversación 1:1 con otro usuario"""
    other_user = get_object_or_404(User, id=user_id, is_active=True)
    
    # Verificar que no es el mismo usuario
    if other_user == request.user:
        messages.error(request, 'No puedes iniciar una conversación contigo mismo.')
        return redirect('chat_list')
    
    # Buscar si ya existe una sala 1:1 entre estos usuarios
    existing_room = ChatRoom.objects.filter(
        is_group=False,
        participants=request.user
    ).filter(participants=other_user).first()
    
    if existing_room:
        return redirect('chat_room', room_id=existing_room.id)
    
    # Crear nueva sala 1:1
    room = ChatRoom.objects.create(is_group=False)
    room.participants.add(request.user, other_user)
    
    return redirect('chat_room', room_id=room.id)


@login_required
def chat_create_group(request):
    """Vista para crear un chat grupal"""
    if request.method == 'POST':
        form = ChatRoomForm(request.POST, user=request.user)
        if form.is_valid():
            room = form.save()
            messages.success(request, f'Grupo "{room}" creado exitosamente.')
            return redirect('chat_room', room_id=room.id)
    else:
        form = ChatRoomForm(user=request.user)
    
    context = {
        'form': form,
        'page_title': 'Crear Grupo de Chat',
    }
    
    return render(request, 'tickets/chat_create_group.html', context)


@login_required
def chat_users_list(request):
    """Vista para listar usuarios disponibles para chat"""
    users = User.objects.filter(is_active=True).exclude(id=request.user.id).order_by(
        'first_name', 'last_name', 'username'
    )
    
    context = {
        'users': users,
        'page_title': 'Usuarios - Chat',
    }
    
    return render(request, 'tickets/chat_users.html', context)


@login_required
def chat_ajax_load_messages(request, room_id):
    """Vista AJAX para cargar mensajes de una sala"""
    room = get_object_or_404(ChatRoom, id=room_id, participants=request.user)
    
    # Obtener mensajes después de cierto timestamp (para polling)
    last_message_id = request.GET.get('last_message_id', 0)
    messages = room.messages.filter(id__gt=last_message_id).order_by('created_at')
    
    messages_data = []
    for message in messages:
        messages_data.append({
            'id': message.id,
            'sender_name': message.sender.get_full_name() or message.sender.username,
            'sender_id': message.sender.id,
            'message': message.message,
            'attachment_name': message.get_attachment_name(),
            'attachment_url': message.attachment.url if message.attachment else None,
            'attachment_size': message.get_attachment_size(),
            'created_at': message.created_at.strftime('%H:%M'),
            'is_own': message.sender == request.user,
        })
    
    return JsonResponse({
        'messages': messages_data,
        'room_id': room.id,
    })


@login_required
def chat_ajax_send_message(request, room_id):
    """Vista AJAX para enviar mensajes"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    room = get_object_or_404(ChatRoom, id=room_id, participants=request.user)
    
    form = ChatMessageForm(request.POST, request.FILES)
    if form.is_valid():
        message = form.save(commit=False)
        message.room = room
        message.sender = request.user
        message.save()
        
        # Actualizar última actividad
        room.last_activity = timezone.now()
        room.save()
        
        return JsonResponse({
            'success': True,
            'message': {
                'id': message.id,
                'sender_name': message.sender.get_full_name() or message.sender.username,
                'sender_id': message.sender.id,
                'message': message.message,
                'attachment_name': message.get_attachment_name(),
                'attachment_url': message.attachment.url if message.attachment else None,
                'attachment_size': message.get_attachment_size(),
                'created_at': message.created_at.strftime('%H:%M'),
                'is_own': True,
            }
        })
    else:
        return JsonResponse({
            'error': 'Formulario inválido',
            'errors': form.errors
        }, status=400)


@login_required
@user_passes_test(is_agent)
def pdf_get_pages_view(request):
    """Vista para extraer páginas específicas de un PDF"""
    if request.method == 'POST':
        if 'pdf_file' not in request.FILES:
            messages.error(request, 'No se ha seleccionado ningún archivo PDF.')
            return render(request, 'tickets/pdf_get_pages.html')
        
        pdf_file = request.FILES['pdf_file']
        pages_input = request.POST.get('pages', '').strip()
        
        # Validar que el archivo sea PDF
        if not pdf_file.name.lower().endswith('.pdf'):
            messages.error(request, 'El archivo debe ser un PDF.')
            return render(request, 'tickets/pdf_get_pages.html')
        
        # Validar entrada de páginas
        if not pages_input:
            messages.error(request, 'Debe especificar las páginas a extraer.')
            return render(request, 'tickets/pdf_get_pages.html')
        
        try:
            # Importar PyPDF2
            from PyPDF2 import PdfReader, PdfWriter
            import io
            
            # Leer el PDF
            pdf_reader = PdfReader(pdf_file)
            total_pages = len(pdf_reader.pages)
            
            # Procesar la entrada de páginas (ej: "1,3,5-7")
            pages_to_extract = []
            for part in pages_input.split(','):
                part = part.strip()
                if '-' in part:
                    # Rango de páginas
                    start, end = part.split('-')
                    start = int(start.strip())
                    end = int(end.strip())
                    if start > end:
                        raise ValueError("Rango inválido")
                    pages_to_extract.extend(range(start, end + 1))
                else:
                    # Página individual
                    pages_to_extract.append(int(part))
            
            # Validar que las páginas estén en el rango válido
            for page_num in pages_to_extract:
                if page_num < 1 or page_num > total_pages:
                    messages.error(request, f'La página {page_num} no existe. El PDF tiene {total_pages} páginas.')
                    return render(request, 'tickets/pdf_get_pages.html')
            
            # Crear nuevo PDF con las páginas seleccionadas
            pdf_writer = PdfWriter()
            for page_num in pages_to_extract:
                page = pdf_reader.pages[page_num - 1]  # PyPDF2 usa índice base 0
                pdf_writer.add_page(page)
            
            # Crear respuesta HTTP con el PDF
            output = io.BytesIO()
            pdf_writer.write(output)
            output.seek(0)
            
            # Preparar nombre del archivo de salida
            original_name = pdf_file.name.rsplit('.', 1)[0]
            pages_str = pages_input.replace(',', '_').replace('-', '_')
            output_filename = f"{original_name}_pages_{pages_str}.pdf"
            
            response = HttpResponse(output.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            
            messages.success(request, f'PDF generado exitosamente con las páginas: {pages_input}')
            return response
            
        except ValueError as e:
            messages.error(request, f'Error en el formato de páginas: {str(e)}. Use formato como "1,3,5-7".')
        except Exception as e:
            messages.error(request, f'Error al procesar el PDF: {str(e)}')
        
        return render(request, 'tickets/pdf_get_pages.html')
    
    return render(request, 'tickets/pdf_get_pages.html')


@login_required
@user_passes_test(is_agent)
def pdf_join_view(request):
    """Vista para unir múltiples archivos PDF en uno solo"""
    if request.method == 'POST':
        pdf_files = request.FILES.getlist('pdf_files')
        
        if not pdf_files:
            messages.error(request, 'No se han seleccionado archivos PDF.')
            return render(request, 'tickets/pdf_join.html')
        
        if len(pdf_files) < 2:
            messages.error(request, 'Debe seleccionar al menos 2 archivos PDF para unir.')
            return render(request, 'tickets/pdf_join.html')
        
        # Validar que todos los archivos sean PDF
        for pdf_file in pdf_files:
            if not pdf_file.name.lower().endswith('.pdf'):
                messages.error(request, f'El archivo "{pdf_file.name}" no es un PDF válido.')
                return render(request, 'tickets/pdf_join.html')
        
        try:
            # Importar PyPDF2
            from PyPDF2 import PdfReader, PdfWriter
            import io
            
            # Crear nuevo PDF writer
            pdf_writer = PdfWriter()
            
            # Leer y agregar páginas de cada PDF
            for pdf_file in pdf_files:
                try:
                    pdf_reader = PdfReader(pdf_file)
                    
                    # Agregar todas las páginas del PDF actual
                    for page in pdf_reader.pages:
                        pdf_writer.add_page(page)
                        
                except Exception as e:
                    messages.error(request, f'Error al procesar el archivo "{pdf_file.name}": {str(e)}')
                    return render(request, 'tickets/pdf_join.html')
            
            # Crear respuesta HTTP con el PDF unido
            output = io.BytesIO()
            pdf_writer.write(output)
            output.seek(0)
            
            # Preparar nombre del archivo ZIP
            current_time = timezone.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"PDF_unido_{current_time}.pdf"
            
            response = HttpResponse(output.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            
            messages.success(request, f'PDF unido exitosamente. Se combinaron {len(pdf_files)} archivos.')
            return response
            
        except Exception as e:
            messages.error(request, f'Error al unir los PDFs: {str(e)}')
        
        return render(request, 'tickets/pdf_join.html')
    
    return render(request, 'tickets/pdf_join.html')


@login_required
@user_passes_test(is_agent)
def pdf_split_view(request):
    """Vista para dividir un PDF en archivos individuales por página"""
    if request.method == 'POST':
        if 'pdf_file' not in request.FILES:
            messages.error(request, 'No se ha seleccionado ningún archivo PDF.')
            return render(request, 'tickets/pdf_split.html')
        
        pdf_file = request.FILES['pdf_file']
        
        # Validar que el archivo sea PDF
        if not pdf_file.name.lower().endswith('.pdf'):
            messages.error(request, 'El archivo debe ser un PDF.')
            return render(request, 'tickets/pdf_split.html')
        
        try:
            # Importar PyPDF2 y zipfile
            from PyPDF2 import PdfReader, PdfWriter
            import io
            import zipfile
            from django.utils import timezone
            
            # Leer el PDF
            pdf_reader = PdfReader(pdf_file)
            total_pages = len(pdf_reader.pages)
            
            if total_pages == 0:
                messages.error(request, 'El PDF no contiene páginas válidas.')
                return render(request, 'tickets/pdf_split.html')
            
            if total_pages == 1:
                messages.warning(request, 'El PDF solo tiene una página. No es necesario dividirlo.')
                return render(request, 'tickets/pdf_split.html')
            
            # Crear un archivo ZIP en memoria
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Crear un PDF para cada página
                for page_num in range(total_pages):
                    pdf_writer = PdfWriter()
                    page = pdf_reader.pages[page_num]
                    pdf_writer.add_page(page)
                    
                    # Crear PDF en memoria
                    page_buffer = io.BytesIO()
                    pdf_writer.write(page_buffer)
                    page_buffer.seek(0)
                    
                    # Preparar nombre del archivo de página
                    original_name = pdf_file.name.rsplit('.', 1)[0]
                    page_filename = f"{original_name}_pagina_{page_num + 1:03d}.pdf"
                    
                    # Agregar al ZIP
                    zip_file.writestr(page_filename, page_buffer.read())
            
            zip_buffer.seek(0)
            
            # Preparar nombre del archivo ZIP
            current_time = timezone.now().strftime('%Y%m%d_%H%M%S')
            original_name = pdf_file.name.rsplit('.', 1)[0]
            zip_filename = f"{original_name}_dividido_{current_time}.zip"
            
            # Crear respuesta HTTP con el ZIP
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
            
            messages.success(request, f'PDF dividido exitosamente en {total_pages} archivos individuales.')
            return response
            
        except Exception as e:
            messages.error(request, f'Error al dividir el PDF: {str(e)}')
        
        return render(request, 'tickets/pdf_split.html')
    
    return render(request, 'tickets/pdf_split.html')


@login_required
@user_passes_test(is_agent)
def calculator_view(request):
    """Vista para la calculadora con historial"""
    return render(request, 'tickets/calculator.html')


@login_required
@user_passes_test(is_agent)
def command_library_view(request):
    """Vista principal de la biblioteca de comandos con búsqueda"""
    from .forms import CommandSearchForm
    from .models import Command
    from django.db.models import Q
    
    form = CommandSearchForm(request.GET)
    commands = Command.objects.all()
    
    if form.is_valid():
        query = form.cleaned_data.get('query')
        category = form.cleaned_data.get('category')
        favorites_only = form.cleaned_data.get('favorites_only')
        dangerous_only = form.cleaned_data.get('dangerous_only')
        
        # Búsqueda por texto
        if query:
            commands = commands.filter(
                Q(title__icontains=query) |
                Q(description__icontains=query) |
                Q(command__icontains=query) |
                Q(tags__icontains=query) |
                Q(example_usage__icontains=query) |
                Q(notes__icontains=query)
            )
        
        # Filtrar por categoría
        if category:
            commands = commands.filter(category=category)
        
        # Filtrar solo favoritos
        if favorites_only:
            commands = commands.filter(is_favorite=True)
        
        # Filtrar solo comandos peligrosos
        if dangerous_only:
            commands = commands.filter(is_dangerous=True)
    
    # Estadísticas
    total_commands = Command.objects.count()
    favorites_count = Command.objects.filter(is_favorite=True).count()
    dangerous_count = Command.objects.filter(is_dangerous=True).count()
    
    # Paginación
    paginator = Paginator(commands, 12)  # 12 comandos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'form': form,
        'page_obj': page_obj,
        'total_commands': total_commands,
        'favorites_count': favorites_count,
        'dangerous_count': dangerous_count,
        'query': request.GET.get('query', ''),
    }
    
    return render(request, 'tickets/command_library.html', context)


@login_required
@user_passes_test(is_agent)
def command_create_view(request):
    """Vista para crear un nuevo comando"""
    from .forms import CommandForm
    from .models import Command
    
    if request.method == 'POST':
        form = CommandForm(request.POST)
        if form.is_valid():
            command = form.save(commit=False)
            command.created_by = request.user
            command.save()
            messages.success(request, 'Comando creado exitosamente.')
            return redirect('command_detail', pk=command.pk)
    else:
        form = CommandForm()
    
    return render(request, 'tickets/command_form.html', {
        'form': form,
        'title': 'Crear Comando',
        'action': 'Crear'
    })


@login_required
@user_passes_test(is_agent)
def command_detail_view(request, pk):
    """Vista de detalle de un comando"""
    from .models import Command
    
    command = get_object_or_404(Command, pk=pk)
    
    context = {
        'command': command,
    }
    
    return render(request, 'tickets/command_detail.html', context)


@login_required
@user_passes_test(is_agent)
def command_edit_view(request, pk):
    """Vista para editar un comando"""
    from .forms import CommandForm
    from .models import Command
    
    command = get_object_or_404(Command, pk=pk)
    
    if request.method == 'POST':
        form = CommandForm(request.POST, instance=command)
        if form.is_valid():
            form.save()
            messages.success(request, 'Comando actualizado exitosamente.')
            return redirect('command_detail', pk=command.pk)
    else:
        form = CommandForm(instance=command)
    
    return render(request, 'tickets/command_form.html', {
        'form': form,
        'command': command,
        'title': 'Editar Comando',
        'action': 'Actualizar'
    })


@login_required
@user_passes_test(is_agent)
def command_delete_view(request, pk):
    """Vista para eliminar un comando"""
    from .models import Command
    
    command = get_object_or_404(Command, pk=pk)
    
    if request.method == 'POST':
        command.delete()
        messages.success(request, 'Comando eliminado exitosamente.')
        return redirect('command_library')
    
    return render(request, 'tickets/command_delete.html', {
        'command': command,
    })


@login_required
@user_passes_test(is_agent)
def command_copy_view(request, pk):
    """Vista AJAX para copiar un comando e incrementar contador de uso"""
    from .models import Command
    
    if request.method == 'POST':
        command = get_object_or_404(Command, pk=pk)
        command.increment_usage()
        
        return JsonResponse({
            'success': True,
            'command': command.command,
            'usage_count': command.usage_count,
            'message': 'Comando copiado al portapapeles'
        })
    
    return JsonResponse({'success': False})


@login_required
@user_passes_test(is_agent)
def command_toggle_favorite_view(request, pk):
    """Vista AJAX para alternar favorito de un comando"""
    from .models import Command
    
    if request.method == 'POST':
        command = get_object_or_404(Command, pk=pk)
        command.is_favorite = not command.is_favorite
        command.save()
        
        return JsonResponse({
            'success': True,
            'is_favorite': command.is_favorite,
            'message': 'Favorito actualizado'
        })
    
    return JsonResponse({'success': False})


# ===== VISTAS DE FORMULARIO DE CONTACTO PÚBLICO =====

def public_contact_form_view(request, token):
    """Vista pública para formulario de contacto"""
    from .models import UserProfile, ContactFormSubmission
    from .forms import PublicContactForm
    
    # Buscar el perfil del usuario con el token
    try:
        profile = UserProfile.objects.get(public_token=token, enable_public_contact_form=True)
    except UserProfile.DoesNotExist:
        return render(request, 'tickets/public_contact_error.html', {
            'error_message': 'Formulario de contacto no encontrado o no disponible'
        })
    
    if request.method == 'POST':
        form = PublicContactForm(request.POST)
        if form.is_valid():
            # Crear la submission del formulario
            submission = form.save(commit=False)
            submission.submitted_by_user = profile.user
            
            # Obtener información adicional de la petición
            submission.ip_address = get_client_ip(request)
            submission.user_agent = request.META.get('HTTP_USER_AGENT', '')
            
            submission.save()
            
            # Redireccionar a página de éxito
            return redirect('contact_form_success')
    else:
        form = PublicContactForm()
    
    context = {
        'form': form,
        'profile': profile,
        'user_owner': profile.user,
        'page_title': f'Formulario de Contacto - {profile.user.get_full_name() or profile.user.username}'
    }
    
    return render(request, 'tickets/public_contact_form.html', context)


def contact_form_success(request):
    """Vista de éxito después de enviar formulario de contacto"""
    return render(request, 'tickets/contact_form_success.html')


@login_required
def contact_submissions_list(request):
    """Vista para listar formularios de contacto recibidos"""
    from .models import ContactFormSubmission
    from django.core.paginator import Paginator
    
    # Solo mostrar formularios enviados a este usuario
    submissions = ContactFormSubmission.objects.filter(
        submitted_by_user=request.user
    ).order_by('-submitted_at')
    
    # Filtros
    status_filter = request.GET.get('status')
    if status_filter:
        submissions = submissions.filter(status=status_filter)
    
    # Paginación
    paginator = Paginator(submissions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    stats = {
        'total': submissions.count(),
        'pending': submissions.filter(status='pending').count(),
        'approved': submissions.filter(status='approved').count(),
        'rejected': submissions.filter(status='rejected').count(),
        'company_created': submissions.filter(status='company_created').count(),
    }
    
    context = {
        'page_obj': page_obj,
        'submissions': page_obj,
        'stats': stats,
        'status_filter': status_filter,
        'status_choices': ContactFormSubmission.STATUS_CHOICES,
        'page_title': 'Formularios de Contacto Recibidos'
    }
    
    return render(request, 'tickets/contact_submissions_list.html', context)


@login_required
def contact_submission_detail(request, pk):
    """Vista detallada de un formulario de contacto"""
    from .models import ContactFormSubmission
    from .forms import ContactFormManagementForm
    
    submission = get_object_or_404(
        ContactFormSubmission, 
        pk=pk, 
        submitted_by_user=request.user
    )
    
    if request.method == 'POST':
        if 'update_notes' in request.POST:
            # Actualizar solo las notas administrativas
            admin_notes = request.POST.get('admin_notes', '')
            submission.admin_notes = admin_notes
            submission.save()
            messages.success(request, 'Notas administrativas actualizadas exitosamente.')
            return redirect('contact_submission_detail', pk=pk)
        else:
            # Actualizar el formulario completo
            form = ContactFormManagementForm(request.POST, instance=submission)
            if form.is_valid():
                updated_submission = form.save(commit=False)
                if updated_submission.status != submission.status:
                    updated_submission.processed_by = request.user
                    updated_submission.processed_at = timezone.now()
                updated_submission.save()
                
                messages.success(request, 'Formulario de contacto actualizado exitosamente.')
                return redirect('contact_submission_detail', pk=pk)
    else:
        form = ContactFormManagementForm(instance=submission)
    
    context = {
        'submission': submission,
        'form': form,
        'page_title': f'Formulario de Contacto - {submission.company_name}'
    }
    
    return render(request, 'tickets/contact_submission_detail.html', context)


@login_required
def contact_submission_approve(request, pk):
    """Aprobar un formulario de contacto y crear empresa"""
    from .models import ContactFormSubmission, Company
    from .forms import CompanyFromContactForm
    
    submission = get_object_or_404(
        ContactFormSubmission, 
        pk=pk, 
        submitted_by_user=request.user
    )
    
    if request.method == 'POST':
        form = CompanyFromContactForm(request.POST, contact_form=submission)
        
        if form.is_valid():
            # Crear la empresa
            company = form.save()
            
            # Marcar el formulario como procesado
            submission.mark_as_processed(
                user=request.user,
                status='company_created',
                company=company
            )
            
            messages.success(request, f'Empresa "{company.name}" creada exitosamente desde el formulario de contacto.')
            return redirect('company_detail', company_id=company.id)
        else:
            messages.error(request, f'Error al crear la empresa: {form.errors}')
    else:
        form = CompanyFromContactForm(contact_form=submission)
    
    context = {
        'submission': submission,
        'form': form,
        'page_title': f'Crear Empresa - {submission.company_name}'
    }
    
    return render(request, 'tickets/contact_submission_approve.html', context)


@login_required
def contact_submission_reject(request, pk):
    """Rechazar un formulario de contacto"""
    from .models import ContactFormSubmission
    
    submission = get_object_or_404(
        ContactFormSubmission, 
        pk=pk, 
        submitted_by_user=request.user
    )
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        
        # Actualizar el formulario
        submission.mark_as_processed(
            user=request.user,
            status='rejected'
        )
        
        if reason:
            submission.admin_notes = f"Rechazado: {reason}"
            submission.save()
        
        messages.success(request, 'Formulario de contacto rechazado.')
        return redirect('contact_submissions_list')
    
    context = {
        'submission': submission,
        'page_title': f'Rechazar Formulario - {submission.company_name}'
    }
    
    return render(request, 'tickets/contact_submission_reject.html', context)


def get_client_ip(request):
    """Obtener la IP del cliente"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


@login_required
def tetris_view(request):
    """Vista para el juego Tetris - para todos los usuarios autenticados"""
    context = {
        'page_title': 'Tetris - Juegos',
        'user_role': get_user_role(request.user)
    }
    return render(request, 'tickets/tetris.html', context)


def public_company_stats(request, token):
    """Vista pública para mostrar estadísticas de una empresa usando su token o ID"""
    try:
        # Intentar primero como UUID token
        company = Company.objects.get(public_token=token, is_active=True)
    except (Company.DoesNotExist, ValueError):
        try:
            # Si falla, intentar como ID numérico
            company = Company.objects.get(id=int(token), is_active=True)
        except (Company.DoesNotExist, ValueError):
            raise Http404("Empresa no encontrada o token inválido")
    
    # Obtener estadísticas públicas
    stats = company.get_public_stats()
    
    # Obtener tickets abiertos de la empresa (limitados a los primeros 10)
    open_tickets = company.tickets.filter(status='open').order_by('-created_at')[:10]
    
    context = {
        'company': company,
        'stats': stats,
        'open_tickets': open_tickets,
        'page_title': f'Estadísticas - {company.name}'
    }
    
    return render(request, 'tickets/public_company_stats.html', context)


# ============= VISTAS CRM =============

@login_required
def crm_dashboard(request):
    """Dashboard principal del CRM"""
    # Obtener configuración del sistema
    system_config = SystemConfiguration.get_config()
    
    # Obtener oportunidades del usuario
    if is_agent(request.user):
        opportunities = Opportunity.objects.all()
    else:
        opportunities = Opportunity.objects.filter(
            Q(created_by=request.user) | Q(assigned_to=request.user)
        )
    
    # Estadísticas generales
    total_opportunities = opportunities.count()
    total_value = opportunities.aggregate(total=Sum('value'))['total'] or 0
    total_expected_value = sum(opp.expected_value for opp in opportunities)
    
    # Oportunidades por estado
    status_stats = {}
    for status in OpportunityStatus.objects.filter(is_active=True):
        count = opportunities.filter(status=status).count()
        value = opportunities.filter(status=status).aggregate(total=Sum('value'))['total'] or 0
        status_stats[status] = {'count': count, 'value': value}
    
    # Oportunidades próximas a vencer (próximos 7 días)
    next_week = timezone.now().date() + timedelta(days=7)
    upcoming_opportunities = opportunities.filter(
        expected_close_date__lte=next_week,
        expected_close_date__gte=timezone.now().date(),
        status__is_final=False
    ).order_by('expected_close_date')[:5]
    
    # Oportunidades vencidas
    overdue_opportunities = opportunities.filter(
        expected_close_date__lt=timezone.now().date(),
        status__is_final=False
    ).count()
    
    context = {
        'total_opportunities': total_opportunities,
        'total_value': total_value,
        'total_expected_value': total_expected_value,
        'status_stats': status_stats,
        'upcoming_opportunities': upcoming_opportunities,
        'overdue_opportunities': overdue_opportunities,
        'page_title': 'CRM Dashboard',
        'system_config': system_config,
    }
    
    return render(request, 'tickets/crm_dashboard.html', context)


@login_required
def opportunity_list(request):
    """Lista de oportunidades"""
    # Obtener oportunidades del usuario
    if is_agent(request.user):
        opportunities = Opportunity.objects.all()
    else:
        opportunities = Opportunity.objects.filter(
            Q(created_by=request.user) | Q(assigned_to=request.user)
        )
    
    # Filtros
    status_filter = request.GET.get('status')
    company_filter = request.GET.get('company')
    assigned_filter = request.GET.get('assigned')
    search = request.GET.get('search')
    overdue_filter = request.GET.get('overdue')
    
    if status_filter:
        opportunities = opportunities.filter(status_id=status_filter)
    
    if company_filter:
        opportunities = opportunities.filter(company_id=company_filter)
    
    if assigned_filter:
        opportunities = opportunities.filter(assigned_to_id=assigned_filter)
    
    if overdue_filter == 'true':
        from django.utils import timezone
        today = timezone.now().date()
        opportunities = opportunities.filter(
            expected_close_date__lt=today,
            status__is_final=False
        )
    
    if search:
        opportunities = opportunities.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(company__name__icontains=search) |
            Q(contact_name__icontains=search)
        )
    
    # Ordenamiento
    order_by = request.GET.get('order_by', '-created_at')
    opportunities = opportunities.order_by(order_by)
    
    # Paginación
    paginator = Paginator(opportunities, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Agregar información de permisos a cada oportunidad
    for opportunity in page_obj:
        opportunity.user_can_edit = opportunity.can_be_edited_by(request.user)
    
    # Para los filtros
    statuses = OpportunityStatus.objects.filter(is_active=True).order_by('order')
    companies = Company.objects.all()
    users = User.objects.filter(is_active=True)
    
    context = {
        'page_obj': page_obj,
        'statuses': statuses,
        'companies': companies,
        'users': users,
        'current_filters': {
            'status': status_filter,
            'company': company_filter,
            'assigned': assigned_filter,
            'search': search,
            'order_by': order_by,
        },
        'page_title': 'Oportunidades'
    }
    
    return render(request, 'tickets/opportunity_list.html', context)


@login_required
def opportunity_detail(request, pk):
    """Detalle de oportunidad"""
    opportunity = get_object_or_404(Opportunity, pk=pk)
    
    # Verificar permisos
    from . import utils
    if not utils.is_agent(request.user):
        if opportunity.created_by != request.user and opportunity.assigned_to != request.user:
            messages.error(request, 'No tienes permisos para ver esta oportunidad.')
            return redirect('opportunity_list')
    
    # Obtener notas y historial
    notes = opportunity.notes.all()
    history = opportunity.status_history.all()
    
    # Verificar si puede ser editada
    can_be_edited = opportunity.can_be_edited_by(request.user)
    
    context = {
        'opportunity': opportunity,
        'notes': notes,
        'history': history,
        'can_be_edited': can_be_edited,
        'page_title': f'Oportunidad: {opportunity.name}'
    }
    
    return render(request, 'tickets/opportunity_detail.html', context)


@login_required
def opportunity_create(request):
    """Crear nueva oportunidad"""
    if request.method == 'POST':
        # Procesar formulario
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        company_id = request.POST.get('company')
        contact_name = request.POST.get('contact_name')
        contact_position = request.POST.get('contact_position', '')
        contact_email = request.POST.get('contact_email', '')
        contact_phone = request.POST.get('contact_phone', '')
        value = request.POST.get('value')
        probability = request.POST.get('probability', 20)
        status_id = request.POST.get('status')
        expected_close_date = request.POST.get('expected_close_date')
        assigned_to_id = request.POST.get('assigned_to')
        source = request.POST.get('source', '')
        
        try:
            company = Company.objects.get(pk=company_id)
            status = OpportunityStatus.objects.get(pk=status_id)
            assigned_to = User.objects.get(pk=assigned_to_id) if assigned_to_id else None
            
            opportunity = Opportunity.objects.create(
                name=name,
                description=description,
                company=company,
                contact_name=contact_name,
                contact_position=contact_position,
                contact_email=contact_email,
                contact_phone=contact_phone,
                value=value,
                probability=probability,
                status=status,
                expected_close_date=expected_close_date,
                assigned_to=assigned_to,
                created_by=request.user,
                source=source
            )
            
            # Crear historial inicial
            OpportunityStatusHistory.objects.create(
                opportunity=opportunity,
                new_status=status,
                changed_by=request.user,
                comment="Oportunidad creada"
            )
            
            messages.success(request, 'Oportunidad creada exitosamente.')
            return redirect('opportunity_detail', pk=opportunity.pk)
            
        except Exception as e:
            messages.error(request, f'Error al crear la oportunidad: {str(e)}')
    
    # Para GET request
    companies = Company.objects.all()
    statuses = OpportunityStatus.objects.filter(is_active=True).order_by('order')
    users = User.objects.filter(is_active=True)
    
    # Valores por defecto
    default_close_date = (timezone.now().date() + timedelta(days=90)).strftime('%Y-%m-%d')  # 3 meses
    default_status = statuses.first()  # Estado con menor secuencia (order)
    default_assigned_to = request.user  # Usuario que crea la oportunidad
    
    # Si viene un contact_id, precargar datos del contacto
    contact_data = {}
    contact_id = request.GET.get('contact_id')
    if contact_id:
        try:
            contact = Contact.objects.get(pk=contact_id)
            contact_data = {
                'contact_name': contact.name,
                'contact_email': contact.email or '',
                'contact_phone': contact.phone or '',
                'contact_position': contact.position or '',
                'source': f"Contacto: {contact.source}" if contact.source else "Contacto previo",
            }
            # Si el contacto tiene empresa, buscar si existe en Company
            if contact.company:
                try:
                    company = Company.objects.filter(name__iexact=contact.company).first()
                    if company:
                        contact_data['default_company'] = company
                except:
                    pass
        except Contact.DoesNotExist:
            pass
    
    # Opciones de probabilidad
    probability_choices = [
        (0, '0%'), (10, '10%'), (20, '20%'), (30, '30%'), (40, '40%'),
        (50, '50%'), (60, '60%'), (70, '70%'), (80, '80%'), (90, '90%'), (100, '100%')
    ]
    
    context = {
        'companies': companies,
        'statuses': statuses,
        'users': users,
        'probability_choices': probability_choices,
        'page_title': 'Nueva Oportunidad',
        'default_close_date': default_close_date,
        'default_status': default_status,
        'default_assigned_to': default_assigned_to,
        'contact_data': contact_data,
    }
    
    return render(request, 'tickets/opportunity_form.html', context)


@login_required
def opportunity_edit(request, pk):
    """Editar oportunidad"""
    opportunity = get_object_or_404(Opportunity, pk=pk)
    
    # Verificar permisos
    if not opportunity.can_be_edited_by(request.user):
        messages.error(request, 'No tienes permisos para editar esta oportunidad.')
        return redirect('opportunity_detail', pk=pk)
    
    if request.method == 'POST':
        # Procesar formulario
        previous_status = opportunity.status
        
        opportunity.name = request.POST.get('name')
        opportunity.description = request.POST.get('description', '')
        opportunity.company_id = request.POST.get('company')
        opportunity.contact_name = request.POST.get('contact_name')
        opportunity.contact_position = request.POST.get('contact_position', '')
        opportunity.contact_email = request.POST.get('contact_email', '')
        opportunity.contact_phone = request.POST.get('contact_phone', '')
        opportunity.value = request.POST.get('value')
        opportunity.probability = request.POST.get('probability', 20)
        opportunity.status_id = request.POST.get('status')
        opportunity.expected_close_date = request.POST.get('expected_close_date')
        opportunity.assigned_to_id = request.POST.get('assigned_to') or None
        opportunity.source = request.POST.get('source', '')
        
        # Si el estado cambió, crear historial
        if previous_status.pk != int(request.POST.get('status')):
            comment = request.POST.get('status_comment', '')
            OpportunityStatusHistory.objects.create(
                opportunity=opportunity,
                previous_status=previous_status,
                new_status=opportunity.status,
                changed_by=request.user,
                comment=comment
            )
        
        opportunity.save()
        messages.success(request, 'Oportunidad actualizada exitosamente.')
        return redirect('opportunity_detail', pk=pk)
    
    # Para GET request
    companies = Company.objects.all()
    statuses = OpportunityStatus.objects.filter(is_active=True)
    users = User.objects.filter(is_active=True)
    
    # Opciones de probabilidad
    probability_choices = [
        (0, '0%'), (10, '10%'), (20, '20%'), (30, '30%'), (40, '40%'),
        (50, '50%'), (60, '60%'), (70, '70%'), (80, '80%'), (90, '90%'), (100, '100%')
    ]
    
    context = {
        'opportunity': opportunity,
        'companies': companies,
        'statuses': statuses,
        'users': users,
        'probability_choices': probability_choices,
        'page_title': f'Editar: {opportunity.name}'
    }
    
    return render(request, 'tickets/opportunity_form.html', context)


@login_required
def opportunity_delete(request, pk):
    """Eliminar oportunidad"""
    opportunity = get_object_or_404(Opportunity, pk=pk)
    
    # Verificar permisos
    from . import utils
    if not utils.is_agent(request.user) and opportunity.created_by != request.user:
        messages.error(request, 'No tienes permisos para eliminar esta oportunidad.')
        return redirect('opportunity_detail', pk=pk)
    
    if request.method == 'POST':
        opportunity.delete()
        messages.success(request, 'Oportunidad eliminada exitosamente.')
        return redirect('opportunity_list')
    
    context = {
        'opportunity': opportunity,
        'page_title': f'Eliminar: {opportunity.name}'
    }
    
    return render(request, 'tickets/opportunity_delete.html', context)


@login_required
def opportunity_status_list(request):
    """Lista de estados de oportunidades"""
    from . import utils
    if not utils.is_agent(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('crm_dashboard')
    
    statuses = OpportunityStatus.objects.all().order_by('order', 'name')
    
    context = {
        'statuses': statuses,
        'page_title': 'Estados de Oportunidades'
    }
    
    return render(request, 'tickets/opportunity_status_list.html', context)


@login_required
def opportunity_status_create(request):
    """Crear estado de oportunidad"""
    from . import utils
    if not utils.is_agent(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('crm_dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        color = request.POST.get('color', '#007bff')
        is_final = request.POST.get('is_final') == 'on'
        is_won = request.POST.get('is_won') == 'on'
        order = request.POST.get('order', 0)
        
        try:
            status = OpportunityStatus.objects.create(
                name=name,
                description=description,
                color=color,
                is_final=is_final,
                is_won=is_won,
                order=order,
                created_by=request.user
            )
            messages.success(request, 'Estado creado exitosamente.')
            return redirect('opportunity_status_list')
        except Exception as e:
            messages.error(request, f'Error al crear el estado: {str(e)}')
    
    context = {
        'page_title': 'Nuevo Estado'
    }
    
    return render(request, 'tickets/opportunity_status_form.html', context)


@login_required
def opportunity_status_edit(request, pk):
    """Editar estado de oportunidad"""
    from .models import OpportunityStatus
    from . import utils
    
    status = get_object_or_404(OpportunityStatus, pk=pk)
    
    if not utils.is_agent(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('crm_dashboard')
    
    if request.method == 'POST':
        status.name = request.POST.get('name')
        status.description = request.POST.get('description', '')
        status.color = request.POST.get('color', '#007bff')
        status.is_final = request.POST.get('is_final') == 'on'
        status.is_won = request.POST.get('is_won') == 'on'
        status.order = request.POST.get('order', 0)
        status.is_active = request.POST.get('is_active') == 'on'
        
        status.save()
        messages.success(request, 'Estado actualizado exitosamente.')
        return redirect('opportunity_status_list')
    
    context = {
        'status': status,
        'page_title': f'Editar: {status.name}'
    }
    
    return render(request, 'tickets/opportunity_status_form.html', context)


@login_required
def opportunity_status_delete(request, pk):
    """Eliminar estado de oportunidad"""
    from .models import OpportunityStatus
    
    status = get_object_or_404(OpportunityStatus, pk=pk)
    
    # Verificar si hay oportunidades usando este estado
    if status.opportunities.exists():
        messages.error(request, 'No se puede eliminar este estado porque hay oportunidades que lo utilizan.')
        return redirect('opportunity_status_list')
    
    if request.method == 'POST':
        status.delete()
        messages.success(request, 'Estado eliminado exitosamente.')
        return redirect('opportunity_status_list')
    
    context = {
        'status': status,
        'page_title': f'Eliminar: {status.name}'
    }
    
    return render(request, 'tickets/opportunity_status_delete.html', context)


@login_required
def opportunity_add_note(request, pk):
    """Agregar nota a oportunidad"""
    opportunity = get_object_or_404(Opportunity, pk=pk)
    
    # Verificar permisos
    from . import utils
    if not utils.is_agent(request.user):
        if opportunity.created_by != request.user and opportunity.assigned_to != request.user:
            messages.error(request, 'No tienes permisos para agregar notas a esta oportunidad.')
            return redirect('opportunity_detail', pk=pk)
    
    if request.method == 'POST':
        content = request.POST.get('content')
        is_important = request.POST.get('is_important') == 'on'
        
        if content:
            OpportunityNote.objects.create(
                opportunity=opportunity,
                content=content,
                is_important=is_important,
                created_by=request.user
            )
            messages.success(request, 'Nota agregada exitosamente.')
        else:
            messages.error(request, 'El contenido de la nota es requerido.')
    
    return redirect('opportunity_detail', pk=pk)


# ============================
# VISTAS DE ACTIVIDADES DE OPORTUNIDADES
# ============================

@login_required
def opportunity_activity_create(request, opportunity_id):
    """Crear nueva actividad para oportunidad"""
    from .models import Opportunity, OpportunityActivity
    from .forms import OpportunityActivityForm
    
    opportunity = get_object_or_404(Opportunity, pk=opportunity_id)
    
    # Verificar permisos
    from . import utils
    if not utils.is_agent(request.user):
        if opportunity.created_by != request.user and opportunity.assigned_to != request.user:
            messages.error(request, 'No tienes permisos para crear actividades en esta oportunidad.')
            return redirect('opportunity_detail', pk=opportunity_id)
    
    if request.method == 'POST':
        form = OpportunityActivityForm(request.POST, opportunity=opportunity, current_user=request.user)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.opportunity = opportunity
            activity.created_by = request.user
            activity.save()
            
            messages.success(request, f'Actividad "{activity.title}" creada exitosamente.')
            return redirect('opportunity_detail', pk=opportunity_id)
    else:
        form = OpportunityActivityForm(opportunity=opportunity, current_user=request.user)
    
    context = {
        'form': form,
        'opportunity': opportunity,
        'page_title': f'Nueva Actividad - {opportunity.name}'
    }
    
    return render(request, 'tickets/opportunity_activity_form.html', context)


@login_required
def opportunity_activity_list(request, opportunity_id):
    """Listar actividades de una oportunidad"""
    from .models import Opportunity
    
    opportunity = get_object_or_404(Opportunity, pk=opportunity_id)
    
    # Verificar permisos
    from . import utils
    if not utils.is_agent(request.user):
        if opportunity.created_by != request.user and opportunity.assigned_to != request.user:
            messages.error(request, 'No tienes permisos para ver las actividades de esta oportunidad.')
            return redirect('opportunity_detail', pk=opportunity_id)
    
    activities = opportunity.activities.all()
    
    # Filtros
    status = request.GET.get('status')
    if status:
        activities = activities.filter(status=status)
    
    activity_type = request.GET.get('type')
    if activity_type:
        activities = activities.filter(activity_type=activity_type)
    
    context = {
        'opportunity': opportunity,
        'activities': activities,
        'page_title': f'Actividades - {opportunity.name}'
    }
    
    return render(request, 'tickets/opportunity_activity_list.html', context)


@login_required
def opportunity_activity_detail(request, pk):
    """Detalle de actividad"""
    from .models import OpportunityActivity
    
    activity = get_object_or_404(OpportunityActivity, pk=pk)
    
    # Verificar permisos
    from . import utils
    if not utils.is_agent(request.user):
        if (activity.created_by != request.user and 
            activity.assigned_to != request.user):
            messages.error(request, 'No tienes permisos para ver esta actividad.')
            return redirect('opportunity_detail', pk=activity.opportunity.pk)
    
    context = {
        'activity': activity,
        'page_title': f'Actividad: {activity.title}'
    }
    
    return render(request, 'tickets/opportunity_activity_detail.html', context)


@login_required
def opportunity_activity_edit(request, pk):
    """Editar actividad"""
    from .models import OpportunityActivity
    from .forms import OpportunityActivityForm
    
    activity = get_object_or_404(OpportunityActivity, pk=pk)
    
    # Verificar permisos
    from . import utils
    if not utils.is_agent(request.user):
        if (activity.created_by != request.user and 
            activity.assigned_to != request.user):
            messages.error(request, 'No tienes permisos para editar esta actividad.')
            return redirect('opportunity_activity_detail', pk=pk)
    
    if request.method == 'POST':
        form = OpportunityActivityForm(request.POST, instance=activity, opportunity=activity.opportunity)
        if form.is_valid():
            form.save()
            messages.success(request, 'Actividad actualizada exitosamente.')
            return redirect('opportunity_activity_detail', pk=pk)
    else:
        form = OpportunityActivityForm(instance=activity, opportunity=activity.opportunity)
    
    context = {
        'form': form,
        'activity': activity,
        'opportunity': activity.opportunity,
        'page_title': f'Editar: {activity.title}'
    }
    
    return render(request, 'tickets/opportunity_activity_form.html', context)


@login_required
def opportunity_activity_complete(request, pk):
    """Completar actividad"""
    from .models import OpportunityActivity
    from .forms import OpportunityActivityCompleteForm
    
    activity = get_object_or_404(OpportunityActivity, pk=pk)
    
    # Verificar permisos
    if activity.assigned_to != request.user:
        messages.error(request, 'Solo el asignado puede completar esta actividad.')
        return redirect('opportunity_activity_detail', pk=pk)
    
    if request.method == 'POST':
        form = OpportunityActivityCompleteForm(request.POST, instance=activity)
        if form.is_valid():
            activity = form.save(commit=False)
            if activity.status == 'completed':
                activity.completed_date = timezone.now()
            activity.save()
            
            messages.success(request, 'Actividad completada exitosamente.')
            return redirect('opportunity_activity_detail', pk=pk)
    else:
        form = OpportunityActivityCompleteForm(instance=activity)
    
    context = {
        'form': form,
        'activity': activity,
        'page_title': f'Completar: {activity.title}'
    }
    
    return render(request, 'tickets/opportunity_activity_complete.html', context)


@login_required
def opportunity_activity_delete(request, pk):
    """Eliminar actividad"""
    from .models import OpportunityActivity
    
    activity = get_object_or_404(OpportunityActivity, pk=pk)
    opportunity_id = activity.opportunity.pk
    
    # Verificar permisos
    from . import utils
    if not utils.is_agent(request.user):
        if activity.created_by != request.user:
            messages.error(request, 'No tienes permisos para eliminar esta actividad.')
            return redirect('opportunity_activity_detail', pk=pk)
    
    if request.method == 'POST':
        activity_title = activity.title
        activity.delete()
        messages.success(request, f'Actividad "{activity_title}" eliminada exitosamente.')
        return redirect('opportunity_detail', pk=opportunity_id)
    
    context = {
        'activity': activity,
        'page_title': f'Eliminar: {activity.title}'
    }
    
    return render(request, 'tickets/opportunity_activity_delete.html', context)


@login_required
def my_activities_dashboard(request):
    """Panel de actividades del usuario"""
    from .models import OpportunityActivity
    from django.db.models import Q
    
    # Actividades asignadas al usuario
    activities = OpportunityActivity.objects.filter(assigned_to=request.user)
    
    # Filtros
    status = request.GET.get('status')
    if status:
        activities = activities.filter(status=status)
    
    # Separar por estado y fecha
    today = timezone.now().date()
    
    overdue_activities = activities.filter(
        status__in=['pending', 'in_progress'],
        scheduled_date__date__lt=today
    )
    
    today_activities = activities.filter(
        scheduled_date__date=today
    )
    
    upcoming_activities = activities.filter(
        status__in=['pending', 'in_progress'],
        scheduled_date__date__gt=today
    ).order_by('scheduled_date')[:10]
    
    completed_recent = activities.filter(
        status='completed',
        completed_date__gte=today - timezone.timedelta(days=7)
    ).order_by('-completed_date')[:5]
    
    # Estadísticas
    stats = {
        'total_pending': activities.filter(status='pending').count(),
        'total_in_progress': activities.filter(status='in_progress').count(),
        'total_overdue': overdue_activities.count(),
        'total_today': today_activities.count(),
        'total_completed_week': completed_recent.count(),
    }
    
    context = {
        'overdue_activities': overdue_activities,
        'today_activities': today_activities,
        'upcoming_activities': upcoming_activities,
        'completed_recent': completed_recent,
        'stats': stats,
        'page_title': 'Mis Actividades'
    }
    
    return render(request, 'tickets/my_activities_dashboard.html', context)


# ============================
# VISTAS DE REUNIONES
# ============================

@login_required
@user_passes_test(is_agent, login_url='/')
def meeting_list_view(request):
    """Vista para listar reuniones"""
    from .models import Meeting
    
    meetings = Meeting.objects.filter(organizer=request.user)
    
    # Filtros
    status = request.GET.get('status')
    if status:
        meetings = meetings.filter(status=status)
    
    search = request.GET.get('search')
    if search:
        meetings = meetings.filter(
            models.Q(title__icontains=search) | 
            models.Q(description__icontains=search)
        )
    
    meetings = meetings.order_by('-date')
    
    # Estadísticas
    stats = {
        'total': Meeting.objects.filter(organizer=request.user).count(),
        'scheduled': Meeting.objects.filter(organizer=request.user, status='scheduled').count(),
        'in_progress': Meeting.objects.filter(organizer=request.user, status='in_progress').count(),
        'finished': Meeting.objects.filter(organizer=request.user, status='finished').count(),
    }
    
    context = {
        'page_title': 'Reuniones',
        'meetings': meetings,
        'stats': stats,
        'current_status': status,
        'search_query': search,
    }
    return render(request, 'tickets/meeting_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def meeting_create_view(request):
    """Vista para crear reuniones"""
    from .forms import MeetingForm
    
    if request.method == 'POST':
        form = MeetingForm(request.POST, user=request.user)
        if form.is_valid():
            meeting = form.save(commit=False)
            meeting.organizer = request.user
            meeting.save()
            messages.success(request, f'Reunión "{meeting.title}" creada exitosamente.')
            return redirect('meeting_detail', pk=meeting.pk)
    else:
        form = MeetingForm(user=request.user)
    
    context = {
        'page_title': 'Nueva Reunión',
        'form': form,
    }
    return render(request, 'tickets/meeting_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def meeting_detail_view(request, pk):
    """Vista detallada de una reunión"""
    from .models import Meeting
    
    meeting = get_object_or_404(Meeting, pk=pk, organizer=request.user)
    
    # Obtener asistentes y preguntas
    attendees = meeting.meetingattendee_set.all().order_by('registered_at')
    questions = meeting.meetingquestion_set.all().order_by('-asked_at')
    
    # Estadísticas
    stats = {
        'total_attendees': attendees.count(),
        'total_questions': questions.count(),
        'pending_questions': questions.filter(status='pending').count(),
        'answered_questions': questions.filter(status='answered').count(),
    }
    
    context = {
        'page_title': f'Reunión: {meeting.title}',
        'meeting': meeting,
        'attendees': attendees,
        'questions': questions,
        'stats': stats,
    }
    return render(request, 'tickets/meeting_detail.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def meeting_print_view(request, pk):
    """Vista para imprimir reunión"""
    from .models import Meeting
    
    meeting = get_object_or_404(Meeting, pk=pk, organizer=request.user)
    
    # Obtener asistentes y preguntas
    attendees = meeting.meetingattendee_set.all().order_by('registered_at')
    questions = meeting.meetingquestion_set.all().order_by('-asked_at')
    
    # Estadísticas
    stats = {
        'total_attendees': attendees.count(),
        'total_questions': questions.count(),
        'pending_questions': questions.filter(status='pending').count(),
        'answered_questions': questions.filter(status='answered').count(),
    }
    
    context = {
        'page_title': f'Informe: {meeting.title}',
        'meeting': meeting,
        'attendees': attendees,
        'questions': questions,
        'stats': stats,
    }
    return render(request, 'tickets/meeting_print.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def meeting_edit_view(request, pk):
    """Vista para editar reuniones"""
    from .forms import MeetingForm
    from .models import Meeting
    
    meeting = get_object_or_404(Meeting, pk=pk, organizer=request.user)
    
    if request.method == 'POST':
        form = MeetingForm(request.POST, instance=meeting, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f'Reunión "{meeting.title}" actualizada exitosamente.')
            return redirect('meeting_detail', pk=meeting.pk)
    else:
        form = MeetingForm(instance=meeting, user=request.user)
    
    context = {
        'page_title': f'Editar: {meeting.title}',
        'form': form,
        'meeting': meeting,
    }
    return render(request, 'tickets/meeting_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def meeting_delete_view(request, pk):
    """Vista para eliminar reuniones"""
    from .models import Meeting
    
    meeting = get_object_or_404(Meeting, pk=pk, organizer=request.user)
    
    if request.method == 'POST':
        meeting_title = meeting.title
        meeting.delete()
        messages.success(request, f'Reunión "{meeting_title}" eliminada exitosamente.')
        return redirect('meeting_list')
    
    context = {
        'page_title': f'Eliminar: {meeting.title}',
        'meeting': meeting,
    }
    return render(request, 'tickets/meeting_delete.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def meeting_questions_view(request, pk):
    """Vista para gestionar preguntas de una reunión"""
    from .models import Meeting
    
    meeting = get_object_or_404(Meeting, pk=pk, organizer=request.user)
    questions = meeting.meetingquestion_set.all().order_by('-asked_at')
    
    # Filtros
    status = request.GET.get('status')
    if status:
        questions = questions.filter(status=status)
    
    context = {
        'page_title': f'Preguntas: {meeting.title}',
        'meeting': meeting,
        'questions': questions,
        'current_status': status,
    }
    return render(request, 'tickets/meeting_questions.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def meeting_attendees_view(request, pk):
    """Vista para gestionar asistentes de una reunión"""
    from .models import Meeting
    
    meeting = get_object_or_404(Meeting, pk=pk, organizer=request.user)
    attendees = meeting.meetingattendee_set.all().order_by('registered_at')
    
    context = {
        'page_title': f'Asistentes: {meeting.title}',
        'meeting': meeting,
        'attendees': attendees,
    }
    return render(request, 'tickets/meeting_attendees.html', context)


# ============================
# VISTAS PÚBLICAS DE REUNIONES
# ============================

def meeting_public_view(request, token):
    """Vista pública de una reunión"""
    from .models import Meeting, MeetingAttendee, MeetingQuestion
    from .forms import MeetingAttendeeForm, MeetingQuestionForm
    
    meeting = get_object_or_404(Meeting, public_token=token, is_active=True)
    
    # Verificar si hay un usuario registrado en la sesión para esta reunión
    session_key = f'meeting_{meeting.id}_attendee_id'
    user_attendee = None
    user_registered = False
    user_questions = []
    
    if session_key in request.session:
        try:
            attendee_id = request.session[session_key]
            user_attendee = MeetingAttendee.objects.get(id=attendee_id, meeting=meeting)
            user_registered = True
            user_questions = MeetingQuestion.objects.filter(
                meeting=meeting, 
                attendee=user_attendee
            ).order_by('-asked_at')
        except MeetingAttendee.DoesNotExist:
            # Limpiar la sesión si el asistente no existe
            del request.session[session_key]
    
    # Solo mostrar el formulario de registro si no hay nadie registrado en la sesión
    register_form = None
    if not user_registered:
        register_form = MeetingAttendeeForm(meeting=meeting)
    
    question_form = MeetingQuestionForm() if meeting.allow_questions and user_registered else None
    
    context = {
        'page_title': f'Reunión: {meeting.title}',
        'meeting': meeting,
        'user_registered': user_registered,
        'user_attendee': user_attendee,
        'user_questions': user_questions,
        'register_form': register_form,
        'question_form': question_form,
    }
    return render(request, 'tickets/meeting_public.html', context)


def meeting_register_view(request, token):
    """Vista pública para registrarse en una reunión"""
    from .models import Meeting
    from .forms import MeetingAttendeeForm
    
    meeting = get_object_or_404(Meeting, public_token=token, is_active=True)
    
    # Verificar si ya hay un usuario registrado en esta sesión
    session_key = f'meeting_{meeting.id}_attendee_id'
    if session_key in request.session:
        messages.warning(request, 'Ya hay una persona registrada en esta sesión. Solo se permite un registro por sesión.')
        return redirect('meeting_public', token=token)
    
    if request.method == 'POST':
        form = MeetingAttendeeForm(request.POST, meeting=meeting)
        if form.is_valid():
            attendee = form.save(commit=False)
            attendee.meeting = meeting
            attendee.ip_address = request.META.get('REMOTE_ADDR')
            attendee.save()
            
            # Guardar el ID del asistente en la sesión
            request.session[session_key] = attendee.id
            
            messages.success(request, '¡Te has registrado exitosamente en la reunión!')
            return redirect('meeting_public', token=token)
    else:
        form = MeetingAttendeeForm(meeting=meeting)
    
    context = {
        'page_title': f'Registrarse: {meeting.title}',
        'meeting': meeting,
        'form': form,
    }
    return render(request, 'tickets/meeting_register.html', context)


def meeting_questions_public_view(request, token):
    """Vista pública para ver preguntas de una reunión"""
    from .models import Meeting
    
    meeting = get_object_or_404(Meeting, public_token=token, is_active=True)
    
    if not meeting.allow_questions:
        messages.error(request, 'Esta reunión no permite preguntas.')
        return redirect('meeting_public', token=token)
    
    questions = meeting.meetingquestion_set.filter(status__in=['pending', 'answered']).order_by('-asked_at')
    
    context = {
        'page_title': f'Preguntas: {meeting.title}',
        'meeting': meeting,
        'questions': questions,
    }
    return render(request, 'tickets/meeting_questions_public.html', context)


def meeting_ask_question_view(request, token):
    """Vista pública para hacer una pregunta"""
    from .models import Meeting, MeetingAttendee
    from .forms import MeetingQuestionForm
    
    meeting = get_object_or_404(Meeting, public_token=token, is_active=True)
    
    if not meeting.allow_questions:
        messages.error(request, 'Esta reunión no permite preguntas.')
        return redirect('meeting_public', token=token)
    
    # Verificar si hay un asistente registrado en la sesión
    session_key = f'meeting_{meeting.id}_attendee_id'
    user_attendee = None
    
    if session_key in request.session:
        try:
            attendee_id = request.session[session_key]
            user_attendee = MeetingAttendee.objects.get(id=attendee_id, meeting=meeting)
        except MeetingAttendee.DoesNotExist:
            pass
    
    if request.method == 'POST':
        form = MeetingQuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.meeting = meeting
            question.attendee = user_attendee  # Asociar con el asistente de la sesión
            question.ip_address = request.META.get('REMOTE_ADDR')
            
            # Si no hay asistente en sesión, usar los datos del formulario
            if not user_attendee:
                question.asker_name = question.asker_name or 'Anónimo'
                question.asker_email = question.asker_email or ''
            else:
                question.asker_name = user_attendee.name
                question.asker_email = user_attendee.email
            
            question.save()
            messages.success(request, '¡Tu pregunta ha sido enviada!')
            return redirect('meeting_public', token=token)
    else:
        form = MeetingQuestionForm()
    
    context = {
        'page_title': f'Hacer Pregunta: {meeting.title}',
        'meeting': meeting,
        'form': form,
    }
    return render(request, 'tickets/meeting_ask_question.html', context)
    
    context = {
        'page_title': f'Hacer Pregunta: {meeting.title}',
        'meeting': meeting,
        'form': form,
    }
    return render(request, 'tickets/meeting_ask_question.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def meeting_answer_question_view(request, pk, question_id):
    """Vista AJAX para responder preguntas"""
    from .models import Meeting, MeetingQuestion
    import json
    
    meeting = get_object_or_404(Meeting, pk=pk, organizer=request.user)
    question = get_object_or_404(MeetingQuestion, pk=question_id, meeting=meeting)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            answer = data.get('answer', '').strip()
            
            if not answer:
                return JsonResponse({
                    'success': False,
                    'error': 'La respuesta no puede estar vacía.'
                })
            
            question.answer = answer
            question.answered_by = request.user
            question.answered_at = timezone.now()
            question.status = 'answered'
            question.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Pregunta respondida exitosamente.'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Datos JSON inválidos.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al procesar la respuesta: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'})


# ========================
# SISTEMA DE CAPACITACIÓN
# ========================

@login_required
def debug_user_access(request):
    """Vista de debug para verificar acceso de usuarios a cursos"""
    from .models import Course, Company
    from django.http import HttpResponse
    
    debug_info = []
    debug_info.append(f"Usuario actual: {request.user.username}")
    debug_info.append(f"Es agente: {request.user.groups.filter(name='Agentes').exists()}")
    
    try:
        profile = request.user.userprofile
        debug_info.append(f"UserProfile existe: Sí")
        debug_info.append(f"Empresa del usuario: {profile.company}")
        debug_info.append(f"ID de empresa del usuario: {profile.company.id if profile.company else 'None'}")
    except Exception as e:
        debug_info.append(f"Error con UserProfile: {e}")
    
    debug_info.append("\n--- CURSOS ---")
    courses = Course.objects.filter(is_active=True)
    for course in courses:
        debug_info.append(f"\nCurso: {course.title}")
        debug_info.append(f"Empresa del curso: {course.company}")
        debug_info.append(f"ID empresa del curso: {course.company.id if course.company else 'None'}")
        debug_info.append(f"Puede acceder: {course.can_user_access(request.user)}")
    
    debug_info.append("\n--- EMPRESAS ---")
    companies = Company.objects.all()
    for company in companies:
        debug_info.append(f"Empresa: {company.name} (ID: {company.id})")
    
    return HttpResponse("<pre>" + "\n".join(debug_info) + "</pre>")

@login_required
def course_generate_public_token(request, pk):
    """Generar token público para un curso"""
    from .models import Course
    from django.contrib import messages
    from django.http import JsonResponse
    
    course = get_object_or_404(Course, pk=pk, is_active=True)
    
    # Verificar permisos - solo el creador o superusuario
    if course.created_by != request.user and not request.user.is_superuser:
        # Verificar si es una petición AJAX usando headers
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos para esta acción'})
        messages.error(request, 'No tienes permisos para generar enlaces públicos.')
        return redirect('course_detail', pk=pk)
    
    # Generar token
    token = course.generate_public_token()
    public_url = course.get_public_url(request)
    
    # Verificar si es una petición AJAX usando headers
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True, 
            'token': str(token),
            'public_url': public_url,
            'message': 'Enlace público generado exitosamente'
        })
    
    messages.success(request, f'Enlace público generado: {public_url}')
    return redirect('course_detail', pk=pk)

@login_required
def course_disable_public_access(request, pk):
    """Deshabilitar acceso público a un curso"""
    from .models import Course
    from django.contrib import messages
    from django.http import JsonResponse
    
    course = get_object_or_404(Course, pk=pk, is_active=True)
    
    # Verificar permisos
    if course.created_by != request.user and not request.user.is_superuser:
        # Verificar si es una petición AJAX usando headers
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos para esta acción'})
        messages.error(request, 'No tienes permisos para deshabilitar el acceso público.')
        return redirect('course_detail', pk=pk)
    
    course.disable_public_access()
    
    # Verificar si es una petición AJAX usando headers
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': 'Acceso público deshabilitado'})
    
    messages.success(request, 'Acceso público deshabilitado correctamente.')
    return redirect('course_detail', pk=pk)

def course_public(request, token):
    """Vista pública del curso (sin autenticación requerida)"""
    from .models import Course, CourseClassView
    from django.shortcuts import get_object_or_404
    
    course = get_object_or_404(
        Course, 
        public_token=token, 
        is_public_enabled=True, 
        is_active=True
    )
    
    classes = course.classes.filter(is_active=True).order_by('order', 'title')
    
    # Agregar información de visualizaciones para cada clase
    classes_with_views = []
    for course_class in classes:
        view_count = course_class.get_view_count()
        user_has_viewed = False
        
        # Si hay usuario autenticado, verificar si ha visto la clase
        if request.user.is_authenticated:
            user_has_viewed = course_class.is_viewed_by_user(request.user)
        
        classes_with_views.append({
            'class': course_class,
            'view_count': view_count,
            'user_has_viewed': user_has_viewed,
            'status': 'vista' if user_has_viewed else 'sin ver'
        })
    
    context = {
        'course': course,
        'classes_with_views': classes_with_views,
        'page_title': f'Curso: {course.title}',
        'is_public_view': True
    }
    
    return render(request, 'tickets/course_public.html', context)

def course_class_public(request, token, class_id):
    """Vista pública de una clase específica"""
    from .models import Course, CourseClass, CourseClassView
    
    course = get_object_or_404(
        Course, 
        public_token=token, 
        is_public_enabled=True, 
        is_active=True
    )
    course_class = get_object_or_404(CourseClass, pk=class_id, course=course, is_active=True)
    
    # Verificar si el usuario autenticado ha visto la clase
    user_has_viewed = False
    if request.user.is_authenticated:
        user_has_viewed = course_class.is_viewed_by_user(request.user)
    
    # Registrar la visualización solo si hay usuario autenticado
    if request.user.is_authenticated:
        view_obj, created = CourseClassView.objects.get_or_create(
            course_class=course_class,
            user=request.user,
            defaults={
                'ip_address': get_client_ip(request)
            }
        )
        
        # Incrementar contador de vistas solo si es una nueva visualización
        if created:
            course_class.save()
    
    # Obtener clases siguientes y anteriores
    all_classes = course.classes.filter(is_active=True).order_by('order', 'title')
    current_index = None
    
    for i, class_item in enumerate(all_classes):
        if class_item.pk == course_class.pk:
            current_index = i
            break
    
    prev_class = None
    next_class = None
    
    if current_index is not None:
        if current_index > 0:
            prev_class = all_classes[current_index - 1]
        if current_index < len(all_classes) - 1:
            next_class = all_classes[current_index + 1]
    
    context = {
        'course': course,
        'course_class': course_class,
        'page_title': f'{course.title} - {course_class.title}',
        'user_has_viewed': user_has_viewed,
        'prev_class': prev_class,
        'next_class': next_class,
    }
    
    return render(request, 'tickets/course_class_public.html', context)

@login_required
def course_list(request):
    """Lista todos los cursos disponibles"""
    from .models import Course
    from . import utils
    
    # Verificar si el usuario puede gestionar cursos (agentes y profesores)
    can_manage = utils.can_manage_courses(request.user)
    
    # Filtrar cursos según la empresa del usuario
    courses = Course.objects.filter(is_active=True)
    
    if can_manage:
        # Los gestores (agentes y profesores) siempre ven todos los cursos (públicos y empresariales)
        user_accessible_courses = list(courses)
    else:
        # Filtrar según empresa del usuario para usuarios no gestores
        user_accessible_courses = []
        for course in courses:
            if course.can_user_access(request.user):
                user_accessible_courses.append(course)
    
    context = {
        'courses': user_accessible_courses,
        'can_manage_courses': can_manage,
        'page_title': 'Cursos de Capacitación'
    }
    
    return render(request, 'tickets/course_list.html', context)


@login_required
def course_detail(request, pk):
    """Detalle de un curso con sus clases"""
    from .models import Course
    from . import utils
    
    course = get_object_or_404(Course, pk=pk, is_active=True)
    
    # Verificar si el usuario puede gestionar cursos (agentes y profesores)
    can_manage = utils.can_manage_courses(request.user)
    
    # Verificar si el usuario puede acceder a este curso
    if not can_manage and not course.can_user_access(request.user):
        messages.error(request, 'No tienes permisos para acceder a este curso.')
        return redirect('course_list')
    
    classes = course.classes.filter(is_active=True).order_by('order', 'title')
    
    # Agregar información de visualizaciones para cada clase
    classes_with_views = []
    for course_class in classes:
        view_count = course_class.get_view_count()
        user_has_viewed = course_class.is_viewed_by_user(request.user)
        
        classes_with_views.append({
            'class': course_class,
            'view_count': view_count,
            'user_has_viewed': user_has_viewed,
        })
    
    context = {
        'course': course,
        'classes': classes,
        'classes_with_views': classes_with_views,
        'can_manage_courses': can_manage,
        'page_title': f'Curso: {course.title}'
    }
    
    return render(request, 'tickets/course_detail.html', context)


@login_required
def course_create(request):
    """Crear nuevo curso (agentes y profesores)"""
    from .models import Course
    from .forms import CourseForm
    from . import utils
    
    if not utils.can_manage_courses(request.user):
        messages.error(request, 'Solo los agentes y profesores pueden crear cursos.')
        return redirect('course_list')
    
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save(commit=False)
            course.created_by = request.user
            course.save()
            messages.success(request, f'Curso "{course.title}" creado exitosamente.')
            return redirect('course_detail', pk=course.pk)
    else:
        form = CourseForm()
    
    context = {
        'form': form,
        'page_title': 'Crear Nuevo Curso'
    }
    
    return render(request, 'tickets/course_form.html', context)


@login_required
def course_edit(request, pk):
    """Editar curso (agentes y profesores)"""
    from .models import Course
    from .forms import CourseForm
    from . import utils
    
    course = get_object_or_404(Course, pk=pk)
    
    if not utils.can_manage_courses(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('crm_dashboard')
    
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, f'Curso "{course.title}" actualizado exitosamente.')
            return redirect('course_detail', pk=pk)
    else:
        form = CourseForm(instance=course)
    
    context = {
        'form': form,
        'course': course,
        'page_title': f'Editar: {course.title}'
    }
    
    return render(request, 'tickets/course_form.html', context)


@login_required
def course_delete(request, pk):
    """Eliminar curso (agentes y profesores)"""
    from .models import Course
    from . import utils
    
    course = get_object_or_404(Course, pk=pk)
    
    if not utils.can_manage_courses(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('crm_dashboard')
    
    if request.method == 'POST':
        course_title = course.title
        course.delete()
        messages.success(request, f'Curso "{course_title}" eliminado exitosamente.')
        return redirect('course_list')
    
    context = {
        'course': course,
        'page_title': f'Eliminar: {course.title}'
    }
    
    return render(request, 'tickets/course_delete.html', context)


@login_required
def course_class_create(request, course_id):
    """Crear nueva clase para un curso (agentes y profesores)"""
    from .models import Course, CourseClass
    from .forms import CourseClassForm
    from . import utils
    
    course = get_object_or_404(Course, pk=course_id)
    
    if not utils.can_manage_courses(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('crm_dashboard')
    
    if request.method == 'POST':
        form = CourseClassForm(request.POST)
        if form.is_valid():
            course_class = form.save(commit=False)
            course_class.course = course
            course_class.created_by = request.user
            course_class.save()
            messages.success(request, f'Clase "{course_class.title}" agregada exitosamente.')
            return redirect('course_detail', pk=course_id)
    else:
        # Sugerir el siguiente número de orden
        next_order = course.classes.count() + 1
        form = CourseClassForm(initial={'order': next_order})
    
    context = {
        'form': form,
        'course': course,
        'page_title': f'Nueva Clase - {course.title}'
    }
    
    return render(request, 'tickets/course_class_form.html', context)


@login_required
def course_class_edit(request, course_id, pk):
    """Editar clase (agentes y profesores)"""
    from .models import Course, CourseClass
    from .forms import CourseClassForm
    from . import utils
    
    course = get_object_or_404(Course, pk=course_id)
    course_class = get_object_or_404(CourseClass, pk=pk, course=course)
    
    if not utils.can_manage_courses(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('crm_dashboard')
    
    if request.method == 'POST':
        form = CourseClassForm(request.POST, instance=course_class)
        if form.is_valid():
            form.save()
            messages.success(request, f'Clase "{course_class.title}" actualizada exitosamente.')
            return redirect('course_detail', pk=course_id)
    else:
        form = CourseClassForm(instance=course_class)
    
    context = {
        'form': form,
        'course': course,
        'course_class': course_class,
        'page_title': f'Editar: {course_class.title}'
    }
    
    return render(request, 'tickets/course_class_form.html', context)


@login_required
def course_class_delete(request, course_id, pk):
    """Eliminar clase (agentes y profesores)"""
    from .models import Course, CourseClass
    from . import utils
    
    course = get_object_or_404(Course, pk=course_id)
    course_class = get_object_or_404(CourseClass, pk=pk, course=course)
    
    if not utils.can_manage_courses(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('crm_dashboard')
    
    if request.method == 'POST':
        class_title = course_class.title
        course_class.delete()
        messages.success(request, f'Clase "{class_title}" eliminada exitosamente.')
        return redirect('course_detail', pk=course_id)
    
    context = {
        'course': course,
        'course_class': course_class,
        'page_title': f'Eliminar: {course_class.title}'
    }
    
    return render(request, 'tickets/course_class_delete.html', context)


@login_required
def course_class_detail(request, course_id, pk):
    """Ver detalle de una clase específica"""
    from .models import Course, CourseClass, CourseClassView
    
    course = get_object_or_404(Course, pk=course_id, is_active=True)
    course_class = get_object_or_404(CourseClass, pk=pk, course=course, is_active=True)
    
    # Verificar si el usuario es agente
    is_agent = request.user.groups.filter(name='Agentes').exists()
    
    # Verificar si el usuario puede acceder a este curso
    if not is_agent and not course.can_user_access(request.user):
        messages.error(request, 'No tienes permisos para acceder a este curso.')
        return redirect('course_list')
    
    # Registrar la visualización si no existe
    view_obj, created = CourseClassView.objects.get_or_create(
        course_class=course_class,
        user=request.user,
        defaults={
            'ip_address': get_client_ip(request)
        }
    )
    
    # Obtener clases anterior y siguiente para navegación
    all_classes = course.classes.filter(is_active=True).order_by('order', 'title')
    current_index = None
    
    for i, class_item in enumerate(all_classes):
        if class_item.pk == course_class.pk:
            current_index = i
            break
    
    prev_class = None
    next_class = None
    
    if current_index is not None:
        if current_index > 0:
            prev_class = all_classes[current_index - 1]
        if current_index < len(all_classes) - 1:
            next_class = all_classes[current_index + 1]
    
    context = {
        'course': course,
        'course_class': course_class,
        'page_title': f'{course.title} - {course_class.title}',
        'user_has_viewed': True,  # Siempre True ya que acabamos de registrar la vista
        'prev_class': prev_class,
        'next_class': next_class,
    }
    
    return render(request, 'tickets/course_class_detail.html', context)


def get_client_ip(request):
    """Obtiene la IP del cliente"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


@login_required
def mark_class_as_viewed(request, course_id, class_id):
    """Marcar una clase como vista via AJAX"""
    from .models import Course, CourseClass, CourseClassView
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        course = get_object_or_404(Course, pk=course_id, is_active=True)
        course_class = get_object_or_404(CourseClass, pk=class_id, course=course, is_active=True)
        
        # Verificar si el usuario es agente
        is_agent = request.user.groups.filter(name='Agentes').exists()
        
        # Verificar si el usuario puede acceder a este curso
        if not is_agent and not course.can_user_access(request.user):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para acceder a este curso'})
        
        # Crear o actualizar la visualización
        view_obj, created = CourseClassView.objects.get_or_create(
            course_class=course_class,
            user=request.user,
            defaults={
                'ip_address': get_client_ip(request)
            }
        )
        
        return JsonResponse({
            'success': True,
            'created': created,
            'view_count': course_class.get_view_count(),
            'message': 'Clase marcada como vista' if created else 'Ya habías visto esta clase'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ===== VISTAS DE CONTACTOS =====

@login_required
def contact_list(request):
    """Lista de contactos con filtros"""
    contacts = Contact.objects.all()
    
    # Filtros
    status = request.GET.get('status')
    if status:
        contacts = contacts.filter(status=status)
    
    search = request.GET.get('search')
    if search:
        contacts = contacts.filter(
            Q(name__icontains=search) |
            Q(email__icontains=search) |
            Q(company__icontains=search) |
            Q(phone__icontains=search)
        )
    
    date_from = request.GET.get('date_from')
    if date_from:
        contacts = contacts.filter(contact_date__date__gte=date_from)
    
    date_to = request.GET.get('date_to')
    if date_to:
        contacts = contacts.filter(contact_date__date__lte=date_to)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(contacts, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    total_contacts = Contact.objects.count()
    positive_contacts = Contact.objects.filter(status='positive').count()
    negative_contacts = Contact.objects.filter(status='negative').count()
    today_contacts = Contact.objects.filter(contact_date__date=timezone.now().date()).count()
    
    context = {
        'page_obj': page_obj,
        'contacts': page_obj.object_list,
        'total_contacts': total_contacts,
        'positive_contacts': positive_contacts,
        'negative_contacts': negative_contacts,
        'today_contacts': today_contacts,
        'page_title': 'Contactos'
    }
    
    return render(request, 'tickets/contact_list.html', context)


@login_required
def contact_create(request):
    """Crear nuevo contacto"""
    from .forms import ContactForm
    
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            contact = form.save(commit=False)
            contact.created_by = request.user
            contact.save()
            
            messages.success(request, f'Contacto "{contact.name}" creado exitosamente.')
            return redirect('contact_detail', pk=contact.pk)
    else:
        form = ContactForm()
    
    context = {
        'form': form,
        'page_title': 'Nuevo Contacto'
    }
    
    return render(request, 'tickets/contact_form.html', context)


@login_required
def contact_detail(request, pk):
    """Ver detalles de un contacto"""
    contact = get_object_or_404(Contact, pk=pk)
    
    context = {
        'contact': contact,
        'page_title': f'Contacto: {contact.name}'
    }
    
    return render(request, 'tickets/contact_detail.html', context)


@login_required
def contact_edit(request, pk):
    """Editar contacto"""
    from .forms import ContactForm
    
    contact = get_object_or_404(Contact, pk=pk)
    
    if request.method == 'POST':
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()
            
            messages.success(request, f'Contacto "{contact.name}" actualizado exitosamente.')
            return redirect('contact_detail', pk=contact.pk)
    else:
        form = ContactForm(instance=contact)
    
    context = {
        'form': form,
        'contact': contact,
        'page_title': f'Editar: {contact.name}'
    }
    
    return render(request, 'tickets/contact_form.html', context)


@login_required
def contact_delete(request, pk):
    """Eliminar contacto"""
    contact = get_object_or_404(Contact, pk=pk)
    
    if request.method == 'POST':
        contact_name = contact.name
        contact.delete()
        messages.success(request, f'Contacto "{contact_name}" eliminado exitosamente.')
        return redirect('contact_list')
    
    context = {
        'contact': contact,
        'page_title': f'Eliminar: {contact.name}'
    }
    
    return render(request, 'tickets/contact_delete.html', context)


# ===== VISTAS DEL BLOG =====

def blog_list(request):
    """Lista pública de artículos del blog"""
    posts = BlogPost.objects.filter(status='published').select_related('category', 'created_by')
    
    # Filtros
    category_slug = request.GET.get('category')
    if category_slug:
        posts = posts.filter(category__slug=category_slug)
    
    search = request.GET.get('search')
    if search:
        posts = posts.filter(
            Q(title__icontains=search) |
            Q(excerpt__icontains=search) |
            Q(content__icontains=search) |
            Q(tags__icontains=search)
        )
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(posts, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Categorías para el menú
    categories = BlogCategory.objects.filter(is_active=True)
    
    # Artículos destacados
    featured_posts = BlogPost.objects.filter(
        status='published', 
        is_featured=True
    ).select_related('category')[:3]
    
    context = {
        'page_obj': page_obj,
        'posts': page_obj.object_list,
        'categories': categories,
        'featured_posts': featured_posts,
        'current_category': category_slug,
        'page_title': 'Blog',
        'is_authenticated': request.user.is_authenticated,
    }
    
    return render(request, 'tickets/blog_list.html', context)


def blog_post_detail(request, slug):
    """Vista detalle de un artículo del blog"""
    post = get_object_or_404(BlogPost, slug=slug, status='published')
    
    # Incrementar contador de visualizaciones
    post.increment_views()
    
    # Procesar comentarios
    if request.method == 'POST':
        from .forms import BlogCommentForm
        comment_form = BlogCommentForm(request.POST)
        if comment_form.is_valid():
            comment = comment_form.save(commit=False)
            comment.post = post
            
            # Obtener IP del usuario
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                comment.ip_address = x_forwarded_for.split(',')[0]
            else:
                comment.ip_address = request.META.get('REMOTE_ADDR')
            
            comment.save()
            messages.success(request, 'Tu comentario ha sido enviado y está pendiente de aprobación.')
            return redirect('blog_post_detail', slug=slug)
    else:
        from .forms import BlogCommentForm
        comment_form = BlogCommentForm()
    
    # Comentarios aprobados
    comments = post.comments.filter(is_approved=True).order_by('created_at')
    
    # Artículos relacionados
    related_posts = BlogPost.objects.filter(
        status='published',
        category=post.category
    ).exclude(pk=post.pk)[:3]
    
    context = {
        'post': post,
        'comments': comments,
        'comment_form': comment_form,
        'related_posts': related_posts,
        'page_title': post.title,
        'is_authenticated': request.user.is_authenticated,
    }
    
    return render(request, 'tickets/blog_post_detail.html', context)


@login_required
def blog_category_list(request):
    """Lista de categorías del blog (admin)"""
    categories = BlogCategory.objects.all().order_by('name')
    
    context = {
        'categories': categories,
        'page_title': 'Categorías de Blog'
    }
    
    return render(request, 'tickets/blog_category_list.html', context)


@login_required
def blog_category_create(request):
    """Crear nueva categoría de blog"""
    from .forms import BlogCategoryForm
    
    if request.method == 'POST':
        print(f"POST data: {request.POST}")
        form = BlogCategoryForm(request.POST)
        print(f"Form is valid: {form.is_valid()}")
        if not form.is_valid():
            print(f"Form errors: {form.errors}")
        
        if form.is_valid():
            try:
                category = form.save(commit=False)
                category.created_by = request.user
                
                # Auto-generar slug único
                from django.utils.text import slugify
                base_slug = slugify(category.name)
                slug = base_slug
                counter = 1
                
                # Verificar si el slug ya existe y crear uno único
                while BlogCategory.objects.filter(slug=slug).exists():
                    slug = f"{base_slug}-{counter}"
                    counter += 1
                
                category.slug = slug
                category.save()
                print(f"Categoría creada: {category.name} - {category.slug}")
                
                messages.success(request, f'Categoría "{category.name}" creada exitosamente.')
                return redirect('blog_category_list')
                
            except Exception as e:
                print(f"Exception al crear categoría: {str(e)}")
                messages.error(request, f'Error al crear la categoría: {str(e)}')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = BlogCategoryForm()
    
    context = {
        'form': form,
        'page_title': 'Nueva Categoría'
    }
    
    return render(request, 'tickets/blog_category_form.html', context)


@login_required
def blog_post_list_admin(request):
    """Lista de artículos del blog (admin)"""
    posts = BlogPost.objects.all().select_related('category', 'created_by')
    
    # Filtros
    status = request.GET.get('status')
    if status:
        posts = posts.filter(status=status)
    
    category_id = request.GET.get('category')
    if category_id:
        posts = posts.filter(category_id=category_id)
    
    search = request.GET.get('search')
    if search:
        posts = posts.filter(
            Q(title__icontains=search) |
            Q(excerpt__icontains=search)
        )
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(posts, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    categories = BlogCategory.objects.filter(is_active=True)
    
    context = {
        'page_obj': page_obj,
        'posts': page_obj.object_list,
        'categories': categories,
        'page_title': 'Gestión de Blog'
    }
    
    return render(request, 'tickets/blog_post_list_admin.html', context)


@login_required
def blog_post_create(request):
    """Crear nuevo artículo de blog"""
    from .forms import BlogPostForm
    
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES)
        if form.is_valid():
            post = form.save(commit=False)
            post.created_by = request.user
            post.save()
            
            messages.success(request, f'Artículo "{post.title}" creado exitosamente.')
            return redirect('blog_post_detail_admin', pk=post.pk)
    else:
        form = BlogPostForm()
    
    context = {
        'form': form,
        'page_title': 'Nuevo Artículo'
    }
    
    return render(request, 'tickets/blog_post_form.html', context)


@login_required
def blog_post_detail_admin(request, pk):
    """Vista detalle de artículo (admin)"""
    post = get_object_or_404(BlogPost, pk=pk)
    
    # Comentarios pendientes
    pending_comments = post.comments.filter(is_approved=False).order_by('-created_at')
    approved_comments = post.comments.filter(is_approved=True).order_by('-created_at')
    
    context = {
        'post': post,
        'pending_comments': pending_comments,
        'approved_comments': approved_comments,
        'page_title': f'Artículo: {post.title}'
    }
    
    return render(request, 'tickets/blog_post_detail_admin.html', context)


@login_required
def blog_comment_approve(request, pk):
    """Aprobar comentario"""
    comment = get_object_or_404(BlogComment, pk=pk)
    comment.is_approved = True
    comment.save()
    
    messages.success(request, 'Comentario aprobado exitosamente.')
    return redirect('blog_post_detail_admin', pk=comment.post.pk)


@login_required
def blog_comment_delete(request, pk):
    """Eliminar comentario"""
    comment = get_object_or_404(BlogComment, pk=pk)
    post_pk = comment.post.pk
    comment.delete()
    
    messages.success(request, 'Comentario eliminado exitosamente.')
    return redirect('blog_post_detail_admin', pk=post_pk)


@login_required
def blog_post_edit(request, pk):
    """Editar artículo de blog"""
    from .forms import BlogPostForm
    
    post = get_object_or_404(BlogPost, pk=pk)
    
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            post = form.save(commit=False)
            
            # Si se está publicando por primera vez, establecer fecha de publicación
            if post.status == 'published' and not post.published_at:
                from django.utils import timezone
                post.published_at = timezone.now()
            
            post.save()
            messages.success(request, f'Artículo "{post.title}" actualizado exitosamente.')
            return redirect('blog_post_detail_admin', pk=post.pk)
    else:
        form = BlogPostForm(instance=post)
    
    context = {
        'form': form,
        'post': post,
        'page_title': f'Editar: {post.title}',
        'action': 'Actualizar'
    }
    
    return render(request, 'tickets/blog_post_form.html', context)


@login_required
def blog_post_toggle_status(request, pk):
    """Cambiar estado de publicación del artículo"""
    post = get_object_or_404(BlogPost, pk=pk)
    
    if post.status == 'published':
        post.status = 'draft'
        post.published_at = None
        messages.success(request, f'Artículo "{post.title}" despublicado.')
    else:
        post.status = 'published'
        if not post.published_at:
            from django.utils import timezone
            post.published_at = timezone.now()
        messages.success(request, f'Artículo "{post.title}" publicado exitosamente.')
    
    post.save()
    return redirect('blog_post_detail_admin', pk=post.pk)


@login_required
def blog_post_delete(request, pk):
    """Eliminar artículo de blog"""
    post = get_object_or_404(BlogPost, pk=pk)
    
    if request.method == 'POST':
        title = post.title
        post.delete()
        messages.success(request, f'Artículo "{title}" eliminado exitosamente.')
        return redirect('blog_post_list_admin')
    
    context = {
        'post': post,
        'page_title': f'Eliminar: {post.title}'
    }
    
    return render(request, 'tickets/blog_post_delete.html', context)


# ===========================================
# VISTAS PARA CHAT CON IA
# ===========================================

@login_required
def ai_chat_list_view(request):
    """Vista principal del chat con IA - lista de sesiones"""
    config = SystemConfiguration.get_config()
    
    # Verificar si el chat IA está habilitado
    if not config.ai_chat_enabled:
        messages.warning(request, 'El chat con IA no está habilitado en el sistema.')
        return redirect('dashboard')
    
    # Verificar si hay API key configurada
    if not config.openai_api_key:
        messages.warning(request, 'El chat con IA no está configurado. Contacta al administrador.')
        return redirect('dashboard')
    
    # Obtener sesiones del usuario
    sessions = AIChatSession.objects.filter(user=request.user, is_active=True)
    
    # Formulario para nueva sesión
    if request.method == 'POST':
        form = AIChatSessionForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.user = request.user
            session.save()
            messages.success(request, f'Nueva conversación creada: {session.title}')
            return redirect('ai_chat_live', session_id=session.id)
    else:
        form = AIChatSessionForm(initial={'ai_model': config.openai_model})
    
    context = {
        'sessions': sessions,
        'form': form,
        'config': config,
    }
    
    return render(request, 'tickets/ai_chat_list.html', context)


@login_required
def ai_chat_session_view(request, session_id):
    """Vista para una sesión específica de chat con IA"""
    config = SystemConfiguration.get_config()
    
    # Verificar si el chat IA está habilitado
    if not config.ai_chat_enabled or not config.openai_api_key:
        messages.warning(request, 'El chat con IA no está disponible.')
        return redirect('dashboard')
    
    # Obtener la sesión
    session = get_object_or_404(AIChatSession, id=session_id, user=request.user, is_active=True)
    
    # Obtener mensajes de la sesión
    messages_list = session.messages.all()
    
    # Formulario para nuevo mensaje
    if request.method == 'POST':
        print(f"DEBUG: POST recibido para sesión {session_id}")
        print(f"DEBUG: request.POST completo: {dict(request.POST)}")
        print(f"DEBUG: request.POST.get('message'): '{request.POST.get('message', 'NO_ENCONTRADO')}'")
        print(f"DEBUG: Content-Type: {request.content_type}")
        print(f"DEBUG: POST keys: {list(request.POST.keys())}")
        
        form = AIChatMessageForm(request.POST)
        print(f"DEBUG: Formulario creado con datos: {form.data}")
        print(f"DEBUG: Formulario es válido: {form.is_valid()}")
        
        if form.is_valid():
            user_message = form.cleaned_data['message']
            print(f"DEBUG: Mensaje del usuario: {user_message}")
            
            # Guardar mensaje del usuario
            user_msg = AIChatMessage.objects.create(
                session=session,
                role='user',
                content=user_message
            )
            print(f"DEBUG: Mensaje del usuario guardado con ID: {user_msg.id}")
            
            # Llamar a la API de OpenAI
            try:
                print("DEBUG: Llamando a OpenAI API...")
                ai_response = call_openai_api(session, user_message)
                print(f"DEBUG: Respuesta de OpenAI recibida: {ai_response['content'][:50]}...")
                
                # Guardar respuesta de la IA
                ai_msg = AIChatMessage.objects.create(
                    session=session,
                    role='assistant',
                    content=ai_response['content'],
                    tokens_used=ai_response.get('tokens_used', 0)
                )
                print(f"DEBUG: Respuesta de IA guardada con ID: {ai_msg.id}")
                
                messages.success(request, 'Mensaje enviado correctamente.')
                
            except Exception as e:
                print(f"DEBUG: Error en OpenAI API: {str(e)}")
                messages.error(request, f'Error al comunicarse con la IA: {str(e)}')
            
            return redirect('ai_chat_session', session_id=session.id)
        else:
            print(f"DEBUG: Formulario no válido. Errores: {form.errors}")
            print(f"DEBUG: Errores por campo: {dict(form.errors)}")
            for field, errors in form.errors.items():
                print(f"DEBUG: Campo '{field}': {errors}")
    else:
        form = AIChatMessageForm()
    
    context = {
        'session': session,
        'messages_list': messages_list,
        'form': form,
        'config': config,
    }
    
    return render(request, 'tickets/ai_chat_session.html', context)


@login_required
def ai_chat_delete_session_view(request, session_id):
    """Vista para eliminar una sesión de chat"""
    session = get_object_or_404(AIChatSession, id=session_id, user=request.user)
    
    if request.method == 'POST':
        session.is_active = False
        session.save()
        messages.success(request, f'Conversación "{session.title}" eliminada.')
        return redirect('ai_chat_list')
    
    context = {
        'session': session,
    }
    
    return render(request, 'tickets/ai_chat_delete_session.html', context)


@login_required
def ai_chat_debug_view(request, session_id):
    """Vista de debug para chat IA"""
    config = SystemConfiguration.get_config()
    
    # Obtener la sesión
    session = get_object_or_404(AIChatSession, id=session_id, user=request.user, is_active=True)
    
    # Obtener mensajes de la sesión
    messages_list = session.messages.all()
    
    # Formulario para nuevo mensaje
    if request.method == 'POST':
        print(f"DEBUG: POST recibido para sesión {session_id}")
        print(f"DEBUG: Datos POST: {request.POST}")
        
        form = AIChatMessageForm(request.POST)
        if form.is_valid():
            user_message = form.cleaned_data['message']
            print(f"DEBUG: Mensaje del usuario: {user_message}")
            
            # Guardar mensaje del usuario
            user_msg = AIChatMessage.objects.create(
                session=session,
                role='user',
                content=user_message
            )
            print(f"DEBUG: Mensaje del usuario guardado con ID: {user_msg.id}")
            
            # Llamar a la API de OpenAI
            try:
                print("DEBUG: Llamando a OpenAI API...")
                ai_response = call_openai_api(session, user_message)
                print(f"DEBUG: Respuesta de OpenAI recibida: {ai_response['content'][:50]}...")
                
                # Guardar respuesta de la IA
                ai_msg = AIChatMessage.objects.create(
                    session=session,
                    role='assistant',
                    content=ai_response['content'],
                    tokens_used=ai_response.get('tokens_used', 0)
                )
                print(f"DEBUG: Respuesta de IA guardada con ID: {ai_msg.id}")
                
                messages.success(request, 'Mensaje enviado correctamente.')
                
            except Exception as e:
                print(f"DEBUG: Error en OpenAI API: {str(e)}")
                messages.error(request, f'Error al comunicarse con la IA: {str(e)}')
            
            return redirect('ai_chat_debug', session_id=session.id)
        else:
            print(f"DEBUG: Formulario no válido: {form.errors}")
    else:
        form = AIChatMessageForm()
    
    context = {
        'session': session,
        'messages_list': messages_list,
        'form': form,
        'page_title': f'Debug Chat IA - {session.title}',
    }
    
    return render(request, 'tickets/ai_chat_debug.html', context)


@login_required
def ai_chat_simple_view(request, session_id):
    """Vista simple del chat IA para debugging"""
    config = SystemConfiguration.get_config()
    
    # Verificar configuración
    if not config.ai_chat_enabled or not config.openai_api_key:
        messages.warning(request, 'El chat con IA no está disponible.')
        return redirect('dashboard')
    
    # Obtener sesión
    session = get_object_or_404(AIChatSession, id=session_id, user=request.user, is_active=True)
    messages_list = session.messages.all()
    
    # Procesar formulario
    if request.method == 'POST':
        print(f"=== POST SIMPLE RECIBIDO ===")
        print(f"request.POST: {dict(request.POST)}")
        print(f"message value: '{request.POST.get('message', 'NO_ENCONTRADO')}'")
        
        message_content = request.POST.get('message', '').strip()
        
        if len(message_content) >= 5:
            try:
                # Guardar mensaje del usuario
                user_msg = AIChatMessage.objects.create(
                    session=session,
                    role='user',
                    content=message_content
                )
                print(f"Usuario mensaje guardado: {user_msg.id}")
                
                # Llamar a OpenAI
                ai_response = call_openai_api(session, message_content)
                
                # Guardar respuesta IA
                ai_msg = AIChatMessage.objects.create(
                    session=session,
                    role='assistant',
                    content=ai_response['content'],
                    tokens_used=ai_response.get('tokens_used', 0)
                )
                print(f"IA mensaje guardado: {ai_msg.id}")
                
                messages.success(request, 'Mensaje enviado correctamente.')
                return redirect('ai_chat_simple', session_id=session.id)
                
            except Exception as e:
                print(f"Error: {e}")
                messages.error(request, f'Error: {str(e)}')
        else:
            messages.error(request, 'El mensaje debe tener al menos 5 caracteres.')
    
    # Crear formulario vacío para los errores
    form = AIChatMessageForm()
    
    context = {
        'session': session,
        'messages_list': messages_list,
        'form': form,
        'page_title': f'Chat Simple - {session.title}',
    }
    
    return render(request, 'tickets/ai_chat_simple.html', context)


@login_required
def ai_chat_ajax_send_message(request, session_id):
    """Vista AJAX para enviar mensajes al chat IA"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        config = SystemConfiguration.get_config()
        
        # Verificar configuración
        if not config.ai_chat_enabled or not config.openai_api_key:
            return JsonResponse({'error': 'Chat IA no disponible'}, status=400)
        
        # Obtener sesión
        try:
            session = AIChatSession.objects.get(id=session_id, user=request.user, is_active=True)
        except AIChatSession.DoesNotExist:
            return JsonResponse({'error': 'Sesión no encontrada'}, status=404)
        
        # Obtener mensaje
        message_content = request.POST.get('message', '').strip()
        
        if len(message_content) < 5:
            return JsonResponse({'error': 'El mensaje debe tener al menos 5 caracteres'}, status=400)
        
        # Guardar mensaje del usuario
        user_msg = AIChatMessage.objects.create(
            session=session,
            role='user',
            content=message_content
        )
        
        # Llamar a OpenAI
        ai_response = call_openai_api(session, message_content)
        
        # Guardar respuesta IA
        ai_msg = AIChatMessage.objects.create(
            session=session,
            role='assistant',
            content=ai_response['content'],
            tokens_used=ai_response.get('tokens_used', 0)
        )
        
        return JsonResponse({
            'success': True,
            'user_message': {
                'id': user_msg.id,
                'content': user_msg.content,
                'created_at': user_msg.created_at.strftime('%H:%M')
            },
            'ai_message': {
                'id': ai_msg.id,
                'content': ai_msg.content,
                'created_at': ai_msg.created_at.strftime('%H:%M'),
                'tokens_used': ai_msg.tokens_used
            }
        })
        
    except Exception as e:
        import traceback
        print(f"Error completo: {traceback.format_exc()}")
        return JsonResponse({'error': f'Error al procesar mensaje: {str(e)}'}, status=500)


@login_required
def ai_chat_live_view(request, session_id):
    """Vista mejorada del chat IA con AJAX y UX mejorada"""
    config = SystemConfiguration.get_config()
    
    # Verificar configuración
    if not config.ai_chat_enabled:
        messages.warning(request, 'El chat con IA no está habilitado en el sistema.')
        return redirect('dashboard')
    
    if not config.openai_api_key:
        messages.warning(request, 'El chat con IA no está configurado. Contacta al administrador.')
        return redirect('dashboard')
    
    # Obtener sesión
    session = get_object_or_404(AIChatSession, id=session_id, user=request.user, is_active=True)
    
    # Obtener mensajes
    messages_list = session.messages.all()
    
    context = {
        'session': session,
        'messages_list': messages_list,
        'page_title': f'{session.title} - Chat IA',
    }
    
    return render(request, 'tickets/ai_chat_live.html', context)


@login_required
def ai_chat_modern_view(request, session_id):
    """Vista del chat IA con diseño moderno (Telegram/Discord style)"""
    config = SystemConfiguration.get_config()
    
    # Verificar configuración
    if not config.ai_chat_enabled:
        messages.warning(request, 'El chat con IA no está habilitado en el sistema.')
        return redirect('dashboard')
    
    if not config.openai_api_key:
        messages.warning(request, 'El chat con IA no está configurado. Contacta al administrador.')
        return redirect('dashboard')
    
    # Obtener sesión
    session = get_object_or_404(AIChatSession, id=session_id, user=request.user, is_active=True)
    
    # Obtener mensajes
    messages_list = session.messages.all()
    
    context = {
        'session': session,
        'messages': messages_list,  # Cambio de nombre para el template moderno
        'page_title': f'{session.title} - Chat IA',
    }
    
    return render(request, 'tickets/ai_chat_modern.html', context)


# Helper function for AI analysis
def call_openai_api(session, message_content):
    """
    Función auxiliar para llamar a la API de OpenAI
    """
    config = SystemConfiguration.get_config()
    
    if not config.openai_api_key:
        raise Exception('API key de OpenAI no configurada')
    
    from openai import OpenAI
    client = OpenAI(api_key=config.openai_api_key)
    
    # Obtener historial de mensajes para contexto
    previous_messages = session.messages.order_by('created_at')[:10]  # Últimos 10 mensajes
    messages = [{"role": "system", "content": "Eres un asistente IA útil y conversacional."}]
    
    # Agregar mensajes previos para contexto
    for msg in previous_messages:
        messages.append({
            "role": msg.role,
            "content": msg.content
        })
    
    # Agregar mensaje actual
    messages.append({
        "role": "user",
        "content": message_content
    })
    
    response = client.chat.completions.create(
        model=config.openai_model or 'gpt-4o-mini',
        messages=messages,
        max_tokens=2000,
        temperature=0.7
    )
    
    return {
        'content': response.choices[0].message.content,
        'tokens_used': response.usage.total_tokens
    }


@login_required
def improve_ticket_with_ai(request, ticket_id):
    """Vista AJAX para mejorar el título y descripción de un ticket usando IA"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Verificar permisos del ticket
        if is_agent(request.user):
            ticket = get_object_or_404(Ticket, pk=ticket_id)
        else:
            # Los usuarios pueden solo editar sus propios tickets o tickets de su empresa/proyectos
            user_projects = request.user.assigned_projects.all()
            user_company = None
            
            try:
                user_company = request.user.profile.company
            except:
                pass
            
            query_conditions = Q(created_by=request.user)
            if user_projects.exists():
                query_conditions |= Q(project__in=user_projects)
            if user_company:
                query_conditions |= Q(company=user_company)
            
            ticket = get_object_or_404(Ticket, pk=ticket_id)
            if not Ticket.objects.filter(pk=ticket_id).filter(query_conditions).exists():
                return JsonResponse({'error': 'No tienes permisos para mejorar este ticket'}, status=403)
        
        # Verificar configuración de OpenAI
        config = SystemConfiguration.get_config()
        if not config.openai_api_key:
            return JsonResponse({'error': 'API key de OpenAI no configurada'}, status=500)
        
        # Obtener tipo de mejora solicitada
        improvement_type = request.POST.get('type', 'both')  # 'title', 'description', 'both'
        
        # Preparar el prompt según el tipo de mejora
        if improvement_type == 'title':
            prompt = f"""
            Mejora el siguiente título de ticket de soporte técnico para que sea más claro, específico y profesional:

            Título actual: "{ticket.title}"
            Descripción del problema: "{ticket.description}"
            
            Proporciona ÚNICAMENTE un título mejorado, sin explicaciones adicionales.
            El título debe ser:
            - Claro y específico
            - Profesional pero comprensible
            - Máximo 100 caracteres
            - Debe reflejar exactamente el problema descrito
            """
        elif improvement_type == 'description':
            prompt = f"""
            Mejora la siguiente descripción de ticket de soporte técnico para que sea más clara, detallada y estructurada:

            Título: "{ticket.title}"
            Descripción actual: "{ticket.description}"
            
            Proporciona ÚNICAMENTE una descripción mejorada, sin explicaciones adicionales.
            La descripción debe ser:
            - Clara y bien estructurada
            - Incluir pasos para reproducir si es relevante
            - Mencionar el impacto del problema
            - Ser específica y detallada
            - Mantener el contenido original pero mejor organizado
            """
        else:  # both
            prompt = f"""
            Mejora el siguiente ticket de soporte técnico para que el título y descripción sean más claros, específicos y profesionales:

            Título actual: "{ticket.title}"
            Descripción actual: "{ticket.description}"
            
            Responde ÚNICAMENTE con el siguiente formato JSON sin texto adicional:
            {{
                "title": "título mejorado aquí",
                "description": "descripción mejorada aquí"
            }}
            
            Criterios:
            - Título: Claro, específico, profesional, máximo 100 caracteres
            - Descripción: Clara, estructurada, detallada, bien organizada
            - Mantener el contenido original pero mejorado
            """
        
        # Llamar a OpenAI
        try:
            from openai import OpenAI
            client = OpenAI(api_key=config.openai_api_key)
        except Exception as e:
            return JsonResponse({'error': f'Error al inicializar cliente OpenAI: {str(e)}'}, status=500)
        
        try:
            response = client.chat.completions.create(
                model=config.openai_model or 'gpt-4o-mini',
                messages=[
                    {"role": "system", "content": "Eres un experto en soporte técnico que mejora la comunicación de tickets para hacerlos más claros y profesionales."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=1000,
                temperature=0.3
            )
            
            ai_response = response.choices[0].message.content.strip()
        except Exception as e:
            return JsonResponse({'error': f'Error al llamar a OpenAI: {str(e)}'}, status=500)
        
        # Procesar respuesta según el tipo
        if improvement_type == 'both':
            try:
                import json
                # Limpiar respuesta si tiene marcadores de código
                if ai_response.startswith('```json'):
                    ai_response = ai_response.replace('```json', '').replace('```', '').strip()
                elif ai_response.startswith('```'):
                    ai_response = ai_response.replace('```', '').strip()
                
                result = json.loads(ai_response)
                return JsonResponse({
                    'success': True,
                    'title': result.get('title', ''),
                    'description': result.get('description', ''),
                    'tokens_used': response.usage.total_tokens
                })
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Error al procesar la respuesta de IA'}, status=500)
        else:
            return JsonResponse({
                'success': True,
                improvement_type: ai_response,
                'tokens_used': response.usage.total_tokens
            })
            
    except Exception as e:
        return JsonResponse({'error': f'Error al mejorar ticket: {str(e)}'}, status=500)


@login_required
@user_passes_test(is_agent, login_url='/')
def search_company_info_with_ai(request, company_id):
    """Vista AJAX para buscar información de una empresa en la web usando IA"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        company = get_object_or_404(Company, pk=company_id)
        
        # Verificar configuración de OpenAI
        config = SystemConfiguration.get_config()
        if not config.openai_api_key:
            return JsonResponse({'error': 'API key de OpenAI no configurada'}, status=500)
        
        # Obtener nombre de la empresa del request o usar el actual
        company_name = request.POST.get('company_name', company.name).strip()
        
        if not company_name:
            return JsonResponse({'error': 'Nombre de empresa requerido'}, status=400)
        
        # Prompt para buscar información de la empresa
        prompt = f"""
        Busca información detallada sobre la empresa "{company_name}". 
        
        Proporciona la información en el siguiente formato JSON exacto (sin texto adicional):
        {{
            "name": "nombre oficial completo de la empresa",
            "website": "sitio web principal (con https://)",
            "email": "email principal de contacto",
            "phone": "teléfono principal",
            "address": "dirección completa de la sede principal",
            "description": "descripción detallada de la empresa, sus servicios y actividades",
            "industry": "sector o industria principal",
            "founded": "año de fundación si está disponible",
            "employees": "número aproximado de empleados si está disponible",
            "social_media": {{
                "linkedin": "URL de LinkedIn",
                "twitter": "URL de Twitter",
                "facebook": "URL de Facebook"
            }},
            "additional_info": "información adicional relevante como premios, certificaciones, etc."
        }}
        
        IMPORTANTE:
        - Si no encuentras información específica, usa "No disponible" para ese campo
        - Asegúrate de que el website tenga formato completo (https://)
        - La descripción debe ser informativa y profesional
        - Incluye solo información verificable y actualizada
        """
        
        # Llamar a OpenAI
        from openai import OpenAI
        client = OpenAI(api_key=config.openai_api_key)
        
        response = client.chat.completions.create(
            model=config.openai_model or 'gpt-4o-mini',
            messages=[
                {"role": "system", "content": "Eres un experto investigador de empresas que busca información precisa y actualizada sobre compañías usando fuentes públicas disponibles. Respondes únicamente con JSON válido."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1500,
            temperature=0.3
        )
        
        ai_response = response.choices[0].message.content.strip()
        
        # Limpiar respuesta si tiene marcadores de código
        if ai_response.startswith('```json'):
            ai_response = ai_response.replace('```json', '').replace('```', '').strip()
        elif ai_response.startswith('```'):
            ai_response = ai_response.replace('```', '').strip()
        
        try:
            import json
            result = json.loads(ai_response)
            
            # Validar que los campos requeridos estén presentes
            required_fields = ['name', 'website', 'email', 'phone', 'address', 'description']
            for field in required_fields:
                if field not in result:
                    result[field] = 'No disponible'
            
            return JsonResponse({
                'success': True,
                'data': result,
                'tokens_used': response.usage.total_tokens
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Error al procesar la respuesta de IA'}, status=500)
            
    except Exception as e:
        return JsonResponse({'error': f'Error al buscar información: {str(e)}'}, status=500)


@login_required
@user_passes_test(is_agent, login_url='/')
def search_company_info_general(request):
    """Vista AJAX para buscar información de una empresa sin requerir ID específico"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Verificar configuración de OpenAI
        config = SystemConfiguration.get_config()
        if not config.openai_api_key:
            return JsonResponse({'error': 'API key de OpenAI no configurada'}, status=500)
        
        # Obtener nombre de la empresa del request
        company_name = request.POST.get('company_name', '').strip()
        
        if not company_name:
            return JsonResponse({'error': 'Nombre de empresa requerido'}, status=400)
        
        # Prompt para buscar información de la empresa
        prompt = f"""
        Busca información detallada sobre la empresa "{company_name}". 
        
        Proporciona la información en el siguiente formato JSON exacto (sin texto adicional):
        {{
            "name": "nombre oficial completo de la empresa",
            "website": "sitio web principal (con https://)",
            "email": "email principal de contacto",
            "phone": "teléfono principal",
            "address": "dirección completa de la sede principal",
            "description": "descripción detallada de la empresa, sus servicios y actividades",
            "industry": "sector o industria principal",
            "founded": "año de fundación si está disponible",
            "employees": "número aproximado de empleados si está disponible",
            "social_media": {{
                "linkedin": "URL de LinkedIn",
                "twitter": "URL de Twitter",
                "facebook": "URL de Facebook"
            }},
            "additional_info": "información adicional relevante como premios, certificaciones, etc."
        }}
        
        IMPORTANTE:
        - Si no encuentras información específica, usa "No disponible" para ese campo
        - Asegúrate de que el website tenga formato completo (https://)
        - La descripción debe ser informativa y profesional
        - Incluye solo información verificable y actualizada
        """
        
        # Llamar a OpenAI
        from openai import OpenAI
        client = OpenAI(api_key=config.openai_api_key)
        
        response = client.chat.completions.create(
            model=config.openai_model or 'gpt-4o-mini',
            messages=[
                {"role": "system", "content": "Eres un experto investigador de empresas que busca información precisa y actualizada sobre compañías usando fuentes públicas disponibles. Respondes únicamente con JSON válido."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1500,
            temperature=0.3
        )
        
        ai_response = response.choices[0].message.content.strip()
        
        # Limpiar respuesta si tiene marcadores de código
        if ai_response.startswith('```json'):
            ai_response = ai_response.replace('```json', '').replace('```', '').strip()
        elif ai_response.startswith('```'):
            ai_response = ai_response.replace('```', '').strip()
        
        try:
            import json
            result = json.loads(ai_response)
            
            # Validar que los campos requeridos estén presentes
            required_fields = ['name', 'website', 'email', 'phone', 'address', 'description']
            for field in required_fields:
                if field not in result:
                    result[field] = 'No disponible'
            
            return JsonResponse({
                'success': True,
                'data': result,
                'tokens_used': response.usage.total_tokens
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Error al procesar la respuesta de IA'}, status=500)
            
    except Exception as e:
        return JsonResponse({'error': f'Error al buscar información: {str(e)}'}, status=500)


@login_required
@user_passes_test(is_agent, login_url='/')
def enhance_contact_with_ai(request, contact_id):
    """Vista AJAX para mejorar información de un contacto usando IA"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        contact = get_object_or_404(Contact, pk=contact_id)
        
        # Verificar configuración de OpenAI
        config = SystemConfiguration.get_config()
        if not config.openai_api_key:
            return JsonResponse({'error': 'API key de OpenAI no configurada'}, status=500)
        
        # Obtener datos actuales del contacto del request
        import json
        contact_data = json.loads(request.body)
        
        # Construir información existente
        existing_info = []
        if contact_data.get('name'):
            existing_info.append(f"Nombre: {contact_data['name']}")
        if contact_data.get('email'):
            existing_info.append(f"Email: {contact_data['email']}")
        if contact_data.get('phone'):
            existing_info.append(f"Teléfono: {contact_data['phone']}")
        if contact_data.get('position'):
            existing_info.append(f"Posición: {contact_data['position']}")
        if contact_data.get('company'):
            existing_info.append(f"Empresa: {contact_data['company']}")
        if contact_data.get('notes'):
            existing_info.append(f"Notas actuales: {contact_data['notes']}")
        
        if not existing_info:
            return JsonResponse({'error': 'No hay información suficiente para mejorar'}, status=400)
        
        existing_info_text = "\n".join(existing_info)
        
        # Prompt para mejorar información del contacto
        prompt = f"""
        Tienes la siguiente información parcial de un contacto profesional:
        
        {existing_info_text}
        
        Basándote en esta información, busca y proporciona datos adicionales que podrían ser útiles para completar el perfil del contacto. 
        
        Proporciona la respuesta en el siguiente formato JSON exacto (sin texto adicional):
        {{
            "full_name": "nombre completo si es diferente o más completo que el actual",
            "position": "título o posición profesional más específica o completa",
            "professional_title": "título profesional o especialización",
            "email": "email profesional si no está disponible",
            "phone": "teléfono profesional si no está disponible",
            "linkedin": "URL del perfil de LinkedIn si está disponible",
            "company_info": {{
                "name": "nombre completo y correcto de la empresa",
                "industry": "industria o sector de la empresa", 
                "size": "tamaño aproximado de la empresa",
                "location": "ubicación principal de la empresa"
            }},
            "additional_info": "información adicional relevante sobre el contacto, su rol, experiencia, etc.",
            "suggested_notes": "sugerencias de notas profesionales para agregar al contacto basadas en la información encontrada"
        }}
        
        IMPORTANTE:
        - Si no encuentras información específica o no puedes mejorar un campo, usa "No disponible"
        - Proporciona solo información que sea profesionalmente relevante
        - Las sugerencias de notas deben ser útiles para seguimiento comercial o profesional
        - Usa fuentes públicas y profesionales para la información
        - Sé preciso y evita especulaciones
        """
        
        # Llamar a OpenAI
        from openai import OpenAI
        client = OpenAI(api_key=config.openai_api_key)
        
        response = client.chat.completions.create(
            model=config.openai_model or 'gpt-4o-mini',
            messages=[
                {"role": "system", "content": "Eres un experto asistente de CRM que ayuda a completar información de contactos profesionales usando fuentes públicas y conocimiento general del mundo empresarial. Respondes únicamente con JSON válido."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1200,
            temperature=0.3
        )
        
        ai_response = response.choices[0].message.content.strip()
        
        # Limpiar respuesta si tiene marcadores de código
        if ai_response.startswith('```json'):
            ai_response = ai_response.replace('```json', '').replace('```', '').strip()
        elif ai_response.startswith('```'):
            ai_response = ai_response.replace('```', '').strip()
        
        try:
            result = json.loads(ai_response)
            
            # Validar que los campos requeridos estén presentes
            required_fields = ['full_name', 'position', 'email', 'phone', 'additional_info', 'suggested_notes']
            for field in required_fields:
                if field not in result:
                    result[field] = 'No disponible'
            
            return JsonResponse({
                'success': True,
                'data': result,
                'tokens_used': response.usage.total_tokens
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Error al procesar la respuesta de IA'}, status=500)
            
    except Exception as e:
        return JsonResponse({'error': f'Error al mejorar contacto: {str(e)}'}, status=500)


# =============================================
# VISTAS DE CONCEPTOS
# =============================================

@login_required
@user_passes_test(is_agent)
def concept_list_view(request):
    """Vista para listar todos los conceptos"""
    concepts = Concept.objects.all().order_by('order', 'term')
    
    context = {
        'concepts': concepts,
        'title': 'Gestión de Conceptos'
    }
    return render(request, 'tickets/concept_list.html', context)


@login_required
@user_passes_test(is_agent)
def concept_create_view(request):
    """Vista para crear un nuevo concepto"""
    if request.method == 'POST':
        form = ConceptForm(request.POST)
        if form.is_valid():
            concept = form.save(commit=False)
            concept.created_by = request.user
            concept.save()
            messages.success(request, f'Concepto "{concept.term}" creado exitosamente.')
            return redirect('concept_list')
    else:
        form = ConceptForm()
    
    context = {
        'form': form,
        'title': 'Crear Concepto'
    }
    return render(request, 'tickets/concept_form.html', context)


@login_required
@user_passes_test(is_agent)
def concept_edit_view(request, pk):
    """Vista para editar un concepto existente"""
    concept = get_object_or_404(Concept, pk=pk)
    
    if request.method == 'POST':
        form = ConceptForm(request.POST, instance=concept)
        if form.is_valid():
            form.save()
            messages.success(request, f'Concepto "{concept.term}" actualizado exitosamente.')
            return redirect('concept_list')
    else:
        form = ConceptForm(instance=concept)
    
    context = {
        'form': form,
        'concept': concept,
        'title': 'Editar Concepto'
    }
    return render(request, 'tickets/concept_form.html', context)


@login_required
@user_passes_test(is_agent)
def concept_delete_view(request, pk):
    """Vista para eliminar un concepto"""
    concept = get_object_or_404(Concept, pk=pk)
    
    if request.method == 'POST':
        concept_term = concept.term
        concept.delete()
        messages.success(request, f'Concepto "{concept_term}" eliminado exitosamente.')
        return redirect('concept_list')
    
    context = {
        'concept': concept,
        'title': 'Eliminar Concepto'
    }
    return render(request, 'tickets/concept_delete.html', context)


@login_required
def concept_detail_view(request, pk):
    """Vista para ver los detalles de un concepto"""
    concept = get_object_or_404(Concept, pk=pk)
    
    context = {
        'concept': concept,
        'title': f'Concepto: {concept.term}'
    }
    return render(request, 'tickets/concept_detail.html', context)


def public_concepts_view(request):
    """Vista pública para mostrar el glosario de conceptos"""
    concepts = Concept.objects.filter(is_active=True).order_by('order', 'term')
    
    context = {
        'concepts': concepts,
        'title': 'Glosario de Conceptos'
    }
    return render(request, 'tickets/public_concepts.html', context)


# =====================================================
# FUNCIONALIDADES DE IA PARA BLOG
# =====================================================

@login_required
@require_http_methods(["POST"])
def blog_ai_improve_content(request, pk):
    """Mejorar contenido del artículo con IA"""
    try:
        from .ai_utils import AIContentOptimizer
        from .utils import is_agent
        
        post = get_object_or_404(BlogPost, pk=pk)
        
        # Verificar permisos
        if not is_agent(request.user):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        optimizer = AIContentOptimizer()
        result = optimizer.improve_content(post.title, post.content)
        
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def blog_ai_generate_meta_description(request, pk):
    """Generar meta descripción con IA"""
    try:
        from .ai_utils import AIContentOptimizer
        from .utils import is_agent
        
        post = get_object_or_404(BlogPost, pk=pk)
        
        # Verificar permisos
        if not is_agent(request.user):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        optimizer = AIContentOptimizer()
        result = optimizer.generate_meta_description(post.title, post.content)
        
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def blog_ai_suggest_titles(request, pk):
    """Sugerir títulos alternativos con IA"""
    try:
        from .ai_utils import AIContentOptimizer
        from .utils import is_agent
        
        post = get_object_or_404(BlogPost, pk=pk)
        
        # Verificar permisos
        if not is_agent(request.user):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        optimizer = AIContentOptimizer()
        result = optimizer.suggest_titles(post.content)
        
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def blog_ai_generate_tags(request, pk):
    """Generar etiquetas con IA"""
    try:
        from .ai_utils import AIContentOptimizer
        from .utils import is_agent
        
        post = get_object_or_404(BlogPost, pk=pk)
        
        # Verificar permisos
        if not is_agent(request.user):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        optimizer = AIContentOptimizer()
        result = optimizer.generate_tags(post.title, post.content)
        
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def blog_ai_analyze_readability(request, pk):
    """Analizar legibilidad del contenido con IA"""
    try:
        from .ai_utils import AIContentOptimizer
        from .utils import is_agent
        
        post = get_object_or_404(BlogPost, pk=pk)
        
        # Verificar permisos
        if not is_agent(request.user):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        optimizer = AIContentOptimizer()
        result = optimizer.analyze_readability(post.content)
        
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


@login_required
def blog_ai_test_connection(request):
    """Probar conexión con IA"""
    try:
        from .ai_utils import test_ai_connection
        from .utils import is_agent
        
        # Verificar permisos
        if not is_agent(request.user):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        result = test_ai_connection()
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error interno: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def blog_ai_improve_complete_article(request, pk):
    """Mejorar completamente el artículo (título + contenido) con IA"""
    try:
        from .ai_utils import AIContentOptimizer
        from .utils import is_agent
        
        post = get_object_or_404(BlogPost, pk=pk)
        
        # Verificar permisos
        if not is_agent(request.user):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        optimizer = AIContentOptimizer()
        result = optimizer.improve_complete_article(post.title, post.content)
        
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def blog_ai_generate_image(request, pk):
    """Generar imagen para el artículo con IA basándose en el título"""
    try:
        from .ai_utils import AIContentOptimizer
        from .utils import is_agent
        
        post = get_object_or_404(BlogPost, pk=pk)
        
        # Verificar permisos
        if not is_agent(request.user):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        optimizer = AIContentOptimizer()
        result = optimizer.generate_article_image(post.title, post.content)
        
        if result.get("success"):
            # Actualizar el post con la nueva imagen
            post.featured_image = result["file_path"]
            post.save()
            
            return JsonResponse({
                "success": True,
                "image_url": result["local_url"],
                "file_path": result["file_path"],
                "message": "Imagen generada y guardada exitosamente"
            })
        else:
            return JsonResponse(result)
            
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def blog_ai_apply_improvements(request, pk):
    """Aplicar mejoras sugeridas por IA al artículo"""
    try:
        from .utils import is_agent
        
        post = get_object_or_404(BlogPost, pk=pk)
        
        # Verificar permisos
        if not is_agent(request.user):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        data = json.loads(request.body)
        
        # Aplicar cambios si están presentes
        if 'content' in data and data['content']:
            post.content = data['content']
        
        if 'meta_description' in data and data['meta_description']:
            post.meta_description = data['meta_description']
        
        if 'tags' in data and data['tags']:
            post.tags = data['tags']
        
        if 'title' in data and data['title']:
            post.title = data['title']
            # Regenerar slug si se cambia el título
            from django.utils.text import slugify
            post.slug = slugify(post.title)
        
        post.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Artículo actualizado exitosamente con las mejoras de IA'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


# =============================================================================
# VISTAS DE EXÁMENES
# =============================================================================

@login_required
def exam_list(request):
    """Lista de exámenes"""
    exams = Exam.objects.all().order_by('-created_at')
    
    # Agregar información del último intento del usuario para cada examen
    for exam in exams:
        exam.user_last_attempt = exam.attempts.filter(
            user=request.user
        ).order_by('-completed_at').first()
    
    context = {
        'exams': exams
    }
    return render(request, 'tickets/exam_list.html', context)


@login_required
def exam_create(request):
    """Crear nuevo examen (agentes y profesores)"""
    from .utils import can_manage_courses
    
    if not can_manage_courses(request.user):
        messages.error(request, 'No tienes permisos para crear exámenes')
        return redirect('exam_list')
    
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        passing_score = request.POST.get('passing_score', 70)
        time_limit = request.POST.get('time_limit') or None
        is_public = request.POST.get('is_public') == 'on'
        
        exam = Exam.objects.create(
            title=title,
            description=description,
            passing_score=int(passing_score),
            time_limit=int(time_limit) if time_limit else None,
            is_public=is_public,
            created_by=request.user
        )
        
        messages.success(request, f'Examen "{exam.title}" creado exitosamente')
        return redirect('exam_detail', pk=exam.pk)
    
    return render(request, 'tickets/exam_form.html')


@login_required
def exam_detail(request, pk):
    """Detalle del examen"""
    from .utils import is_agent
    
    exam = get_object_or_404(Exam, pk=pk)
    questions = exam.questions.all().order_by('order')
    
    # Filtrar intentos según el tipo de usuario
    if is_agent(request.user) or exam.created_by == request.user:
        # Agentes y creadores ven todos los intentos
        attempts = exam.attempts.all().order_by('-completed_at')[:10]
    else:
        # Usuarios normales solo ven sus propios intentos
        attempts = exam.attempts.filter(user=request.user).order_by('-completed_at')[:5]
    
    # Obtener el último intento del usuario actual
    user_last_attempt = None
    if request.user.is_authenticated:
        user_last_attempt = exam.attempts.filter(
            user=request.user
        ).order_by('-completed_at').first()
    
    context = {
        'exam': exam,
        'questions': questions,
        'attempts': attempts,
        'user_last_attempt': user_last_attempt,
        'can_edit': is_agent(request.user) or exam.created_by == request.user
    }
    return render(request, 'tickets/exam_detail.html', context)


@login_required
def exam_edit(request, pk):
    """Editar examen"""
    from .utils import can_manage_courses
    
    exam = get_object_or_404(Exam, pk=pk)
    
    if not (can_manage_courses(request.user) or exam.created_by == request.user):
        messages.error(request, 'No tienes permisos para editar este examen')
        return redirect('exam_detail', pk=exam.pk)
    
    if request.method == 'POST':
        exam.title = request.POST.get('title')
        exam.description = request.POST.get('description', '')
        exam.passing_score = int(request.POST.get('passing_score', 70))
        exam.time_limit = int(request.POST.get('time_limit')) if request.POST.get('time_limit') else None
        exam.is_public = request.POST.get('is_public') == 'on'
        exam.save()
        
        messages.success(request, 'Examen actualizado exitosamente')
        return redirect('exam_detail', pk=exam.pk)
    
    context = {'exam': exam}
    return render(request, 'tickets/exam_form.html', context)


@login_required
def exam_delete(request, pk):
    """Eliminar examen"""
    from .utils import can_manage_courses
    
    exam = get_object_or_404(Exam, pk=pk)
    
    if not (can_manage_courses(request.user) or exam.created_by == request.user):
        messages.error(request, 'No tienes permisos para eliminar este examen')
        return redirect('exam_detail', pk=exam.pk)
    
    if request.method == 'POST':
        exam_title = exam.title
        exam.delete()
        messages.success(request, f'Examen "{exam_title}" eliminado exitosamente')
        return redirect('exam_list')
    
    context = {'exam': exam}
    return render(request, 'tickets/exam_delete.html', context)


def exam_take_public(request, token):
    """Tomar examen público (sin autenticación)"""
    exam = get_object_or_404(Exam, public_token=token, is_public=True)
    
    if request.method == 'POST':
        # Registrar datos del participante
        participant_name = request.POST.get('participant_name')
        participant_email = request.POST.get('participant_email')
        
        if not participant_name or not participant_email:
            messages.error(request, 'Nombre y email son requeridos')
            return render(request, 'tickets/exam_public_start.html', {'exam': exam})
        
        # Redirigir al examen con los datos en sesión
        request.session['exam_participant'] = {
            'name': participant_name,
            'email': participant_email,
            'exam_id': exam.id,
            'started_at': timezone.now().isoformat()
        }
        
        return redirect('exam_take_questions', token=token)
    
    return render(request, 'tickets/exam_public_start.html', {'exam': exam})


def exam_take_questions(request, token):
    """Mostrar preguntas del examen público"""
    exam = get_object_or_404(Exam, public_token=token, is_public=True)
    
    # Verificar datos del participante en sesión
    participant_data = request.session.get('exam_participant')
    if not participant_data or participant_data.get('exam_id') != exam.id:
        return redirect('exam_take_public', token=token)
    
    questions = exam.questions.all().order_by('order')
    
    if request.method == 'POST':
        # Procesar respuestas
        from django.utils import timezone
        from datetime import datetime
        
        started_at = datetime.fromisoformat(participant_data['started_at'])
        completed_at = timezone.now()
        time_taken = int((completed_at - started_at).total_seconds())
        
        # Crear intento de examen
        attempt = ExamAttempt.objects.create(
            exam=exam,
            participant_name=participant_data['name'],
            participant_email=participant_data['email'],
            total_questions=questions.count(),
            started_at=started_at,
            completed_at=completed_at,
            time_taken=time_taken,
            score=0,  # Se calculará después
            correct_answers=0
        )
        
        # Procesar respuestas
        correct_count = 0
        for question in questions:
            selected_option = request.POST.get(f'question_{question.id}')
            if selected_option:
                is_correct = selected_option == question.correct_option
                if is_correct:
                    correct_count += 1
                
                ExamAnswer.objects.create(
                    attempt=attempt,
                    question=question,
                    selected_option=selected_option,
                    is_correct=is_correct
                )
        
        # Calcular puntuación
        score = (correct_count / questions.count()) * 100 if questions.count() > 0 else 0
        attempt.correct_answers = correct_count
        attempt.score = score
        attempt.passed = score >= exam.passing_score
        attempt.save()
        
        # Limpiar sesión
        del request.session['exam_participant']
        
        return redirect('exam_results', attempt_id=attempt.id)
    
    context = {
        'exam': exam,
        'questions': questions,
        'participant_data': participant_data
    }
    return render(request, 'tickets/exam_take.html', context)


@login_required
def exam_take_authenticated(request, pk):
    """Tomar examen autenticado"""
    exam = get_object_or_404(Exam, pk=pk)
    questions = exam.questions.all().order_by('order')
    
    if request.method == 'POST':
        # Procesar respuestas
        from django.utils import timezone
        
        started_at = timezone.now() - timezone.timedelta(
            seconds=int(request.POST.get('time_taken', 0))
        )
        completed_at = timezone.now()
        time_taken = int(request.POST.get('time_taken', 0))
        
        # Crear intento de examen
        attempt = ExamAttempt.objects.create(
            exam=exam,
            user=request.user,
            participant_name=f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
            participant_email=request.user.email,
            total_questions=questions.count(),
            started_at=started_at,
            completed_at=completed_at,
            time_taken=time_taken,
            score=0,
            correct_answers=0
        )
        
        # Procesar respuestas
        correct_count = 0
        for question in questions:
            selected_option = request.POST.get(f'question_{question.id}')
            if selected_option:
                is_correct = selected_option == question.correct_option
                if is_correct:
                    correct_count += 1
                
                ExamAnswer.objects.create(
                    attempt=attempt,
                    question=question,
                    selected_option=selected_option,
                    is_correct=is_correct
                )
        
        # Calcular puntuación
        score = (correct_count / questions.count()) * 100 if questions.count() > 0 else 0
        attempt.correct_answers = correct_count
        attempt.score = score
        attempt.passed = score >= exam.passing_score
        attempt.save()
        
        return redirect('exam_results', attempt_id=attempt.id)
    
    context = {
        'exam': exam,
        'questions': questions
    }
    return render(request, 'tickets/exam_take.html', context)


def exam_results(request, attempt_id):
    """Mostrar resultados del examen"""
    from .utils import is_agent
    
    attempt = get_object_or_404(ExamAttempt, id=attempt_id)
    answers = attempt.answers.all().select_related('question')
    
    # Solo los agentes pueden ver las respuestas correctas
    can_view_answers = request.user.is_authenticated and is_agent(request.user)
    
    # Generar token del certificado si es necesario
    if attempt.passed and not attempt.certificate_token:
        attempt.generate_certificate_token()
    
    context = {
        'attempt': attempt,
        'answers': answers,
        'exam': attempt.exam,
        'can_view_answers': can_view_answers
    }
    return render(request, 'tickets/exam_results.html', context)


def download_certificate(request, attempt_id):
    """Descargar certificado de examen aprobado"""
    from .certificate_utils import create_certificate_response
    
    attempt = get_object_or_404(ExamAttempt, id=attempt_id)
    
    # Verificar que el usuario puede descargar este certificado
    # Para exámenes públicos, cualquiera puede descargar el certificado si tiene el ID
    # Para exámenes privados, solo el usuario o agentes pueden descargarlo
    if not attempt.exam.is_public and request.user.is_authenticated:
        if not (attempt.user == request.user or 
                request.user.groups.filter(name='Agentes').exists()):
            messages.error(request, 'No tienes permisos para descargar este certificado')
            return redirect('exam_results', attempt_id=attempt_id)
    elif not attempt.exam.is_public and not request.user.is_authenticated:
        messages.error(request, 'Debes iniciar sesión para descargar este certificado')
        return redirect('exam_results', attempt_id=attempt_id)
    
    # Verificar que el examen fue aprobado
    if not attempt.passed:
        messages.error(request, 'Solo se pueden generar certificados para exámenes aprobados')
        return redirect('exam_results', attempt_id=attempt_id)
    
    # Generar y retornar el certificado
    response = create_certificate_response(attempt)
    if response:
        return response
    else:
        messages.error(request, 'Error al generar el certificado')
        return redirect('exam_results', attempt_id=attempt_id)


def verify_certificate(request, token):
    """Verificar autenticidad de un certificado"""
    from .certificate_utils import verify_certificate_data
    
    cert_data = verify_certificate_data(token)
    
    context = {
        'cert_data': cert_data,
        'token': token
    }
    return render(request, 'tickets/certificate_verify.html', context)


@login_required
def question_create(request, exam_pk):
    """Crear pregunta para examen"""
    from .utils import can_manage_courses
    
    exam = get_object_or_404(Exam, pk=exam_pk)
    
    if not (can_manage_courses(request.user) or exam.created_by == request.user):
        messages.error(request, 'No tienes permisos para agregar preguntas')
        return redirect('exam_detail', pk=exam.pk)
    
    if request.method == 'POST':
        question = ExamQuestion.objects.create(
            exam=exam,
            question_text=request.POST.get('question_text'),
            option_a=request.POST.get('option_a'),
            option_b=request.POST.get('option_b'),
            option_c=request.POST.get('option_c'),
            correct_option=request.POST.get('correct_option'),
            order=exam.questions.count() + 1
        )
        
        messages.success(request, 'Pregunta agregada exitosamente')
        return redirect('exam_detail', pk=exam.pk)
    
    context = {'exam': exam}
    return render(request, 'tickets/question_form.html', context)


@login_required
def question_edit(request, pk):
    """Editar pregunta"""
    from .utils import can_manage_courses
    
    question = get_object_or_404(ExamQuestion, pk=pk)
    exam = question.exam
    
    if not (can_manage_courses(request.user) or exam.created_by == request.user):
        messages.error(request, 'No tienes permisos para editar esta pregunta')
        return redirect('exam_detail', pk=exam.pk)
    
    if request.method == 'POST':
        question.question_text = request.POST.get('question_text')
        question.option_a = request.POST.get('option_a')
        question.option_b = request.POST.get('option_b')
        question.option_c = request.POST.get('option_c')
        question.correct_option = request.POST.get('correct_option')
        question.save()
        
        messages.success(request, 'Pregunta actualizada exitosamente')
        return redirect('exam_detail', pk=exam.pk)
    
    context = {'question': question, 'exam': exam}
    return render(request, 'tickets/question_form.html', context)


@login_required
def question_delete(request, pk):
    """Eliminar pregunta"""
    from .utils import can_manage_courses
    
    question = get_object_or_404(ExamQuestion, pk=pk)
    exam = question.exam
    
    if not (can_manage_courses(request.user) or exam.created_by == request.user):
        messages.error(request, 'No tienes permisos para eliminar esta pregunta')
        return redirect('exam_detail', pk=exam.pk)
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Pregunta eliminada exitosamente')
        return redirect('exam_detail', pk=exam.pk)
    
    context = {'question': question, 'exam': exam}
    return render(request, 'tickets/question_delete.html', context)


@login_required
def exam_attempts_list(request):
    """Listado de todos los intentos de exámenes (solo para agentes)"""
    from .utils import is_agent
    
    if not is_agent(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    # Obtener parámetros de filtrado
    exam_filter = request.GET.get('exam', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search = request.GET.get('search', '')
    
    # Filtrar intentos
    attempts = ExamAttempt.objects.select_related('exam', 'user').filter(
        completed_at__isnull=False
    ).order_by('-completed_at')
    
    # Aplicar filtros
    if exam_filter:
        attempts = attempts.filter(exam_id=exam_filter)
    
    if status_filter == 'passed':
        attempts = attempts.filter(passed=True)
    elif status_filter == 'failed':
        attempts = attempts.filter(passed=False)
    
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            attempts = attempts.filter(completed_at__date__gte=date_from_obj.date())
        except ValueError:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            attempts = attempts.filter(completed_at__date__lte=date_to_obj.date())
        except ValueError:
            pass
    
    if search:
        from django.db.models import Q
        attempts = attempts.filter(
            Q(participant_name__icontains=search) |
            Q(participant_email__icontains=search) |
            Q(exam__title__icontains=search)
        )
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(attempts, 25)  # 25 intentos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener lista de exámenes para el filtro
    exams = Exam.objects.all().order_by('title')
    
    # Estadísticas generales
    total_attempts = attempts.count()
    passed_attempts = attempts.filter(passed=True).count()
    failed_attempts = attempts.filter(passed=False).count()
    pass_rate = (passed_attempts / total_attempts * 100) if total_attempts > 0 else 0
    
    context = {
        'page_obj': page_obj,
        'attempts': page_obj,
        'exams': exams,
        'current_filters': {
            'exam': exam_filter,
            'status': status_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
        },
        'stats': {
            'total_attempts': total_attempts,
            'passed_attempts': passed_attempts,
            'failed_attempts': failed_attempts,
            'pass_rate': pass_rate,
        }
    }
    
    return render(request, 'tickets/exam_attempts_list.html', context)


# ==================== CONTACTO WEB ====================

def contacto_web(request):
    """Vista pública para formulario de contacto"""
    from .forms import ContactoWebForm
    from .models import ContactoWeb
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    # Obtener IP del usuario para control de spam
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        user_ip = x_forwarded_for.split(',')[0]
    else:
        user_ip = request.META.get('REMOTE_ADDR')
    
    # Verificar límite de envíos (máximo 3 por hora por IP)
    one_hour_ago = timezone.now() - timedelta(hours=1)
    recent_contacts = ContactoWeb.objects.filter(
        ip_address=user_ip,
        fecha_creacion__gte=one_hour_ago
    ).count()
    
    if recent_contacts >= 3:
        messages.error(request, 'Has alcanzado el límite de contactos por hora. Por favor, inténtalo más tarde.')
        form = ContactoWebForm()
        # Generar CAPTCHA aunque no se pueda enviar
        import random
        num1 = random.randint(1, 10)
        num2 = random.randint(1, 10)
        operation = random.choice(['+', '-', '*'])
        
        if operation == '+':
            question = f"¿Cuánto es {num1} + {num2}?"
            correct_answer = num1 + num2
        elif operation == '-':
            if num1 < num2:
                num1, num2 = num2, num1
            question = f"¿Cuánto es {num1} - {num2}?"
            correct_answer = num1 - num2
        else:
            question = f"¿Cuánto es {num1} × {num2}?"
            correct_answer = num1 * num2
        
        request.session['captcha_answer'] = correct_answer
        form.fields['captcha_question'].initial = question
        
        context = {
            'form': form,
            'page_title': 'Contáctanos',
            'rate_limited': True
        }
        return render(request, 'tickets/contacto_web.html', context)
    
    if request.method == 'POST':
        form = ContactoWebForm(request.POST)
        
        # Verificar CAPTCHA
        captcha_answer = request.POST.get('captcha_answer')
        correct_answer = request.session.get('captcha_answer')
        
        if not captcha_answer or not correct_answer:
            form.add_error('captcha_answer', 'Error en la verificación. Recarga la página e inténtalo de nuevo.')
        else:
            try:
                if int(captcha_answer) != correct_answer:
                    form.add_error('captcha_answer', 'La respuesta del CAPTCHA es incorrecta. Inténtalo de nuevo.')
            except (ValueError, TypeError):
                form.add_error('captcha_answer', 'Por favor, introduce solo números en el resultado.')
        
        if form.is_valid():
            # Guardar el contacto con información adicional
            contacto = form.save(commit=False)
            
            # Obtener IP del usuario
            contacto.ip_address = user_ip
            
            # Obtener User Agent
            contacto.user_agent = request.META.get('HTTP_USER_AGENT', '')
            
            contacto.save()
            
            # Enviar notificación por email
            try:
                from .utils import send_contact_notification
                send_contact_notification(contacto)
            except Exception as e:
                # Log del error pero no interrumpir el flujo
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error enviando notificación de contacto: {str(e)}")
            
            # Limpiar CAPTCHA de la sesión después del envío exitoso
            if 'captcha_answer' in request.session:
                del request.session['captcha_answer']
            
            messages.success(request, '¡Gracias por contactarnos! Hemos recibido tu mensaje y te responderemos pronto.')
            return redirect('contacto_web')
    else:
        form = ContactoWebForm()
        
    # Generar nuevo CAPTCHA y guardarlo en la sesión
    import random
    num1 = random.randint(1, 10)
    num2 = random.randint(1, 10)
    operation = random.choice(['+', '-', '*'])
    
    if operation == '+':
        question = f"¿Cuánto es {num1} + {num2}?"
        correct_answer = num1 + num2
    elif operation == '-':
        # Asegurar que el resultado sea positivo
        if num1 < num2:
            num1, num2 = num2, num1
        question = f"¿Cuánto es {num1} - {num2}?"
        correct_answer = num1 - num2
    else:  # multiplicación
        question = f"¿Cuánto es {num1} × {num2}?"
        correct_answer = num1 * num2
    
    # Guardar la respuesta correcta en la sesión
    request.session['captcha_answer'] = correct_answer
    
    # Establecer la pregunta en el formulario
    form.fields['captcha_question'].initial = question
    
    context = {
        'form': form,
        'page_title': 'Contáctanos'
    }
    
    return render(request, 'tickets/contacto_web.html', context)


@login_required
def contactos_web_list(request):
    """Vista para listar contactos web - solo para agentes"""
    from .utils import is_agent
    from .models import ContactoWeb
    
    if not is_agent(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Obtener contactos
    contactos = ContactoWeb.objects.all().order_by('-fecha_creacion')
    
    # Aplicar filtros
    if search:
        from django.db.models import Q
        contactos = contactos.filter(
            Q(nombre__icontains=search) |
            Q(email__icontains=search) |
            Q(empresa__icontains=search) |
            Q(asunto__icontains=search) |
            Q(mensaje__icontains=search)
        )
    
    if status_filter == 'no_leido':
        contactos = contactos.filter(leido=False)
    elif status_filter == 'leido':
        contactos = contactos.filter(leido=True)
    elif status_filter == 'respondido':
        contactos = contactos.filter(respondido=True)
    elif status_filter == 'pendiente':
        contactos = contactos.filter(respondido=False)
    
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            contactos = contactos.filter(fecha_creacion__date__gte=date_from_obj.date())
        except ValueError:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            contactos = contactos.filter(fecha_creacion__date__lte=date_to_obj.date())
        except ValueError:
            pass
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(contactos, 25)  # 25 contactos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    total_contactos = ContactoWeb.objects.count()
    no_leidos = ContactoWeb.objects.filter(leido=False).count()
    respondidos = ContactoWeb.objects.filter(respondido=True).count()
    pendientes = ContactoWeb.objects.filter(respondido=False).count()
    
    context = {
        'page_obj': page_obj,
        'contactos': page_obj,
        'current_filters': {
            'search': search,
            'status': status_filter,
            'date_from': date_from,
            'date_to': date_to,
        },
        'stats': {
            'total_contactos': total_contactos,
            'no_leidos': no_leidos,
            'respondidos': respondidos,
            'pendientes': pendientes,
        },
        'page_title': 'Contactos Web'
    }
    
    return render(request, 'tickets/contactos_web_list.html', context)


@login_required
def contacto_web_detail(request, pk):
    """Vista para ver detalle de un contacto web"""
    from .utils import is_agent
    from .models import ContactoWeb
    
    if not is_agent(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('dashboard')
    
    contacto = get_object_or_404(ContactoWeb, pk=pk)
    
    # Marcar como leído si no lo estaba
    if not contacto.leido:
        contacto.leido = True
        contacto.save()
    
    # Si es POST, actualizar estado de respondido
    if request.method == 'POST':
        if 'marcar_respondido' in request.POST:
            contacto.respondido = True
            contacto.save()
            messages.success(request, 'Contacto marcado como respondido')
        elif 'marcar_pendiente' in request.POST:
            contacto.respondido = False
            contacto.save()
            messages.success(request, 'Contacto marcado como pendiente')
        
        return redirect('contacto_web_detail', pk=pk)
    
    context = {
        'contacto': contacto,
        'page_title': f'Contacto: {contacto.nombre}'
    }
    
    return render(request, 'tickets/contacto_web_detail.html', context)


# =================================
# URLS PÚBLICAS DE SUBIDA DE DOCUMENTOS
# =================================

@login_required
@user_passes_test(is_agent)
def public_upload_url_list(request):
    """Lista las URLs públicas de subida de documentos"""
    upload_urls = PublicDocumentUpload.objects.filter(created_by=request.user)
    
    context = {
        'upload_urls': upload_urls,
        'page_title': 'URLs Públicas de Subida'
    }
    
    return render(request, 'tickets/public_upload_url_list.html', context)


@login_required
@user_passes_test(is_agent)
def public_upload_url_create(request):
    """Crea una nueva URL pública de subida"""
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        company_id = request.POST.get('company')
        expires_at = request.POST.get('expires_at')
        max_uploads = request.POST.get('max_uploads')
        
        # Crear la URL pública
        upload_url = PublicDocumentUpload(
            title=title,
            description=description,
            created_by=request.user
        )
        
        if company_id:
            upload_url.company_id = company_id
        
        if expires_at:
            upload_url.expires_at = expires_at
        
        if max_uploads:
            upload_url.max_uploads = int(max_uploads)
        
        upload_url.save()
        
        messages.success(request, f'URL pública creada exitosamente. Token: {upload_url.upload_token}')
        return redirect('public_upload_url_detail', pk=upload_url.pk)
    
    companies = Company.objects.all()
    
    context = {
        'companies': companies,
        'page_title': 'Crear URL Pública de Subida'
    }
    
    return render(request, 'tickets/public_upload_url_form.html', context)


@login_required
@user_passes_test(is_agent)
def public_upload_url_detail(request, pk):
    """Detalle de una URL pública de subida"""
    upload_url = get_object_or_404(PublicDocumentUpload, pk=pk, created_by=request.user)
    
    # Obtener documentos subidos usando esta URL
    uploaded_documents = Document.objects.filter(
        tags__icontains=f'upload-token:{upload_url.upload_token}'
    ).order_by('-created_at')
    
    # Generar URL completa
    public_url = request.build_absolute_uri(upload_url.get_public_url())
    
    context = {
        'upload_url': upload_url,
        'uploaded_documents': uploaded_documents,
        'public_url': public_url,
        'page_title': f'URL Pública: {upload_url.title}'
    }
    
    return render(request, 'tickets/public_upload_url_detail.html', context)


@login_required
@user_passes_test(is_agent)
def public_upload_url_toggle(request, pk):
    """Activa/desactiva una URL pública"""
    upload_url = get_object_or_404(PublicDocumentUpload, pk=pk, created_by=request.user)
    
    upload_url.is_active = not upload_url.is_active
    upload_url.save()
    
    status = "activada" if upload_url.is_active else "desactivada"
    messages.success(request, f'URL pública {status} exitosamente')
    
    return redirect('public_upload_url_detail', pk=pk)


def public_document_upload(request, token):
    """Vista pública para subir documentos (no requiere login)"""
    upload_url = get_object_or_404(PublicDocumentUpload, upload_token=token)
    
    # Verificar si la URL es válida
    if not upload_url.is_valid():
        context = {
            'error': 'Esta URL de subida ha expirado o ya no está disponible.',
            'upload_url': upload_url
        }
        return render(request, 'tickets/public_upload_error.html', context)
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        uploaded_file = request.FILES.get('file')
        
        if not title:
            messages.error(request, 'El título es obligatorio')
        elif not uploaded_file:
            messages.error(request, 'Debe seleccionar un archivo')
        else:
            try:
                # Crear el documento
                document = Document(
                    title=title,
                    description=description,
                    file=uploaded_file,
                    created_by=upload_url.created_by,
                    company=upload_url.company,
                    tags=f'upload-token:{upload_url.upload_token},public-upload'
                )
                
                # Establecer metadatos del archivo
                document.file_size = uploaded_file.size
                document.file_type = uploaded_file.content_type or 'unknown'
                
                document.save()
                
                # Actualizar contador y última fecha de uso
                upload_url.upload_count += 1
                upload_url.last_used_at = timezone.now()
                upload_url.save()
                
                messages.success(request, '¡Archivo subido exitosamente!')
                
                # Redirigir a una página de éxito
                context = {
                    'document': document,
                    'upload_url': upload_url,
                    'success': True
                }
                return render(request, 'tickets/public_upload_success.html', context)
                
            except Exception as e:
                messages.error(request, f'Error al subir el archivo: {str(e)}')
    
    context = {
        'upload_url': upload_url,
        'page_title': f'Subir Documento: {upload_url.title}'
    }
    
    return render(request, 'tickets/public_upload_form.html', context)


# === VISTAS DE REGISTRO PÚBLICO DE CURSOS ===

@login_required
def course_generate_registration_token(request, pk):
    """Generar token de registro público para un curso"""
    from .models import Course, CourseRegistrationToken
    from . import utils
    
    course = get_object_or_404(Course, pk=pk)
    
    if not utils.can_manage_courses(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('course_detail', pk=pk)
    
    if request.method == 'POST':
        # Obtener parámetros del formulario
        expires_in_days = request.POST.get('expires_in_days')
        max_registrations = request.POST.get('max_registrations')
        
        # Calcular fecha de expiración
        expires_at = None
        if expires_in_days:
            try:
                days = int(expires_in_days)
                expires_at = timezone.now() + timezone.timedelta(days=days)
            except ValueError:
                pass
        
        # Validar max_registrations
        max_regs = None
        if max_registrations:
            try:
                max_regs = int(max_registrations)
            except ValueError:
                pass
        
        # Crear el token
        token = CourseRegistrationToken.objects.create(
            course=course,
            created_by=request.user,
            expires_at=expires_at,
            max_registrations=max_regs
        )
        
        messages.success(request, f'Token de registro creado exitosamente.')
        return redirect('course_detail', pk=pk)
    
    context = {
        'course': course,
        'page_title': f'Generar Token de Registro - {course.title}'
    }
    
    return render(request, 'tickets/course_generate_registration_token.html', context)


@login_required
def course_disable_registration_token(request, pk, token_id):
    """Desactivar token de registro público"""
    from .models import Course, CourseRegistrationToken
    from . import utils
    
    course = get_object_or_404(Course, pk=pk)
    token = get_object_or_404(CourseRegistrationToken, pk=token_id, course=course)
    
    if not utils.can_manage_courses(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('course_detail', pk=pk)
    
    token.is_active = False
    token.save()
    
    messages.success(request, 'Token de registro desactivado exitosamente.')
    return redirect('course_detail', pk=pk)


def course_public_register(request, token):
    """Formulario de registro público para cursos"""
    from .models import CourseRegistrationToken, UserProfile
    from django.contrib.auth.models import User
    from django.contrib.auth import login
    
    # Buscar el token
    try:
        registration_token = CourseRegistrationToken.objects.get(token=token)
    except CourseRegistrationToken.DoesNotExist:
        messages.error(request, 'Token de registro no válido.')
        return redirect('course_public')
    
    # Verificar si el token es válido
    is_valid, message = registration_token.is_valid()
    if not is_valid:
        messages.error(request, f'Token no válido: {message}')
        return redirect('course_public')
    
    course = registration_token.course
    
    if request.method == 'POST':
        # Obtener datos del formulario
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        
        # Validaciones
        errors = []
        
        if not first_name:
            errors.append('El nombre es obligatorio.')
        
        if not last_name:
            errors.append('El apellido es obligatorio.')
        
        if not email:
            errors.append('El correo electrónico es obligatorio.')
        elif User.objects.filter(email=email).exists():
            errors.append('Ya existe un usuario con este correo electrónico.')
        
        if not password:
            errors.append('La contraseña es obligatoria.')
        elif len(password) < 8:
            errors.append('La contraseña debe tener al menos 8 caracteres.')
        
        if password != password_confirm:
            errors.append('Las contraseñas no coinciden.')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Crear el usuario
                username = email.split('@')[0]
                counter = 1
                original_username = username
                
                # Asegurar que el username sea único
                while User.objects.filter(username=username).exists():
                    username = f"{original_username}{counter}"
                    counter += 1
                
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
                
                # Crear o actualizar el perfil del usuario
                user_profile, created = UserProfile.objects.get_or_create(user=user)
                
                # Asignar la empresa del profesor/agente que creó el token
                try:
                    creator_profile = UserProfile.objects.get(user=registration_token.created_by)
                    if creator_profile.company:
                        user_profile.company = creator_profile.company
                        user_profile.save()
                except UserProfile.DoesNotExist:
                    # Si el creador del token no tiene perfil, usar la empresa del curso como fallback
                    if course.company:
                        user_profile.company = course.company
                        user_profile.save()
                
                # Incrementar contador de registros
                registration_token.increment_registration_count()
                
                # Iniciar sesión automáticamente
                login(request, user)
                
                messages.success(request, f'¡Registro exitoso! Bienvenido al curso "{course.title}".')
                return redirect('course_detail', pk=course.pk)
                
            except Exception as e:
                messages.error(request, f'Error al crear el usuario: {str(e)}')
    
    context = {
        'course': course,
        'token': registration_token,
        'page_title': f'Registro para {course.title}'
    }
    
    return render(request, 'tickets/course_public_register.html', context)


@login_required
def course_registration_tokens_list(request, pk):
    """Lista de tokens de registro para un curso"""
    from .models import Course, CourseRegistrationToken
    from . import utils
    
    course = get_object_or_404(Course, pk=pk)
    
    if not utils.can_manage_courses(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('course_detail', pk=pk)
    
    tokens = CourseRegistrationToken.objects.filter(course=course).order_by('-created_at')
    
    context = {
        'course': course,
        'tokens': tokens,
        'page_title': f'Tokens de Registro - {course.title}'
    }
    
    return render(request, 'tickets/course_registration_tokens_list.html', context)


# =============================================================================
# VISTAS DE EMPLEADOS Y APLICACIONES DE TRABAJO
# =============================================================================

@login_required
def employee_list(request):
    """Lista de empleados"""
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    # Filtrar por empresa del usuario
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company:
        employees = Employee.objects.filter(company=user_profile.company).order_by('-created_at')
    else:
        employees = Employee.objects.all().order_by('-created_at')
    
    # Filtros
    status_filter = request.GET.get('status')
    search = request.GET.get('search')
    
    if status_filter:
        employees = employees.filter(status=status_filter)
    
    if search:
        employees = employees.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search) |
            Q(position__icontains=search)
        )
    
    # Paginación
    paginator = Paginator(employees, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'status_choices': Employee.STATUS_CHOICES,
        'current_status': status_filter,
        'search': search,
        'page_title': 'Gestión de Empleados'
    }
    
    return render(request, 'tickets/employee_list.html', context)


@login_required
def employee_detail(request, pk):
    """Detalle de un empleado"""
    employee = get_object_or_404(Employee, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('employee_list')
    
    # Verificar que el usuario puede ver este empleado (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and employee.company != user_profile.company:
        messages.error(request, 'No tienes permisos para ver este empleado.')
        return redirect('employee_list')
    
    if request.method == 'POST' and 'analyze_ai' in request.POST:
        # Obtener configuración del sistema
        config = SystemConfiguration.objects.first()
        if not config or not config.openai_api_key:
            messages.error(request, 'La configuración de IA no está disponible. Contacta al administrador.')
            return redirect('employee_detail', pk=pk)
            
        # Preparar datos para IA
        datos = f"Nombre: {employee.get_full_name()}\nEmail: {employee.email}\nTeléfono: {employee.phone}\nCargo: {employee.position}"
        if employee.salary_euros:
            datos += f"\nSalario propuesto: €{employee.salary_euros}"
        if employee.description:
            datos += f"\nDescripción/Experiencia: {employee.description}"
        if employee.resume_file:
            datos += f"\nCurrículo adjunto: {employee.resume_file.name}"
            
        # Usar el prompt personalizado de la configuración
        prompt = config.ai_employee_analysis_prompt.format(datos=datos)
        
        # Llamada a OpenAI
        try:
            from openai import OpenAI
            
            client = OpenAI(api_key=config.openai_api_key)
            
            response = client.chat.completions.create(
                model=config.openai_model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=500,
                temperature=0.7
            )
            
            texto = response.choices[0].message.content.strip()
            resumen = ""
            puntuacion = None
            
            # Extraer resumen y puntuación
            import re
            resumen_match = re.search(r"Resumen:\s*(.*?)(?=Puntuación:|$)", texto, re.DOTALL)
            puntuacion_match = re.search(r"Puntuación:\s*(\d+)", texto)
            
            if resumen_match:
                resumen = resumen_match.group(1).strip()
            else:
                # Si no encuentra el formato esperado, usar todo el texto
                resumen = texto
                
            if puntuacion_match:
                puntuacion = int(puntuacion_match.group(1))
                # Validar que la puntuación esté entre 1 y 10
                if puntuacion < 1 or puntuacion > 10:
                    puntuacion = None
                    
            # Guardar resultados
            employee.ai_analysis = resumen if resumen else texto
            employee.ai_score = puntuacion
            employee.save()
            
            messages.success(request, "Análisis de IA completado y guardado correctamente.")
            
        except Exception as e:
            messages.error(request, f"Error al analizar con IA: {str(e)}")
            
        return redirect('employee_detail', pk=pk)
    context = {
        'employee': employee,
        'page_title': f'Empleado: {employee.get_full_name()}',
        'config': SystemConfiguration.objects.first()
    }
    return render(request, 'tickets/employee_detail.html', context)


@login_required
def employee_change_status(request, pk):
    """Cambiar el estado de un empleado"""
    employee = get_object_or_404(Employee, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('employee_list')
    
    # Verificar que el usuario puede modificar este empleado (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and employee.company != user_profile.company:
        messages.error(request, 'No tienes permisos para modificar este empleado.')
        return redirect('employee_list')
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in ['candidate', 'in_process', 'employee', 'not_employee']:
            employee.status = new_status
            employee.save()
            messages.success(request, f'Estado actualizado a {employee.get_status_display()}')
            return redirect('employee_detail', pk=pk)
        else:
            messages.error(request, 'Estado no válido.')
    
    return redirect('employee_detail', pk=pk)


@login_required
def employee_edit_hiring_opinion(request, pk):
    """Editar opinión de reunión de contratación"""
    employee = get_object_or_404(Employee, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('employee_list')
    
    # Verificar que el usuario puede modificar este empleado (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and employee.company != user_profile.company:
        messages.error(request, 'No tienes permisos para modificar este empleado.')
        return redirect('employee_list')
    
    if request.method == 'POST':
        form = EmployeeHiringOpinionForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, 'Opinión de reunión actualizada correctamente.')
            
            # Redirigir según el tipo de empleado
            if employee.status in ['candidate', 'in_process']:
                return redirect('candidate_detail', pk=pk)
            else:
                return redirect('active_employee_detail', pk=pk)
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = EmployeeHiringOpinionForm(instance=employee)
    
    context = {
        'form': form,
        'employee': employee,
        'page_title': f'Editar Opinión de Reunión - {employee.get_full_name()}',
    }
    
    return render(request, 'tickets/employee_edit_hiring_opinion.html', context)


@login_required
def employee_payroll_list(request, pk):
    """Lista de nóminas de un empleado"""
    employee = get_object_or_404(Employee, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('employee_list')
    
    # Verificar que el usuario puede ver este empleado (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and employee.company != user_profile.company:
        messages.error(request, 'No tienes permisos para ver este empleado.')
        return redirect('employee_list')
    
    # Obtener nóminas del empleado
    payrolls = employee.payrolls.all().order_by('-period_year', '-period_month')
    
    context = {
        'employee': employee,
        'payrolls': payrolls,
        'page_title': f'Nóminas - {employee.get_full_name()}',
    }
    
    return render(request, 'tickets/employee_payroll_list.html', context)


@login_required
def employee_payroll_create(request, pk):
    """Agregar nueva nómina a un empleado"""
    employee = get_object_or_404(Employee, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('employee_list')
    
    # Verificar que el usuario puede modificar este empleado (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and employee.company != user_profile.company:
        messages.error(request, 'No tienes permisos para modificar este empleado.')
        return redirect('employee_list')
    
    if request.method == 'POST':
        form = EmployeePayrollForm(request.POST, request.FILES)
        if form.is_valid():
            payroll = form.save(commit=False)
            payroll.employee = employee
            payroll.created_by = request.user
            try:
                payroll.save()
                messages.success(request, f'Nómina de {payroll.get_period_display()} agregada correctamente.')
                return redirect('employee_payroll_list', pk=employee.pk)
            except Exception as e:
                if 'UNIQUE constraint failed' in str(e) or 'unique_together' in str(e):
                    messages.error(request, f'Ya existe una nómina para {payroll.get_period_display()}.')
                else:
                    messages.error(request, f'Error al guardar la nómina: {str(e)}')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = EmployeePayrollForm()
    
    context = {
        'form': form,
        'employee': employee,
        'page_title': f'Nueva Nómina - {employee.get_full_name()}',
    }
    
    return render(request, 'tickets/employee_payroll_form.html', context)


@login_required
def employee_payroll_detail(request, employee_pk, payroll_pk):
    """Ver detalles de una nómina específica"""
    employee = get_object_or_404(Employee, pk=employee_pk)
    payroll = get_object_or_404(EmployeePayroll, pk=payroll_pk, employee=employee)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('employee_list')
    
    # Verificar que el usuario puede ver este empleado (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and employee.company != user_profile.company:
        messages.error(request, 'No tienes permisos para ver este empleado.')
        return redirect('employee_list')
    
    context = {
        'employee': employee,
        'payroll': payroll,
        'page_title': f'Nómina {payroll.get_period_display()} - {employee.get_full_name()}',
    }
    
    return render(request, 'tickets/employee_payroll_detail.html', context)


@login_required
def employee_payroll_edit(request, employee_pk, payroll_pk):
    """Editar una nómina existente"""
    employee = get_object_or_404(Employee, pk=employee_pk)
    payroll = get_object_or_404(EmployeePayroll, pk=payroll_pk, employee=employee)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('employee_list')
    
    # Verificar que el usuario puede modificar este empleado (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and employee.company != user_profile.company:
        messages.error(request, 'No tienes permisos para modificar este empleado.')
        return redirect('employee_list')
    
    if request.method == 'POST':
        # Debug: Verificar archivos recibidos
        print(f"DEBUG - FILES recibidos: {request.FILES}")
        print(f"DEBUG - POST data: {request.POST}")
        
        form = EmployeePayrollForm(request.POST, request.FILES, instance=payroll)
        if form.is_valid():
            try:
                saved_payroll = form.save()
                print(f"DEBUG - Nómina guardada - Comprobante: {saved_payroll.payment_receipt}")
                messages.success(request, f'Nómina de {payroll.get_period_display()} actualizada correctamente.')
                return redirect('employee_payroll_detail', employee_pk=employee.pk, payroll_pk=payroll.pk)
            except Exception as e:
                print(f"DEBUG - Error al guardar: {str(e)}")
                messages.error(request, f'Error al actualizar la nómina: {str(e)}')
        else:
            print(f"DEBUG - Errores del formulario: {form.errors}")
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = EmployeePayrollForm(instance=payroll)
    
    context = {
        'form': form,
        'employee': employee,
        'payroll': payroll,
        'page_title': f'Editar Nómina {payroll.get_period_display()} - {employee.get_full_name()}',
    }
    
    return render(request, 'tickets/employee_payroll_form.html', context)


@login_required
def employee_payroll_delete(request, employee_pk, payroll_pk):
    """Eliminar una nómina"""
    employee = get_object_or_404(Employee, pk=employee_pk)
    payroll = get_object_or_404(EmployeePayroll, pk=payroll_pk, employee=employee)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('employee_list')
    
    # Verificar que el usuario puede modificar este empleado (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and employee.company != user_profile.company:
        messages.error(request, 'No tienes permisos para modificar este empleado.')
        return redirect('employee_list')
    
    if request.method == 'POST':
        period_display = payroll.get_period_display()
        payroll.delete()
        messages.success(request, f'Nómina de {period_display} eliminada correctamente.')
        return redirect('employee_payroll_list', pk=employee.pk)
    
    context = {
        'employee': employee,
        'payroll': payroll,
        'page_title': f'Eliminar Nómina {payroll.get_period_display()} - {employee.get_full_name()}',
    }
    
    return render(request, 'tickets/employee_payroll_delete.html', context)


@login_required
def candidate_list(request):
    """Lista de candidatos (estado: candidato e in_process)"""
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    # Filtrar por empresa del usuario
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company:
        employees = Employee.objects.filter(
            company=user_profile.company,
            status__in=['candidate', 'in_process']
        ).order_by('-created_at')
    else:
        employees = Employee.objects.filter(status__in=['candidate', 'in_process']).order_by('-created_at')
    
    # Búsqueda
    search = request.GET.get('search')
    if search:
        employees = employees.filter(
            models.Q(first_name__icontains=search) |
            models.Q(last_name__icontains=search) |
            models.Q(email__icontains=search) |
            models.Q(position__icontains=search)
        )
    
    # Paginación
    paginator = Paginator(employees, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'employees': page_obj,
        'search': search,
        'page_title': 'Candidatos',
        'section': 'candidates'
    }
    
    return render(request, 'tickets/candidate_list.html', context)


@login_required
def candidate_detail(request, pk):
    """Detalle de un candidato"""
    employee = get_object_or_404(Employee, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('candidate_list')
    
    # Verificar que el usuario puede ver este candidato (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and employee.company != user_profile.company:
        messages.error(request, 'No tienes permisos para ver este candidato.')
        return redirect('candidate_list')
    
    # Análisis de IA (mismo código que employee_detail)
    if request.method == 'POST' and 'analyze_ai' in request.POST:
        config = SystemConfiguration.objects.first()
        if not config or not config.openai_api_key:
            messages.error(request, 'La configuración de IA no está disponible. Contacta al administrador.')
            return redirect('candidate_detail', pk=pk)
            
        # Preparar datos para IA
        datos = f"Nombre: {employee.get_full_name()}\nEmail: {employee.email}\nTeléfono: {employee.phone}\nCargo: {employee.position}"
        if employee.salary_euros:
            datos += f"\nSalario propuesto: €{employee.salary_euros}"
        if employee.description:
            datos += f"\nDescripción/Experiencia: {employee.description}"
        if employee.resume_file:
            datos += f"\nCurrículo adjunto: {employee.resume_file.name}"
            
        # Usar el prompt personalizado de la configuración
        prompt = config.ai_employee_analysis_prompt.format(datos=datos)
        
        # Llamada a OpenAI
        try:
            from openai import OpenAI
            
            client = OpenAI(api_key=config.openai_api_key)
            
            response = client.chat.completions.create(
                model=config.openai_model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=500,
                temperature=0.7
            )
            
            texto = response.choices[0].message.content.strip()
            resumen = ""
            puntuacion = None
            
            # Extraer resumen y puntuación
            import re
            resumen_match = re.search(r"Resumen:\s*(.*?)(?=Puntuación:|$)", texto, re.DOTALL)
            puntuacion_match = re.search(r"Puntuación:\s*(\d+)", texto)
            
            if resumen_match:
                resumen = resumen_match.group(1).strip()
            else:
                # Si no encuentra el formato esperado, usar todo el texto
                resumen = texto
                
            if puntuacion_match:
                puntuacion = int(puntuacion_match.group(1))
                # Validar que la puntuación esté entre 1 y 10
                if puntuacion < 1 or puntuacion > 10:
                    puntuacion = None
                    
            # Guardar resultados
            employee.ai_analysis = resumen if resumen else texto
            employee.ai_score = puntuacion
            employee.save()
            
            messages.success(request, "Análisis de IA completado y guardado correctamente.")
            
        except Exception as e:
            messages.error(request, f"Error al analizar con IA: {str(e)}")
            
        return redirect('candidate_detail', pk=pk)
    
    context = {
        'employee': employee,
        'page_title': f'Candidato: {employee.get_full_name()}',
        'config': SystemConfiguration.objects.first(),
        'section': 'candidates'
    }
    
    return render(request, 'tickets/candidate_detail.html', context)


@login_required
def active_employee_list(request):
    """Lista de empleados activos (estado: employee)"""
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    # Filtrar por empresa del usuario
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company:
        employees = Employee.objects.filter(
            company=user_profile.company,
            status='employee'
        ).order_by('-created_at')
    else:
        employees = Employee.objects.filter(status='employee').order_by('-created_at')
    
    # Búsqueda
    search = request.GET.get('search')
    if search:
        employees = employees.filter(
            models.Q(first_name__icontains=search) |
            models.Q(last_name__icontains=search) |
            models.Q(email__icontains=search) |
            models.Q(position__icontains=search)
        )
    
    # Paginación
    paginator = Paginator(employees, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'employees': page_obj,
        'search': search,
        'page_title': 'Empleados Activos',
        'section': 'employees'
    }
    
    return render(request, 'tickets/active_employee_list.html', context)


@login_required
def active_employee_detail(request, pk):
    """Detalle de un empleado activo"""
    employee = get_object_or_404(Employee, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('active_employee_list')
    
    # Verificar que el usuario puede ver este empleado (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and employee.company != user_profile.company:
        messages.error(request, 'No tienes permisos para ver este empleado.')
        return redirect('active_employee_list')
    
    context = {
        'employee': employee,
        'page_title': f'Empleado: {employee.get_full_name()}',
        'config': SystemConfiguration.objects.first(),
        'section': 'employees'
    }
    
    return render(request, 'tickets/active_employee_detail.html', context)


@login_required
def job_application_token_list(request):
    """Lista de tokens de aplicación de trabajo"""
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    # Filtrar por empresa del usuario
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company:
        tokens = JobApplicationToken.objects.filter(company=user_profile.company).order_by('-created_at')
    else:
        tokens = JobApplicationToken.objects.all().order_by('-created_at')
    
    # Paginación
    paginator = Paginator(tokens, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'page_title': 'Enlaces de Aplicación de Empleo'
    }
    
    return render(request, 'tickets/job_application_token_list.html', context)


@login_required
def job_application_token_create(request):
    """Crear nuevo token de aplicación de trabajo"""
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            job_title = request.POST.get('job_title')
            job_description = request.POST.get('job_description')
            proposed_salary_euros = request.POST.get('proposed_salary_euros')
            max_applications = request.POST.get('max_applications')
            expires_at = request.POST.get('expires_at')
            
            if not job_title or not job_description:
                messages.error(request, 'El título y descripción del trabajo son obligatorios.')
                return render(request, 'tickets/job_application_token_form.html')
            
            # Obtener empresa del usuario
            user_profile = UserProfile.objects.filter(user=request.user).first()
            if not user_profile or not user_profile.company:
                messages.error(request, 'Debes estar asociado a una empresa para crear enlaces de aplicación.')
                return redirect('job_application_token_list')
            
            # Crear el token
            token = JobApplicationToken.objects.create(
                job_title=job_title,
                job_description=job_description,
                proposed_salary_euros=float(proposed_salary_euros) if proposed_salary_euros else None,
                max_applications=int(max_applications) if max_applications else None,
                expires_at=datetime.strptime(expires_at, '%Y-%m-%dT%H:%M') if expires_at else None,
                company=user_profile.company,
                created_by=request.user
            )
            
            messages.success(request, f'Enlace de aplicación creado exitosamente para "{job_title}".')
            return redirect('job_application_token_detail', pk=token.pk)
            
        except Exception as e:
            messages.error(request, f'Error al crear el enlace: {str(e)}')
    
    context = {
        'page_title': 'Crear Enlace de Aplicación de Empleo'
    }
    
    return render(request, 'tickets/job_application_token_form.html', context)


@login_required
def job_application_token_detail(request, pk):
    """Detalle de un token de aplicación de trabajo"""
    token = get_object_or_404(JobApplicationToken, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes', 'Profesores']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('job_application_token_list')
    
    # Verificar que el usuario puede ver este token (misma empresa)
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and token.company != user_profile.company:
        messages.error(request, 'No tienes permisos para ver este enlace.')
        return redirect('job_application_token_list')
    
    # Obtener aplicaciones relacionadas
    applications = Employee.objects.filter(
        company=token.company,
        created_at__gte=token.created_at
    ).order_by('-created_at')
    
    context = {
        'token': token,
        'applications': applications,
        'page_title': f'Enlace: {token.job_title}'
    }
    
    return render(request, 'tickets/job_application_token_detail.html', context)


def public_job_application(request, token):
    """Vista pública para aplicar a un empleo"""
    try:
        application_token = get_object_or_404(JobApplicationToken, application_token=token)
    except Exception:
        return render(request, 'tickets/public_job_application_error.html', {
            'error': 'El enlace de aplicación no es válido.'
        })
    
    if not application_token.is_valid():
        error_msg = 'Este enlace de aplicación ya no está disponible.'
        if application_token.expires_at and timezone.now() > application_token.expires_at:
            error_msg = 'Este enlace de aplicación ha expirado.'
        elif application_token.max_applications and application_token.application_count >= application_token.max_applications:
            error_msg = 'Este enlace ha alcanzado el máximo de aplicaciones permitidas.'
        
        return render(request, 'tickets/public_job_application_error.html', {
            'error': error_msg
        })
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            phone = request.POST.get('phone', '').strip()
            description = request.POST.get('description', '').strip()
            resume_file = request.FILES.get('resume_file')
            
            # Validaciones básicas
            if not all([first_name, last_name, email, description]):
                messages.error(request, 'Todos los campos marcados con * son obligatorios.')
                return render(request, 'tickets/public_job_application.html', {
                    'token': application_token,
                    'page_title': f'Aplicar a: {application_token.job_title}'
                })
            
            # Verificar que el email no exista ya
            if Employee.objects.filter(email=email, company=application_token.company).exists():
                messages.error(request, 'Ya existe una aplicación con este correo electrónico.')
                return render(request, 'tickets/public_job_application.html', {
                    'token': application_token,
                    'page_title': f'Aplicar a: {application_token.job_title}'
                })
            
            # Crear el empleado/aplicante
            employee = Employee.objects.create(
                first_name=first_name,
                last_name=last_name,
                email=email,
                phone=phone,
                position=application_token.job_title,
                salary_euros=application_token.proposed_salary_euros,
                description=description,
                resume_file=resume_file,
                company=application_token.company,
                status='applied'
            )
            
            # Incrementar contador de aplicaciones
            application_token.application_count += 1
            application_token.save()
            
            # Procesar CV con IA si está configurada
            try:
                analyze_resume_with_ai(employee)
            except Exception as e:
                print(f"Error al analizar CV con IA: {e}")
            
            return render(request, 'tickets/public_job_application_success.html', {
                'employee': employee,
                'token': application_token,
                'page_title': 'Aplicación Enviada'
            })
            
        except Exception as e:
            messages.error(request, f'Error al enviar la aplicación: {str(e)}')
    
    context = {
        'token': application_token,
        'page_title': f'Aplicar a: {application_token.job_title}'
    }
    
    return render(request, 'tickets/public_job_application.html', context)


def analyze_resume_with_ai(employee):
    """Analizar currículo con IA si está configurada"""
    try:
        # Obtener configuración de IA
        ai_config = SystemConfiguration.objects.first()
        if not ai_config or not hasattr(ai_config, 'openai_api_key') or not ai_config.openai_api_key:
            return
        
        # Si no hay archivo de CV, analizar solo la descripción
        content_to_analyze = employee.description
        
        if employee.resume_file:
            try:
                # Leer el contenido del archivo (solo si es texto plano o PDF simple)
                file_path = employee.resume_file.path
                if file_path.endswith('.txt'):
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content_to_analyze += f"\n\nContenido del CV:\n{f.read()}"
            except Exception:
                pass  # Si no se puede leer el archivo, solo usar la descripción
        
        if not content_to_analyze.strip():
            return
        
        # Configurar OpenAI
        import openai
        openai.api_key = ai_config.openai_api_key
        
        # Prompt para análisis de CV
        prompt = f"""
        Analiza el siguiente currículo/perfil profesional y proporciona:
        
        1. Un resumen ejecutivo del candidato (máximo 3 párrafos)
        2. Fortalezas principales
        3. Áreas de mejora o faltantes
        4. Puntuación del 1 al 10 basada en:
           - Experiencia relevante
           - Habilidades técnicas
           - Comunicación escrita
           - Potencial de crecimiento
        
        Puesto aplicado: {employee.position}
        
        Perfil del candidato:
        {content_to_analyze}
        
        Responde en formato estructurado y profesional.
        """
        
        # Llamar a la API de OpenAI
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Eres un experto en recursos humanos y análisis de currículos."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=800,
            temperature=0.3
        )
        
        ai_analysis = response.choices[0].message.content
        
        # Extraer puntuación del análisis (buscar números del 1-10)
        import re
        score_match = re.search(r'(\d+)/10|(\d+)\s*de\s*10|puntuación.*?(\d+)', ai_analysis.lower())
        ai_score = None
        if score_match:
            for group in score_match.groups():
                if group and 1 <= int(group) <= 10:
                    ai_score = int(group)
                    break
        
        # Guardar análisis
        employee.ai_analysis = ai_analysis
        employee.ai_score = ai_score
        employee.save()
        
    except Exception as e:
        print(f"Error en análisis de IA: {e}")
        # No fallar silenciosamente, pero tampoco interrumpir el proceso


# =============================================================================
# VISTAS DE ACUERDOS/CONTRATOS
# =============================================================================

@login_required
def agreement_list(request):
    """Lista de acuerdos"""
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    # Filtrar por empresa del usuario si es necesario
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company:
        agreements = Agreement.objects.filter(company=user_profile.company)
    else:
        agreements = Agreement.objects.all()
    
    # Filtros de búsqueda
    search = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    
    if search:
        agreements = agreements.filter(
            models.Q(title__icontains=search) |
            models.Q(body__icontains=search)
        )
    
    if status_filter:
        agreements = agreements.filter(status=status_filter)
    
    agreements = agreements.order_by('-created_at')
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(agreements, 20)
    page_number = request.GET.get('page')
    agreements = paginator.get_page(page_number)
    
    context = {
        'agreements': agreements,
        'search': search,
        'status_filter': status_filter,
        'page_title': 'Gestión de Acuerdos',
    }
    
    return render(request, 'tickets/agreement_list.html', context)


@login_required
def agreement_create(request):
    """Crear nuevo acuerdo"""
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = AgreementForm(request.POST)
        if form.is_valid():
            agreement = form.save(commit=False)
            agreement.created_by = request.user
            
            # Asignar empresa del usuario
            user_profile = UserProfile.objects.filter(user=request.user).first()
            if user_profile and user_profile.company:
                agreement.company = user_profile.company
            
            # Generar contenido con IA si se solicitó
            if form.cleaned_data.get('generate_ai_content'):
                ai_content = agreement.generate_ai_content(
                    form.cleaned_data.get('ai_prompt_addition', '')
                )
                if ai_content:
                    agreement.body = ai_content
                    messages.success(request, 'Contenido generado con IA exitosamente.')
                else:
                    messages.warning(request, 'No se pudo generar contenido con IA. Verifica la configuración.')
            
            agreement.save()
            messages.success(request, f'Acuerdo "{agreement.title}" creado correctamente.')
            return redirect('agreement_detail', pk=agreement.pk)
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = AgreementForm()
    
    context = {
        'form': form,
        'page_title': 'Nuevo Acuerdo',
    }
    
    return render(request, 'tickets/agreement_form.html', context)


@login_required
def agreement_detail(request, pk):
    """Ver detalles de un acuerdo"""
    agreement = get_object_or_404(Agreement, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    # Verificar empresa
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and agreement.company != user_profile.company:
        messages.error(request, 'No tienes permisos para ver este acuerdo.')
        return redirect('agreement_list')
    
    signatures = agreement.signatures.all().order_by('-signed_at')
    
    context = {
        'agreement': agreement,
        'signatures': signatures,
        'page_title': f'Acuerdo - {agreement.title}',
    }
    
    return render(request, 'tickets/agreement_detail.html', context)


@login_required
def agreement_edit(request, pk):
    """Editar acuerdo existente"""
    agreement = get_object_or_404(Agreement, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    # Verificar empresa
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and agreement.company != user_profile.company:
        messages.error(request, 'No tienes permisos para editar este acuerdo.')
        return redirect('agreement_list')
    
    # Verificar si es editable
    if not agreement.is_editable:
        messages.error(request, 'Este acuerdo ya no se puede editar porque tiene firmas válidas.')
        return redirect('agreement_detail', pk=agreement.pk)
    
    if request.method == 'POST':
        form = AgreementForm(request.POST, instance=agreement)
        if form.is_valid():
            agreement = form.save()
            
            # Generar contenido con IA si se solicitó
            if form.cleaned_data.get('generate_ai_content'):
                ai_content = agreement.generate_ai_content(
                    form.cleaned_data.get('ai_prompt_addition', '')
                )
                if ai_content:
                    agreement.body = ai_content
                    agreement.save()
                    messages.success(request, 'Contenido actualizado con IA exitosamente.')
                else:
                    messages.warning(request, 'No se pudo generar contenido con IA.')
            
            messages.success(request, f'Acuerdo "{agreement.title}" actualizado correctamente.')
            return redirect('agreement_detail', pk=agreement.pk)
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = AgreementForm(instance=agreement)
    
    context = {
        'form': form,
        'agreement': agreement,
        'page_title': f'Editar - {agreement.title}',
    }
    
    return render(request, 'tickets/agreement_form.html', context)


@login_required
def agreement_delete(request, pk):
    """Eliminar acuerdo"""
    agreement = get_object_or_404(Agreement, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    # Verificar empresa
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and agreement.company != user_profile.company:
        messages.error(request, 'No tienes permisos para eliminar este acuerdo.')
        return redirect('agreement_list')
    
    if request.method == 'POST':
        title = agreement.title
        agreement.delete()
        messages.success(request, f'Acuerdo "{title}" eliminado correctamente.')
        return redirect('agreement_list')
    
    context = {
        'agreement': agreement,
        'page_title': f'Eliminar - {agreement.title}',
    }
    
    return render(request, 'tickets/agreement_delete.html', context)


@login_required
def agreement_publish(request, pk):
    """Publicar acuerdo para firma"""
    agreement = get_object_or_404(Agreement, pk=pk)
    
    # Verificar permisos
    if not request.user.groups.filter(name__in=['Administradores', 'Agentes']).exists():
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')
    
    # Verificar empresa
    user_profile = UserProfile.objects.filter(user=request.user).first()
    if user_profile and user_profile.company and agreement.company != user_profile.company:
        messages.error(request, 'No tienes permisos para publicar este acuerdo.')
        return redirect('agreement_list')
    
    if request.method == 'POST':
        form = AgreementPublicForm(request.POST)
        if form.is_valid():
            agreement.status = 'published'
            agreement.published_at = timezone.now()
            agreement.save()
            
            messages.success(request, f'Acuerdo "{agreement.title}" publicado correctamente.')
            messages.info(request, f'URL pública: {agreement.get_public_url(request)}')
            
            # TODO: Enviar notificaciones por email si está configurado
            email_list = form.cleaned_data.get('email_list', [])
            if email_list and form.cleaned_data.get('send_notifications'):
                # Implementar envío de emails
                pass
            
            return redirect('agreement_detail', pk=agreement.pk)
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = AgreementPublicForm()
    
    context = {
        'form': form,
        'agreement': agreement,
        'page_title': f'Publicar - {agreement.title}',
    }
    
    return render(request, 'tickets/agreement_publish.html', context)


# =============================================================================
# VISTAS PÚBLICAS DE ACUERDOS (SIN LOGIN)
# =============================================================================

def public_agreement_sign(request, token):
    """Vista pública para firmar un acuerdo"""
    agreement = get_object_or_404(Agreement, public_token=token)
    
    # Verificar si el acuerdo puede ser firmado
    if not agreement.can_be_signed:
        error_msg = "Este acuerdo no está disponible para firmar."
        if agreement.status != 'published':
            error_msg = "Este acuerdo aún no ha sido publicado."
        elif agreement.is_expired:
            error_msg = "Este acuerdo ha expirado."
        elif agreement.max_signers and agreement.signature_count >= agreement.max_signers:
            error_msg = "Este acuerdo ha alcanzado el número máximo de firmantes."
        
        context = {
            'agreement': agreement,
            'error_message': error_msg,
            'page_title': f'Acuerdo - {agreement.title}',
        }
        return render(request, 'tickets/public_agreement_error.html', context)
    
    if request.method == 'POST':
        form = AgreementSignatureForm(request.POST)
        form.agreement = agreement  # Para validaciones
        
        if form.is_valid():
            signature = form.save(commit=False)
            signature.agreement = agreement
            
            # Capturar información adicional
            signature.ip_address = request.META.get('REMOTE_ADDR')
            signature.user_agent = request.META.get('HTTP_USER_AGENT', '')
            
            # Si el acuerdo requiere aprobación, marcar como no aprobado
            if agreement.requires_approval:
                signature.is_approved = False
            
            signature.save()
            
            messages.success(request, '¡Acuerdo firmado correctamente!')
            return redirect('public_agreement_success', token=token, signature_id=signature.pk)
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = AgreementSignatureForm()
    
    context = {
        'agreement': agreement,
        'form': form,
        'signatures_count': agreement.signature_count,
        'page_title': f'Firmar - {agreement.title}',
    }
    
    return render(request, 'tickets/public_agreement_sign.html', context)


def public_agreement_success(request, token, signature_id):
    """Vista de confirmación después de firmar"""
    agreement = get_object_or_404(Agreement, public_token=token)
    signature = get_object_or_404(AgreementSignature, pk=signature_id, agreement=agreement)
    
    context = {
        'agreement': agreement,
        'signature': signature,
        'download_url': signature.get_download_url(request),
        'page_title': f'Firmado - {agreement.title}',
    }
    
    return render(request, 'tickets/public_agreement_success.html', context)


def download_signed_agreement(request, agreement_token, signature_id):
    """Descargar PDF del acuerdo firmado"""
    agreement = get_object_or_404(Agreement, public_token=agreement_token)
    signature = get_object_or_404(AgreementSignature, pk=signature_id, agreement=agreement)
    
    # Actualizar estadísticas de descarga
    signature.pdf_downloaded_at = timezone.now()
    signature.download_count += 1
    signature.save()
    
    # Generar PDF del acuerdo firmado
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from io import BytesIO
    
    # Crear PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    title = Paragraph(f"<b>{agreement.title}</b>", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Contenido del acuerdo
    body_paragraphs = agreement.body.split('\n')
    for para in body_paragraphs:
        if para.strip():
            p = Paragraph(para, styles['Normal'])
            story.append(p)
            story.append(Spacer(1, 10))
    
    # Información de la firma
    story.append(Spacer(1, 30))
    story.append(Paragraph("<b>INFORMACIÓN DE LA FIRMA:</b>", styles['Heading2']))
    story.append(Spacer(1, 10))
    
    signature_info = f"""
    <b>Firmante:</b> {signature.signer_name}<br/>
    <b>Email:</b> {signature.signer_email}<br/>
    <b>Fecha de Firma:</b> {signature.signed_at.strftime('%d/%m/%Y %H:%M')}<br/>
    <b>IP:</b> {signature.ip_address or 'No disponible'}<br/>
    """
    
    story.append(Paragraph(signature_info, styles['Normal']))
    
    # Construir PDF
    doc.build(story)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{agreement.title}_firmado_{signature.signer_name}.pdf"'
    
    return response


# =============================================================================
# VISTAS PARA LANDING PAGES
# =============================================================================

def is_agent_or_superuser(user):
    """
    Verifica si un usuario es agente o superusuario
    """
    if not user.is_authenticated:
        return False
    return user.is_superuser or is_agent(user)

@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def landing_page_list(request):
    """Vista para listar todas las landing pages"""
    landing_pages = LandingPage.objects.all().order_by('-created_at')
    
    context = {
        'page_title': 'Landing Pages',
        'landing_pages': landing_pages,
    }
    
    return render(request, 'tickets/landing_page_list.html', context)


@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def landing_page_create(request):
    """Vista para crear una nueva landing page"""
    if request.method == 'POST':
        form = LandingPageForm(request.POST, request.FILES)
        if form.is_valid():
            landing_page = form.save(commit=False)
            landing_page.created_by = request.user
            landing_page.save()
            messages.success(request, 'Landing Page creada exitosamente.')
            return redirect('landing_page_detail', pk=landing_page.pk)
        else:
            # Agregar errores del formulario a los mensajes
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = LandingPageForm()
    
    context = {
        'page_title': 'Crear Landing Page',
        'form': form,
    }
    
    return render(request, 'tickets/landing_page_form.html', context)


@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def landing_page_detail(request, pk):
    """Vista para ver detalles de una landing page"""
    landing_page = get_object_or_404(LandingPage, pk=pk)
    submissions = landing_page.submissions.all().order_by('-created_at')[:20]
    
    context = {
        'page_title': f'Landing Page: {landing_page.nombre_producto}',
        'landing_page': landing_page,
        'submissions': submissions,
    }
    
    return render(request, 'tickets/landing_page_detail.html', context)


@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def landing_page_edit(request, pk):
    """Vista para editar una landing page"""
    landing_page = get_object_or_404(LandingPage, pk=pk)
    
    if request.method == 'POST':
        form = LandingPageForm(request.POST, request.FILES, instance=landing_page)
        if form.is_valid():
            form.save()
            messages.success(request, 'Landing Page actualizada exitosamente.')
            return redirect('landing_page_detail', pk=landing_page.pk)
    else:
        form = LandingPageForm(instance=landing_page)
    
    context = {
        'page_title': f'Editar Landing Page: {landing_page.nombre_producto}',
        'form': form,
        'landing_page': landing_page,
    }
    
    return render(request, 'tickets/landing_page_form.html', context)


@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def landing_page_delete(request, pk):
    """Vista para eliminar una landing page"""
    landing_page = get_object_or_404(LandingPage, pk=pk)
    
    if request.method == 'POST':
        landing_page.delete()
        messages.success(request, 'Landing Page eliminada exitosamente.')
        return redirect('landing_page_list')
    
    context = {
        'page_title': f'Eliminar Landing Page: {landing_page.nombre_producto}',
        'landing_page': landing_page,
    }
    
    return render(request, 'tickets/landing_page_delete.html', context)


def landing_page_public(request, slug):
    """Vista pública de la landing page"""
    landing_page = get_object_or_404(LandingPage, slug=slug, is_active=True)
    
    # Incrementar contador de visitas
    landing_page.increment_views()
    
    if request.method == 'POST':
        form = LandingPageSubmissionForm(request.POST)
        if form.is_valid():
            submission = form.save(commit=False)
            submission.landing_page = landing_page
            
            # Capturar información adicional
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                submission.ip_address = x_forwarded_for.split(',')[0]
            else:
                submission.ip_address = request.META.get('REMOTE_ADDR')
            
            submission.user_agent = request.META.get('HTTP_USER_AGENT', '')
            submission.utm_source = request.GET.get('utm_source', '')
            submission.utm_medium = request.GET.get('utm_medium', '')
            submission.utm_campaign = request.GET.get('utm_campaign', '')
            
            submission.save()
            
            # Crear contacto web automáticamente
            try:
                # En vista pública no hay usuario logueado, así que created_by será None
                contact = create_contact_from_submission(submission, landing_page, created_by=None)
                
                # Enviar notificación de creación de contacto
                try:
                    from .utils import send_contact_creation_notification
                    send_contact_creation_notification(contact, landing_page)
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Error enviando notificación de contacto: {str(e)}")
                    
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error creando contacto desde landing page: {str(e)}")
            
            # Incrementar contador de envíos
            landing_page.increment_submissions()
            
            # Enviar notificación por email si está configurada
            try:
                from .utils import send_landing_page_notification
                send_landing_page_notification(submission)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error enviando notificación de landing page: {str(e)}")
            
            # Enviar notificación por Telegram si está configurada
            try:
                from .utils import send_telegram_notification
                send_telegram_notification(landing_page, submission)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error enviando notificación de Telegram: {str(e)}")
            
            # Redirigir a página de gracias
            return render(request, 'tickets/landing_page_thanks.html', {
                'landing_page': landing_page,
                'submission': submission,
            })
    else:
        form = LandingPageSubmissionForm()
    
    context = {
        'landing_page': landing_page,
        'form': form,
    }
    
    return render(request, 'tickets/landing_page_public.html', context)


def create_contact_from_submission(submission, landing_page, created_by=None):
    """Crear un contacto desde un envío de landing page"""
    from .models import Contact
    from django.contrib.auth.models import User
    
    # Si no hay usuario logueado, usar un usuario sistema
    if created_by is None:
        # Buscar o crear un usuario sistema para contacts automáticos
        system_user, created = User.objects.get_or_create(
            username='system_landing_pages',
            defaults={
                'first_name': 'Sistema',
                'last_name': 'Landing Pages',
                'email': 'system@landingpages.local',
                'is_active': False  # Usuario inactivo para que no pueda loguearse
            }
        )
        created_by = system_user
    
    # Verificar si ya existe un contacto con el mismo email
    existing_contact = Contact.objects.filter(
        email=submission.email
    ).first()
    
    # Preparar notas con información del landing page y mensaje del usuario
    notas_base = f"Contacto generado desde landing page: {landing_page.nombre_producto}\n\n"
    
    # Agregar mensaje del usuario si existe
    if hasattr(submission, 'mensaje') and submission.mensaje:
        notas_base += f"Mensaje del cliente:\n{submission.mensaje}\n\n"
    
    # Agregar información de seguimiento
    notas_base += f"Datos de seguimiento:\n" \
                  f"- UTM Source: {submission.utm_source or 'N/A'}\n" \
                  f"- UTM Medium: {submission.utm_medium or 'N/A'}\n" \
                  f"- UTM Campaign: {submission.utm_campaign or 'N/A'}\n" \
                  f"- IP Address: {submission.ip_address or 'N/A'}\n" \
                  f"- Fecha de envío: {submission.created_at.strftime('%d/%m/%Y %H:%M')}"
    
    if existing_contact:
        # Actualizar el contacto existente con nueva información
        existing_contact.name = f"{submission.nombre} {submission.apellido}".strip()
        if submission.telefono:
            existing_contact.phone = submission.telefono
        if submission.empresa:
            existing_contact.company = submission.empresa
        existing_contact.source = f"Landing Page: {landing_page.nombre_producto}"
        existing_contact.notes = notas_base
        existing_contact.status = 'positive'  # Nuevo lead es positivo
        existing_contact.save()
        
        return existing_contact
    else:
        # Crear nuevo contacto
        contact = Contact.objects.create(
            name=f"{submission.nombre} {submission.apellido}".strip(),
            email=submission.email,
            phone=submission.telefono or '',
            company=submission.empresa or '',
            source=f"Landing Page: {landing_page.nombre_producto}",
            notes=notas_base,
            status='positive',  # Nuevo lead es positivo por defecto
            created_by=created_by  # Asignar el usuario que creó el contacto
        )
        
        return contact


@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def landing_page_submissions(request, pk):
    """Vista para ver todos los envíos de una landing page"""
    landing_page = get_object_or_404(LandingPage, pk=pk)
    submissions = landing_page.submissions.all().order_by('-created_at')
    
    # Agregar información de contactos existentes
    from .models import Contact
    for submission in submissions:
        # Buscar contacto existente por email
        existing_contact = Contact.objects.filter(
            email=submission.email
        ).first()
        submission.existing_contact = existing_contact
    
    context = {
        'page_title': f'Envíos: {landing_page.nombre_producto}',
        'landing_page': landing_page,
        'submissions': submissions,
    }
    
    return render(request, 'tickets/landing_page_submissions.html', context)


@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def landing_page_submission_detail(request, submission_id):
    """Vista para ver los detalles de un envío específico de landing page"""
    from .models import LandingPageSubmission, Contact
    
    submission = get_object_or_404(LandingPageSubmission, pk=submission_id)
    
    # Buscar contacto existente por email
    existing_contact = Contact.objects.filter(
        email=submission.email
    ).first()
    
    context = {
        'page_title': f'Detalle del Envío #{submission.id}',
        'submission': submission,
        'landing_page': submission.landing_page,
        'existing_contact': existing_contact,
    }
    
    return render(request, 'tickets/landing_page_submission_detail.html', context)


@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def create_contact_from_submission_view(request, submission_id):
    """Vista para crear un contacto manualmente desde un envío de landing page"""
    from .models import LandingPageSubmission
    
    submission = get_object_or_404(LandingPageSubmission, pk=submission_id)
    
    if request.method == 'POST':
        try:
            contact = create_contact_from_submission(submission, submission.landing_page, created_by=request.user)
            messages.success(request, f'Contacto creado exitosamente para {contact.name}')
            
            # Enviar notificación
            try:
                from .utils import send_contact_creation_notification
                send_contact_creation_notification(contact, submission.landing_page)
            except Exception as e:
                messages.warning(request, 'Contacto creado pero no se pudo enviar la notificación por email.')
                
        except Exception as e:
            messages.error(request, f'Error al crear el contacto: {str(e)}')
    
    return redirect('landing_page_submissions', pk=submission.landing_page.pk)


@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def landing_page_contacts(request):
    """Vista para ver todos los contactos creados desde landing pages"""
    from .models import Contact
    
    # Filtrar contactos que contienen "landing page" en la fuente o notas
    contacts = Contact.objects.filter(
        models.Q(source__icontains='landing page') |
        models.Q(notes__icontains='landing page')
    ).order_by('-created_at')
    
    # Estadísticas
    total_contacts = contacts.count()
    positive_contacts = contacts.filter(status='positive').count()
    negative_contacts = contacts.filter(status='negative').count()
    
    context = {
        'page_title': 'Contactos desde Landing Pages',
        'contacts': contacts,
        'total_contacts': total_contacts,
        'positive_contacts': positive_contacts,
        'negative_contacts': negative_contacts,
    }
    
    return render(request, 'tickets/landing_page_contacts.html', context)


@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def ajax_create_contact_from_submission(request, submission_id):
    """Vista AJAX para crear un contacto desde un envío"""
    if request.method == 'POST':
        try:
            from .models import LandingPageSubmission
            submission = get_object_or_404(LandingPageSubmission, pk=submission_id)
            
            contact = create_contact_from_submission(submission, submission.landing_page, created_by=request.user)
            
            if not contact:
                return JsonResponse({
                    'success': False,
                    'message': 'No se pudo crear el contacto'
                })
            
            # Verificar si es un contacto nuevo o actualizado
            from .models import ContactoWeb
            existing_contact_count = ContactoWeb.objects.filter(email=submission.email).count()
            
            if existing_contact_count > 1:
                message = f'Contacto actualizado para {contact.name} ({contact.email})'
            else:
                message = f'Contacto creado exitosamente para {contact.name}'
            
            # Log para debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"Contacto procesado: {message}")
            
            # Enviar notificación (opcional)
            try:
                from .utils import send_contact_creation_notification
                send_contact_creation_notification(contact, submission.landing_page)
            except Exception as e:
                logger.warning(f"Error enviando notificación de contacto: {str(e)}")
            
            response_data = {
                'success': True,
                'message': message,
                'contact_id': contact.id
            }
            
            logger.info(f"Enviando respuesta: {response_data}")
            return JsonResponse(response_data)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error creando contacto desde submission {submission_id}: {str(e)}")
            
            return JsonResponse({
                'success': False,
                'message': f'Error al crear el contacto: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})


@login_required
@user_passes_test(is_agent_or_superuser, login_url='/')
def ajax_create_contacts_batch(request):
    """Vista AJAX para crear contactos en lote desde múltiples envíos"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            submission_ids = data.get('submission_ids', [])
            
            if not submission_ids:
                return JsonResponse({'success': False, 'message': 'No se seleccionaron envíos'})
            
            from .models import LandingPageSubmission
            contacts_created = 0
            errors = []
            
            for submission_id in submission_ids:
                try:
                    submission = LandingPageSubmission.objects.get(pk=submission_id)
                    contact = create_contact_from_submission(submission, submission.landing_page, created_by=request.user)
                    if contact:
                        contacts_created += 1
                except Exception as e:
                    errors.append(f'Error en envío {submission_id}: {str(e)}')
            
            return JsonResponse({
                'success': True,
                'message': f'Se crearon {contacts_created} contactos exitosamente',
                'contacts_created': contacts_created,
                'errors': errors
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error en la operación: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})


@login_required  
@user_passes_test(is_agent_or_superuser, login_url='/')
def ajax_submission_details(request, submission_id):
    """Vista AJAX para obtener detalles de un envío"""
    try:
        from .models import LandingPageSubmission
        submission = get_object_or_404(LandingPageSubmission, pk=submission_id)
        
        data = {
            'success': True,
            'submission': {
                'id': submission.id,
                'nombre': submission.nombre,
                'apellido': submission.apellido,
                'email': submission.email,
                'telefono': submission.telefono or '',
                'empresa': submission.empresa or '',
                'ip_address': submission.ip_address or '',
                'user_agent': submission.user_agent or '',
                'utm_source': submission.utm_source or '',
                'utm_medium': submission.utm_medium or '',
                'utm_campaign': submission.utm_campaign or '',
                'created_at': submission.created_at.strftime('%d/%m/%Y %H:%M:%S'),
                'landing_page': submission.landing_page.nombre_producto
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener detalles: {str(e)}'
        })


@require_http_methods(["DELETE"])
@login_required
def ajax_landing_submission_delete(request, submission_id):
    """Vista AJAX para eliminar un envío de landing page - Solo para agentes"""
    if not is_agent(request.user):
        return JsonResponse({
            'success': False,
            'message': 'No tienes permisos para realizar esta acción.'
        }, status=403)
    
    try:
        from .models import LandingPageSubmission
        submission = get_object_or_404(LandingPageSubmission, pk=submission_id)
        
        # Guardar información para el mensaje de confirmación
        submission_info = f"{submission.nombre} {submission.apellido} ({submission.email})"
        
        # Eliminar el envío
        submission.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Envío de {submission_info} eliminado correctamente.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al eliminar el envío: {str(e)}'
        }, status=500)


def public_company_ticket_create(request, public_token):
    """Vista pública para crear tickets desde una empresa usando su token público"""
    try:
        company = get_object_or_404(Company, public_token=public_token, is_active=True)
    except (Company.DoesNotExist, ValueError):
        return render(request, 'tickets/public_company_ticket_error.html', {
            'error_message': 'Esta empresa no existe o no tiene habilitada la creación pública de tickets.'
        })
    
    if request.method == 'POST':
        form = PublicCompanyTicketForm(request.POST, company=company)
        if form.is_valid():
            try:
                # Crear o obtener usuario del sistema para tickets públicos
                system_user, created = User.objects.get_or_create(
                    username='system_public_tickets',
                    defaults={
                        'email': 'system@ticketproo.com',
                        'first_name': 'Sistema',
                        'last_name': 'Tickets Públicos',
                        'is_active': True,
                    }
                )
                
                # Crear el ticket
                ticket = form.save(commit=False)
                ticket.created_by = system_user
                ticket.ticket_type = 'desarrollo'  # Tipo por defecto para tickets públicos
                ticket.status = 'open'
                ticket.save()
                
                return render(request, 'tickets/public_company_ticket_success.html', {
                    'ticket': ticket,
                    'company': company,
                    'customer_name': form.cleaned_data['customer_name'],
                    'customer_email': form.cleaned_data['customer_email'],
                })
                
            except Exception as e:
                form.add_error(None, f'Error al crear el ticket: {str(e)}')
    else:
        form = PublicCompanyTicketForm(company=company)
    
    return render(request, 'tickets/public_company_ticket_create.html', {
        'form': form,
        'company': company,
    })


# ===========================================
# VISTAS PARA TAREAS DE ÓRDENES DE TRABAJO
# ===========================================

@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_task_list_view(request, work_order_pk):
    """Vista para listar tareas de una orden de trabajo"""
    work_order = get_object_or_404(WorkOrder, pk=work_order_pk)
    tasks = work_order.tasks.all().order_by('order', 'created_at')
    
    context = {
        'work_order': work_order,
        'tasks': tasks,
        'page_title': f'Tareas - {work_order.order_number}',
        'total_estimated_hours': sum(task.estimated_hours or 0 for task in tasks),
        'total_logged_hours': work_order.get_total_hours_logged(),
        'progress_percentage': work_order.get_progress_percentage(),
    }
    
    return render(request, 'tickets/work_order_task_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_task_create_view(request, work_order_pk):
    """Vista para crear una nueva tarea en una orden de trabajo"""
    work_order = get_object_or_404(WorkOrder, pk=work_order_pk)
    
    if request.method == 'POST':
        from .forms import WorkOrderTaskForm
        form = WorkOrderTaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.work_order = work_order
            task.created_by = request.user
            task.save()
            messages.success(request, f'Tarea "{task.title}" creada exitosamente.')
            return redirect('work_order_task_list', work_order_pk=work_order.pk)
    else:
        from .forms import WorkOrderTaskForm
        form = WorkOrderTaskForm()
    
    context = {
        'form': form,
        'work_order': work_order,
        'page_title': f'Nueva Tarea - {work_order.order_number}'
    }
    
    return render(request, 'tickets/work_order_task_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_task_edit_view(request, work_order_pk, task_pk):
    """Vista para editar una tarea"""
    work_order = get_object_or_404(WorkOrder, pk=work_order_pk)
    from .models import WorkOrderTask
    task = get_object_or_404(WorkOrderTask, pk=task_pk, work_order=work_order)
    
    if request.method == 'POST':
        from .forms import WorkOrderTaskForm
        form = WorkOrderTaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tarea "{task.title}" actualizada exitosamente.')
            return redirect('work_order_task_list', work_order_pk=work_order.pk)
    else:
        from .forms import WorkOrderTaskForm
        form = WorkOrderTaskForm(instance=task)
    
    context = {
        'form': form,
        'work_order': work_order,
        'task': task,
        'page_title': f'Editar Tarea - {task.title}'
    }
    
    return render(request, 'tickets/work_order_task_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_task_delete_view(request, work_order_pk, task_pk):
    """Vista para eliminar una tarea"""
    work_order = get_object_or_404(WorkOrder, pk=work_order_pk)
    from .models import WorkOrderTask
    task = get_object_or_404(WorkOrderTask, pk=task_pk, work_order=work_order)
    
    if request.method == 'POST':
        title = task.title
        task.delete()
        messages.success(request, f'Tarea "{title}" eliminada exitosamente.')
        return redirect('work_order_task_list', work_order_pk=work_order.pk)
    
    context = {
        'work_order': work_order,
        'task': task,
        'page_title': f'Eliminar Tarea - {task.title}'
    }
    
    return render(request, 'tickets/work_order_task_delete.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def work_order_task_bulk_create_view(request, work_order_pk):
    """Vista para crear múltiples tareas de una vez"""
    work_order = get_object_or_404(WorkOrder, pk=work_order_pk)
    
    if request.method == 'POST':
        from .forms import WorkOrderTaskBulkForm
        form = WorkOrderTaskBulkForm(request.POST)
        if form.is_valid():
            tasks_text = form.cleaned_data['tasks_text']
            tasks_created = 0
            
            for i, line in enumerate(tasks_text.strip().split('\n')):
                line = line.strip()
                if line:
                    from .models import WorkOrderTask
                    WorkOrderTask.objects.create(
                        work_order=work_order,
                        title=line,
                        description=f'Tarea creada automáticamente',
                        status='pending',
                        order=i + 1,
                        created_by=request.user
                    )
                    tasks_created += 1
            
            messages.success(request, f'{tasks_created} tareas creadas exitosamente.')
            return redirect('work_order_task_list', work_order_pk=work_order.pk)
    else:
        from .forms import WorkOrderTaskBulkForm
        form = WorkOrderTaskBulkForm()
    
    context = {
        'form': form,
        'work_order': work_order,
        'page_title': f'Crear Tareas Múltiples - {work_order.order_number}'
    }
    
    return render(request, 'tickets/work_order_task_bulk_form.html', context)


def work_order_task_status_update_view(request, work_order_pk, task_pk):
    """Vista AJAX para actualizar el estado de una tarea"""
    if request.method == 'POST':
        work_order = get_object_or_404(WorkOrder, pk=work_order_pk)
        from .models import WorkOrderTask
        task = get_object_or_404(WorkOrderTask, pk=task_pk, work_order=work_order)
        
        new_status = request.POST.get('status')
        
        if new_status in ['pending', 'in_progress', 'completed']:
            old_status = task.get_status_display()
            task.status = new_status
            
            # Si se marca como completada, establecer fecha de finalización
            if new_status == 'completed':
                task.completed_at = timezone.now()
            elif new_status == 'in_progress':
                task.started_at = timezone.now()
                task.completed_at = None
            else:  # pending
                task.started_at = None
                task.completed_at = None
            
            task.save()
            
            new_status_display = task.get_status_display()
            
            return JsonResponse({
                'success': True,
                'message': f'Estado actualizado a "{new_status_display}"',
                'new_status': new_status,
                'new_status_display': new_status_display,
                'progress_percentage': work_order.get_progress_percentage()
            })
        
        return JsonResponse({
            'success': False,
            'error': 'Estado inválido'
        })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def work_order_task_time_start_view(request, work_order_pk, task_pk):
    """Vista para iniciar el tiempo de una tarea"""
    print(f"=== INICIO work_order_task_time_start_view ===")
    print(f"work_order_pk={work_order_pk}, task_pk={task_pk}, method={request.method}")
    print(f"User: {request.user}")
    
    if request.method == 'POST':
        try:
            work_order = get_object_or_404(WorkOrder, pk=work_order_pk)
            print(f"Work order encontrada: {work_order.order_number}")
            
            from .models import WorkOrderTask, WorkOrderTaskTimeSession
            task = get_object_or_404(WorkOrderTask, pk=task_pk, work_order=work_order)
            print(f"Tarea encontrada: {task.title}, Estado: {task.status}")
            
            # Verificar si ya hay una sesión activa para este usuario
            active_session = task.get_active_time_session(request.user)
            print(f"Sesión activa existente: {active_session}")
            
            if active_session:
                print(f"Ya existe una sesión activa: {active_session}")
                return JsonResponse({
                    'success': False,
                    'error': 'Ya tienes una sesión de tiempo activa para esta tarea'
                })
            
            # Crear nueva sesión de tiempo
            print("Creando nueva sesión...")
            session = WorkOrderTaskTimeSession.objects.create(
                task=task,
                user=request.user
            )
            print(f"Sesión creada exitosamente: {session} (ID: {session.id})")
            
            # Actualizar estado de la tarea a "en progreso" si está pendiente
            if task.status == 'pending':
                task.status = 'in_progress'
                task.started_at = timezone.now()
                task.save()
                print(f"Tarea actualizada a 'in_progress'")
            
            response_data = {
                'success': True,
                'message': 'Tiempo iniciado correctamente',
                'session_id': session.id,
                'started_at': session.start_time.strftime('%H:%M:%S')
            }
            print(f"Enviando respuesta exitosa: {response_data}")
            return JsonResponse(response_data)
            
        except Exception as e:
            print(f"ERROR EXCEPCIÓN: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': f'Error interno: {str(e)}'
            })
    
    print("Método no es POST, enviando error")
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def work_order_task_time_stop_view(request, work_order_pk, task_pk):
    """Vista para detener el tiempo de una tarea"""
    if request.method == 'POST':
        work_order = get_object_or_404(WorkOrder, pk=work_order_pk)
        from .models import WorkOrderTask
        task = get_object_or_404(WorkOrderTask, pk=task_pk, work_order=work_order)
        
        # Buscar sesión activa para este usuario
        active_session = task.get_active_time_session(request.user)
        if not active_session:
            return JsonResponse({
                'success': False,
                'error': 'No hay una sesión de tiempo activa para esta tarea'
            })
        
        # Obtener descripción opcional del POST
        description = request.POST.get('description', '')
        
        # Detener la sesión
        time_entry = active_session.stop_session(description)
        
        if time_entry:
            return JsonResponse({
                'success': True,
                'message': f'Tiempo registrado: {time_entry.hours}h',
                'total_time': task.get_total_time_logged(),
                'session_duration': active_session.get_duration_hours()
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Error al registrar el tiempo'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def work_order_public_task_view(request, token):
    """Vista pública para tareas de órdenes de trabajo compartidas"""
    work_order = get_object_or_404(WorkOrder, public_share_token=token, is_public=True)
    tasks = work_order.tasks.all().order_by('order', 'created_at')
    
    # Manejar actualización de estado de tarea por POST
    if request.method == 'POST':
        task_id = request.POST.get('task_id')
        new_status = request.POST.get('status')
        
        if task_id and new_status in ['pending', 'in_progress', 'completed']:
            from .models import WorkOrderTask
            try:
                task = WorkOrderTask.objects.get(id=task_id, work_order=work_order)
                old_status = task.get_status_display()
                task.status = new_status
                
                # Actualizar fechas según el estado
                if new_status == 'completed':
                    task.completed_at = timezone.now()
                    
                    # Si la tarea se marca como completada y no tiene tiempo registrado,
                    # calcular tiempo basado en started_at y completed_at
                    if task.get_total_time_logged() == 0 and task.started_at:
                        from .models import WorkOrderTaskTimeEntry
                        from decimal import Decimal
                        
                        # Calcular duración en horas
                        duration = task.completed_at - task.started_at
                        hours = Decimal(str(duration.total_seconds() / 3600))
                        
                        # Crear entrada de tiempo automática
                        WorkOrderTaskTimeEntry.objects.create(
                            task=task,
                            hours=round(hours, 2),
                            description="Tiempo automático calculado al completar la tarea",
                            date=task.completed_at.date()
                        )
                        
                        # Actualizar horas reales de la tarea
                        task.update_actual_hours()
                        
                elif new_status == 'in_progress':
                    if not task.started_at:  # Solo si no se había iniciado antes
                        task.started_at = timezone.now()
                    task.completed_at = None
                else:  # pending
                    task.started_at = None
                    task.completed_at = None
                
                task.save()
                
                # Actualizar las horas del work order después de cualquier cambio de estado
                work_order.update_actual_hours()
                
                new_status_display = task.get_status_display()
                messages.success(request, f'Tarea "{task.title}" actualizada de "{old_status}" a "{new_status_display}".')
            except WorkOrderTask.DoesNotExist:
                messages.error(request, 'Tarea no encontrada.')
        
        return redirect('public_work_order_tasks', token=token)
    
    context = {
        'work_order': work_order,
        'tasks': tasks,
        'page_title': f'Tareas - {work_order.order_number}',
        'is_public_view': True,
        'total_estimated_hours': sum(task.estimated_hours or 0 for task in tasks),
        'total_logged_hours': work_order.get_total_hours_logged(),
        'progress_percentage': work_order.get_progress_percentage(),
    }
    
    return render(request, 'tickets/work_order_public_tasks.html', context)


def work_order_public_task_time_start_view(request, token, task_pk):
    """Vista pública para iniciar el tiempo de una tarea"""
    if request.method == 'POST':
        work_order = get_object_or_404(WorkOrder, public_share_token=token, is_public=True)
        from .models import WorkOrderTask, WorkOrderTaskTimeSession
        task = get_object_or_404(WorkOrderTask, pk=task_pk, work_order=work_order)
        
        # Para vista pública, crear sesión anónima (sin usuario)
        # Verificar si ya hay una sesión activa para esta tarea
        active_session = task.active_sessions.filter(is_active=True, user__isnull=True).first()
        if active_session:
            return JsonResponse({
                'success': False,
                'error': 'Ya hay una sesión de tiempo activa para esta tarea'
            })
        
        # Crear nueva sesión de tiempo sin usuario (pública)
        session = WorkOrderTaskTimeSession.objects.create(
            task=task,
            user=None  # Para sesiones públicas
        )
        
        # Actualizar estado de la tarea a "en progreso" si está pendiente
        if task.status == 'pending':
            task.status = 'in_progress'
            task.started_at = timezone.now()
            task.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Tiempo iniciado correctamente',
            'session_id': session.id,
            'started_at': session.start_time.strftime('%H:%M:%S')
        })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


def work_order_public_task_time_stop_view(request, token, task_pk):
    """Vista pública para detener el tiempo de una tarea"""
    if request.method == 'POST':
        work_order = get_object_or_404(WorkOrder, public_share_token=token, is_public=True)
        from .models import WorkOrderTask
        task = get_object_or_404(WorkOrderTask, pk=task_pk, work_order=work_order)
        
        # Buscar sesión activa pública (sin usuario)
        active_session = task.active_sessions.filter(is_active=True, user__isnull=True).first()
        if not active_session:
            return JsonResponse({
                'success': False,
                'error': 'No hay una sesión de tiempo activa para esta tarea'
            })
        
        # Obtener descripción opcional del POST
        description = request.POST.get('description', '')
        
        # Detener la sesión
        time_entry = active_session.stop_session(description)
        
        if time_entry:
            return JsonResponse({
                'success': True,
                'message': f'Tiempo registrado: {time_entry.hours}h',
                'total_time': task.get_total_time_logged(),
                'session_duration': active_session.get_duration_hours()
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Error al registrar el tiempo'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


# ===========================================
# VISTAS PARA EVALUACIÓN DE IA DE LANDING PAGES
# ===========================================

@login_required
@user_passes_test(is_agent, login_url='/')
def landing_submissions_list_view(request):
    """Vista para listar envíos de landing pages con evaluación de IA"""
    from .models import LandingPageSubmission
    
    # Filtros
    evaluated_filter = request.GET.get('evaluated', 'all')
    priority_filter = request.GET.get('priority', 'all')
    score_filter = request.GET.get('score', 'all')
    search = request.GET.get('search', '')
    
    # Base queryset
    submissions = LandingPageSubmission.objects.select_related('landing_page').order_by('-created_at')
    
    # Aplicar filtros
    if evaluated_filter == 'yes':
        submissions = submissions.filter(ai_evaluated=True)
    elif evaluated_filter == 'no':
        submissions = submissions.filter(ai_evaluated=False)
    
    if priority_filter != 'all':
        submissions = submissions.filter(ai_priority_level=priority_filter)
    
    if score_filter == 'high':
        submissions = submissions.filter(ai_overall_score__gte=7.0)
    elif score_filter == 'medium':
        submissions = submissions.filter(ai_overall_score__gte=4.0, ai_overall_score__lt=7.0)
    elif score_filter == 'low':
        submissions = submissions.filter(ai_overall_score__lt=4.0)
    
    if search:
        submissions = submissions.filter(
            Q(nombre__icontains=search) |
            Q(apellido__icontains=search) |
            Q(email__icontains=search) |
            Q(empresa__icontains=search) |
            Q(mensaje__icontains=search)
        )
    
    # Paginación
    paginator = Paginator(submissions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    from .ai_landing_evaluator import get_evaluation_stats
    try:
        stats = get_evaluation_stats()
    except:
        stats = {}
    
    context = {
        'page_title': 'Evaluación IA - Landing Pages',
        'page_obj': page_obj,
        'evaluated_filter': evaluated_filter,
        'priority_filter': priority_filter,
        'score_filter': score_filter,
        'search': search,
        'stats': stats,
        'priority_choices': [
            ('low', 'Baja'),
            ('medium', 'Media'),
            ('high', 'Alta'),
            ('urgent', 'Urgente'),
        ]
    }
    
    return render(request, 'tickets/landing_submissions_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def landing_submission_detail_view(request, submission_id):
    """Vista detallada de un envío de landing page con evaluación IA"""
    from .models import LandingPageSubmission
    
    submission = get_object_or_404(LandingPageSubmission, id=submission_id)
    
    context = {
        'page_title': f'Evaluación: {submission.nombre} {submission.apellido}',
        'submission': submission,
    }
    
    return render(request, 'tickets/landing_submission_detail.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def evaluate_submission_ai_view(request, submission_id):
    """Vista para evaluar un envío específico con IA"""
    from .models import LandingPageSubmission
    
    submission = get_object_or_404(LandingPageSubmission, id=submission_id)
    
    if request.method == 'POST':
        try:
            success = submission.evaluate_with_ai()
            if success:
                messages.success(request, f'Evaluación IA completada. Puntuación: {submission.ai_overall_score}/10')
            else:
                messages.error(request, 'Error al evaluar con IA')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    return redirect('landing_submission_detail', submission_id=submission.id)


@login_required
@user_passes_test(is_agent, login_url='/')
def landing_submission_delete_view(request, submission_id):
    """Vista para eliminar una submission de landing page"""
    from .models import LandingPageSubmission
    
    submission = get_object_or_404(LandingPageSubmission, id=submission_id)
    
    if request.method == 'POST':
        try:
            submission.delete()
            messages.success(request, f'Submission de "{submission.nombre} {submission.apellido}" eliminada correctamente.')
            return redirect('landing_submissions_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar la submission: {str(e)}')
            return redirect('landing_submission_detail', submission_id=submission.id)
    
    context = {
        'submission': submission,
        'page_title': f'Eliminar Submission - {submission.nombre} {submission.apellido}',
    }
    
    return render(request, 'tickets/landing_submission_delete.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def batch_evaluate_submissions_view(request):
    """Vista para evaluación en lote con IA"""
    if request.method == 'POST':
        limit = int(request.POST.get('limit', 10))
        
        try:
            from .ai_landing_evaluator import batch_evaluate_submissions
            results = batch_evaluate_submissions(limit=limit)
            
            messages.success(
                request, 
                f'Evaluación completada: {results["processed"]} procesados, {results["errors"]} errores'
            )
            
            # Mostrar detalles de errores si los hay
            if results["errors"] > 0:
                error_details = [r for r in results["results"] if r["status"] == "error"]
                for error in error_details[:3]:  # Mostrar máximo 3 errores
                    messages.warning(request, f'Error en ID {error["id"]}: {error["error"]}')
                    
        except Exception as e:
            messages.error(request, f'Error en evaluación en lote: {str(e)}')
    
    return redirect('landing_submissions_list')


@login_required
@user_passes_test(is_agent, login_url='/')
def ai_evaluation_dashboard_view(request):
    """Dashboard con estadísticas de evaluación IA"""
    from .ai_landing_evaluator import get_evaluation_stats
    from .models import LandingPageSubmission
    from django.db.models import Count
    from datetime import datetime, timedelta
    
    # Estadísticas generales
    try:
        stats = get_evaluation_stats()
    except:
        stats = {}
    
    # Top submissions por puntuación
    top_submissions = LandingPageSubmission.objects.filter(
        ai_evaluated=True
    ).order_by('-ai_overall_score')[:10]
    
    # Submissions recientes no evaluados
    pending_submissions = LandingPageSubmission.objects.filter(
        ai_evaluated=False
    ).order_by('-created_at')[:10]
    
    # Submissions urgentes
    urgent_submissions = LandingPageSubmission.objects.filter(
        ai_priority_level='urgent'
    ).order_by('-ai_overall_score')[:5]
    
    # Estadísticas por mes
    last_30_days = timezone.now() - timedelta(days=30)
    monthly_stats = LandingPageSubmission.objects.filter(
        created_at__gte=last_30_days
    ).extra(
        select={'month': 'DATE(created_at)'}
    ).values('month').annotate(
        total=Count('id'),
        evaluated=Count('id', filter=Q(ai_evaluated=True))
    ).order_by('-month')[:10]
    
    context = {
        'page_title': 'Dashboard Evaluación IA',
        'stats': stats,
        'top_submissions': top_submissions,
        'pending_submissions': pending_submissions,
        'urgent_submissions': urgent_submissions,
        'monthly_stats': monthly_stats,
    }
    
    return render(request, 'tickets/ai_evaluation_dashboard.html', context)


# ====== VISTAS DE ARCHIVOS COMPARTIDOS ======

@login_required
def shared_files_list_view(request):
    """Vista para listar archivos compartidos"""
    
    user = request.user
    user_profile = getattr(user, 'userprofile', None)
    
    # Si es agente, puede ver todos los archivos
    if is_agent(user):
        files = SharedFile.objects.all()
    else:
        # Si no es agente, solo ve archivos de su empresa y los que él subió
        if user_profile and user_profile.company:
            files = SharedFile.objects.filter(
                Q(company=user_profile.company) | Q(uploaded_by=user)
            )
        else:
            # Si no tiene empresa, solo ve los que él subió
            files = SharedFile.objects.filter(uploaded_by=user)
    
    # Filtros
    search = request.GET.get('search', '')
    company_id = request.GET.get('company', '')
    file_type = request.GET.get('file_type', '')
    
    if search:
        files = files.filter(
            Q(title__icontains=search) | 
            Q(description__icontains=search)
        )
    
    if company_id:
        files = files.filter(company_id=company_id)
    
    if file_type:
        files = files.filter(file_type__icontains=file_type)
    
    files = files.select_related('company', 'uploaded_by').order_by('-created_at')
    
    # Paginación
    paginator = Paginator(files, 20)
    page_number = request.GET.get('page')
    files_page = paginator.get_page(page_number)
    
    # Empresas para el filtro
    companies = Company.objects.filter(is_active=True).order_by('name')
    
    # Tipos de archivo únicos para el filtro
    file_types = SharedFile.objects.values_list('file_type', flat=True).distinct()
    
    context = {
        'page_title': 'Archivos Compartidos',
        'files': files_page,
        'companies': companies,
        'file_types': file_types,
        'search': search,
        'selected_company': company_id,
        'selected_file_type': file_type,
    }
    
    return render(request, 'tickets/shared_files_list.html', context)


@login_required
def shared_file_upload_view(request):
    """Vista para subir archivos compartidos"""
    
    if request.method == 'POST':
        form = SharedFileForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            shared_file = form.save(commit=False)
            shared_file.uploaded_by = request.user
            
            # Si el usuario no es agente y no tiene empresa asignada, usar la primera empresa activa
            if not shared_file.company:
                user_profile = getattr(request.user, 'userprofile', None)
                if user_profile and user_profile.company:
                    shared_file.company = user_profile.company
                else:
                    # Usar la primera empresa activa disponible
                    first_company = Company.objects.filter(is_active=True).first()
                    if first_company:
                        shared_file.company = first_company
            
            shared_file.save()
            messages.success(request, 'Archivo subido exitosamente.')
            return redirect('shared_files_list')
    else:
        form = SharedFileForm(user=request.user)
    
    context = {
        'page_title': 'Subir Archivo',
        'form': form,
    }
    
    return render(request, 'tickets/shared_file_upload.html', context)


@login_required
def shared_file_detail_view(request, file_id):
    """Vista para ver detalles de un archivo compartido"""
    
    shared_file = get_object_or_404(SharedFile, id=file_id)
    
    # Verificar permisos
    user = request.user
    user_profile = getattr(user, 'userprofile', None)
    
    if not is_agent(user):
        # Si no es agente, verificar que tenga acceso
        if not (shared_file.uploaded_by == user or 
                (user_profile and user_profile.company == shared_file.company)):
            messages.error(request, 'No tienes permisos para ver este archivo.')
            return redirect('shared_files_list')
    
    context = {
        'page_title': f'Archivo: {shared_file.title}',
        'shared_file': shared_file,
    }
    
    return render(request, 'tickets/shared_file_detail.html', context)


@login_required
def shared_file_download_view(request, file_id):
    """Vista para descargar archivos compartidos"""
    from django.http import FileResponse
    
    shared_file = get_object_or_404(SharedFile, id=file_id)
    
    # Verificar permisos
    user = request.user
    user_profile = getattr(user, 'userprofile', None)
    
    if not is_agent(user):
        # Si no es agente, verificar que tenga acceso
        if not (shared_file.uploaded_by == user or 
                (user_profile and user_profile.company == shared_file.company)):
            messages.error(request, 'No tienes permisos para descargar este archivo.')
            return redirect('shared_files_list')
    
    # Registrar la descarga
    SharedFileDownload.objects.create(
        shared_file=shared_file,
        downloaded_by=user,
        ip_address=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    # Incrementar contador de descargas
    shared_file.download_count += 1
    shared_file.save(update_fields=['download_count'])
    
    try:
        response = FileResponse(
            shared_file.file.open(),
            as_attachment=True,
            filename=os.path.basename(shared_file.file.name)
        )
        return response
    except FileNotFoundError:
        messages.error(request, 'El archivo no se encontró en el servidor.')
        return redirect('shared_file_detail', file_id=file_id)


def public_file_upload_view(request):
    """Vista pública para subir archivos"""
    
    if request.method == 'POST':
        form = PublicSharedFileForm(request.POST, request.FILES)
        if form.is_valid():
            shared_file = form.save()
            
            # Verificar si se encontró una empresa para el usuario
            uploader_email = form.cleaned_data['uploader_email']
            try:
                existing_user = User.objects.get(email=uploader_email)
                if hasattr(existing_user, 'userprofile') and existing_user.userprofile.company:
                    messages.success(
                        request, 
                        f'Archivo subido exitosamente y asignado automáticamente a la empresa: {existing_user.userprofile.company.name}. Gracias por compartir.'
                    )
                else:
                    messages.success(
                        request, 
                        'Archivo subido exitosamente. Tu cuenta no tiene empresa asignada, por lo que el archivo estará disponible para revisión de los agentes.'
                    )
            except User.DoesNotExist:
                messages.success(
                    request, 
                    'Archivo subido exitosamente. No se encontró una cuenta con este email, por lo que el archivo estará disponible para revisión de los agentes.'
                )
            
            return redirect('public_file_upload')
    else:
        form = PublicSharedFileForm()
    
    context = {
        'page_title': 'Subir Archivo Público',
        'form': form,
    }
    
    return render(request, 'tickets/public_file_upload.html', context)


@login_required 
@user_passes_test(is_agent)
def shared_file_delete_view(request, file_id):
    """Vista para eliminar archivos compartidos (solo agentes)"""
    
    shared_file = get_object_or_404(SharedFile, id=file_id)
    
    if request.method == 'POST':
        # Eliminar archivo físico
        if shared_file.file:
            try:
                shared_file.file.delete()
            except:
                pass
        
        shared_file.delete()
        messages.success(request, 'Archivo eliminado exitosamente.')
        return redirect('shared_files_list')
    
    context = {
        'page_title': 'Eliminar Archivo',
        'shared_file': shared_file,
    }
    
    return render(request, 'tickets/shared_file_delete.html', context)


@login_required
@user_passes_test(is_agent)
def shared_files_stats_view(request):
    """Vista de estadísticas de archivos compartidos (solo agentes)"""
    from django.db.models import Count, Sum
    
    # Estadísticas generales
    total_files = SharedFile.objects.count()
    total_downloads = SharedFileDownload.objects.count()
    total_size = SharedFile.objects.aggregate(
        total=Sum('file_size')
    )['total'] or 0
    
    # Calcular promedio de descargas por archivo
    avg_downloads_per_file = round(total_downloads / total_files, 1) if total_files > 0 else 0
    
    # Archivos más descargados
    top_files = SharedFile.objects.filter(
        download_count__gt=0
    ).order_by('-download_count')[:10]
    
    # Archivos por empresa
    files_by_company = SharedFile.objects.values(
        'company__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Archivos por tipo
    files_by_type = SharedFile.objects.values(
        'file_type'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Actividad reciente
    recent_uploads = SharedFile.objects.select_related(
        'company', 'uploaded_by'
    ).order_by('-created_at')[:10]
    
    recent_downloads = SharedFileDownload.objects.select_related(
        'shared_file', 'downloaded_by'
    ).order_by('-downloaded_at')[:10]
    
    context = {
        'page_title': 'Estadísticas de Archivos Compartidos',
        'total_files': total_files,
        'total_downloads': total_downloads,
        'total_size': total_size,
        'avg_downloads_per_file': avg_downloads_per_file,
        'top_files': top_files,
        'files_by_company': files_by_company,
        'files_by_type': files_by_type,
        'recent_uploads': recent_uploads,
        'recent_downloads': recent_downloads,
    }
    
    return render(request, 'tickets/shared_files_stats.html', context)


# ========================== GRABACIONES VIEWS ==========================

@login_required
def recordings_list_view(request):
    """Vista para listar grabaciones"""
    from .models import Recording
    from .utils import is_agent
    
    # Filtrar grabaciones según el rol del usuario
    if is_agent(request.user):
        # Los agentes pueden ver todas las grabaciones
        recordings = Recording.objects.select_related('company', 'uploaded_by').all()
    else:
        # Los usuarios solo ven grabaciones de su empresa o públicas
        user_company = getattr(request.user.userprofile, 'company', None) if hasattr(request.user, 'userprofile') else None
        if user_company:
            recordings = Recording.objects.select_related('company', 'uploaded_by').filter(
                models.Q(company=user_company) | models.Q(is_public=True, company__isnull=True)
            )
        else:
            recordings = Recording.objects.select_related('company', 'uploaded_by').filter(
                uploaded_by=request.user
            )
    
    # Paginación
    paginator = Paginator(recordings, 12)
    page_number = request.GET.get('page')
    recordings_page = paginator.get_page(page_number)
    
    context = {
        'page_title': 'Grabaciones',
        'recordings': recordings_page,
        'is_agent': is_agent(request.user),
    }
    
    return render(request, 'tickets/recordings_list.html', context)


@login_required
def recording_upload_view(request):
    """Vista para grabar audio desde el navegador"""
    from .forms import RecordingForm
    from .models import Recording
    import base64
    import tempfile
    import os
    from django.core.files.base import ContentFile
    
    if request.method == 'POST':
        form = RecordingForm(request.POST)
        if form.is_valid():
            recording = form.save(commit=False)
            
            # Asignar el usuario que sube
            recording.uploaded_by = request.user
            
            # Asignar empresa del usuario si tiene
            if hasattr(request.user, 'userprofile') and request.user.userprofile.company:
                recording.company = request.user.userprofile.company
            
            # Procesar datos de audio
            audio_data = form.cleaned_data['audio_data']
            if audio_data:
                try:
                    # Remover el prefijo 'data:audio/webm;codecs=opus;base64,' si existe
                    if ',' in audio_data:
                        audio_data = audio_data.split(',')[1]
                    
                    # Decodificar base64
                    audio_binary = base64.b64decode(audio_data)
                    
                    # Crear archivo temporal
                    audio_file = ContentFile(audio_binary, name=f"{recording.title}.webm")
                    recording.audio_file = audio_file
                    
                    recording.save()
                    
                    # Iniciar transcripción automática
                    try:
                        from .ai_utils import transcribe_recording
                        import threading
                        
                        # Ejecutar transcripción en hilo separado para no bloquear la respuesta
                        def transcribe_async():
                            transcribe_recording(recording.id)
                        
                        thread = threading.Thread(target=transcribe_async)
                        thread.daemon = True
                        thread.start()
                        
                        messages.success(request, 'Grabación guardada exitosamente. La transcripción se procesará automáticamente.')
                    except Exception as e:
                        messages.success(request, 'Grabación guardada exitosamente. La transcripción se procesará manualmente.')
                        print(f"Error iniciando transcripción automática: {e}")
                    
                    return redirect('recordings_list')
                    
                except Exception as e:
                    messages.error(request, f'Error al procesar la grabación: {str(e)}')
            else:
                messages.error(request, 'No se recibieron datos de audio.')
    else:
        form = RecordingForm()
    
    context = {
        'page_title': 'Grabar Audio',
        'form': form,
    }
    
    return render(request, 'tickets/recording_upload.html', context)


@login_required
def recording_detail_view(request, recording_id):
    """Vista para ver detalles de una grabación"""
    from .models import Recording, RecordingPlayback
    from .utils import is_agent
    
    recording = get_object_or_404(Recording, id=recording_id)
    
    # Verificar permisos
    user_can_access = False
    if is_agent(request.user):
        user_can_access = True
    elif recording.company and hasattr(request.user, 'userprofile') and request.user.userprofile.company == recording.company:
        user_can_access = True
    elif recording.uploaded_by == request.user:
        user_can_access = True
    elif recording.is_public:
        user_can_access = True
    
    if not user_can_access:
        messages.error(request, 'No tienes permisos para ver esta grabación.')
        return redirect('recordings_list')
    
    # Registrar reproducción
    RecordingPlayback.objects.create(
        recording=recording,
        played_by=request.user,
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    context = {
        'page_title': f'Grabación: {recording.title}',
        'recording': recording,
        'is_agent': is_agent(request.user),
    }
    
    return render(request, 'tickets/recording_detail.html', context)


@login_required
def recording_play_view(request, recording_id):
    """Vista para reproducir grabación (streaming)"""
    from .models import Recording
    from .utils import is_agent
    import mimetypes
    from django.http import HttpResponse, Http404
    from wsgiref.util import FileWrapper
    
    recording = get_object_or_404(Recording, id=recording_id)
    
    # Verificar permisos
    user_can_access = False
    if is_agent(request.user):
        user_can_access = True
    elif recording.company and hasattr(request.user, 'userprofile') and request.user.userprofile.company == recording.company:
        user_can_access = True
    elif recording.uploaded_by == request.user:
        user_can_access = True
    elif recording.is_public:
        user_can_access = True
    
    if not user_can_access:
        raise Http404("Grabación no encontrada")
    
    # Servir archivo de audio
    try:
        audio_file = recording.audio_file.open('rb')
        content_type, _ = mimetypes.guess_type(recording.audio_file.name)
        
        response = HttpResponse(FileWrapper(audio_file), content_type=content_type or 'audio/mpeg')
        response['Content-Disposition'] = f'inline; filename="{recording.title}.{recording.get_audio_extension().lower()}"'
        response['Content-Length'] = recording.file_size or recording.audio_file.size
        
        return response
    except Exception as e:
        raise Http404(f"Error al reproducir grabación: {str(e)}")


def public_recording_upload_view(request):
    """Vista pública para grabar audio desde el navegador"""
    from .forms import PublicRecordingForm
    import base64
    from django.core.files.base import ContentFile
    
    if request.method == 'POST':
        form = PublicRecordingForm(request.POST)
        if form.is_valid():
            recording = form.save(commit=False)
            
            # Buscar usuario por email para asignar empresa automáticamente
            uploader_email = form.cleaned_data['uploader_email']
            try:
                existing_user = User.objects.get(email=uploader_email)
                # Si el usuario tiene perfil con empresa, usar esa empresa
                if hasattr(existing_user, 'userprofile') and existing_user.userprofile.company:
                    recording.company = existing_user.userprofile.company
                else:
                    # Si no tiene empresa asignada, dejarlo sin empresa
                    recording.company = None
            except User.DoesNotExist:
                # Si no existe el usuario, dejarlo sin empresa
                recording.company = None
            
            # Guardar información del uploader en la descripción si no hay descripción
            uploader_info = f"\nSubido por: {form.cleaned_data['uploader_name']} ({uploader_email})"
            if recording.description:
                recording.description += uploader_info
            else:
                recording.description = f"Grabación subida públicamente.{uploader_info}"
            
            # Marcar como público
            recording.is_public = True
            
            # Procesar datos de audio
            audio_data = form.cleaned_data['audio_data']
            if audio_data:
                try:
                    # Remover el prefijo 'data:audio/webm;codecs=opus;base64,' si existe
                    if ',' in audio_data:
                        audio_data = audio_data.split(',')[1]
                    
                    # Decodificar base64
                    audio_binary = base64.b64decode(audio_data)
                    
                    # Crear archivo temporal
                    audio_file = ContentFile(audio_binary, name=f"{recording.title}.webm")
                    recording.audio_file = audio_file
                    
                    recording.save()
                    
                    # Iniciar transcripción automática
                    try:
                        from .ai_utils import transcribe_recording
                        import threading
                        
                        # Ejecutar transcripción en hilo separado para no bloquear la respuesta
                        def transcribe_async():
                            transcribe_recording(recording.id)
                        
                        thread = threading.Thread(target=transcribe_async)
                        thread.daemon = True
                        thread.start()
                    except Exception as e:
                        print(f"Error iniciando transcripción automática: {e}")
                    
                    # Verificar si se encontró una empresa para el usuario
                    try:
                        existing_user = User.objects.get(email=uploader_email)
                        if hasattr(existing_user, 'userprofile') and existing_user.userprofile.company:
                            messages.success(
                                request, 
                                f'Grabación guardada exitosamente y asignada automáticamente a la empresa: {existing_user.userprofile.company.name}. La transcripción se procesará automáticamente.'
                            )
                        else:
                            messages.success(
                                request, 
                                'Grabación guardada exitosamente. Tu cuenta no tiene empresa asignada, por lo que la grabación estará disponible para revisión de los agentes.'
                            )
                    except User.DoesNotExist:
                        messages.success(
                            request, 
                            'Grabación guardada exitosamente. No se encontró una cuenta con este email, por lo que la grabación estará disponible para revisión de los agentes.'
                        )
                    
                    return redirect('public_recording_upload')
                    
                except Exception as e:
                    messages.error(request, f'Error al procesar la grabación: {str(e)}')
            else:
                messages.error(request, 'No se recibieron datos de audio.')
    else:
        form = PublicRecordingForm()
    
    context = {
        'page_title': 'Grabar Audio Público',
        'form': form,
    }
    
    return render(request, 'tickets/public_recording_upload.html', context)


@login_required 
@user_passes_test(is_agent)
def recording_delete_view(request, recording_id):
    """Vista para eliminar grabaciones (solo agentes)"""
    from .models import Recording
    
    recording = get_object_or_404(Recording, id=recording_id)
    
    if request.method == 'POST':
        # Eliminar archivo físico
        if recording.audio_file:
            try:
                recording.audio_file.delete()
            except:
                pass
        
        recording.delete()
        messages.success(request, 'Grabación eliminada exitosamente.')
        return redirect('recordings_list')
    
    context = {
        'page_title': 'Eliminar Grabación',
        'recording': recording,
    }
    
    return render(request, 'tickets/recording_delete.html', context)


@login_required 
@user_passes_test(is_agent)
def recordings_stats_view(request):
    """Vista de estadísticas de grabaciones (solo agentes)"""
    from django.db.models import Count, Sum, Avg
    from .models import Recording, RecordingPlayback
    
    # Estadísticas generales
    total_recordings = Recording.objects.count()
    total_playbacks = RecordingPlayback.objects.count()
    total_size = Recording.objects.aggregate(
        total=Sum('file_size')
    )['total'] or 0
    
    # Calcular promedio de reproducciones por grabación
    avg_playbacks_per_recording = round(total_playbacks / total_recordings, 1) if total_recordings > 0 else 0
    
    # Grabaciones más reproducidas
    top_recordings = Recording.objects.annotate(
        playback_count=Count('playbacks')
    ).filter(playback_count__gt=0).order_by('-playback_count')[:10]
    
    # Grabaciones por empresa
    recordings_by_company = Recording.objects.values(
        'company__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Grabaciones por estado de transcripción
    recordings_by_status = Recording.objects.values(
        'transcription_status'
    ).annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Actividad reciente
    recent_uploads = Recording.objects.select_related(
        'company', 'uploaded_by'
    ).order_by('-created_at')[:10]
    
    recent_playbacks = RecordingPlayback.objects.select_related(
        'recording', 'played_by'
    ).order_by('-played_at')[:10]
    
    context = {
        'page_title': 'Estadísticas de Grabaciones',
        'total_recordings': total_recordings,
        'total_playbacks': total_playbacks,
        'total_size': total_size,
        'avg_playbacks_per_recording': avg_playbacks_per_recording,
        'top_recordings': top_recordings,
        'recordings_by_company': recordings_by_company,
        'recordings_by_status': recordings_by_status,
        'recent_uploads': recent_uploads,
        'recent_playbacks': recent_playbacks,
    }
    
    return render(request, 'tickets/recordings_stats.html', context)


@login_required
def recording_detail_view(request, recording_id):
    """Vista detallada de una grabación con transcripción editable"""
    from .models import Recording, RecordingPlayback
    from .forms import TranscriptionEditForm, TranscriptionActionForm
    from .ai_utils import transcribe_recording, AudioTranscriber
    
    recording = get_object_or_404(Recording, id=recording_id)
    
    # Verificar permisos usando la función is_agent de utils
    if not is_agent(request.user):
        # Si no es agente, verificar que tiene empresa y es la misma que la grabación
        user_profile = getattr(request.user, 'userprofile', None)
        if not user_profile:
            raise PermissionDenied("No tienes perfil de usuario configurado")
        if recording.company != user_profile.company:
            raise PermissionDenied("No tienes permiso para ver esta grabación")
    
    # Registrar reproducción si es una visualización completa
    if request.method == 'GET' and 'transcript_only' not in request.GET:
        RecordingPlayback.objects.create(
            recording=recording,
            played_by=request.user,
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
    
    transcription_form = TranscriptionEditForm(instance=recording)
    action_form = TranscriptionActionForm(recording=recording)
    
    # Procesar formularios
    if request.method == 'POST':
        if 'save_transcription' in request.POST:
            transcription_form = TranscriptionEditForm(request.POST, instance=recording)
            if transcription_form.is_valid():
                transcription_form.save()
                messages.success(request, 'Transcripción guardada exitosamente.')
                return redirect('recording_detail', recording_id=recording.id)
        
        elif 'process_ai' in request.POST:
            action_form = TranscriptionActionForm(request.POST, recording=recording)
            if action_form.is_valid():
                action = action_form.cleaned_data['action']
                context_hint = action_form.cleaned_data['context_hint']
                
                try:
                    if action == 'generate' or action == 'regenerate':
                        # Transcribir desde audio
                        result = transcribe_recording(recording.id)
                        if result['success']:
                            messages.success(request, 'Transcripción generada exitosamente con IA.')
                        else:
                            messages.error(request, f'Error al generar transcripción: {result["error"]}')
                    
                    elif action == 'improve':
                        # Mejorar transcripción existente
                        if recording.transcription_text:
                            transcriber = AudioTranscriber()
                            context = f"Título: {recording.title}. Descripción: {recording.description}"
                            if context_hint:
                                context += f". Contexto adicional: {context_hint}"
                            
                            improved_text = transcriber.improve_transcription_with_ai(
                                recording.transcription_text, 
                                context
                            )
                            
                            if improved_text != recording.transcription_text:
                                recording.transcription_text = improved_text
                                recording.save()
                                messages.success(request, 'Transcripción mejorada exitosamente con IA.')
                            else:
                                messages.info(request, 'La transcripción ya estaba optimizada.')
                        else:
                            messages.error(request, 'No hay transcripción para mejorar.')
                    
                    return redirect('recording_detail', recording_id=recording.id)
                    
                except Exception as e:
                    messages.error(request, f'Error procesando con IA: {str(e)}')
    
    # Obtener estadísticas de reproducción
    playback_count = recording.playbacks.count()
    recent_playbacks = recording.playbacks.select_related('played_by').order_by('-played_at')[:5]
    
    # Calcular valores estadísticos para el template
    bitrate_kbps = None
    if recording.file_size and recording.duration_seconds and recording.duration_seconds > 0:
        # Bitrate en kbps (file_size en bytes / duration_seconds / 128)
        bitrate_kbps = round(recording.file_size / recording.duration_seconds / 128, 0)
    
    context = {
        'page_title': f'Grabación: {recording.title}',
        'recording': recording,
        'transcription_form': transcription_form,
        'action_form': action_form,
        'playback_count': playback_count,
        'recent_playbacks': recent_playbacks,
        'bitrate_kbps': bitrate_kbps,
    }
    
    return render(request, 'tickets/recording_detail.html', context)


@login_required
@user_passes_test(is_agent)
def recording_transcribe_async_view(request, recording_id):
    """Vista para iniciar transcripción asíncrona (solo agentes)"""
    from .models import Recording
    from .ai_utils import transcribe_recording
    import json
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        recording = Recording.objects.get(id=recording_id)
        
        # Verificar que no esté ya procesando
        if recording.transcription_status == 'processing':
            return JsonResponse({
                'success': False, 
                'error': 'La transcripción ya está en proceso'
            })
        
        # Iniciar transcripción
        result = transcribe_recording(recording_id)
        
        return JsonResponse({
            'success': result['success'],
            'transcription': result.get('transcription', ''),
            'confidence': result.get('confidence', 0),
            'error': result.get('error', '')
        })
        
    except Recording.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Grabación no encontrada'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@user_passes_test(is_agent)
def recording_bulk_transcribe_view(request):
    """Vista para transcribir múltiples grabaciones en lote (solo agentes)"""
    from .models import Recording
    from .ai_utils import transcribe_recording
    import json
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        data = json.loads(request.body)
        recording_ids = data.get('recording_ids', [])
        
        if not recording_ids:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron IDs de grabaciones'})
        
        results = []
        for recording_id in recording_ids:
            try:
                recording = Recording.objects.get(id=recording_id)
                
                # Solo procesar si no está ya transcrito o si hay error previo
                if recording.transcription_status in ['pending', 'failed']:
                    result = transcribe_recording(recording_id)
                    results.append({
                        'id': recording_id,
                        'title': recording.title,
                        'success': result['success'],
                        'error': result.get('error', '')
                    })
                else:
                    results.append({
                        'id': recording_id,
                        'title': recording.title,
                        'success': True,
                        'error': 'Ya transcrito'
                    })
                    
            except Recording.DoesNotExist:
                results.append({
                    'id': recording_id,
                    'title': 'Desconocido',
                    'success': False,
                    'error': 'Grabación no encontrada'
                })
        
        success_count = sum(1 for r in results if r['success'])
        
        return JsonResponse({
            'success': True,
            'processed': len(results),
            'successful': success_count,
            'failed': len(results) - success_count,
            'results': results
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def voice_command_create_ticket_view(request):
    """Vista para procesar comandos de voz y crear tickets automáticamente"""
    from .ai_utils import VoiceCommandProcessor
    import tempfile
    import os
    import logging
    
    logger = logging.getLogger(__name__)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        # Verificar que se envió un archivo de audio
        if 'audio' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No se envió archivo de audio'})
        
        audio_file = request.FILES['audio']
        
        # Validar formato de audio
        allowed_formats = ['.wav', '.mp3', '.m4a', '.webm', '.ogg']
        file_extension = os.path.splitext(audio_file.name)[1].lower()
        
        if file_extension not in allowed_formats:
            return JsonResponse({
                'success': False, 
                'error': f'Formato de audio no soportado. Use: {", ".join(allowed_formats)}'
            })
        
        # Crear archivo temporal para procesar el audio
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as temp_file:
            # Escribir el contenido del archivo subido al archivo temporal
            for chunk in audio_file.chunks():
                temp_file.write(chunk)
            temp_file_path = temp_file.name
        
        try:
            # Procesar el comando de voz
            processor = VoiceCommandProcessor()
            result = processor.process_voice_command(temp_file_path, request.user)
            
            if result['success']:
                # Comando procesado exitosamente
                ticket_number = result["ticket"].get("ticket_number", result["ticket"].get("ticket_id", "Sin número"))
                return JsonResponse({
                    'success': True,
                    'message': f'Ticket creado exitosamente: #{ticket_number}',
                    'ticket': result['ticket'],
                    'transcription': result['transcription'],
                    'ticket_info': result['ticket_info']
                })
            else:
                # Error procesando el comando
                return JsonResponse({
                    'success': False,
                    'error': result.get('error', 'Error desconocido procesando comando')
                })
        
        finally:
            # Limpiar archivo temporal
            try:
                os.unlink(temp_file_path)
            except OSError:
                pass  # Ignorar errores al eliminar archivo temporal
    
    except Exception as e:
        logger.error(f"Error en comando de voz: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Error interno del servidor: {str(e)}'})


@login_required 
def voice_command_interface_view(request):
    """Vista para mostrar la interfaz de comandos de voz"""
    from .models import Ticket
    
    # Obtener estadísticas de tickets recientes del usuario
    recent_tickets = Ticket.objects.filter(
        created_by=request.user
    ).order_by('-created_at')[:5]
    
    # Obtener estadísticas generales
    total_tickets = Ticket.objects.filter(created_by=request.user).count()
    open_tickets = Ticket.objects.filter(created_by=request.user, status='open').count()
    
    context = {
        'page_title': 'Comandos de Voz IA - TicketProo',
        'recent_tickets': recent_tickets,
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
    }
    
    return render(request, 'tickets/voice_command_interface.html', context)


# ====================================
# VISTAS DE ANALYTICS Y REPORTES
# ====================================

@login_required
@require_http_methods(["GET"])
def page_visits_analytics_view(request):
    """Vista principal de analytics de visitas a páginas públicas"""
    from django.db.models import Count, Q
    from django.utils import timezone
    from datetime import datetime, timedelta
    from .models import PageVisit
    
    # Filtros de fecha
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)  # Por defecto últimos 30 días
    
    # Obtener fechas desde parámetros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if fecha_desde:
        try:
            start_date = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            end_date = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Filtros adicionales
    page_type_filter = request.GET.get('page_type')
    country_filter = request.GET.get('country')
    exclude_bots = request.GET.get('exclude_bots', 'true') == 'true'
    
    # Query base
    visits = PageVisit.objects.filter(
        visited_at__date__gte=start_date,
        visited_at__date__lte=end_date
    )
    
    if page_type_filter:
        visits = visits.filter(page_type=page_type_filter)
    
    if country_filter:
        visits = visits.filter(country_code=country_filter)
    
    if exclude_bots:
        visits = visits.filter(is_bot=False)
    
    # Estadísticas generales
    total_visits = visits.count()
    unique_ips = visits.values('ip_address').distinct().count()
    unique_countries = visits.exclude(country_code='').values('country_code').distinct().count()
    
    # Visitas por página
    visits_by_page = visits.values('page_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Visitas por día (últimos 30 días)
    visits_by_day = []
    current_date = start_date
    while current_date <= end_date:
        day_visits = visits.filter(visited_at__date=current_date).count()
        visits_by_day.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'visits': day_visits
        })
        current_date += timedelta(days=1)
    
    # Top países
    top_countries = visits.exclude(country='').values('country', 'country_code').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Top navegadores
    top_browsers = visits.exclude(browser='').values('browser').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Dispositivos móviles vs desktop
    mobile_visits = visits.filter(is_mobile=True).count()
    desktop_visits = visits.filter(is_mobile=False).count()
    
    # Asegurar que hay al menos un dispositivo para el gráfico
    if mobile_visits == 0 and desktop_visits == 0:
        desktop_visits = 0
        mobile_visits = 0
    
    # Páginas más visitadas (detalle)
    top_pages = visits.values('page_url', 'page_title', 'page_type').annotate(
        count=Count('id')
    ).order_by('-count')[:20]
    
    # Landing pages específicas
    landing_pages_stats = visits.filter(page_type='landing').values('page_url', 'page_title').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Estadísticas específicas de landing pages
    landing_total_visits = visits.filter(page_type='landing').count()
    landing_unique_visitors = visits.filter(page_type='landing').values('ip_address').distinct().count()
    
    # Conversión de landing pages (si hay parámetros UTM)
    landing_utm_sources = visits.filter(
        page_type='landing',
        utm_source__isnull=False
    ).exclude(utm_source='').values('utm_source').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Fuentes de tráfico (referrers)
    referrers = visits.exclude(referrer='').values('referrer').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Visitas recientes
    recent_visits = visits.order_by('-visited_at')[:50]
    
    # Para filtros
    page_types = PageVisit.PAGE_CHOICES
    countries = visits.exclude(country_code='').values('country_code', 'country').distinct().order_by('country')
    
    context = {
        'page_title': 'Analytics de Páginas Públicas',
        'start_date': start_date,
        'end_date': end_date,
        'total_visits': total_visits,
        'unique_ips': unique_ips,
        'unique_countries': unique_countries,
        'visits_by_page': visits_by_page,
        'visits_by_day': visits_by_day,
        'top_countries': top_countries,
        'top_browsers': top_browsers,
        'mobile_visits': mobile_visits,
        'desktop_visits': desktop_visits,
        'top_pages': top_pages,
        'landing_pages_stats': landing_pages_stats,
        'landing_total_visits': landing_total_visits,
        'landing_unique_visitors': landing_unique_visitors,
        'landing_utm_sources': landing_utm_sources,
        'referrers': referrers,
        'recent_visits': recent_visits,
        'page_types': page_types,
        'countries': countries,
        # Filtros aplicados
        'page_type_filter': page_type_filter,
        'country_filter': country_filter,
        'exclude_bots': exclude_bots,
    }
    
    return render(request, 'tickets/page_visits_analytics.html', context)


@login_required
@require_http_methods(["GET"])
def page_visits_detail_view(request):
    """Vista detallada de visitas a páginas con filtros avanzados"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    from .models import PageVisit
    
    # Query base
    visits = PageVisit.objects.all().order_by('-visited_at')
    
    # Filtros
    search = request.GET.get('search')
    page_type = request.GET.get('page_type')
    country = request.GET.get('country')
    is_mobile = request.GET.get('is_mobile')
    exclude_bots = request.GET.get('exclude_bots', 'true') == 'true'
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if search:
        visits = visits.filter(
            Q(page_url__icontains=search) |
            Q(page_title__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(country__icontains=search) |
            Q(browser__icontains=search)
        )
    
    if page_type:
        visits = visits.filter(page_type=page_type)
    
    if country:
        visits = visits.filter(country_code=country)
    
    if is_mobile:
        visits = visits.filter(is_mobile=is_mobile == 'true')
    
    if exclude_bots:
        visits = visits.filter(is_bot=False)
    
    if fecha_desde:
        try:
            from datetime import datetime
            start_date = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            visits = visits.filter(visited_at__date__gte=start_date)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            from datetime import datetime
            end_date = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            visits = visits.filter(visited_at__date__lte=end_date)
        except ValueError:
            pass
    
    # Paginación
    paginator = Paginator(visits, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Para filtros
    page_types = PageVisit.PAGE_CHOICES
    countries = PageVisit.objects.exclude(country_code='').values('country_code', 'country').distinct().order_by('country')
    
    context = {
        'page_title': 'Detalle de Visitas',
        'page_obj': page_obj,
        'page_types': page_types,
        'countries': countries,
        # Filtros aplicados
        'search': search,
        'page_type': page_type,
        'country': country,
        'is_mobile': is_mobile,
        'exclude_bots': exclude_bots,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'tickets/page_visits_detail.html', context)


@login_required
@require_http_methods(["GET"])
def page_visits_export_view(request):
    """Vista para exportar datos de visitas en CSV"""
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    from .models import PageVisit
    
    # Aplicar los mismos filtros que en la vista principal
    visits = PageVisit.objects.all().order_by('-visited_at')
    
    # Filtros (copiar lógica de page_visits_detail_view)
    search = request.GET.get('search')
    page_type = request.GET.get('page_type')
    country = request.GET.get('country')
    exclude_bots = request.GET.get('exclude_bots', 'true') == 'true'
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if search:
        from django.db.models import Q
        visits = visits.filter(
            Q(page_url__icontains=search) |
            Q(page_title__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(country__icontains=search) |
            Q(browser__icontains=search)
        )
    
    if page_type:
        visits = visits.filter(page_type=page_type)
    
    if country:
        visits = visits.filter(country_code=country)
    
    if exclude_bots:
        visits = visits.filter(is_bot=False)
    
    if fecha_desde:
        try:
            from datetime import datetime
            start_date = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            visits = visits.filter(visited_at__date__gte=start_date)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            from datetime import datetime
            end_date = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            visits = visits.filter(visited_at__date__lte=end_date)
        except ValueError:
            pass
    
    # Limitar a 10000 registros para evitar problemas de memoria
    visits = visits[:10000]
    
    # Crear respuesta CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="visitas_paginas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Encabezados
    writer.writerow([
        'Fecha/Hora',
        'Tipo de Página',
        'URL',
        'Título',
        'IP',
        'País',
        'Ciudad',
        'Navegador',
        'Sistema Operativo',
        'Es Móvil',
        'Es Bot',
        'Referrer',
        'UTM Source',
        'UTM Medium',
        'UTM Campaign'
    ])
    
    # Datos
    for visit in visits:
        writer.writerow([
            visit.visited_at.strftime('%Y-%m-%d %H:%M:%S'),
            visit.get_page_type_display(),
            visit.page_url,
            visit.page_title,
            visit.ip_address,
            visit.country,
            visit.city,
            visit.browser_info,
            visit.operating_system,
            'Sí' if visit.is_mobile else 'No',
            'Sí' if visit.is_bot else 'No',
            visit.referrer,
            visit.utm_source,
            visit.utm_medium,
            visit.utm_campaign
        ])
    
    return response


# ====================================
# CONFIGURADOR DE IA PARA BLOG
# ====================================

@login_required
@require_http_methods(["GET"])
def ai_blog_configurators_list_view(request):
    """Vista lista de configuradores de IA para blog"""
    from .models import AIBlogConfigurator
    
    configurators = AIBlogConfigurator.objects.all().order_by('-created_at')
    
    # Agregar información de próxima ejecución y estadísticas
    for config in configurators:
        config.next_run_formatted = config.next_run_time.strftime('%d/%m/%Y %H:%M') if config.next_run_time else 'No programado'
        config.last_run_formatted = config.last_run.strftime('%d/%m/%Y %H:%M') if config.last_run else 'Nunca'
        
        # Calcular estadísticas desde los logs
        logs = config.generation_logs.all()
        config.successful_runs = logs.filter(generation_status='success').count()
        config.failed_runs = logs.filter(generation_status='error').count()
        
        # Contar posts realmente generados desde los logs
        posts_generated = logs.filter(
            generation_status='success',
            generated_post__isnull=False
        ).count()
        config.posts_generated = posts_generated
    
    context = {
        'page_title': 'Configuradores de IA para Blog',
        'configurators': configurators,
    }
    
    return render(request, 'tickets/ai_blog_configurators_list.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def ai_blog_configurator_create_view(request):
    """Vista para crear configurador de IA"""
    from .models import AIBlogConfigurator, BlogCategory
    
    if request.method == 'POST':
        try:
            configurator = AIBlogConfigurator.objects.create(
                name=request.POST.get('name'),
                description=request.POST.get('description', ''),
                keywords=request.POST.get('keywords'),
                topic_template=request.POST.get('topic_template'),
                content_length=request.POST.get('content_length'),
                content_style=request.POST.get('content_style'),
                is_active=request.POST.get('is_active') == 'on',
                schedule_time=request.POST.get('schedule_time'),
                frequency_days=int(request.POST.get('frequency_days', 1)),
                max_posts_per_run=int(request.POST.get('max_posts_per_run', 1)),
                default_category_id=request.POST.get('default_category') or None,
            )
            
            messages.success(request, f'Configurador "{configurator.name}" creado exitosamente.')
            return redirect('ai_blog_configurators_list')
            
        except Exception as e:
            messages.error(request, f'Error creando configurador: {str(e)}')
    
    # Para el formulario
    categories = BlogCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'page_title': 'Crear Configurador de IA',
        'categories': categories,
    }
    
    return render(request, 'tickets/ai_blog_configurator_form.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def ai_blog_configurator_edit_view(request, pk):
    """Vista para editar configurador de IA"""
    from .models import AIBlogConfigurator, BlogCategory
    
    configurator = get_object_or_404(AIBlogConfigurator, pk=pk)
    
    if request.method == 'POST':
        try:
            configurator.name = request.POST.get('name')
            configurator.description = request.POST.get('description', '')
            configurator.keywords = request.POST.get('keywords')
            configurator.topic_template = request.POST.get('topic_template')
            configurator.content_length = request.POST.get('content_length')
            configurator.content_style = request.POST.get('content_style')
            configurator.is_active = request.POST.get('is_active') == 'on'
            configurator.schedule_time = request.POST.get('schedule_time')
            configurator.frequency_days = int(request.POST.get('frequency_days', 1))
            configurator.max_posts_per_run = int(request.POST.get('max_posts_per_run', 1))
            configurator.default_category_id = request.POST.get('default_category') or None
            configurator.save()
            
            messages.success(request, f'Configurador "{configurator.name}" actualizado exitosamente.')
            return redirect('ai_blog_configurators_list')
            
        except Exception as e:
            messages.error(request, f'Error actualizando configurador: {str(e)}')
    
    # Para el formulario
    categories = BlogCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'page_title': 'Editar Configurador de IA',
        'configurator': configurator,
        'categories': categories,
    }
    
    return render(request, 'tickets/ai_blog_configurator_form.html', context)


@login_required
@require_http_methods(["POST"])
def ai_blog_configurator_delete_view(request, pk):
    """Vista para eliminar configurador de IA"""
    from .models import AIBlogConfigurator
    
    configurator = get_object_or_404(AIBlogConfigurator, pk=pk)
    configurator_name = configurator.name
    configurator.delete()
    
    messages.success(request, f'Configurador "{configurator_name}" eliminado exitosamente.')
    return redirect('ai_blog_configurators_list')


@login_required
@require_http_methods(["POST"])
def ai_blog_configurator_toggle_view(request, pk):
    """Vista para activar/desactivar configurador de IA"""
    from .models import AIBlogConfigurator
    
    configurator = get_object_or_404(AIBlogConfigurator, pk=pk)
    configurator.is_active = not configurator.is_active
    configurator.save()
    
    status = 'activado' if configurator.is_active else 'desactivado'
    messages.success(request, f'Configurador "{configurator.name}" {status} exitosamente.')
    
    return redirect('ai_blog_configurators_list')


@login_required
@require_http_methods(["POST"])
def ai_blog_configurator_run_now_view(request, pk):
    """Vista para ejecutar manualmente un configurador de IA"""
    from .models import AIBlogConfigurator
    
    configurator = get_object_or_404(AIBlogConfigurator, pk=pk)
    
    try:
        # Importar la función de generación
        from .ai_blog_generator import run_ai_blog_generation
        
        # Ejecutar manualmente con force=True para ignorar frecuencia programada
        result = run_ai_blog_generation(configurator, force=True)
        
        if result['success']:
            messages.success(request, f'✅ Generación exitosa: {result["posts_created"]} post(s) creado(s)')
        else:
            messages.error(request, f'❌ Error en generación: {result["error"]}')
            
    except Exception as e:
        messages.error(request, f'Error ejecutando configurador: {str(e)}')
    
    return redirect('ai_blog_configurators_list')


@login_required
@require_http_methods(["GET"])
def ai_blog_generation_logs_view(request, pk):
    """Vista para ver logs de generación de un configurador"""
    from .models import AIBlogConfigurator, AIBlogGenerationLog
    
    configurator = get_object_or_404(AIBlogConfigurator, pk=pk)
    logs = AIBlogGenerationLog.objects.filter(configurator=configurator).order_by('-created_at')
    
    # Calcular estadísticas
    total_logs = logs.count()
    successful_logs = logs.filter(generation_status='success').count()
    failed_logs = logs.filter(generation_status='error').count()
    posts_generated = logs.filter(generated_post__isnull=False).count()
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': f'Logs de Generación - {configurator.name}',
        'configurator': configurator,
        'page_obj': page_obj,
        'stats': {
            'total_logs': total_logs,
            'successful_logs': successful_logs,
            'failed_logs': failed_logs,
            'posts_generated': posts_generated,
        }
    }
    
    return render(request, 'tickets/ai_blog_generation_logs.html', context)


# ============= VISTAS DE DOCUMENTACIÓN MÚLTIPLE =============

@login_required
@user_passes_test(is_agent, login_url='/')
def multiple_documentation_list_view(request):
    """Vista para listar todas las documentaciones múltiples"""
    documentations = MultipleDocumentation.objects.all().order_by('-created_at')
    
    # Búsqueda
    search = request.GET.get('search')
    if search:
        documentations = documentations.filter(
            models.Q(title__icontains=search) |
            models.Q(description__icontains=search)
        )
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(documentations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': 'Documentación Múltiple',
        'page_obj': page_obj,
        'search': search,
    }
    
    return render(request, 'tickets/multiple_documentation_list.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def multiple_documentation_create_view(request):
    """Vista para crear una nueva documentación múltiple"""
    from .forms import MultipleDocumentationForm
    
    if request.method == 'POST':
        form = MultipleDocumentationForm(request.POST)
        if form.is_valid():
            documentation = form.save(commit=False)
            documentation.created_by = request.user
            documentation.save()
            
            messages.success(request, f'Documentación "{documentation.title}" creada correctamente.')
            return redirect('multiple_documentation_detail', pk=documentation.pk)
    else:
        form = MultipleDocumentationForm()
    
    context = {
        'page_title': 'Crear Documentación Múltiple',
        'form': form,
        'is_create': True,
    }
    
    return render(request, 'tickets/multiple_documentation_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def multiple_documentation_detail_view(request, pk):
    """Vista para ver detalles de una documentación múltiple"""
    from .models import MultipleDocumentation
    
    documentation = get_object_or_404(MultipleDocumentation, pk=pk)
    items = documentation.items.all().order_by('number')
    
    context = {
        'page_title': f'Documentación: {documentation.title}',
        'documentation': documentation,
        'items': items,
    }
    
    return render(request, 'tickets/multiple_documentation_detail.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def multiple_documentation_edit_view(request, pk):
    """Vista para editar una documentación múltiple"""
    from .models import MultipleDocumentation
    from .forms import MultipleDocumentationForm
    
    documentation = get_object_or_404(MultipleDocumentation, pk=pk)
    
    if request.method == 'POST':
        form = MultipleDocumentationForm(request.POST, instance=documentation)
        if form.is_valid():
            form.save()
            messages.success(request, f'Documentación "{documentation.title}" actualizada correctamente.')
            return redirect('multiple_documentation_detail', pk=documentation.pk)
    else:
        form = MultipleDocumentationForm(instance=documentation)
    
    context = {
        'page_title': f'Editar: {documentation.title}',
        'documentation': documentation,
        'form': form,
        'is_create': False,
    }
    
    return render(request, 'tickets/multiple_documentation_form.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def multiple_documentation_delete_view(request, pk):
    """Vista para eliminar una documentación múltiple"""
    from .models import MultipleDocumentation
    
    documentation = get_object_or_404(MultipleDocumentation, pk=pk)
    
    if request.method == 'POST':
        try:
            title = documentation.title
            documentation.delete()
            messages.success(request, f'Documentación "{title}" eliminada correctamente.')
            return redirect('multiple_documentation_list')
            
        except Exception as e:
            messages.error(request, f'Error al eliminar la documentación: {str(e)}')
            return redirect('multiple_documentation_detail', pk=pk)
    
    context = {
        'page_title': f'Eliminar: {documentation.title}',
        'documentation': documentation,
    }
    
    return render(request, 'tickets/multiple_documentation_delete.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def multiple_documentation_add_item_view(request, pk):
    """Vista para agregar un item a la documentación múltiple"""
    from .models import MultipleDocumentation, MultipleDocumentationItem
    
    documentation = get_object_or_404(MultipleDocumentation, pk=pk)
    
    if request.method == 'POST':
        try:
            # Obtener el siguiente número disponible
            last_item = documentation.items.order_by('-number').first()
            next_number = (last_item.number + 1) if last_item else 1
            
            # Permitir que el usuario especifique el número
            number = int(request.POST.get('number', next_number))
            
            item = MultipleDocumentationItem.objects.create(
                documentation=documentation,
                number=number,
                name=request.POST['name'],
                description=request.POST.get('description', ''),
                file=request.FILES['file']
            )
            
            messages.success(request, f'Documento "{item.name}" agregado correctamente.')
            return redirect('multiple_documentation_detail', pk=documentation.pk)
            
        except Exception as e:
            messages.error(request, f'Error al agregar el documento: {str(e)}')
    
    # Obtener el siguiente número sugerido
    last_item = documentation.items.order_by('-number').first()
    suggested_number = (last_item.number + 1) if last_item else 1
    
    context = {
        'page_title': f'Agregar Documento - {documentation.title}',
        'documentation': documentation,
        'suggested_number': suggested_number,
    }
    
    return render(request, 'tickets/multiple_documentation_add_item.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def multiple_documentation_edit_item_view(request, pk, item_id):
    """Vista para editar un item de documentación"""
    from .models import MultipleDocumentation, MultipleDocumentationItem
    
    documentation = get_object_or_404(MultipleDocumentation, pk=pk)
    item = get_object_or_404(MultipleDocumentationItem, pk=item_id, documentation=documentation)
    
    if request.method == 'POST':
        try:
            item.number = int(request.POST['number'])
            item.name = request.POST['name']
            item.description = request.POST.get('description', '')
            
            # Solo actualizar el archivo si se proporciona uno nuevo
            if 'file' in request.FILES:
                # Eliminar el archivo anterior
                if item.file:
                    item.file.delete(save=False)
                item.file = request.FILES['file']
            
            item.save()
            
            messages.success(request, f'Documento "{item.name}" actualizado correctamente.')
            return redirect('multiple_documentation_detail', pk=documentation.pk)
            
        except Exception as e:
            messages.error(request, f'Error al actualizar el documento: {str(e)}')
    
    context = {
        'page_title': f'Editar Documento - {item.name}',
        'documentation': documentation,
        'item': item,
    }
    
    return render(request, 'tickets/multiple_documentation_edit_item.html', context)


@login_required
@user_passes_test(is_agent, login_url='/')
def multiple_documentation_delete_item_view(request, pk, item_id):
    """Vista para eliminar un item de documentación"""
    from .models import MultipleDocumentation, MultipleDocumentationItem
    
    documentation = get_object_or_404(MultipleDocumentation, pk=pk)
    item = get_object_or_404(MultipleDocumentationItem, pk=item_id, documentation=documentation)
    
    if request.method == 'POST':
        try:
            name = item.name
            # Eliminar el archivo físico
            if item.file:
                item.file.delete(save=False)
            item.delete()
            
            messages.success(request, f'Documento "{name}" eliminado correctamente.')
            return redirect('multiple_documentation_detail', pk=documentation.pk)
            
        except Exception as e:
            messages.error(request, f'Error al eliminar el documento: {str(e)}')
    
    context = {
        'page_title': f'Eliminar Documento - {item.name}',
        'documentation': documentation,
        'item': item,
    }
    
    return render(request, 'tickets/multiple_documentation_delete_item.html', context)


def multiple_documentation_public_view(request, token):
    """Vista pública para ver una documentación múltiple (sin autenticación)"""
    from .models import (
        MultipleDocumentation, MultipleDocumentationStats, 
        MultipleDocumentationVisit
    )
    from .forms import DocumentationPasswordForm
    
    try:
        documentation = get_object_or_404(
            MultipleDocumentation, 
            public_token=token, 
            is_active=True
        )
        
        # Verificar si está protegida con contraseña
        if documentation.password_protected:
            # Verificar si ya se validó la contraseña en esta sesión
            session_key = f'doc_access_{token}'
            if not request.session.get(session_key, False):
                # Procesar formulario de contraseña
                if request.method == 'POST':
                    form = DocumentationPasswordForm(request.POST, documentation=documentation)
                    if form.is_valid():
                        # Contraseña correcta, marcar en sesión
                        request.session[session_key] = True
                        request.session.set_expiry(3600)  # Expirar en 1 hora
                        return redirect('multiple_documentation_public', token=token)
                else:
                    form = DocumentationPasswordForm(documentation=documentation)
                
                # Mostrar formulario de contraseña
                context = {
                    'page_title': f'Acceso a {documentation.title}',
                    'documentation': documentation,
                    'form': form,
                }
                return render(request, 'tickets/multiple_documentation_password.html', context)
        
        # Obtener IP del visitante
        ip_address = request.META.get('REMOTE_ADDR')
        if request.META.get('HTTP_X_FORWARDED_FOR'):
            ip_address = request.META.get('HTTP_X_FORWARDED_FOR').split(',')[0]
        
        # Registrar la visita
        visit = MultipleDocumentationVisit.objects.create(
            documentation=documentation,
            ip_address=ip_address,
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            referer=request.META.get('HTTP_REFERER')
        )
        
        # Actualizar o crear estadísticas
        stats, created = MultipleDocumentationStats.objects.get_or_create(
            documentation=documentation,
            defaults={
                'first_view_date': timezone.now(),
                'last_view_date': timezone.now(),
                'page_views': 1,
                'unique_visitors': 1,
            }
        )
        
        if not created:
            # Actualizar estadísticas existentes
            stats.page_views += 1
            stats.last_view_date = timezone.now()
            
            # Verificar si es un visitante único (no ha visitado en las últimas 24 horas)
            yesterday = timezone.now() - timezone.timedelta(days=1)
            if not MultipleDocumentationVisit.objects.filter(
                documentation=documentation,
                ip_address=ip_address,
                timestamp__gte=yesterday
            ).exclude(pk=visit.pk).exists():
                stats.unique_visitors += 1
            
            stats.save()
        
        items = documentation.items.all().order_by('number')
        
        context = {
            'documentation': documentation,
            'items': items,
            'page_title': documentation.title,
        }
        
        return render(request, 'tickets/multiple_documentation_public.html', context)
        
    except Exception as e:
        from django.http import Http404
        raise Http404("La documentación solicitada no existe o no está disponible.")


def multiple_documentation_download_item_view(request, token, item_id):
    """Vista para descargar un archivo específico de la documentación pública"""
    from .models import (
        MultipleDocumentation, MultipleDocumentationItem, 
        MultipleDocumentationStats, MultipleDocumentationItemStats,
        MultipleDocumentationDownload
    )
    from django.http import HttpResponse, Http404
    
    try:
        documentation = get_object_or_404(
            MultipleDocumentation, 
            public_token=token, 
            is_active=True
        )
        
        # Verificar acceso con contraseña si es necesario
        if documentation.password_protected:
            session_key = f'doc_access_{token}'
            if not request.session.get(session_key, False):
                raise Http404("Acceso no autorizado")
        
        item = get_object_or_404(
            MultipleDocumentationItem, 
            pk=item_id, 
            documentation=documentation
        )
        
        if not item.file:
            raise Http404("Archivo no encontrado")
        
        # Obtener IP del usuario
        ip_address = request.META.get('REMOTE_ADDR')
        if request.META.get('HTTP_X_FORWARDED_FOR'):
            ip_address = request.META.get('HTTP_X_FORWARDED_FOR').split(',')[0]
        
        # Registrar la descarga
        download = MultipleDocumentationDownload.objects.create(
            item=item,
            ip_address=ip_address,
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            referer=request.META.get('HTTP_REFERER')
        )
        
        # Actualizar estadísticas del archivo
        item_stats, created = MultipleDocumentationItemStats.objects.get_or_create(
            item=item,
            defaults={
                'download_count': 1,
                'unique_downloaders': 1,
                'first_download_date': timezone.now(),
                'last_download_date': timezone.now(),
            }
        )
        
        if not created:
            item_stats.download_count += 1
            item_stats.last_download_date = timezone.now()
            
            # Verificar si es un descargador único (no ha descargado este archivo antes)
            if not MultipleDocumentationDownload.objects.filter(
                item=item,
                ip_address=ip_address
            ).exclude(pk=download.pk).exists():
                item_stats.unique_downloaders += 1
            
            item_stats.save()
        
        # Actualizar estadísticas de la documentación
        doc_stats, doc_created = MultipleDocumentationStats.objects.get_or_create(
            documentation=documentation,
            defaults={
                'total_downloads': 1,
                'last_download_date': timezone.now(),
            }
        )
        
        if not doc_created:
            doc_stats.total_downloads += 1
            doc_stats.last_download_date = timezone.now()
            doc_stats.save()
        
        # Obtener la respuesta del archivo
        response = HttpResponse(item.file.read(), content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{item.file.name.split("/")[-1]}"'
        
        return response
        
    except Exception:
        raise Http404("Archivo no encontrado")


@login_required
@user_passes_test(is_agent, login_url='/')
def multiple_documentation_stats_view(request, pk):
    """Vista para mostrar estadísticas detalladas de una documentación"""
    from .models import (
        MultipleDocumentation, MultipleDocumentationStats,
        MultipleDocumentationVisit, MultipleDocumentationDownload
    )
    from django.db.models import Count, Q
    from django.utils import timezone
    from datetime import timedelta
    
    documentation = get_object_or_404(MultipleDocumentation, pk=pk)
    
    # Obtener o crear estadísticas generales
    stats, created = MultipleDocumentationStats.objects.get_or_create(
        documentation=documentation,
        defaults={
            'page_views': 0,
            'unique_visitors': 0,
            'total_downloads': 0,
        }
    )
    
    # Estadísticas por archivo
    items_stats = []
    for item in documentation.items.all().order_by('number'):
        item_stat = getattr(item, 'stats', None)
        if not item_stat:
            from .models import MultipleDocumentationItemStats
            item_stat, _ = MultipleDocumentationItemStats.objects.get_or_create(
                item=item,
                defaults={'download_count': 0, 'unique_downloaders': 0}
            )
        
        items_stats.append({
            'item': item,
            'stats': item_stat,
            'popularity': item_stat.get_popularity_percentage() if item_stat else 0,
        })
    
    # Estadísticas de los últimos 30 días
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_visits = MultipleDocumentationVisit.objects.filter(
        documentation=documentation,
        timestamp__gte=thirty_days_ago
    ).count()
    
    recent_downloads = MultipleDocumentationDownload.objects.filter(
        item__documentation=documentation,
        timestamp__gte=thirty_days_ago
    ).count()
    
    # Estadísticas por día (últimos 7 días)
    daily_stats = []
    for i in range(7):
        date = timezone.now().date() - timedelta(days=i)
        date_start = timezone.make_aware(timezone.datetime.combine(date, timezone.datetime.min.time()))
        date_end = date_start + timedelta(days=1)
        
        visits = MultipleDocumentationVisit.objects.filter(
            documentation=documentation,
            timestamp__gte=date_start,
            timestamp__lt=date_end
        ).count()
        
        downloads = MultipleDocumentationDownload.objects.filter(
            item__documentation=documentation,
            timestamp__gte=date_start,
            timestamp__lt=date_end
        ).count()
        
        daily_stats.append({
            'date': date,
            'visits': visits,
            'downloads': downloads,
        })
    
    daily_stats.reverse()  # Mostrar desde el día más antiguo
    
    # Top 5 archivos más descargados
    top_files = sorted(items_stats, key=lambda x: x['stats'].download_count, reverse=True)[:5]
    
    # Últimas visitas y descargas
    recent_visits_list = MultipleDocumentationVisit.objects.filter(
        documentation=documentation
    ).order_by('-timestamp')[:10]
    
    recent_downloads_list = MultipleDocumentationDownload.objects.filter(
        item__documentation=documentation
    ).select_related('item').order_by('-timestamp')[:10]
    
    context = {
        'page_title': f'Estadísticas - {documentation.title}',
        'documentation': documentation,
        'stats': stats,
        'items_stats': items_stats,
        'recent_visits': recent_visits,
        'recent_downloads': recent_downloads,
        'daily_stats': daily_stats,
        'top_files': top_files,
        'recent_visits_list': recent_visits_list,
        'recent_downloads_list': recent_downloads_list,
        'conversion_rate': stats.get_conversion_rate(),
        'avg_downloads_per_day': stats.get_average_downloads_per_day(),
    }
    
    return render(request, 'tickets/multiple_documentation_stats.html', context)


# ====================
# Task Schedule Views
# ====================

def get_user_company(user):
    """Obtiene la empresa del usuario"""
    if hasattr(user, 'profile') and user.profile.company:
        return user.profile.company
    return None

def user_has_company_access(user, company):
    """Verifica si el usuario tiene acceso a una empresa"""
    user_company = get_user_company(user)
    if not user_company:
        return False
    return user_company == company

def get_user_companies_queryset(user):
    """Obtiene el queryset de empresas accesibles por el usuario"""
    company = get_user_company(user)
    if company:
        return Company.objects.filter(id=company.id)
    return Company.objects.none()

@login_required
def task_schedule_list(request):
    """Vista para listar todos los cronogramas"""
    # Obtener la empresa del usuario
    user_company = None
    if hasattr(request.user, 'profile') and request.user.profile.company:
        user_company = request.user.profile.company
    
    if user_company:
        schedules = TaskSchedule.objects.filter(
            company=user_company
        ).select_related('company', 'created_by').annotate(
            total_tasks=models.Count('tasks'),
            completed_tasks=models.Count('tasks', filter=models.Q(tasks__is_completed=True))
        ).order_by('-created_at')
    else:
        # Si no tiene empresa, mostrar solo los que creó
        schedules = TaskSchedule.objects.filter(
            created_by=request.user
        ).select_related('company', 'created_by').annotate(
            total_tasks=models.Count('tasks'),
            completed_tasks=models.Count('tasks', filter=models.Q(tasks__is_completed=True))
        ).order_by('-created_at')
    
    context = {
        'page_title': 'Planificación de Tareas',
        'schedules': schedules,
    }
    
    return render(request, 'tickets/task_schedule_list.html', context)


@login_required
def task_schedule_create(request):
    """Vista para crear un nuevo cronograma"""
    if request.method == 'POST':
        form = TaskScheduleForm(request.POST)
        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.created_by = request.user
            schedule.save()
            messages.success(request, 'Cronograma creado exitosamente. Las fechas de inicio y fin se calcularán automáticamente basándose en las tareas que agregues.')
            return redirect('task_schedule_detail', pk=schedule.pk)
    else:
        form = TaskScheduleForm()
        # Filtrar empresas del usuario
        form.fields['company'].queryset = get_user_companies_queryset(request.user)
    
    context = {
        'page_title': 'Crear Cronograma',
        'form': form,
        'auto_dates_info': True,
    }
    
    return render(request, 'tickets/task_schedule_form.html', context)


@login_required
def task_schedule_detail(request, pk):
    """Vista para ver los detalles de un cronograma"""
    schedule = get_object_or_404(
        TaskSchedule.objects.select_related('company', 'created_by'),
        pk=pk
    )
    
    # Verificar permisos
    if not user_has_company_access(request.user, schedule.company):
        messages.error(request, 'No tienes permiso para ver este cronograma.')
        return redirect('task_schedule_list')
    
    tasks = schedule.tasks.select_related('assigned_to').prefetch_related('dependencies')
    
    # Calcular KPIs
    kpis = {
        'total_tasks': schedule.get_total_tasks(),
        'completed_tasks': schedule.get_completed_tasks(),
        'pending_tasks': schedule.get_pending_tasks(),
        'overdue_tasks': schedule.get_overdue_tasks(),
        'progress_percentage': schedule.get_progress_percentage(),
        'on_time_rate': schedule.get_on_time_completion_rate(),
        'is_overdue': schedule.is_overdue(),
    }
    
    context = {
        'page_title': schedule.title,
        'schedule': schedule,
        'tasks': tasks,
        'kpis': kpis,
    }
    
    return render(request, 'tickets/task_schedule_detail.html', context)


@login_required
def task_schedule_export(request, pk):
    """Vista para exportar tareas de un cronograma a Excel"""
    schedule = get_object_or_404(
        TaskSchedule.objects.select_related('company', 'created_by'),
        pk=pk
    )
    
    # Verificar permisos
    if not user_has_company_access(request.user, schedule.company):
        messages.error(request, 'No tienes permiso para exportar este cronograma.')
        return redirect('task_schedule_list')
    
    # Importar pandas y crear DataFrame
    try:
        import pandas as pd
        from django.http import HttpResponse
        import io
        from datetime import datetime
        
        # Obtener todas las tareas del cronograma
        tasks = schedule.tasks.select_related('assigned_to').prefetch_related('dependencies').order_by('start_date')
        
        # Preparar datos para export
        data = []
        for task in tasks:
            # Obtener dependencias como texto
            dependencies_text = ', '.join([f"{dep.title}" for dep in task.dependencies.all()])
            
            # Calcular días de duración
            duration_days = 0
            if task.start_date and task.end_date:
                duration_days = (task.end_date - task.start_date).days + 1
            
            # Calcular estado de progreso
            progress_status = "A tiempo"
            if task.is_overdue():
                progress_status = "Atrasado"
            elif task.status == 'completed':
                progress_status = "Completado"
            
            data.append({
                'ID': task.id,
                'Título': task.title,
                'Descripción': task.description or '',
                'Estado': dict(task.STATUS_CHOICES).get(task.status, task.status),
                'Prioridad': dict(task.PRIORITY_CHOICES).get(task.priority, task.priority),
                'Asignado a': task.assigned_to.get_full_name() if task.assigned_to else 'Sin asignar',
                'Fecha de Inicio': task.start_date.strftime('%d/%m/%Y') if task.start_date else '',
                'Fecha de Fin': task.end_date.strftime('%d/%m/%Y') if task.end_date else '',
                'Duración (días)': duration_days,
                'Progreso (%)': task.progress,
                'Estado de Progreso': progress_status,
                'Dependencias': dependencies_text,
                'Notas': task.notes or '',
                'Fecha de Creación': task.created_at.strftime('%d/%m/%Y %H:%M'),
                'Fecha de Actualización': task.updated_at.strftime('%d/%m/%Y %H:%M'),
            })
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Crear respuesta HTTP con archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Cronograma_{schedule.title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Escribir Excel al response
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            # Hoja principal con todas las tareas
            df.to_excel(writer, sheet_name='Tareas', index=False)
            
            # Hoja de resumen del cronograma
            summary_data = {
                'Métrica': [
                    'Nombre del Cronograma',
                    'Descripción',
                    'Empresa',
                    'Creado por',
                    'Fecha de Creación',
                    'Fecha de Inicio',
                    'Fecha de Fin',
                    'Total de Tareas',
                    'Tareas Completadas',
                    'Tareas Pendientes',
                    'Tareas Atrasadas',
                    'Porcentaje de Progreso',
                    'Tasa de Cumplimiento',
                ],
                'Valor': [
                    schedule.title,
                    schedule.description or '',
                    schedule.company.name if schedule.company else '',
                    schedule.created_by.get_full_name(),
                    schedule.created_at.strftime('%d/%m/%Y %H:%M'),
                    schedule.start_date.strftime('%d/%m/%Y') if schedule.start_date else '',
                    schedule.end_date.strftime('%d/%m/%Y') if schedule.end_date else '',
                    schedule.get_total_tasks(),
                    schedule.get_completed_tasks(),
                    schedule.get_pending_tasks(),
                    schedule.get_overdue_tasks(),
                    f"{schedule.get_progress_percentage():.1f}%",
                    f"{schedule.get_on_time_completion_rate():.1f}%",
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Ajustar ancho de columnas automáticamente
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return response
        
    except ImportError:
        messages.error(request, 'La funcionalidad de exportación requiere la instalación de pandas. Contacta al administrador.')
        return redirect('task_schedule_detail', pk=pk)
    except Exception as e:
        messages.error(request, f'Error al exportar: {str(e)}')
        return redirect('task_schedule_detail', pk=pk)


@login_required
def task_schedule_edit(request, pk):
    """Vista para editar un cronograma"""
    schedule = get_object_or_404(TaskSchedule, pk=pk)
    
    # Verificar permisos
    if not user_has_company_access(request.user, schedule.company):
        messages.error(request, 'No tienes permiso para editar este cronograma.')
        return redirect('task_schedule_list')
    
    if request.method == 'POST':
        form = TaskScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cronograma actualizado exitosamente.')
            return redirect('task_schedule_detail', pk=schedule.pk)
    else:
        form = TaskScheduleForm(instance=schedule)
        form.fields['company'].queryset = get_user_companies_queryset(request.user)
    
    context = {
        'page_title': f'Editar - {schedule.title}',
        'form': form,
        'schedule': schedule,
        'auto_dates_info': True,
    }
    
    return render(request, 'tickets/task_schedule_form.html', context)


@login_required
def task_schedule_delete(request, pk):
    """Vista para eliminar un cronograma"""
    schedule = get_object_or_404(TaskSchedule, pk=pk)
    
    # Verificar permisos
    if not user_has_company_access(request.user, schedule.company):
        messages.error(request, 'No tienes permiso para eliminar este cronograma.')
        return redirect('task_schedule_list')
    
    if request.method == 'POST':
        schedule.delete()
        messages.success(request, 'Cronograma eliminado exitosamente.')
        return redirect('task_schedule_list')
    
    context = {
        'page_title': f'Eliminar - {schedule.title}',
        'schedule': schedule,
    }
    
    return render(request, 'tickets/task_schedule_delete.html', context)


@login_required
def task_schedule_duplicate(request, pk):
    """Vista para duplicar un cronograma con todas sus tareas"""
    original_schedule = get_object_or_404(TaskSchedule, pk=pk)
    
    # Verificar permisos
    if not user_has_company_access(request.user, original_schedule.company):
        messages.error(request, 'No tienes permiso para duplicar este cronograma.')
        return redirect('task_schedule_list')
    
    if request.method == 'POST':
        try:
            # Crear el cronograma duplicado
            duplicated_schedule = TaskSchedule.objects.create(
                title=f"Copia de {original_schedule.title}",
                description=original_schedule.description,
                company=original_schedule.company,
                start_date=original_schedule.start_date,
                end_date=original_schedule.end_date,
                is_public=False,  # Por seguridad, no hacer público automáticamente
                created_by=request.user,
                # No copiamos el token público - se generará automáticamente
            )
            
            # Duplicar las tareas del cronograma
            original_tasks = ScheduleTask.objects.filter(schedule=original_schedule)
            task_map = {}  # Para mapear dependencias
            
            # Primer paso: crear todas las tareas sin dependencias
            for original_task in original_tasks:
                duplicated_task = ScheduleTask.objects.create(
                    schedule=duplicated_schedule,
                    title=original_task.title,
                    description=original_task.description,
                    start_date=original_task.start_date,
                    end_date=original_task.end_date,
                    priority=original_task.priority,
                    assigned_to=original_task.assigned_to,
                    is_completed=False,  # Las tareas duplicadas empiezan como no completadas
                    progress_percentage=0,  # Progreso en 0
                )
                task_map[original_task.id] = duplicated_task
            
            # Segundo paso: establecer dependencias
            for original_task in original_tasks:
                duplicated_task = task_map[original_task.id]
                for dependency in original_task.dependencies.all():
                    if dependency.id in task_map:
                        duplicated_task.dependencies.add(task_map[dependency.id])
            
            messages.success(
                request, 
                f'Cronograma "{duplicated_schedule.title}" duplicado exitosamente con {len(task_map)} tareas.'
            )
            
            # Redirigir al nuevo cronograma duplicado
            return redirect('task_schedule_detail', pk=duplicated_schedule.pk)
            
        except Exception as e:
            messages.error(request, f'Error al duplicar el cronograma: {str(e)}')
            return redirect('task_schedule_detail', pk=pk)
    
    # Si es GET, mostrar confirmación
    context = {
        'page_title': f'Duplicar Cronograma: {original_schedule.title}',
        'schedule': original_schedule,
        'total_tasks_count': ScheduleTask.objects.filter(schedule=original_schedule).count(),
    }
    
    return render(request, 'tickets/task_schedule_duplicate_confirm.html', context)


@login_required
def task_schedule_reschedule(request, pk):
    """Vista para reprogramar todas las fechas del cronograma"""
    schedule = get_object_or_404(TaskSchedule, pk=pk)
    
    # Verificar permisos
    if not user_has_company_access(request.user, schedule.company):
        messages.error(request, 'No tienes permiso para reprogramar este cronograma.')
        return redirect('task_schedule_list')
    
    if request.method == 'POST':
        try:
            from datetime import datetime, timedelta
            
            new_start_date_str = request.POST.get('new_start_date')
            if not new_start_date_str:
                messages.error(request, 'Debes especificar una nueva fecha de inicio.')
                return redirect('task_schedule_detail', pk=pk)
            
            # Convertir la nueva fecha de inicio
            new_start_date = datetime.strptime(new_start_date_str, '%Y-%m-%d').date()
            
            # Calcular la diferencia de días entre la fecha actual y la nueva
            current_start_date = schedule.start_date
            date_difference = (new_start_date - current_start_date).days
            
            if date_difference == 0:
                messages.info(request, 'La nueva fecha de inicio es la misma que la actual.')
                return redirect('task_schedule_detail', pk=pk)
            
            # Reprogramar todas las tareas
            tasks_updated = 0
            for task in schedule.tasks.all():
                # Calcular las nuevas fechas sumando/restando la diferencia
                task.start_date = task.start_date + timedelta(days=date_difference)
                task.end_date = task.end_date + timedelta(days=date_difference)
                task.save()
                tasks_updated += 1
            
            # Actualizar las fechas del cronograma
            schedule.start_date = new_start_date
            # Calcular la nueva fecha de fin basándose en las tareas actualizadas
            if schedule.tasks.exists():
                schedule.end_date = schedule.tasks.order_by('-end_date').first().end_date
            schedule.save()
            
            messages.success(
                request, 
                f'Cronograma reprogramado exitosamente. Se han actualizado {tasks_updated} tareas. '
                f'{"Adelantado" if date_difference > 0 else "Retrasado"} {abs(date_difference)} días.'
            )
            
        except ValueError:
            messages.error(request, 'Formato de fecha inválido.')
        except Exception as e:
            messages.error(request, f'Error al reprogramar el cronograma: {str(e)}')
    
    return redirect('task_schedule_detail', pk=pk)


# Vista Gantt deshabilitada - funcionalidad removida temporalmente
# @login_required
# def task_schedule_gantt(request, pk):
#     """Vista Gantt del cronograma"""
#     schedule = get_object_or_404(
#         TaskSchedule.objects.select_related('company', 'created_by'),
#         pk=pk
#     )
#     
#     # Verificar permisos
#     if not user_has_company_access(request.user, schedule.company):
#         messages.error(request, 'No tienes permiso para ver este cronograma.')
#         return redirect('task_schedule_list')
#     
#     tasks = schedule.tasks.select_related('assigned_to').prefetch_related('dependencies')
#     
#     # Preparar datos para el Gantt
#     gantt_data = []
#     for task in tasks:
#         gantt_data.append({
#             'id': task.id,
#             'title': task.title,
#             'start': task.start_date.isoformat(),
#             'end': task.end_date.isoformat(),
#             'progress': task.progress_percentage,
#             'is_completed': task.is_completed,
#             'priority': task.priority,
#             'assigned_to': task.assigned_to.get_full_name() if task.assigned_to else 'Sin asignar',
#             'dependencies': [dep.id for dep in task.dependencies.all()],
#             'is_overdue': task.is_overdue(),
#         })
#     
#     # Calcular KPIs
#     kpis = {
#         'total_tasks': schedule.get_total_tasks(),
#         'completed_tasks': schedule.get_completed_tasks(),
#         'pending_tasks': schedule.get_pending_tasks(),
#         'overdue_tasks': schedule.get_overdue_tasks(),
#         'progress_percentage': schedule.get_progress_percentage(),
#         'on_time_rate': schedule.get_on_time_completion_rate(),
#     }
#     
#     context = {
#         'page_title': f'Vista Gantt - {schedule.title}',
#         'schedule': schedule,
#         'gantt_data': json.dumps(gantt_data),
#         'kpis': kpis,
#     }
#     
#     return render(request, 'tickets/task_schedule_gantt.html', context)


# Task CRUD
@login_required
def schedule_task_create(request, schedule_pk):
    """Vista para crear una nueva tarea"""
    schedule = get_object_or_404(TaskSchedule, pk=schedule_pk)
    
    # Verificar permisos
    if not user_has_company_access(request.user, schedule.company):
        messages.error(request, 'No tienes permiso para agregar tareas a este cronograma.')
        return redirect('task_schedule_list')
    
    if request.method == 'POST':
        form = ScheduleTaskForm(request.POST, schedule=schedule)
        if form.is_valid():
            task = form.save(commit=False)
            task.schedule = schedule
            task.save()
            form.save_m2m()  # Guardar relaciones many-to-many
            messages.success(request, 'Tarea creada exitosamente.')
            return redirect('task_schedule_detail', pk=schedule.pk)
    else:
        # Establecer fechas por defecto del cronograma
        initial = {
            'start_date': schedule.start_date,
            'end_date': schedule.end_date,
        }
        form = ScheduleTaskForm(initial=initial, schedule=schedule)
    
    context = {
        'page_title': f'Nueva Tarea - {schedule.title}',
        'form': form,
        'schedule': schedule,
    }
    
    return render(request, 'tickets/schedule_task_form.html', context)


@login_required
def schedule_task_edit(request, pk):
    """Vista para editar una tarea"""
    task = get_object_or_404(ScheduleTask.objects.select_related('schedule'), pk=pk)
    schedule = task.schedule
    
    # Verificar permisos
    if not user_has_company_access(request.user, schedule.company):
        messages.error(request, 'No tienes permiso para editar esta tarea.')
        return redirect('task_schedule_list')
    
    if request.method == 'POST':
        form = ScheduleTaskForm(request.POST, instance=task, schedule=schedule)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tarea actualizada exitosamente.')
            return redirect('task_schedule_detail', pk=schedule.pk)
    else:
        form = ScheduleTaskForm(instance=task, schedule=schedule)
    
    context = {
        'page_title': f'Editar Tarea - {task.title}',
        'form': form,
        'schedule': schedule,
        'task': task,
    }
    
    return render(request, 'tickets/schedule_task_form.html', context)


@login_required
def schedule_task_delete(request, pk):
    """Vista para eliminar una tarea"""
    task = get_object_or_404(ScheduleTask.objects.select_related('schedule'), pk=pk)
    schedule = task.schedule
    
    # Verificar permisos
    if not user_has_company_access(request.user, schedule.company):
        messages.error(request, 'No tienes permiso para eliminar esta tarea.')
        return redirect('task_schedule_list')
    
    if request.method == 'POST':
        task.delete()
        messages.success(request, 'Tarea eliminada exitosamente.')
        return redirect('task_schedule_detail', pk=schedule.pk)
    
    context = {
        'page_title': f'Eliminar Tarea - {task.title}',
        'task': task,
        'schedule': schedule,
    }
    
    return render(request, 'tickets/schedule_task_delete.html', context)


@login_required
def schedule_task_toggle(request, pk):
    """Vista para marcar/desmarcar una tarea como completada (AJAX)"""
    if request.method == 'POST':
        task = get_object_or_404(ScheduleTask, pk=pk)
        
        # Verificar permisos
        if not user_has_company_access(request.user, task.schedule.company):
            return JsonResponse({'success': False, 'error': 'No tienes permiso.'}, status=403)
        
        task.is_completed = not task.is_completed
        # Establecer progreso automáticamente según el estado
        if task.is_completed:
            task.progress_percentage = 100
        else:
            task.progress_percentage = 0
        task.save()
        
        return JsonResponse({
            'success': True,
            'is_completed': task.is_completed,
            'completed_at': task.completed_at.strftime('%d/%m/%Y %H:%M') if task.completed_at else None,
            'progress': task.progress_percentage,
        })
    
    return JsonResponse({'success': False}, status=400)


# Vista pública del cronograma
def task_schedule_public(request, token):
    """Vista pública del cronograma con token"""
    schedule = get_object_or_404(
        TaskSchedule.objects.select_related('company', 'created_by'),
        public_token=token,
        is_public=True
    )
    
    tasks = schedule.tasks.select_related('assigned_to').prefetch_related('dependencies')
    
    # Calcular total de días
    total_days = sum(task.get_duration_days() for task in tasks)
    
    # Calcular KPIs
    kpis = {
        'total_tasks': schedule.get_total_tasks(),
        'completed_tasks': schedule.get_completed_tasks(),
        'pending_tasks': schedule.get_pending_tasks(),
        'overdue_tasks': schedule.get_overdue_tasks(),
        'progress_percentage': schedule.get_progress_percentage(),
        'on_time_rate': schedule.get_on_time_completion_rate(),
    }
    
    context = {
        'page_title': schedule.title,
        'schedule': schedule,
        'tasks': tasks,
        'kpis': kpis,
        'total_days': total_days,
        'is_public_view': True,
    }
    
    return render(request, 'tickets/task_schedule_public.html', context)


def task_schedule_public_gantt(request, token):
    """Vista Gantt pública del cronograma"""
    schedule = get_object_or_404(
        TaskSchedule.objects.select_related('company', 'created_by'),
        public_token=token,
        is_public=True
    )
    
    tasks = schedule.tasks.select_related('assigned_to').prefetch_related('dependencies')
    
    # Preparar datos para el Gantt
    gantt_data = []
    for task in tasks:
        gantt_data.append({
            'id': task.id,
            'title': task.title,
            'start': task.start_date.isoformat(),
            'end': task.end_date.isoformat(),
            'progress': task.progress_percentage,
            'is_completed': task.is_completed,
            'priority': task.priority,
            'assigned_to': task.assigned_to.get_full_name() if task.assigned_to else 'Sin asignar',
            'dependencies': [dep.id for dep in task.dependencies.all()],
            'is_overdue': task.is_overdue(),
        })
    
    # Calcular KPIs
    kpis = {
        'total_tasks': schedule.get_total_tasks(),
        'completed_tasks': schedule.get_completed_tasks(),
        'pending_tasks': schedule.get_pending_tasks(),
        'overdue_tasks': schedule.get_overdue_tasks(),
        'progress_percentage': schedule.get_progress_percentage(),
        'on_time_rate': schedule.get_on_time_completion_rate(),
    }
    
    context = {
        'page_title': f'Vista Gantt - {schedule.title}',
        'schedule': schedule,
        'gantt_data': json.dumps(gantt_data),
        'kpis': kpis,
        'is_public_view': True,
    }
    
    return render(request, 'tickets/task_schedule_public_gantt.html', context)


def schedule_task_toggle_public(request, pk):
    """Vista pública para marcar/desmarcar una tarea como completada (AJAX)"""
    if request.method == 'POST':
        task = get_object_or_404(ScheduleTask, pk=pk)
        
        # Verificar que la tarea pertenece a un cronograma público
        if not task.schedule.is_public:
            return JsonResponse({'success': False, 'error': 'Esta tarea no es pública.'}, status=403)
        
        # Verificar el token si se proporciona
        import json
        try:
            data = json.loads(request.body)
            token = data.get('token')
            if token and str(task.schedule.public_token) != str(token):
                return JsonResponse({'success': False, 'error': 'Token inválido.'}, status=403)
        except json.JSONDecodeError:
            pass
        
        task.is_completed = not task.is_completed
        # Establecer progreso automáticamente según el estado
        if task.is_completed:
            task.progress_percentage = 100
        else:
            task.progress_percentage = 0
        task.save()
        
        return JsonResponse({
            'success': True,
            'is_completed': task.is_completed,
            'completed_at': task.completed_at.strftime('%d/%m/%Y %H:%M') if task.completed_at else None,
            'progress': task.progress_percentage,
        })
    
    return JsonResponse({'success': False}, status=400)

